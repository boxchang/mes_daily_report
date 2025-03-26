import sys
import os
from PIL import Image as PILImage
from matplotlib.ticker import MultipleLocator, FuncFormatter
from openpyxl.utils import get_column_letter

from database import vnedc_database, mes_database, mes_olap_database
from lib.utils import Utils

curPath = os.path.abspath(os.path.dirname(__file__))
rootPath = os.path.split(curPath)[0]
sys.path.append(rootPath)
import pandas as pd
import matplotlib.pyplot as plt
import logging
from datetime import datetime, timedelta
import math
from matplotlib import font_manager
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, NamedStyle, Font, Border, Side, PatternFill
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from email.mime.image import MIMEImage
from openpyxl.drawing.image import Image
from dateutil.relativedelta import relativedelta
import matplotlib.gridspec as gridspec
import numpy as np

class mes_weekly_report(object):
    this_start_date = ""
    this_end_date = ""
    last_end_date = ""
    last_start_date = ""
    db = None
    plant_name = ['NBR', 'PVC']
    mode = ""
    save_path = ""
    date_mark = ""
    file_list = []
    image_buffers = []
    mach_list = ""
    year = None
    month_week = None

    # Define Style
    percent_style = NamedStyle(name='percent_style', number_format='0.00%')
    right_align_style = NamedStyle(name='right_align_style', alignment=Alignment(horizontal='right'))
    center_align_style = NamedStyle(name='center_align_style', alignment=Alignment(horizontal='center'))

    # Define Header
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal='center')
    header_border = Border(bottom=Side(style='thin'))
    header_columns = {
        'WorkDate': '作業日期',
        'belong_to': '作業日期',
        'Name': '機台號',
        'WorkOrderId': '工單',
        'PartNo': '料號',
        'ProductItem': '品項',
        'Line': '線別',
        'Shift': '班別',
        'max_speed': '車速(最高)',
        'min_speed': '車速(最低)',
        'avg_speed': '車速(平均)',
        'sum_qty': '產量(加總)',
        'Ticket_Qty': '入庫量(加總)',
        'ProductionTime': '生產時間',
        'LineSpeedStd': '標準車速',
        'Target': '目標產能',
        'Separate': '隔離',
        'Scrap': '廢品',
        'SecondGrade': '二級品',
        'OverControl': '超內控',
        'WeightValue': 'IPQC克重',
        'WeightLower': '重量下限',
        'WeightUpper': '重量上限',
        'Activation': '稼動率',
        'OpticalNGRate': '光檢不良率',
        'Achievement Rate': '目標達成率',
        'Date': '日期範圍',
        'CountingQty': '點數機數量',
        'FaultyQuantity': '二級品數量',
        'ScrapQuantity': '廢品數量',
        'SeparateQuantity': '隔離品數量',
        'SeparateQty': '隔離品數量',
        'OnlinePacking': '包裝確認量',
        'WIPPacking': '半成品數量',
        'RunTime': '實際運轉時間',
        'StopTime': '停機時間',
        'AllTime': '可運轉時間',
        'StandSpeed': '標準車速',
        'StdSpeed': '標準車速',
        'Capacity': '產能效率',
        'Yield': '良率',
        'SeparateRate': '隔離率',
        'ScrapRate': '報廢率',
        'Plant': '廠別',
        'Year': '年',
        'WorkOrder': '工單',
        'Machine': '機台',
        'MonthWeek': '週別',
        'MaxSpeed': '最大車速',
        'MinSpeed': '最小車速',
        'AvgSpeed': '平均車速',
        'Tensile_Value': '抗拉強度值',
        'Tensile_Limit': '抗拉強度上下限',
        'Tensile_Status': '抗拉強度結果',
        'Elongation_Value': '伸長率值',
        'Elongation_Limit': '伸長率上下限',
        'Elongation_Status': '伸長率結果',
        'Roll_Value': '卷唇厚度值',
        'Roll_Limit': '卷唇厚度上下限',
        'Roll_Status': '卷唇厚度結果',
        'Cuff_Value': '袖厚度值',
        'Cuff_Limit': '袖厚度上下限',
        'Cuff_Status': '袖厚度結果',
        'Palm_Value': '掌厚度值',
        'Palm_Limit': '掌厚度上下限',
        'Palm_Status': '掌厚度結果',
        'Finger_Value': '指厚度值',
        'Finger_Limit': '指厚度上下限',
        'Finger_Status': '指厚度結果',
        'FingerTip_Value': '指尖厚度值',
        'FingerTip_Limit': '指尖厚度上下限',
        'FingerTip_Status': '指尖厚度結果',
        'Length_Value': '長度值',
        'Length_Limit': '長度上下限',
        'Length_Status': '長度結果',
        'Width_Value': '寬度值',
        'Width_Limit': '寬度上下限',
        'Width_Status': '寬度結果',
        'Weight_Value': '重量值',
        'Weight_Limit': '重量上下限',
        'Weight_Status': '重量結果',
        'Weight_Light': '超輕檢驗',
        'Weight_Heavy': '超重檢驗',
        'Pinhole_Value': '針孔值',
        'Pinhole_Limit': '針孔上下限',
        'Pinhole_Status': '針孔結果',
        'StandardAQL': '工單AQL',
        'InspectedAQL': '量測AQL',
        'Length_NG_Count': '長度隔離',
        'Width_NG_Count': '寬度隔離',
        'Weight_Light_NG_Count': '過輕隔離',
        'Weight_Heavy_NG_Count': '過重隔離',
        'Pinhole_NG_Count': '針孔隔離',
        'Tensile_NG_Count': '拉力隔離',
        'Elongation_NG_Count': '伸長率隔離',
        'Thickness_NG_Count': '厚度隔離',
        'us_ng_heavy_rate': '美線重量重',
        'us_ng_light_rate': '美線重量輕',
        'eu_ng_heavy_rate': '歐線重量重',
        'eu_ng_light_rate': '歐線重量輕',
        'jp_ng_heavy_rate': '日線重量重',
        'jp_ng_light_rate': '日線重量輕',
        'Gap': '差異',
        'SalePlaceCode': '銷售地點',
        'runcard': 'Runcard',
        'Pinhole': '針孔數量',
        'Pinhole_Sample': '針孔檢查數量'

    }

    # 配置日志记录器
    logging.basicConfig(
        level=logging.INFO,  # 设置日志级别为 DEBUG，这样所有级别的日志都会被记录
        format='%(asctime)s - %(levelname)s - %(message)s',  # 指定日志格式
        filename='weekly.log',  # 指定日志文件
        filemode='w'  # 写入模式，'w' 表示每次运行程序时会覆盖日志文件
    )

    def get_week_date_df_fix(self):
        vnedc_db = vnedc_database()

        sql = f"""
            SELECT *
            FROM [MES_OLAP].[dbo].[week_date] 
            where  [year] = 2025 and [month] =3 and month_week = 'W1'
             and enable = 1
        """

        print(sql)
        raws = vnedc_db.select_sql_dict(sql)

        return raws[0]

    def __init__(self, mode):
        self.db = vnedc_database()
        self.mes_db = mes_database()
        today = datetime.now().date()
        self.mode = mode

        week_date = Utils().get_week_date_dist()
        #week_date = self.get_week_date_df_fix()
        self.this_end_date = week_date['end_date']
        self.this_start_date = week_date['start_date']
        self.year = week_date['year']
        self.month_week = str(week_date['month']) + week_date['month_week']
        fold_name = str(week_date['year']) + str(week_date['month']).zfill(2) + week_date['month_week']
        save_path = os.path.join("weekly_output", fold_name)

        self.save_path = save_path
        # Check folder to create
        if not os.path.exists(save_path):
            os.makedirs(save_path)

        tmp_start_date = datetime.strptime(self.this_start_date, "%Y-%m-%d").strftime("%m%d")
        tmp_end_date = datetime.strptime(self.this_end_date, "%Y-%m-%d").strftime("%m%d")
        self.date_mark = "{start_date}_{end_date}".format(start_date=tmp_start_date,
                                                          end_date=tmp_end_date)


    def generate_excel(self, writer, df, machine_name):
        column_letter = {'belong_to': 'A', 'Machine': 'B', 'Line': 'C', 'Shift': 'D', 'WorkOrder': 'E',
                 'PartNo': 'F', 'ProductItem': 'G', 'StandardAQL': 'H', 'InspectedAQL': 'I', 'Period': 'J',
                 'MaxSpeed': 'K', 'MinSpeed': 'L', 'AvgSpeed': 'M', 'StdSpeed': 'N',
                 'CountingQty': 'O', 'OnlinePacking': 'P', 'WIPPacking': 'Q', 'Target': 'R', 'ScrapQuantity': 'S', 'FaultyQuantity': 'T',
                 'RunTime': 'U', 'StopTime': 'V', 'AllTime': 'W', 'MonthWeek': 'X',
                 'Tensile_Value': 'Y', 'Tensile_Limit': 'Z', 'Tensile_Status': 'AA',
                 'Elongation_Value': 'AB', 'Elongation_Limit': 'AC', 'Elongation_Status': 'AD',
                 'Roll_Value': 'AE', 'Roll_Limit': 'AF', 'Roll_Status': 'AG',
                 'Cuff_Value': 'AH', 'Cuff_Limit': 'AI', 'Cuff_Status': 'AJ',
                 'Palm_Value': 'AK', 'Palm_Limit': 'AL', 'Palm_Status': 'AM',
                 'Finger_Value': 'AN', 'Finger_Limit': 'AO', 'Finger_Status': 'AP',
                 'FingerTip_Value': 'AQ', 'FingerTip_Limit': 'AR', 'FingerTip_Status': 'AS',
                 'Length_Value': 'AT', 'Length_Limit': 'AU', 'Length_Status': 'AV',
                 'Weight_Value': 'AW', 'Weight_Limit': 'AX', 'Weight_Light': 'AY', 'Weight_Heavy': 'AZ',
                 'Width_Value': 'BA', 'Width_Limit': 'BB', 'Width_Status': 'BC',
                 'Pinhole_Value': 'BD', 'Pinhole_Limit': 'BE', 'Pinhole_Status': 'BF', 'IPQC': 'BG',
                 'SeparateQty': 'BH'
                }

        df['Period'] = df['Period'].apply(lambda x: f"{int(x):02}:00")
        df["Weight_Value"] = df["Weight_Value"].apply(lambda x: f"{x:.2f}")
        df.loc[df['Weight_Value'].isna() | (df['Weight_Value'] == 0), 'Weight_Value'] = ''

        # Rename columns
        df.rename(columns=self.header_columns, inplace=True)
        namesheet = str(machine_name).split('_')[-1]
        save_path = self.save_path

        file_name = f"MES_{machine_name}_Chart.png"
        chart_img = os.path.join(save_path, file_name)
        # if os.path.exists(chart_img):
        #     header_row = 31
        #     data_start_row = 32
        # else:
        #     header_row = 0
        #     data_start_row = 1
        header_row = 0
        data_start_row = 1

        df_copy = df.copy()
        df_copy = df_copy.fillna("")

        # Write data to the Excel sheet
        df_copy.to_excel(writer, sheet_name=namesheet, index=False, startrow=header_row)

        workbook = writer.book
        worksheet = writer.sheets.get(namesheet)

        # Freeze the first row
        # worksheet.freeze_panes = worksheet['A'+str(data_start_row+1)]
        worksheet.freeze_panes = worksheet['A2']

        if not worksheet:
            worksheet = workbook.add_worksheet(namesheet)

        # Apply Header Style
        for cell in worksheet[data_start_row]:
            cell.font = self.header_font
            cell.alignment = self.header_alignment
            cell.border = self.header_border

        # Adjust column formatting
        for col in worksheet.columns:
            max_length = max(len(str(cell.value)) for cell in col)
            col_letter = col[0].column_letter

            worksheet.column_dimensions[col_letter].width = max_length + 5
            for cell in col:
                if col_letter in [column_letter['MaxSpeed'], column_letter['MinSpeed'], column_letter['AvgSpeed'],
                                  column_letter['StdSpeed']]:
                    cell.alignment = self.right_align_style.alignment
                elif col_letter in [column_letter['MaxSpeed'], column_letter['MinSpeed'], column_letter['AvgSpeed'],
                                  column_letter['StdSpeed']]:
                    cell.alignment = self.center_align_style.alignment
                elif col_letter in [column_letter['CountingQty'], column_letter['OnlinePacking'], column_letter['Target'], column_letter['SeparateQty']]:
                    cell.number_format = '#,##0'
                    cell.alignment = self.right_align_style.alignment

        # # 設置欄的 outlineLevel 讓其可以折疊/展開
        hide_columns = ['Tensile_Value','Tensile_Limit','Tensile_Status','Elongation_Value','Elongation_Limit','Elongation_Status',
                        'Roll_Value','Roll_Limit','Roll_Status','Cuff_Value','Cuff_Limit','Cuff_Status','Palm_Value','Palm_Limit','Palm_Status',
                        'Finger_Value','Finger_Limit','Finger_Status','FingerTip_Value','FingerTip_Limit','FingerTip_Status',
                        'Length_Value','Length_Limit','Length_Status', 'Weight_Value', 'Weight_Limit', 'Weight_Light', 'Weight_Heavy',
                        'Width_Value', 'Width_Limit','Width_Status',
                        'Pinhole_Value','Pinhole_Limit','Pinhole_Status']
        for column in hide_columns:
            worksheet.column_dimensions[column_letter[column]].outlineLevel = 1
        worksheet.column_dimensions.group(column_letter['Tensile_Value'], column_letter['Pinhole_Status'], hidden=True)

        hide_columns = ['Roll_Value', 'Roll_Limit', 'Roll_Status', 'Cuff_Value', 'Cuff_Limit', 'Cuff_Status',
                        'Palm_Value', 'Palm_Limit', 'Palm_Status',
                        'Finger_Value', 'Finger_Limit', 'Finger_Status', 'FingerTip_Value', 'FingerTip_Limit',
                        'FingerTip_Status']
        for column in hide_columns:
            worksheet.column_dimensions[column_letter[column]].outlineLevel = 2
        worksheet.column_dimensions.group(column_letter['Roll_Value'], column_letter['FingerTip_Status'], hidden=True)


        try:
            img = Image(chart_img)
            img.height = 6 * 96
            img.width = 16 * 96
            img.anchor = 'A' + str(len(df) + 5)
            worksheet.add_image(img)
        except:
            print('No counting machine data yet!')
            pass

        return workbook

    def generate_raw_excel(self, plant):
        save_path = self.save_path
        date_mark = self.date_mark
        mode = self.mode
        month_week = self.month_week
        year = self.year

        file_name = f'MES_{plant}_{mode}_Report_{date_mark}.xlsx'
        excel_file = os.path.join(save_path, file_name)
        with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
            sql = f"""SELECT WorkDate, Machine, Line, Shift, WorkOrder, PartNo, ProductItem, StandardAQL, InspectedAQL,
                        Period, MaxSpeed, MinSpeed, AvgSpeed, StdSpeed, CountingQty, OnlinePacking, WIPPacking, Target, ScrapQuantity, FaultyQuantity, RunTime, StopTime, 60 as AllTime, c.MonthWeek,
                        Tensile_Value,Tensile_Limit,Tensile_Status,Elongation_Value,Elongation_Limit,Elongation_Status,
                        Roll_Value,Roll_Limit,Roll_Status,Cuff_Value,Cuff_Limit,Cuff_Status,Palm_Value,Palm_Limit,Palm_Status,
                        Finger_Value,Finger_Limit,Finger_Status,FingerTip_Value,FingerTip_Limit,FingerTip_Status,
                        Length_Value,Length_Limit,Length_Status, Weight_Value, Weight_Limit, 
                        CASE WHEN Weight_Defect IS NULL THEN NULL WHEN Weight_Defect = 'LL2' THEN 'NG' ELSE 'PASS' END AS Weight_Light, 
                        CASE WHEN Weight_Defect IS NULL THEN NULL WHEN Weight_Defect = 'LL1' THEN 'NG' ELSE 'PASS' END AS Weight_Heavy,
                        Width_Value,Width_Limit,Width_Status,
                        Pinhole_Value,Pinhole_Limit,Pinhole_Status
                        FROM [MES_OLAP].[dbo].[counting_daily_info_raw] c
                        LEFT JOIN [MES_OLAP].[dbo].[mes_ipqc_data] ipqc on c.Runcard = ipqc.Runcard
                        where c.Year = {year} and c.MonthWeek = '{month_week}'  
                        and c.branch like'%{plant}%'
                        and not (WorkOrder != '' and InspectedAQL ='')
                        --and StandardAQL is not Null and InspectedAQL is not null 
                        order by Machine, WorkDate, Cast(Period as Int), Line
                        """

            data = self.db.select_sql_dict(sql)
            df = pd.DataFrame(data)

            # 設定IPQC欄位判斷條件
            df['IPQC'] = df.apply(lambda row: "" if pd.isna(row['WorkOrder']) or row['WorkOrder'] == ""
            else ('NG' if 'NG' in row[['Tensile_Status', 'Elongation_Status', 'Roll_Status', 'Cuff_Status',
                                       'Palm_Status', 'Finger_Status', 'FingerTip_Status', 'Length_Status',
                                       'Weight_Heavy', 'Weight_Light', 'Width_Status', 'Pinhole_Status']].values
                  else 'PASS'), axis=1)
            try:
                df['SeparateQty'] = df.apply(
                    lambda row: row['OnlinePacking'] + row['WIPPacking'] if row['IPQC'] == 'NG'
                        or (row['StandardAQL'] is not None
                            and row['InspectedAQL'] is not None
                            and row['StandardAQL'] < row['InspectedAQL']) else None, axis=1)

                df.loc[df['Weight_Value'].isna() | (df['Weight_Value'] == 0), 'Target'] = 0
                df.loc[df['Weight_Value'].isna() | (df['Weight_Value'] == 0), 'CountingQty'] = 0
            except Exception as e:
                print(e)


            machine_groups = df.groupby('Machine')
            summary_df = self.generate_summary(writer, machine_groups)

            mach_sum_list = []
            for machine_name, machine_df in machine_groups:
                self.generate_excel(writer, machine_df, machine_name)
                mach_sum_df = self.ipqc_ng_data(machine_df)
                mach_sum_list.append(mach_sum_df)
            all_mach_sum_df = pd.concat(mach_sum_list)
            self.generate_ipqc_ng_data_excel(writer, all_mach_sum_df)

            self.delete_counting_weekly_info_raw(plant, year, month_week)
            self.insert_counting_weekly_info_raw(plant, year, month_week, summary_df)

            self.generate_12aspect_output_excel(writer, plant)
            self.generate_12aspect_cosmetic_excel(writer, plant, year, month_week)
            self.generate_12aspect_cosmetic_summary_excel(writer, plant, year, month_week)

        self.file_list.append(excel_file)

    def generate_ipqc_ng_data_excel(self, writer, df):
        colmn_letter = {'Machine': 'A', 'Line': 'B', 'ProductItem': 'C',
                        'OnlinePacking': 'D', 'WIPPacking': 'E', 'ScrapQty': 'F', 'FaultyQty': 'G',
                        'Length_NG_Count': 'H', 'Width_NG_Count': 'I',
                        'Weight_Light_NG_Count': 'J', 'Weight_Heavy_NG_Count': 'K', 'Pinhole_NG_Count': 'L',
                        'Tensile_NG_Count': 'M', 'Elongation_NG_Count': 'N', 'Thickness_NG_Count': 'O',
                        }

        sheet_name = "隔離數量"

        df.rename(columns=self.header_columns, inplace=True)

        df.to_excel(writer, sheet_name=sheet_name, index=False)

        worksheet = writer.sheets[sheet_name]

        for col in worksheet.columns:
            max_length = max(len(str(cell.value)) for cell in col)
            col_letter = col[0].column_letter

            worksheet.column_dimensions[col_letter].width = max_length + 5

            # Set alignment
            for cell in col:
                if col_letter in [colmn_letter['OnlinePacking'], colmn_letter['WIPPacking'],
                                  colmn_letter['ScrapQty'], colmn_letter['FaultyQty'],
                                  colmn_letter['Length_NG_Count'], colmn_letter['Width_NG_Count'],
                                  colmn_letter['Weight_Light_NG_Count'], colmn_letter['Weight_Heavy_NG_Count'],
                                  colmn_letter['Pinhole_NG_Count'], colmn_letter['Tensile_NG_Count'],
                                  colmn_letter['Elongation_NG_Count'], colmn_letter['Thickness_NG_Count']]:
                    cell.number_format = '#,##0'
                    cell.alignment = self.right_align_style.alignment
                else:
                    cell.alignment = self.center_align_style.alignment

    def ipqc_ng_data(self, df):
        # Only check IPQC data
        df['重量值'] = pd.to_numeric(df['重量值'], errors='coerce')
        df = df[df['重量值'] > 0]

        # 計算 "NG" 的次數
        df["Length_NG_Count"] = (df["包裝確認量"]+df["半成品數量"]).where(df["長度結果"] == "NG", 0)
        df["Width_NG_Count"] = (df["包裝確認量"]+df["半成品數量"]).where(df["寬度結果"] == "NG", 0)
        df["Weight_Light_NG_Count"] = (df["包裝確認量"]+df["半成品數量"]).where(df["超輕檢驗"] == "NG", 0)
        df["Weight_Heavy_NG_Count"] = (df["包裝確認量"] + df["半成品數量"]).where(df["超重檢驗"] == "NG", 0)
        df["Pinhole_NG_Count"] = (df["包裝確認量"]+df["半成品數量"]).where(df["針孔結果"] == "NG", 0)
        df["Tensile_NG_Count"] = (df["包裝確認量"]+df["半成品數量"]).where(df["抗拉強度結果"] == "NG", 0)
        df["Elongation_NG_Count"] = (df["包裝確認量"]+df["半成品數量"]).where(df["伸長率結果"] == "NG", 0)
        df["Roll_NG_Count"] = (df["包裝確認量"]+df["半成品數量"]).where(df["卷唇厚度結果"] == "NG", 0)
        df["Cuff_NG_Count"] = (df["包裝確認量"]+df["半成品數量"]).where(df["袖厚度結果"] == "NG", 0)
        df["Palm_NG_Count"] = (df["包裝確認量"]+df["半成品數量"]).where(df["掌厚度結果"] == "NG", 0)
        df["Finger_NG_Count"] = (df["包裝確認量"]+df["半成品數量"]).where(df["指厚度結果"] == "NG", 0)
        df["FingerTip_NG_Count"] = (df["包裝確認量"]+df["半成品數量"]).where(df["指尖厚度結果"] == "NG", 0)
        df["Thickness_NG_Count"] = df["Roll_NG_Count"] + df["Cuff_NG_Count"] + df["Palm_NG_Count"] + df["Finger_NG_Count"] + df["FingerTip_NG_Count"]

        # 計算總 NG 數量
        df["Total_NG"] = df["Tensile_NG_Count"] + df["Elongation_NG_Count"] + df["Roll_NG_Count"] \
                         + df["Cuff_NG_Count"] + df["Palm_NG_Count"] + df["Finger_NG_Count"] + df["FingerTip_NG_Count"] \
                         + df["Length_NG_Count"] + df["Width_NG_Count"] \
                         + df["Weight_Light_NG_Count"] + df["Weight_Heavy_NG_Count"] + df["Pinhole_NG_Count"]

        sum_list = ["包裝確認量", "半成品數量", "廢品數量", "二級品數量", "Length_NG_Count", "Width_NG_Count", "Weight_Light_NG_Count", "Weight_Heavy_NG_Count",
                    "Pinhole_NG_Count", "Tensile_NG_Count", "Elongation_NG_Count", "Thickness_NG_Count"]
        df_grouped = df.groupby(["機台", "線別", "品項"], as_index=False)[sum_list].sum()

        return df_grouped

    # 外觀週累計
    def generate_12aspect_cosmetic_summary_excel(self, writer, plant, year, month_week):
        column_letter = {'Plant': 'A', 'Year': 'B', 'MonthWeek': 'C',
                         'total_6100': 'D', 'LL1_6100': 'E', 'LL1_6100_rate': 'F', 'LL2_6100': 'G', 'LL2_6100_rate': 'H',
                         'total_6200': 'I', 'LL1_6200': 'J', 'LL1_6200_rate': 'K', 'LL2_6200': 'L', 'LL2_6200_rate': 'M',
                         'total_6300': 'N', 'LL1_6300': 'O', 'LL1_6300_rate': 'P', 'LL2_6300': 'Q', 'LL2_6300_rate': 'R',
                         'total_7000': 'S', 'LL1_7000': 'T', 'LL1_7000_rate': 'U', 'LL2_7000': 'V', 'LL2_7000_rate': 'W',
                         'critical_qty': 'X', 'critical_rate': 'Y', 'critical_dpm': 'Z',
                         'major_qty': 'AA', 'major_rate': 'AB', 'major_dpm': 'AC',
                         'minor_qty': 'AD', 'minor_rate': 'AE', 'minor_dpm': 'AF',
                         'pinhole_qty': 'AG', 'pinhole_rate': 'AH', 'pinhole_dpm': 'AI', 'cosmetic_check_qty': 'AJ',
                         }
        header_column = {'Plant': '廠別', 'Year': '年', 'MonthWeek': '週別',
                         'total_6100': '美線總時數', 'LL1_6100': '美線超重時數', 'LL1_6100_rate': '美線超重比例', 'LL2_6100': '美線超輕時數', 'LL2_6100_rate': '美線超輕比例',
                         'total_6200': '歐線總時數', 'LL1_6200': '歐線超重時數', 'LL1_6200_rate': '歐線超重比例', 'LL2_6200': '歐線超輕時數', 'LL2_6200_rate': '歐線超輕比例',
                         'total_6300': '日線總時數', 'LL1_6300': '日線超重時數', 'LL1_6300_rate': '日線超重比例', 'LL2_6300': '日線超輕時數', 'LL2_6300_rate': '日線超輕比例',
                         'total_7000': 'OBM總時數', 'LL1_7000': 'OBM超重時數', 'LL1_7000_rate': 'OBM超重比例', 'LL2_7000': 'OBM超輕時數', 'LL2_7000_rate': 'OBM超輕比例',
                         'critical_qty': 'Critical數量', 'critical_rate': 'Critical比例', 'critical_dpm': 'Critical DPM',
                         'major_qty': 'Major數量', 'major_rate': 'Major比例', 'major_dpm': 'Major DPM',
                         'minor_qty': 'Minor數量', 'minor_rate': 'Minor比例', 'minor_dpm': 'Minor DPM',
                         'pinhole_qty': '針孔數量', 'pinhole_rate': '針孔比例', 'pinhole_dpm': '針孔DPM', 'cosmetic_check_qty': '外觀檢查總數量',
                         }

        # 外觀週累計
        cosmetic_summary_sql = f"""
                    SELECT [Plant],r.Year,r.MonthWeek
                          ,[total_6100],[LL1_6100],[LL1_6100_rate],[LL2_6100],[LL2_6100_rate]
                          ,[total_6200],[LL1_6200],[LL1_6200_rate],[LL2_6200],[LL2_6200_rate]
                          ,[total_6300],[LL1_6300],[LL1_6300_rate],[LL2_6300],[LL2_6300_rate]
                          ,[total_7000],[LL1_7000],[LL1_7000_rate],[LL2_7000],[LL2_7000_rate]
                          ,[critical_qty],[critical_rate],[critical_dpm]
                          ,[major_qty],[major_rate],[major_dpm]
                          ,[minor_qty],[minor_rate],[minor_dpm]
                          ,[pinhole_qty],[pinhole_rate],[pinhole_dpm]
                          ,[cosmetic_check_qty]
                      FROM [MES_OLAP].[dbo].[appearance_weekly_info_raw] r
                      JOIN [MES_OLAP].[dbo].[week_date] w on r.Year = w.year and r.MonthWeek = Cast(month as varchar)+month_week
                      where Plant = '{plant}' and w.enable = 1
                      order by [MonthWeek]
                    """
        cosmetic_summary_data = self.db.select_sql_dict(cosmetic_summary_sql)
        df = pd.DataFrame(cosmetic_summary_data)

        sheet_name = "外觀週累計"

        df = df.rename(columns=header_column)

        df.to_excel(writer, sheet_name=sheet_name, index=False)

        worksheet = writer.sheets[sheet_name]
        worksheet.freeze_panes = worksheet['A2']

        for col in worksheet.columns:
            max_length = max(len(str(cell.value)) for cell in col)
            col_letter = col[0].column_letter

            worksheet.column_dimensions[col_letter].width = max_length + 5

            # Set alignment
            for cell in col:
                if col_letter in [column_letter['total_6100'], column_letter['LL1_6100'], column_letter['LL2_6100'],
                                  column_letter['total_6200'], column_letter['LL1_6200'], column_letter['LL2_6200'],
                                  column_letter['total_6300'], column_letter['LL1_6300'], column_letter['LL2_6300'],
                                  column_letter['total_7000'], column_letter['LL1_7000'], column_letter['LL2_7000'],
                                  column_letter['critical_qty'], column_letter['critical_dpm'],
                                  column_letter['major_qty'], column_letter['major_dpm'],
                                  column_letter['minor_qty'], column_letter['minor_dpm'],
                                  column_letter['pinhole_qty'], column_letter['pinhole_dpm'],
                                  column_letter['cosmetic_check_qty']]:
                    cell.number_format = '#,##0'
                    cell.alignment = self.right_align_style.alignment
                elif col_letter in [column_letter['LL1_6100_rate'], column_letter['LL2_6100_rate'],
                                  column_letter['LL1_6200_rate'], column_letter['LL2_6200_rate'],
                                  column_letter['LL1_6300_rate'], column_letter['LL2_6300_rate'],
                                  column_letter['LL1_7000_rate'], column_letter['LL2_7000_rate'],
                                  column_letter['critical_rate'], column_letter['major_rate'],
                                  column_letter['minor_rate'], column_letter['pinhole_rate'],
                                  ]:
                    cell.alignment = self.center_align_style.alignment
                    cell.number_format = '0.00%'
                    worksheet.column_dimensions[col_letter].width = 14
                else:
                    cell.alignment = self.center_align_style.alignment

        hide_columns = ['total_6100', 'LL1_6100', 'LL2_6100',
                        'total_6200', 'LL1_6200', 'LL2_6200',
                        'total_6300', 'LL1_6300', 'LL2_6300',
                        'total_7000', 'LL1_7000', 'LL2_7000',
                        'critical_qty', 'major_qty', 'minor_qty', 'pinhole_qty']
        for column in hide_columns:
            worksheet.column_dimensions[column_letter[column]].hidden = True

    def generate_12aspect_cosmetic_excel(self, writer, plant, year, month_week):
        column_letter = {'runcard': 'A', 'belong_to': 'B', 'Machine': 'C', 'Line': 'D', 'Shift': 'E',
                         'WorkOrder': 'F', 'PartNo': 'G', 'ProductItem': 'H',
                        'SalePlaceCode': 'I', 'Period': 'J',
                         '6100LL1': 'K', '6100LL2': 'L',
                         '6200LL1': 'M', '6200LL2': 'N',
                         '6300LL1': 'O', '6300LL2': 'P',
                         '7000LL1': 'Q', '7000LL2': 'R',
                        }

        cosmetic_column_letter = {'AL4': 'S', 'KL3': 'T', 'AL1': 'U', 'IL1': 'V', 'AL2': 'W', 'AL5': 'X', 'AL6': 'Y', 'KL4': 'Z', 'AL7': 'AA', 'IL2': 'AB',
                                  'ML9': 'AC', 'KN7': 'AD', 'EL1': 'AE', 'BL1': 'AF', 'FL2': 'AG', 'GL2': 'AH', 'BL3': 'AI', 'AL3': 'AJ', 'KN6': 'AK',
                                  'BL2': 'AL', 'BN7': 'AM', 'BN1': 'AN', 'HL1': 'AO', 'GN1': 'AP', 'KL2': 'AQ',
                                  'EN2': 'AR', 'CN1': 'AS', 'FN2': 'AT', 'KN4': 'AU', 'CL1': 'AV', 'BN2': 'AW', 'BN6': 'AX', 'DN1': 'AY', 'BN9': 'AZ',
                                  'DL4': 'BA', 'BN8': 'BB', 'EL2': 'BC', 'FN1': 'BD', 'EN1': 'BE', 'BN5': 'BF', 'MX3': 'BG', 'CL3': 'BH', 'KN5': 'BI', 'DL3': 'BJ', 'CHECK QTY': 'BK',
                                  'B': 'BL', 'BT': 'BM', 'BT_1': 'BN', 'C': 'BO', 'C_1': 'BP', 'D1': 'BQ', 'D1_1': 'BR', 'D2': 'BS', 'D2_1': 'BT', 'D3': 'BU', 'D3_1': 'BV', 'D4': 'BW', 'D4_1': 'BX', 'D5': 'BY', 'D5_1': 'BZ',
                                    'K1': 'CA', 'K1_1': 'CB', 'K2': 'CC', 'K2_1': 'CD', 'K3': 'CE', 'K3_1': 'CF', 'K4': 'CG', 'K4_1': 'CH',
                                    'N1': 'CI', 'N1_1': 'CJ', 'N2': 'CK', 'N2_1': 'CL', 'N3': 'CM', 'N3_1': 'CN', 'N4': 'CO', 'N4_1': 'CP', 'N5': 'CQ', 'N5_1': 'CR', 'Pinhole': 'CS', 'Pinhole_Sample': 'CT',
                        }

        try:

            sql = f"""    
              SELECT counting.Runcard runcard,counting.belong_to,counting.Machine,counting.Line,counting.Shift,counting.WorkOrder,counting.PartNo,counting.ProductItem,SalePlaceCode,counting.Period
              FROM [MES_OLAP].[dbo].[counting_daily_info_raw] counting
              LEFT JOIN [MES_OLAP].[dbo].[mes_ipqc_data] ipqc on ipqc.Runcard = counting.Runcard
              LEFT JOIN [MES_OLAP].[dbo].[sap_customer_define] cus on cus.CustomerCode = counting.CustomerCode
              where counting.Year = {year} and counting.MonthWeek = '{month_week}'
              and Machine like '%{plant}%'
              and (OnlinePacking > 0 or WIPPacking > 0)
              order by Machine, WorkDate, Cast(Period as Int)
            
            """
            data = self.db.select_sql_dict(sql)
            df = pd.DataFrame(data)

            weight_sql = f"""
            SELECT ipqc.Runcard runcard, Cast(SalePlaceCode as varchar)+Weight_Defect defect_code, 1 qty 
              FROM [MES_OLAP].[dbo].[counting_daily_info_raw] counting
              LEFT JOIN [MES_OLAP].[dbo].[mes_ipqc_data] ipqc on ipqc.Runcard = counting.Runcard
              LEFT JOIN [MES_OLAP].[dbo].[sap_customer_define] cus on cus.CustomerCode = counting.CustomerCode
              where counting.Year = {year} and counting.MonthWeek = '{month_week}'
              and Weight_Status = 'NG'
              and Machine like '%{plant}%'
              and InspectedAQL is not Null
              
              --6100 美
              --6200 歐
              --6300 日
              --LL1 過重
              --LL2 過輕
            """
            weight_data = self.db.select_sql_dict(weight_sql)
            weight_df = pd.DataFrame(weight_data)

            weight_df = weight_df.pivot(index="runcard", columns="defect_code", values="qty").reset_index()
            order = ["runcard", "6100LL1", "6100LL2", "6200LL1", "6200LL2", "6300LL1", "6300LL2", "7000LL1", "7000LL2"]
            weight_df = weight_df.reindex(columns=order, fill_value='')
            weight_header = {'6100LL1': '美線過重', '6100LL2': '美線過輕', '6200LL1': '歐線過重', '6200LL2': '歐線過輕', '6300LL1': '日線過重', '6300LL2': '日線過輕', '7000LL1': 'OBM過重', '7000LL2': 'OBM過輕'}

            defect_sql = f"""
            SELECT [defect_code],
                  COALESCE(NULLIF(desc1, ''), desc2) AS defect_desc
              FROM [MES_OLAP].[dbo].[mes_defect_define]
            """
            defect_dist = self.db.select_sql_dict(defect_sql)
            defect_dist = {item['defect_code']: item['defect_desc'] for item in defect_dist}

            cosmetic_sql = f"""
             SELECT r.runcard, d.defect_code, sum(qty) sum_qty, max(cos.cosmetic_inspect_qty) inspect_qty
              FROM [MES_OLAP].[dbo].[counting_daily_info_raw] r
              LEFT JOIN [MES_OLAP].[dbo].[mes_ipqc_cosmetic_data] cos on r.Runcard = cos.runcard
              LEFT JOIN [MES_OLAP].[dbo].[mes_defect_define] d on d.defect_code = cos.defect_code
              where r.Year = {year} and r.MonthWeek = '{month_week}'
              group by r.runcard, defect_level, d.defect_code, desc2
            """
            cosmetic_data = self.db.select_sql_dict(cosmetic_sql)
            cosmetic_df = pd.DataFrame(cosmetic_data)

            inspect_qty_df = cosmetic_df[['runcard', 'inspect_qty']].drop_duplicates()
            cosmetic_df = cosmetic_df.pivot(index="runcard", columns="defect_code", values="sum_qty").reset_index()
            order = ["runcard", "AL4","KL3","AL1","IL1","AL2","AL5","AL6","KL4","AL7","IL2","ML9","KN7","EL1","BL1",
                     "FL2","GL2","BL3","AL3","KN6","BL2","BN7","BN1","HL1","GN1","KL2","EN2","CN1","FN2","KN4","CL1",
                     "BN2","BN6","DN1","BN9","DL4","BN8","EL2","FN1","EN1","BN5","MX3","CL3","KN5","DL3"
                     ]
            cosmetic_df = cosmetic_df.reindex(columns=order, fill_value='')

            # 計算針孔Defect Code加總
            pinhole_sql = f"""
            SELECT r.runcard, d.defect_code, sum(qty) sum_qty
              FROM [MES_OLAP].[dbo].[counting_daily_info_raw] r
              JOIN [MES_OLAP].[dbo].[mes_ipqc_pinhole_data] cos on r.Runcard = cos.runcard
              JOIN [MES_OLAP].[dbo].[mes_defect_define] d on d.defect_code = cos.defect_code
              where r.Year = {year} and r.MonthWeek = '{month_week}'
              group by r.runcard, defect_level, d.defect_code, desc2
            """
            pinhole_data = self.db.select_sql_dict(pinhole_sql)
            pinhole_df = pd.DataFrame(pinhole_data)
            pinhole_df = pinhole_df.pivot(index="runcard", columns="defect_code", values="sum_qty").reset_index()
            pinhole_df['Pinhole'] = pinhole_df.iloc[:, 1:-1].notna().any(axis=1).astype(int)

            # 計算針孔樣本數
            pinhole_sample_sql = f"""
            SELECT r.Id runcard,WorkCenterTypeName, AQL AS WO_AQL
              FROM [PMGMES].[dbo].[PMG_MES_RunCard] r
              JOIN [PMGMES].[dbo].[PMG_MES_RunCard_IPQCInspectIOptionMapping] m on r.Id = m.RunCardId
              JOIN [PMGMES].[dbo].[PMG_MES_WorkOrder] w on r.WorkOrderId = w.Id
              where GroupType = 'Pinhole' and r.InspectionDate between '{self.this_start_date}' and '{self.this_end_date}'
            """
            pinhole_sample_data = self.mes_db.select_sql_dict(pinhole_sample_sql)
            pinhole_sample_df = pd.DataFrame(pinhole_sample_data)

            if plant == "PVC":
                pinhole_sample_df['Pinhole_Sample'] = 25
            elif plant == "NBR":
                pinhole_sample_df['Pinhole_Sample'] = np.where(pinhole_sample_df['WO_AQL'] == '1.0', 50, 25)

            df = pd.merge(df, weight_df, on=['runcard'], how='left')
            df = pd.merge(df, cosmetic_df, on=['runcard'], how='left')
            df = pd.merge(df, pinhole_df, on=['runcard'], how='left')
            df = pd.merge(df, inspect_qty_df, on="runcard", how="left")
            df = pd.merge(df, pinhole_sample_df, on="runcard", how="left")
            df = df.fillna('')

            sum_item = [
                '6100LL1', '6100LL2', '6200LL1', '6200LL2', '6300LL1', '6300LL2', '7000LL1', '7000LL2',
                "AL4", "KL3", "AL1", "IL1", "AL2", "AL5", "AL6", "KL4", "AL7", "IL2", "ML9", "KN7", "EL1", "BL1",
                "FL2", "GL2", "BL3", "AL3", "KN6", "BL2", "BN7", "BN1", "HL1", "GN1", "KL2", "EN2", "CN1", "FN2", "KN4",
                "CL1", "BN2", "BN6", "DN1", "BN9", "DL4", "BN8", "EL2", "FN1", "EN1", "BN5", "MX3", "CL3", "KN5", "DL3",
                'inspect_qty',
                'B', 'BT', 'BT_1', 'C', 'C_1', 'D1', 'D1_1', 'D2', 'D2_1', 'D3', 'D3_1', 'D4', 'D4_1', 'D5', 'D5_1',
                'K1', 'K1_1', 'K2', 'K2_1', 'K3', 'K3_1', 'K4', 'K4_1',
                'N1', 'N1_1', 'N2', 'N2_1', 'N3', 'N3_1', 'N4', 'N4_1', 'N5', 'N5_1', 'Pinhole', 'Pinhole_Sample', ]

            # 確保所有 defect_code 欄位都存在，順序與 expected_defect_codes 一致
            all_columns = ['runcard', 'belong_to', 'Machine', 'Line', 'Shift', 'WorkOrder', 'PartNo', 'ProductItem', 'SalePlaceCode', 'Period'] + sum_item  # 保持固定順序
            df = df.reindex(columns=all_columns, fill_value='')

            sum_list = {}
            for item in sum_item:
                try:
                    df[item] = pd.to_numeric(df[item], errors='coerce').fillna(0)
                    sum_list[item] = df[item].sum()
                except Exception as e:
                    print(e)

            sum_LL1_6100 = sum_list['6100LL1']
            sum_LL2_6100 = sum_list['6100LL2']
            count_6100_qty = (df['SalePlaceCode'] == 6100).sum()
            rate_LL1_6100 = round(sum_LL1_6100 / count_6100_qty, 4) if count_6100_qty > 0 else 0
            rate_LL2_6100 = round(sum_LL2_6100 / count_6100_qty, 4) if count_6100_qty > 0 else 0

            sum_LL1_6200 = sum_list['6200LL1']
            sum_LL2_6200 = sum_list['6200LL2']
            count_6200_qty = (df['SalePlaceCode'] == 6200).sum()
            rate_LL1_6200 = round(sum_LL1_6200 / count_6200_qty, 4) if count_6200_qty > 0 else 0
            rate_LL2_6200 = round(sum_LL2_6200 / count_6200_qty, 4) if count_6200_qty > 0 else 0

            sum_LL1_6300 = sum_list['6300LL1']
            sum_LL2_6300 = sum_list['6300LL2']
            count_6300_qty = (df['SalePlaceCode'] == 6300).sum()
            rate_LL1_6300 = round(sum_LL1_6300 / count_6300_qty, 4) if count_6300_qty > 0 else 0
            rate_LL2_6300 = round(sum_LL2_6300 / count_6300_qty, 4) if count_6300_qty > 0 else 0

            sum_LL1_7000 = sum_list['7000LL1']
            sum_LL2_7000 = sum_list['7000LL2']
            count_7000_qty = (df['SalePlaceCode'] == 7000).sum()
            rate_LL1_7000 = round(sum_LL1_7000 / count_7000_qty, 4) if count_7000_qty > 0 else 0
            rate_LL2_7000 = round(sum_LL2_7000 / count_7000_qty, 4) if count_7000_qty > 0 else 0

            sum_inspect_qty = sum_list['inspect_qty']

            sum_critical = sum_list['AL4'] + sum_list['KL3'] + sum_list['AL1'] + sum_list['IL1'] + sum_list['AL2'] + \
                           sum_list['AL5'] + sum_list['AL6'] + sum_list['KL4'] + sum_list['AL7'] + sum_list['IL2']
            critical_rate = round(sum_critical / sum_inspect_qty, 4)
            critical_dpm = round(critical_rate * 1000000, 0)

            sum_major = sum_list['ML9'] + sum_list['KN7'] + sum_list['EL1'] + sum_list['BL1'] + sum_list['FL2'] + \
                           sum_list['GL2'] + sum_list['BL3'] + sum_list['AL3'] + sum_list['KN6'] + sum_list['BL2'] + \
                        sum_list['BN7'] + sum_list['BN1'] + sum_list['HL1'] + sum_list['GN1'] + sum_list['KL2']
            major_rate = round(sum_major / sum_inspect_qty, 4)
            major_dpm = round(major_rate * 1000000, 0)


            sum_minor = sum_list['EN2'] + sum_list['CN1'] + sum_list['FN2'] + sum_list['KN4'] + sum_list['CL1'] + \
                           sum_list['BN2'] + sum_list['BN6'] + sum_list['DN1'] + sum_list['BN9'] + sum_list['DL4'] + \
                        sum_list['BN8'] + sum_list['EL2'] + sum_list['FN1'] + sum_list['EN1'] + \
                        sum_list['MX3'] + sum_list['CL3'] + sum_list['KN5'] + sum_list['DL3']
            minor_rate = round(sum_minor / sum_inspect_qty, 4)
            minor_dpm = round(minor_rate * 1000000, 0)

            sum_pinhole = sum_list['Pinhole']
            sum_pinhole_sample = sum_list['Pinhole_Sample']

            pinhole_rate = round(sum_pinhole / sum_pinhole_sample, 4)
            pinhole_dpm = round(pinhole_rate * 1000000, 0)

            df = df.rename(columns=weight_header)
            df = df.rename(columns=defect_dist)
            df = df.rename(columns={'inspect_qty': '外觀檢查數量'})
            df = df.rename(columns=self.header_columns)

            sheet_name = "外觀"
            df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=1)

            # Read the written Excel file
            workbook = writer.book
            worksheet = writer.sheets[sheet_name]

            # header_alignment = Alignment(horizontal='center', wrap_text=True)
            header_alignment = Alignment(horizontal='center')
            header_border = Border(
                top=Side(style='medium'),
                bottom=Side(style='medium'),
                left=Side(style='medium'),
                right=Side(style='medium')
            )
            fill_style = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

            # Weight Header
            worksheet.merge_cells(f"{column_letter['6100LL1']}1:{column_letter['7000LL2']}1")
            start_row, start_col = worksheet[f"{column_letter['6100LL1']}1"].row, worksheet[f"{column_letter['6100LL1']}1"].column
            end_row, end_col = worksheet[f"{column_letter['7000LL2']}1"].row, worksheet[f"{column_letter['7000LL2']}1"].column
            cell = worksheet[f"{column_letter['6100LL1']}1"]
            cell.value = "重量檢驗"
            cell.alignment = Alignment(horizontal="center", vertical="center")
            # cell.border = header_border
            for row in range(start_row, end_row + 1):
                for col in range(start_col, end_col + 1):
                    cell = worksheet.cell(row=row, column=col)
                    cell.border = header_border
                    cell.fill = fill_style

            # Cosmetic Header
            worksheet.merge_cells(f"{cosmetic_column_letter['AL4']}1:{cosmetic_column_letter['IL2']}1")
            start_row, start_col = worksheet[f"{cosmetic_column_letter['AL4']}1"].row, worksheet[f"{cosmetic_column_letter['AL4']}1"].column
            end_row, end_col = worksheet[f"{cosmetic_column_letter['IL2']}1"].row, worksheet[f"{cosmetic_column_letter['IL2']}1"].column
            cell = worksheet[f"{cosmetic_column_letter['AL4']}1"]
            cell.value = "外觀CRITICAL"
            cell.alignment = Alignment(horizontal="center", vertical="center")
            for row in range(start_row, end_row + 1):
                for col in range(start_col, end_col + 1):
                    cell = worksheet.cell(row=row, column=col)
                    cell.border = header_border
                    cell.fill = fill_style

            worksheet.merge_cells(f"{cosmetic_column_letter['ML9']}1:{cosmetic_column_letter['KL2']}1")
            start_row, start_col = worksheet[f"{cosmetic_column_letter['ML9']}1"].row, worksheet[f"{cosmetic_column_letter['ML9']}1"].column
            end_row, end_col = worksheet[f"{cosmetic_column_letter['KL2']}1"].row, worksheet[f"{cosmetic_column_letter['KL2']}1"].column
            cell = worksheet[f"{cosmetic_column_letter['ML9']}1"]
            cell.value = "外觀MAJOR"
            cell.alignment = Alignment(horizontal="center", vertical="center")
            for row in range(start_row, end_row + 1):
                for col in range(start_col, end_col + 1):
                    cell = worksheet.cell(row=row, column=col)
                    cell.border = header_border
                    cell.fill = fill_style

            worksheet.merge_cells(f"{cosmetic_column_letter['EN2']}1:{cosmetic_column_letter['DL3']}1")
            start_row, start_col = worksheet[f"{cosmetic_column_letter['EN2']}1"].row, worksheet[f"{cosmetic_column_letter['EN2']}1"].column
            end_row, end_col = worksheet[f"{cosmetic_column_letter['DL3']}1"].row, worksheet[f"{cosmetic_column_letter['DL3']}1"].column
            cell = worksheet[f"{cosmetic_column_letter['EN2']}1"]
            cell.value = "外觀MINOR"
            cell.alignment = Alignment(horizontal="center", vertical="center")
            for row in range(start_row, end_row + 1):
                for col in range(start_col, end_col + 1):
                    cell = worksheet.cell(row=row, column=col)
                    cell.border = header_border
                    cell.fill = fill_style

            # Pinhole Header
            worksheet.merge_cells(f"{cosmetic_column_letter['B']}1:{cosmetic_column_letter['N5_1']}1")
            start_row, start_col = worksheet[f"{cosmetic_column_letter['B']}1"].row, worksheet[f"{cosmetic_column_letter['B']}1"].column
            end_row, end_col = worksheet[f"{cosmetic_column_letter['N5_1']}1"].row, worksheet[f"{cosmetic_column_letter['N5_1']}1"].column
            cell = worksheet[f"{cosmetic_column_letter['B']}1"]
            cell.value = "針孔"
            cell.alignment = Alignment(horizontal="center", vertical="center")
            for row in range(start_row, end_row + 1):
                for col in range(start_col, end_col + 1):
                    cell = worksheet.cell(row=row, column=col)
                    cell.border = header_border
                    cell.fill = fill_style

            worksheet.freeze_panes = worksheet['A3']

            for cell in worksheet[2]:  # First line is Header
                cell.font = self.header_font
                cell.alignment = header_alignment
                cell.border = self.header_border
                # Formatting

            for col in worksheet.columns:
                max_length = max(len(str(cell.value)) for cell in col)
                col_letter = col[1].column_letter

                worksheet.column_dimensions[col_letter].width = max_length + 5

                # Set alignment
                for cell in col:
                    if col_letter in [column_letter['runcard'], column_letter['belong_to'], column_letter['Machine'],
                                      column_letter['Line'], column_letter['Shift'], column_letter['WorkOrder'],
                                      column_letter['PartNo'], column_letter['ProductItem'], column_letter['SalePlaceCode'],
                                      column_letter['Period']]:

                        cell.alignment = self.center_align_style.alignment

                    # if col_letter in [colmn_letter['Activation'], colmn_letter['OEE']]:
                    #     worksheet.column_dimensions[col_letter].hidden = True

            # 最後一行加總
            thin_top_border = Border(top=Side(style="thin"))
            last_row = worksheet.max_row+1

            # K 欄 (11) 到 CT 欄 (96)
            for col_idx in range(11, 98 + 1):
                col_letter = get_column_letter(col_idx)

                # 設定 SUM 公式
                sum_formula = f"=SUM({col_letter}2:{col_letter}{last_row-1})"
                worksheet[f"{col_letter}{last_row}"] = sum_formula  # ✅ 填入最後一行

                # 在最後一行上方（倒數第 2 行）加上框線
                if last_row > 1:  # 確保至少有兩行
                    cell_above = worksheet[f"{col_letter}{last_row}"]
                    cell_above.border = thin_top_border  # ✅ 設定上方框線


            delete_sql = f"""
            DELETE FROM [MES_OLAP].[dbo].[appearance_weekly_info_raw]
            WHERE Plant = '{plant}' AND [Year] = {year} AND MonthWeek = '{month_week}'
            """
            self.db.execute_sql(delete_sql)

            insert_sql = f"""
            INSERT INTO [MES_OLAP].[dbo].[appearance_weekly_info_raw] (Plant, Year, MonthWeek, 
            total_6100, LL1_6100, LL1_6100_rate, LL2_6100, LL2_6100_rate, 
            total_6200, LL1_6200, LL1_6200_rate, LL2_6200, LL2_6200_rate,
            total_6300, LL1_6300, LL1_6300_rate, LL2_6300, LL2_6300_rate,
            total_7000, LL1_7000, LL1_7000_rate, LL2_7000, LL2_7000_rate,
            critical_qty, critical_rate, critical_dpm,
            major_qty, major_rate, major_dpm, minor_qty, minor_rate, minor_dpm, pinhole_qty, pinhole_rate, pinhole_dpm,
            cosmetic_check_qty, pinhole_check_qty)
            VALUES('{plant}', {year}, '{month_week}', 
            {count_6100_qty}, {sum_LL1_6100}, {rate_LL1_6100}, {sum_LL2_6100}, {rate_LL2_6100}, 
            {count_6200_qty}, {sum_LL1_6200}, {rate_LL1_6200}, {sum_LL2_6200}, {rate_LL2_6200},
            {count_6300_qty}, {sum_LL1_6300}, {rate_LL1_6300}, {sum_LL2_6300}, {rate_LL2_6300}, 
            {count_7000_qty}, {sum_LL1_7000}, {rate_LL1_7000}, {sum_LL2_7000}, {rate_LL2_7000}, 
            {sum_critical}, {critical_rate}, {critical_dpm},
            {sum_major}, {major_rate}, {major_dpm}, {sum_minor}, {minor_rate}, {minor_dpm}, {sum_pinhole}, {pinhole_rate}, {pinhole_dpm},
            {sum_inspect_qty}, null)
            """
            self.db.execute_sql(insert_sql)

        except Exception as e:
            print(e)

    def generate_12aspect_output_excel(self, writer, plant):
        colmn_letter = {'Plant': 'A', 'Year': 'B', 'MonthWeek': 'C', 'CountingQty': 'D', 'SeparateQty': 'E', 'AvgSpeed': 'F',
                        'Capacity': 'G', 'Yield': 'H', 'Activation': 'I', 'OEE': 'J', 'SeparateRate': 'K', 'ScrapRate': 'L',
                        'Target': 'M', 'OnlinePacking': 'N', 'Gap': 'O'}

        sql = f"""
            SELECT [Plant], r.Year, [MonthWeek]
              ,[CountingQty],[SeparateQty] SeparateQuantity,[AvgSpeed]
              ,[Capacity],[Yield],[Activation]
              ,[OEE],[SeparateRate],[ScrapRate],[Target],[OnlinePacking], Target-OnlinePacking Gap
          FROM [MES_OLAP].[dbo].[counting_weekly_info_raw] r
          JOIN [MES_OLAP].[dbo].[week_date] w on r.Year = w.year and r.MonthWeek = Cast(month as varchar)+month_week
          where Plant = '{plant}' and w.enable = 1
          Order by [MonthWeek]
        """
        data = self.db.select_sql_dict(sql)
        df = pd.DataFrame(data)

        df.rename(columns=self.header_columns, inplace=True)

        sheet_name = "週累計"
        df.to_excel(writer, sheet_name=sheet_name, index=False)



        # Read the written Excel file
        workbook = writer.book
        worksheet = writer.sheets[sheet_name]

        # Freeze the first row
        worksheet.freeze_panes = worksheet['A2']

        # Apply Header Style
        for cell in worksheet[1]:  # First line is Header
            cell.font = self.header_font
            cell.alignment = self.header_alignment
            cell.border = self.header_border
            # Formatting

        for col in worksheet.columns:
            max_length = max(len(str(cell.value)) for cell in col)
            col_letter = col[0].column_letter

            worksheet.column_dimensions[col_letter].width = max_length + 5

            # Set alignment
            for cell in col:
                if col_letter in [colmn_letter['CountingQty'], colmn_letter['SeparateQty'],
                                  colmn_letter['OnlinePacking'], colmn_letter['Target'], colmn_letter['Gap']]:  # Apply right alignment for specific columns
                    cell.number_format = '#,##0'
                    cell.alignment = self.right_align_style.alignment
                elif col_letter in [colmn_letter['Activation'], colmn_letter['Capacity'], colmn_letter['Yield'],
                                    colmn_letter['OEE'], colmn_letter['SeparateRate'], colmn_letter['ScrapRate']]:
                    worksheet.column_dimensions[col_letter].width = 15
                    cell.alignment = self.center_align_style.alignment
                    cell.number_format = '0.0%'
                else:
                    cell.alignment = self.center_align_style.alignment

                if col_letter in [colmn_letter['Activation'], colmn_letter['OEE']]:
                    worksheet.column_dimensions[col_letter].hidden = True

    def generate_summary(self, writer, machine_groups):
        colmn_letter = {'Name': 'A', 'Week': 'B', 'shift': 'C', 'Line': 'D', 'CountingQty': 'E', 'FaultyQty': 'F', 'ScrapQty': 'G',
                        'SeparateQty': 'H', 'OnlinePacking': 'I', 'RunTime': 'J', 'AllTime': 'K',
                        'AvgSpeed': 'L', 'Target': 'M', 'Capacity': 'N', 'Yield': 'O', 'Activation': 'P', 'OEE': 'Q', 'Achievement': 'R', 'SeparateRate': 'S'}
        summary_data = []
        tmp_date = self.date_mark.replace('_', '~')

        tmp_week = f"{self.month_week} ({tmp_date})"

        thin_border_top = Border(top=Side(style="thin"))
        thin_border_bottom = Border(bottom=Side(style="thin"))
        for machine_name, machine_df in machine_groups:
            for shift in machine_df['Shift'].unique():  # Loop through each unique Line
                for line in machine_df['Line'].unique():
                    filtered_df = machine_df[(machine_df['Line'] == line) & (machine_df['Shift'] == shift)]
                    countingQty = filtered_df['CountingQty'].sum()
                    faultyQty = filtered_df['FaultyQuantity'].sum()
                    scrapQty = filtered_df['ScrapQuantity'].sum()
                    separateQty = filtered_df['SeparateQty'].sum()
                    onlinePacking = filtered_df['OnlinePacking'].sum()
                    runTime = filtered_df['RunTime'].sum()
                    stopTime = filtered_df['StopTime'].sum()
                    allTime = filtered_df['AllTime'].sum()
                    avgSpeed = round(filtered_df['StdSpeed'].mean(), 0)
                    target = filtered_df['Target'].sum()
                    rate = round((int(onlinePacking) / int(target)), 3) if int(target) > 0 else 0

                    summary_row = {
                        'Name': machine_name,
                        'Date': tmp_week,
                        'Shift': shift,
                        'Line': line,
                        'CountingQty': countingQty,
                        'FaultyQuantity': faultyQty,
                        'ScrapQuantity': scrapQty,
                        'SeparateQuantity': separateQty,
                        'OnlinePacking': onlinePacking,
                        'RunTime': runTime,
                        'AllTime': allTime,
                        'AvgSpeed': avgSpeed,
                        'Target': target,
                        'Capacity': '',
                        'Yield': '',
                        'Activation': '',
                        'OEE': '',
                        'Achievement Rate': rate
                    }
                    summary_data.append(summary_row)

            # Summary Row
            countingQty = sum(item['CountingQty'] for item in summary_data if item['Name'] == machine_name)
            faultyQty = sum(item['FaultyQuantity'] for item in summary_data if item['Name'] == machine_name)
            scrapQty = sum(item['ScrapQuantity'] for item in summary_data if item['Name'] == machine_name)
            separateQty = sum(item['SeparateQuantity'] for item in summary_data if item['Name'] == machine_name)
            onlinePacking = sum(item['OnlinePacking'] for item in summary_data if item['Name'] == machine_name)
            runTime = sum(item['RunTime'] for item in summary_data if item['Name'] == machine_name)
            allTime = sum(item['AllTime'] for item in summary_data if item['Name'] == machine_name)

            avgSpeed_values = [item['AvgSpeed'] for item in summary_data if item['Name'] == machine_name]
            avgSpeed = round(sum(avgSpeed_values) / len(avgSpeed_values), 0) if avgSpeed_values else 0

            target = sum(item['Target'] for item in summary_data if item['Name'] == machine_name)
            # 稼動率allTime必須扣除計劃性停機時間
            # activation = round(runTime/allTime, 3) if int(allTime) > 0 else 0
            activation = ''
            output = countingQty+faultyQty+scrapQty
            capacity = round(output/target, 3) if int(target) > 0 else 0
            _yield = round((onlinePacking-separateQty)/output, 3) if int(output) > 0 else 0

            # 等有稼動率才能做OEE
            # oee = round(activation * capacity * _yield, 3)
            oee = ''
            rate = round(onlinePacking/target, 3) if int(target) > 0 else 0
            separateRate = round(separateQty/onlinePacking, 3) if int(onlinePacking) > 0 else 0

            summary_data.append({'Name': machine_name, 'Date': tmp_week, 'Shift': '', 'Line': '',
                                 'CountingQty': countingQty, 'FaultyQuantity': faultyQty, 'ScrapQuantity': scrapQty,
                                 'SeparateQuantity': separateQty, 'OnlinePacking': onlinePacking,
                                 'RunTime': runTime, 'AllTime': allTime, 'AvgSpeed': avgSpeed, 'Target': target,
                                 'Capacity': capacity, 'Yield': _yield, 'Activation': activation, 'OEE': oee, 'Achievement Rate': rate, 'SeparateRate': separateRate})
        summary_df = pd.DataFrame(summary_data)
        # Change column names
        summary_df.rename(columns=self.header_columns, inplace=True)
        summary_sheet_name = "Summary"
        summary_df.to_excel(writer, sheet_name=summary_sheet_name, index=False)
        # Read the written Excel file
        workbook = writer.book
        worksheet = writer.sheets[summary_sheet_name]

        # Apply Header Style
        for cell in worksheet[1]:  # First line is Header
            cell.font = self.header_font
            cell.alignment = self.header_alignment
            cell.border = self.header_border
            # Formatting

        for col in worksheet.columns:
            max_length = max(len(str(cell.value)) for cell in col)
            col_letter = col[0].column_letter

            worksheet.column_dimensions[col_letter].width = max_length + 5

            # Set alignment
            for cell in col:
                if col_letter in [colmn_letter['CountingQty'], colmn_letter['FaultyQty'], colmn_letter['ScrapQty'], colmn_letter['SeparateQty'],
                                  colmn_letter['OnlinePacking'], colmn_letter['Target'], colmn_letter['RunTime'], colmn_letter['AllTime']]:  # Apply right alignment for specific columns
                    cell.number_format = '#,##0'
                    cell.alignment = self.right_align_style.alignment
                elif col_letter in [colmn_letter['Activation'], colmn_letter['Capacity'], colmn_letter['Yield'], colmn_letter['OEE'], colmn_letter['Achievement'], colmn_letter['SeparateRate']]:
                    worksheet.column_dimensions[col_letter].width = 15
                    cell.alignment = self.center_align_style.alignment
                    cell.number_format = '0.0%'
                else:
                    cell.alignment = self.center_align_style.alignment

                if col_letter in [colmn_letter['Activation'], colmn_letter['OEE']]:
                    worksheet.column_dimensions[col_letter].hidden = True

        name_col = 1
        current_name = None
        group_start = None

        for row in range(2, len(summary_df) + 2):  # Data starts from row 2 in Excel
            name = worksheet.cell(row=row, column=name_col).value

            if name != current_name:
                # Finalize the previous group
                if group_start is not None:
                    last_row = row - 1
                    for col in range(1, summary_df.shape[1] + 1):  # Apply border to all columns
                        worksheet.cell(row=last_row, column=col).border = thin_border_top + thin_border_bottom
                    if row - group_start > 1:
                        worksheet.row_dimensions.group(group_start, row - 2, hidden=True)

                # Start a new group
                current_name = name
                group_start = row

        # Finalize the last group
        if group_start is not None:
            last_row = len(summary_df) + 1
            for col in range(1, summary_df.shape[1] + 1):
                worksheet.cell(row=last_row, column=col).border = thin_border_top + thin_border_bottom
            if len(summary_df) + 1 - group_start > 1:
                worksheet.row_dimensions.group(group_start, len(summary_df) + 1 - 1, hidden=True)

        # 註解
        comment = Comment(text="有做IPQC的機台實際運轉時間 * 標準車速", author="System")
        comment.width = 600
        worksheet[colmn_letter['Target'] + "1"].comment = comment

        comment = Comment(text="有做IPQC的機台實際運轉時間 / (可運轉時間-計劃性停機時間)", author="System")
        comment.width = 600
        worksheet[colmn_letter['Activation'] + "1"].comment = comment

        comment = Comment(text="(點數機數量+二級品數量+廢品數量) / 目標產能", author="System")
        comment.width = 600
        worksheet[colmn_letter['Capacity'] + "1"].comment = comment

        comment = Comment(text="(包裝確認量-隔離品數量)/(點數機數量+二級品數量+廢品數量)", author="System")
        comment.width = 600
        worksheet[colmn_letter['Yield'] + "1"].comment = comment

        comment = Comment(text="稼動率 * 產能效率 * 良率", author="System")
        comment.width = 600
        worksheet[colmn_letter['OEE'] + "1"].comment = comment

        comment = Comment(text="包裝確認量 / 目標產能", author="System")
        comment.width = 600
        worksheet[colmn_letter['Achievement'] + "1"].comment = comment

        comment = Comment(text="IPQC其中一項判定不合格或AQL不符合工單標準", author="System")
        comment.width = 600
        worksheet[colmn_letter['SeparateQty'] + "1"].comment = comment

        comment = Comment(text="隔離品數量 / 包裝確認量", author="System")
        comment.width = 600
        worksheet[colmn_letter['SeparateRate'] + "1"].comment = comment

        # # 設置欄的 outlineLevel 讓其可以折疊/展開
        hide_columns = ['CountingQty', 'FaultyQty', 'ScrapQty', 'SeparateQty', 'OnlinePacking',
                        'RunTime', 'AllTime', ]
        for column in hide_columns:
            worksheet.column_dimensions[colmn_letter[column]].outlineLevel = 1

        worksheet.column_dimensions.group(colmn_letter['CountingQty'], colmn_letter['AllTime'], hidden=True)

        return summary_df

    def generate_chart(self, plant):
        this_start_date = self.this_start_date
        this_end_date = self.this_end_date
        last_start_date = self.last_start_date
        last_end_date = self.last_end_date
        mode = self.mode
        save_path = self.save_path
        date_mark = self.date_mark
        yticks_labels = []

        plt.rcParams['font.sans-serif'] = ['Microsoft YaHei']

        # Target只看有做IPQC的部分
        sql = f"""SELECT Machine name,
                    sum(case when cast(belong_to as date) between '{this_start_date}' and '{this_end_date}' then OnlinePacking else 0 end) as this_time,
                    sum(case when cast(belong_to as date) between '{this_start_date}' and '{this_end_date}' then Target else 0 end) as target_this_time,
                    (sum(case when cast(belong_to as date) between '{this_start_date}' and '{this_end_date}' then Target else 0 end) -
                    sum(case when cast(belong_to as date) between '{this_start_date}' and '{this_end_date}' then OnlinePacking else 0 end)) as this_unfinish
                    FROM [MES_OLAP].[dbo].[counting_daily_info_raw] c
                    JOIN [MES_OLAP].[dbo].[mes_ipqc_data] ipqc on c.Runcard = ipqc.Runcard
                    where Machine like '%{plant}%' and not (WorkOrder != '' and InspectedAQL ='')
                    and Weight_Value > 0
                    group by Machine
                    order by Machine"""
        data = self.db.select_sql_dict(sql)

        # Output Bar Chart
        x_labels = [str(item['name']).split('_')[-1] for item in data]
        x_range = range(0, len(x_labels) * 2, 2)

        this_data = [int(item['this_time']) for item in data]
        this_unfinish = [int(item['this_unfinish']) if int(item['this_unfinish']) > 0 else 0 for item in data]

        max_data = max(this_data, default=0)
        step_data = 10

        this_rate = [round((item['this_time'] / item['target_this_time']) * 100, 2) if int(
            item['target_this_time']) > 0 else 0 for item in data]

        max_rate = max(this_rate, default=0)
        rounded_max_rate = (math.ceil(max_rate / 10) * 10)
        rounded_step_rate = 20

        bar_width = 0.6
        plt.figure(figsize=(16, 9))
        fig, ax1 = plt.subplots(figsize=(16, 9))

        this_week_bars = ax1.bar([i for i in x_range], this_data, width=bar_width,
                                  label=f"包裝確認量",
                                  align='center', color='#10ba81')
        unfinish_bars = ax1.bar([i for i in x_range], this_unfinish, width=bar_width, bottom=this_data,
                                  label=f"週目標差異",
                                  align='center', color='lightgreen')

        ax1.set_xticks(x_range)
        ax1.set_xticklabels(x_labels, rotation=0, ha="center", fontsize=12)

        yticks_positions = []
        if 'NBR' in plant:
            yticks_positions = [1000000, 2000000, 3000000, 4000000, 5000000, 6000000, 7000000, 10000000]
            yticks_labels = ['1百萬', '2百萬', '3百萬', '4百萬', '5百萬', '6百萬', '7百萬', '']
        elif 'PVC' in plant:
            yticks_positions = [1000000, 2000000, 3000000, 4000000, 5000000]
            yticks_labels = ['1百萬', '2百萬', '3百萬', '4百萬', '']

        for index, bar in enumerate(this_week_bars):
            height = bar.get_height()
            unfinish_height = unfinish_bars[index].get_height()

            if len(str(max_data)) > 7:
                show_text = f'{round(height/(10**(len(str(max_data)) - 2)),2)}'[:4] if height > 0 else ''
            elif 4 < len(str(max_data)) <= 7:
                show_text = f'{round(height/(10**(len(str(max_data)) - 1)),2)}'[:4] if height > 0 else ''

            ax1.text(
                bar.get_x() + bar.get_width() / 2,
                height+unfinish_height,
                show_text,
                ha='center', va='bottom', fontsize=12  # Align the text
            )

        ax1.set_yticks(yticks_positions)
        ax1.set_yticklabels(yticks_labels, fontsize=12)

        # Achievement Rate Line Chart (橘色的線)
        ax2 = ax1.twinx()

        # line_label = f"{this_start_date.strftime('%d/%m')}-{this_end_date.strftime('%d/%m')}"
        sr_achieve_rate = "本週達成率"
        filtered_data = [(x, rate) for x, rate in zip(x_range, this_rate) if rate != 0]  # 折線圖上的文字
        x_filtered, this_rate_filtered = zip(*filtered_data)
        this_rate_line = ax2.plot(x_filtered, this_rate_filtered,
                                  label=sr_achieve_rate,
                                  color='#ED7D31', marker='o', linewidth=1.5)

        # Label Name
        sr_target = "達成率目標%"
        # Chart Label
        ry_label = "達成率(%)"
        ly_label = "Product (百萬)"

        name = self.date_mark
        title = f"\n{plant} {self.year} {self.month_week} ({name})目標達成率\n"

        # Achievement Rate Standard Line
        if "NBR" in plant :
            target_rate = 95
        elif "PVC" in plant:
            target_rate = 98

        yticks_positions = list(range(0, rounded_max_rate + 1 * rounded_step_rate, rounded_step_rate))
        if target_rate not in yticks_positions:
            yticks_positions.append(target_rate)
        yticks_positions.append(rounded_max_rate + 0.25 * rounded_step_rate)
        yticks_positions = sorted(yticks_positions)

        yticks_labels = [f"{i}" + '%' for i in yticks_positions]
        yticks_labels[-1] = ""
        ax2.set_yticks(yticks_positions)
        ax2.set_yticklabels(yticks_labels, fontsize=12)
        ax2.axhline(y=target_rate, color='red', linestyle='--', linewidth=1, label=sr_target)

        ax1.xaxis.set_label_coords(0.975, -0.014)
        ax1.set_ylabel('Product output', fontsize=12)
        ax2.set_ylabel('Archive rates', fontsize=12)

        for i, value in enumerate(this_rate):
            ax2.text(
                x_range[i],
                value + 1.5,
                f"{value:.1f}%" if value != 0 else '',
                ha='center', va='bottom', fontsize=12, color='white',
                bbox=dict(facecolor='#ED7D31', edgecolor='none', boxstyle='round,pad=0.3')
            )

        handles1, labels1 = ax1.get_legend_handles_labels()
        handles2, labels2 = ax2.get_legend_handles_labels()
        fig.legend(
            handles1 + handles2,
            labels1 + labels2,
            loc='center left',
            fontsize=12,
            title="Note",
            title_fontsize=12,
            bbox_to_anchor=(1.0, 0.5),
            ncol=1
        )

        # Customize axes and borders
        ax = plt.gca()  # Get current axes
        ax.spines['top'].set_color('white')
        ax.spines['top'].set_linewidth(1.5)

        ax1.tick_params(axis='y', colors='black')  # Y-axis ticks color
        y_tick_lines = ax1.get_yticklines()
        y_tick_lines[-2].set_visible(False)
        y_tick_lines[-1].set_visible(False)
        ax1.annotate(
            '',
            xy=(0, 1.0),
            xytext=(0, 0.98),
            xycoords='axes fraction',
            arrowprops=dict(facecolor='black', arrowstyle='-|>,widthA=0.4,widthB=1.4', linewidth=0.5)
        )
        ax1.set_ylabel(ly_label, labelpad=20, rotation=0)
        ax1.yaxis.set_label_coords(-0.01, 1.01)

        ax2.tick_params(axis='y', colors='black')
        y_tick_lines = ax2.get_yticklines()
        y_tick_lines[-2].set_visible(False)
        y_tick_lines[-1].set_visible(False)
        ax2.annotate(
            '',
            xy=(1, 1.0),
            xytext=(1, 0.98),
            xycoords='axes fraction',
            arrowprops=dict(facecolor='black', arrowstyle='-|>,widthA=0.4,widthB=1.4', linewidth=0.5)
        )

        ax2.set_ylabel(ry_label, labelpad=20, rotation=0)
        ax2.yaxis.set_label_coords(1.01, 1.03)

        plt.title(title, fontsize=20)

        plt.tight_layout()

        file_name = f'MES_{plant}_{mode}_{date_mark}_Chart.png'
        chart_img = os.path.join(save_path, file_name)

        plt.savefig(f"{chart_img}", dpi=100, bbox_inches="tight", pad_inches=0.45)
        plt.close()
        self.image_buffers.append(chart_img)

    # 過濾掉無效數據 (scrap 和 secondgrade)
    def filter_valid_data(self, x_range, y_values):
        filtered_x = [x for i, x in enumerate(x_range) if
                      y_values[i] is not None and not np.isnan(y_values[i])]
        filtered_y = [y for y in y_values if y is not None and not np.isnan(y)]
        return filtered_x, filtered_y

    def rate_chart(self, plant):
        this_start_date = self.this_start_date
        this_end_date = self.this_end_date
        date_mark = self.date_mark
        save_path = self.save_path

        if plant == 'NBR':
            plant_ = 'NBR'
        else:
            plant_ = 'PVC1'

        data1 = self.mach_list

        sql = f"""WITH raw_data AS (
                        SELECT Machine name, WorkDate, 
                               CAST(ScrapQuantity AS FLOAT) AS Scrap, 
                               CAST(FaultyQuantity AS FLOAT) AS SecondGrade,
                               CAST(CountingQty as Float) as sum_qty
                        FROM [MES_OLAP].[dbo].[counting_daily_info_raw]
                        WHERE belong_to between '{this_start_date}' AND '{this_end_date}'
                    )
                    SELECT 
                        name,
                        SUM(Case when Scrap > 0 then Scrap else 0 end) AS scrap,
                        SUM(case when SecondGrade > 0 then SecondGrade else 0 end) AS secondgrade,
                        SUM(case when sum_qty > 0 then sum_qty else 0 end) as sum_qty
                    FROM raw_data
                    WHERE name like '%{plant_}%'
                    GROUP BY name
                    ORDER BY name;
                """
        data2 = vnedc_database().select_sql_dict(sql)
        data2_dict = {item['name']: item for item in data2}
        data = []
        for item in data1:
            name = item['name']
            if name in data2_dict:
                data.append(data2_dict[name])
            else:
                data.append({'name': name, 'scrap': 0, 'secondgrade': 0, 'sum_qty': 0})

        x_labels = [str(item['name']).split('_')[-1] for item in data]
        x_range = range(len(x_labels))

        scrap = [round((item['scrap'] / (item['sum_qty']+item['scrap']+item['secondgrade'])) * 100, 2) if item['sum_qty'] > 0 else 0 for item in data]
        secondgrade = [round((item['secondgrade'] / (item['sum_qty']+item['scrap']+item['secondgrade'])) * 100, 2) if item['sum_qty'] > 0 else 0 for item in
                       data]

        # 過濾 scrap 數據
        filtered_x_scrap, filtered_scrap = self.filter_valid_data(x_range, scrap)

        # 過濾 secondgrade 數據
        filtered_x_secondgrade, filtered_secondgrade = self.filter_valid_data(x_range, secondgrade)

        # Create the figure and subplots
        plt.figure(figsize=(16, 9))
        fig = plt.figure(figsize=(16, 9))
        gs = gridspec.GridSpec(2, 1, height_ratios=[1, 1], hspace=0)  # No space between subplots

        # Subplot for scrap
        y_max1 = 3
        y_ticks = np.arange(0, y_max1, 0.4)  # 分成10個刻度
        y_ticks = y_ticks[:-1]

        ax1 = fig.add_subplot(gs[0])
        ax1.plot(filtered_x_scrap, filtered_scrap, label="廢品率 (%)", marker='o', linestyle='-', color='#ED7D31',
                 linewidth=2)
        ax1.axhline(y=0.8, color='#ED7D31', linestyle='--', linewidth=1.5, label="廢品標準線(0.8)")
        ax1.set_xticks(x_range)
        ax1.set_xticklabels([])  # Hide x-axis labels for the top plot
        # ax1.set_ylabel('廢品率 (%)', fontsize=12, rotation=0)
        ax1.text(-0.1, 0.6, '廢品率 (%)', fontsize=12, rotation=0, ha='center', va='center', transform=ax1.transAxes)
        ax1.set_ylim(0, y_max1)
        ax1.set_yticks(y_ticks)
        ax1.yaxis.set_major_formatter(FuncFormatter(self.add_percent))

        offset = 0.03
        for i, scrap_val in enumerate(filtered_scrap):
            ax1.text(filtered_x_scrap[i], scrap_val+offset, f"{scrap_val:.2f}%", ha='center', va='bottom', fontsize=12,
                     color='#ED7D31')

        # Subplot for secondgrade
        y_max2 = 1.6
        y_ticks = np.arange(0, y_max2, 0.2)  # 分成10個刻度
        y_ticks = y_ticks[:-1]

        ax2 = fig.add_subplot(gs[1])
        ax2.plot(filtered_x_secondgrade, filtered_secondgrade, label="二級品率 (%)", marker='o', linestyle='-',
                 color='#70AD47', linewidth=2)
        ax2.axhline(y=0.2, color='#70AD47', linestyle='--', linewidth=1.5, label="二級品標準線(0.2)")
        ax2.set_xticks(x_range)
        ax2.set_xticklabels(x_labels, rotation=0, ha="center", fontsize=12)
        # ax2.set_ylabel('二級品率 (%)', fontsize=12, rotation=0)
        ax2.text(-0.1, 0.6, '二級品率 (%)', fontsize=12, rotation=0, ha='center', va='center', transform=ax2.transAxes)
        ax2.set_ylim(0, y_max2)
        ax2.set_yticks(y_ticks)
        ax2.yaxis.set_major_formatter(FuncFormatter(self.add_percent))

        offset = 0.03
        for i, secondgrade_val in enumerate(filtered_secondgrade):
            ax2.text(filtered_x_secondgrade[i], secondgrade_val+offset, f"{secondgrade_val:.2f}%", ha='center', va='bottom',
                     fontsize=12, color='#70AD47')

        # Add a title for the entire figure
        fig.suptitle(f"二級品及廢品率 ({plant})", fontsize=20)

        # Add legend for both plots
        handles1, labels1 = ax1.get_legend_handles_labels()
        handles2, labels2 = ax2.get_legend_handles_labels()
        fig.legend(
            handles1 + handles2,
            labels1 + labels2,
            loc='center left',
            fontsize=10,
            title="Note",
            title_fontsize=12,
            bbox_to_anchor=(1.0, 0.5),
            ncol=1
        )

        plt.tight_layout()

        file_name = f'MES_{plant}_{date_mark}_Rate_Chart_Line.png'
        chart_img = os.path.join(save_path, file_name)

        plt.savefig(chart_img, dpi=100, bbox_inches="tight", pad_inches=0.45)
        plt.close()

        self.image_buffers.append(chart_img)

    def chart_table(self, plant, x_labels, secondgrade, scrap):
        date_mark = self.date_mark
        save_path = self.save_path

        table_data = [
            [plant] + x_labels,
            ["廢品(%)"] + scrap,
            ["二級品(%)"] + secondgrade,
        ]

        # 創建圖表
        fig, ax = plt.subplots(figsize=(12, 3))  # 調整尺寸以適配內容
        ax.axis("tight")
        ax.axis("off")

        # 添加表格
        table = ax.table(cellText=table_data, loc="center", cellLoc="center")

        font_path = font_manager.findfont("Microsoft YaHei")
        font_prop = font_manager.FontProperties(fname=font_path, size=14)
        header_color = "#0A4E9B"
        row_colors = ["#E8F3FF", "#FFFFFF"]

        for (row, col), cell in table.get_celld().items():
            cell.set_width(0.1)
            # Header styling
            if row == 0:
                cell.set_facecolor(header_color)
                cell.set_text_props(color="white", weight="bold", fontproperties=font_prop)
            # Row styling
            else:
                cell.set_facecolor(row_colors[row % 2])
                cell.set_text_props(fontproperties=font_prop)
            # Adjust height for all cells
            cell.set_height(0.2)

        plt.tight_layout(rect=[0, 0, 0, 0])
        file_name = f'MES_{plant}_{date_mark}_Rate_Chart_Table.png'
        chart_img = os.path.join(save_path, file_name)
        plt.savefig(chart_img, bbox_inches="tight", dpi=100, pad_inches=0)
        plt.close()

        self.image_buffers.append(chart_img)

    def send_email(self, file_list, image_buffers):
        mode = self.mode
        date_mark = self.date_mark
        this_start_date = self.this_start_date
        logging.info(f"Start to send Email")
        smtp_config, to_emails, admin_emails = self.read_config('mes_weekly_report_mail.config')

        # SMTP Sever config setting
        smtp_server = smtp_config.get('smtp_server')
        smtp_port = int(smtp_config.get('smtp_port', 587))
        smtp_user = smtp_config.get('smtp_user')
        smtp_password = smtp_config.get('smtp_password')
        sender_alias = "GD Report"
        sender_email = smtp_user
        # Mail Info
        msg = MIMEMultipart()
        msg['From'] = f"{sender_alias} <{sender_email}>"
        msg['To'] = ', '.join(to_emails)

        msg['Subject'] = f'[GD Report] {self.year} {self.month_week}達成率報表 {date_mark}'

        # Mail Content
        html = """\
                <html>
                  <body>
                """
        for i in range(len(image_buffers)):
            html += f'<img src="cid:chart_image{i}"><br>'

        html += """\
                  </body>
                </html>
                """

        msg.attach(MIMEText(html, 'html'))

        # Attach Excel
        for file_path in file_list:
            if os.path.exists(file_path):
                with open(file_path, "rb") as attachment:
                    part = MIMEBase("application", "octet-stream")
                    part.set_payload(attachment.read())
                encoders.encode_base64(part)
                part.add_header(
                    "Content-Disposition",
                    f"attachment; filename={os.path.basename(file_path)}"
                )
                msg.attach(part)
            else:
                print(f"File not found: {file_path}")
        logging.info(f"Attached Excel files")

        # Attach Picture
        for idx, image_path in enumerate(image_buffers):
            if os.path.exists(image_path):
                with open(image_path, "rb") as img_file:
                    img = MIMEImage(img_file.read())
                img.add_header("Content-ID", f"<chart_image{idx}>")
                msg.attach(img)
            else:
                logging.info(f"Image not found: {image_path}")
                print(f"Image not found: {image_path}")
        logging.info(f"Attached Picture")

        # Send Email
        try:
            # server = smtplib.SMTP(smtp_server, smtp_port)
            # server.starttls()  # 启用 TLS 加密
            # server.login(smtp_user, smtp_password)  # 登录到 SMTP 服务器
            # server.sendmail(smtp_user, to_emails, msg.as_string())
            # server.quit()

            # 發送郵件（不進行密碼驗證）
            server = smtplib.SMTP(smtp_server, smtp_port)
            server.ehlo()  # 啟動與伺服器的對話
            server.sendmail(smtp_user, to_emails, msg.as_string())
            print("Sent Email Successfully")
        except Exception as e:
            print(f"Sent Email Fail: {e}")
            logging.info(f"Sent Email Fail: {e}")
        finally:
            attachment.close()

    def read_config(self, config_file):
        smtp_config = {}
        to_emails = []
        admin_emails = []

        current_section = None

        with open(config_file, 'r') as file:
            for line in file:
                line = line.strip()

                # Skip empty lines and comments
                if not line or line.startswith('#'):
                    continue

                # Detect section headers
                if line.startswith('[') and line.endswith(']'):
                    current_section = line[1:-1].lower()
                    continue

                # Read lines based on the current section
                if current_section == 'smtp':
                    if '=' in line:
                        key, value = line.split('=', 1)
                        smtp_config[key.strip()] = value.strip()
                elif current_section == 'recipients':
                    to_emails.append(line)
                elif current_section == 'admin_email':
                    admin_emails.append(line)

        return smtp_config, to_emails, admin_emails

    def add_percent(self, y, pos):
        return f"{y:.1f}%"  # 格式化為整數百分比

    def weekly_chart(self, plant):
        save_path = self.save_path

        data1 = self.mach_list

        for machine in data1:
            try:
                this_data = []
                this_rate = []
                x_labels = []
                week_date = []
                unfinish_data = []

                today = datetime.now()

                if machine['name'] in ['VN_GD_PVC1_L01', 'VN_GD_PVC1_L02']:
                    start_date = date(2025, 1, 6)
                    start_week_number = 1
                elif machine['name'] in ['VN_GD_PVC1_L05', 'VN_GD_PVC1_L06']:
                    start_date = date(2024, 12, 2)
                    start_week_number = 49
                else:
                    # Get the starting week (Sep 30)
                    start_date = date(2024, 9, 30)
                    start_week_number = 40

                x_labels, week_date = Utils().generate_previous_weeks_with_dates(self.this_end_date)

                print(machine['name'])
                for i, item in enumerate(week_date):
                    week_name = x_labels[i]
                    if item[0] < start_date:
                        this_data.append(0)
                        this_rate.append(0)
                        unfinish_data.append(0)
                        continue

                    print(f"Week {week_name} - start {item[0]} - end - {item[1]}")

                    # Target只看有做IPQC的部分
                    sql = f"""
							SELECT name,qty, target, (case when target-qty > 0 then target-qty else 0 end) unfinish_qty FROM 
                            (
									SELECT Machine name, SUM(OnlinePacking) AS qty,
									SUM(Target) AS target
									FROM [MES_OLAP].[dbo].[counting_daily_info_raw] c
									JOIN [MES_OLAP].[dbo].[mes_ipqc_data] ipqc on c.Runcard = ipqc.Runcard
									WHERE Machine = '{machine['name']}'
									AND Weight_Value > 0
									AND belong_to BETWEEN '{item[0]}' AND '{item[1]}'
									GROUP BY Machine
							) A
                            """
                    print(sql)
                    rows = vnedc_database().select_sql_dict(sql)
                    if len(rows) == 0:
                        this_data.append(0)
                        this_rate.append(0)
                        unfinish_data.append(0)
                    else:
                        this_data.append(rows[0]['qty'])
                        unfinish_data.append(rows[0]['unfinish_qty'])

                        try:
                            rate = round(rows[0]['qty'] / rows[0]['target'], 3) * 100 if rows[0]['target'] > 0 else 100
                        except Exception as e:
                            rate = 100
                            print(f"{e} at {machine['name']} at week start {item[0]} - end {item[1]}")
                            pass
                        this_rate.append(rate)
                    # x_labels.append(f'W {week}')

                x_range = range(0, len(x_labels) * 2, 2)

                tmp_data = [qty + unfinish for qty, unfinish in zip(this_data, unfinish_data)]

                max_data = max(tmp_data, default=0)
                step_data = 10
                yticks_positions, yticks_labels = Utils().chart_y_label(max_data, step_data)

                max_rate = max(this_rate, default=0)
                rounded_max_rate = (math.ceil(max_rate / 10) * 10)
                rounded_step_rate = 20

                bar_width = 0.6
                plt.figure(figsize=(24, 9))
                fig, ax1 = plt.subplots(figsize=(24, 9))
                this_week_bars = ax1.bar([i for i in x_range], this_data, width=bar_width,
                                          label=f"包裝確認量週累計", align='center', color='#10ba81')
                unfinish_bars = ax1.bar([i for i in x_range], unfinish_data, width=bar_width, bottom=this_data,
                                        label=f"週目標差異",
                                        align='center', color='lightgreen')
                ax1.set_xticks(x_range)
                ax1.set_xticklabels(x_labels, rotation=0, ha="center", fontsize=12)


                for index, bar in enumerate(this_week_bars):
                    height = bar.get_height()
                    unfinish_height = unfinish_bars[index].get_height()
                    if len(str(max_data)) >= 7:
                        show_text = f'{round(height/1000000, 1)}'[:4] if height > 0 else ''
                    elif 4 < len(str(max_data)) < 7:
                        show_text = f'{round(height/10000, 1)}'[:4] if height > 0 else ''
                    ax1.text(
                        bar.get_x() + bar.get_width() / 2,
                        height + unfinish_height,
                        show_text,
                        ha='center', va='bottom', fontsize=12  # Align the text
                    )

                yticks_labels[-1] = ""
                ax1.set_yticks(yticks_positions)
                ax1.set_yticklabels(yticks_labels, fontsize=12)

                ax2 = ax1.twinx()
                sr_achieve_rate = "達成率"

                filtered_data = [(x, rate) for x, rate in zip(x_range, this_rate) if rate != 0]
                x_filtered, this_rate_filtered = zip(*filtered_data)
                this_rate_line = ax2.plot(x_filtered, this_rate_filtered,
                                          label=sr_achieve_rate,
                                          color='#ED7D31', marker='o', linewidth=1.5)
                sr_target = "達成率目標"

                ry_label = "達成率(%)"
                ly_label = "Product (百萬)"

                name = f"各週產出量"
                title = f"\n{plant} ({machine['name']})\n"

                if plant == "NBR":
                    target_rate = 95
                else:
                    target_rate = 98

                yticks_positions = list(range(0, rounded_max_rate + 1 * rounded_step_rate, rounded_step_rate))
                if target_rate not in yticks_positions:
                    yticks_positions.append(target_rate)
                yticks_positions.append(rounded_max_rate + 0.25 * rounded_step_rate)
                yticks_positions = sorted(yticks_positions)

                yticks_labels = [f"{i}" + '%' for i in yticks_positions]
                yticks_labels[-1] = ""
                ax2.set_yticks(yticks_positions)
                ax2.set_yticklabels(yticks_labels)
                ax2.axhline(y=target_rate, color='red', linestyle='--', linewidth=1, label=sr_target)

                ax1.xaxis.set_label_coords(0.975, -0.014)
                ax1.set_ylabel('Product output', fontsize=12)
                ax2.set_ylabel('Archive rates', fontsize=12)

                for i, value in enumerate(this_rate):
                    ax2.text(
                        x_range[i],
                        value + 1.5,
                        f"{value:.1f}%" if value != 0 else '',
                        ha='center', va='bottom', fontsize=12, color='white',
                        bbox=dict(facecolor='#ED7D31', edgecolor='none', boxstyle='round,pad=0.3')
                    )

                handles1, labels1 = ax1.get_legend_handles_labels()
                handles2, labels2 = ax2.get_legend_handles_labels()
                fig.legend(
                    handles1 + handles2,
                    labels1 + labels2,
                    loc='center left',
                    fontsize=12,
                    title="Note",
                    title_fontsize=12,
                    bbox_to_anchor=(1.0, 0.5),
                    ncol=1
                )

                # Customize axes and borders
                ax = plt.gca()  # Get current axes
                ax.spines['top'].set_color('white')
                ax.spines['top'].set_linewidth(1.5)

                ax1.tick_params(axis='y', colors='black')  # Y-axis ticks color
                y_tick_lines = ax1.get_yticklines()
                y_tick_lines[-2].set_visible(False)
                y_tick_lines[-1].set_visible(False)
                ax1.annotate(
                    '',
                    xy=(0, 1.0),
                    xytext=(0, 0.98),
                    xycoords='axes fraction',
                    arrowprops=dict(facecolor='black', arrowstyle='-|>,widthA=0.4,widthB=1.4', linewidth=0.5)
                )
                ax1.set_ylabel(ly_label, labelpad=20, rotation=0)
                ax1.yaxis.set_label_coords(-0.01, 1.01)

                ax2.tick_params(axis='y', colors='black')
                y_tick_lines = ax2.get_yticklines()
                y_tick_lines[-2].set_visible(False)
                y_tick_lines[-1].set_visible(False)
                ax2.annotate(
                    '',
                    xy=(1, 1.0),
                    xytext=(1, 0.98),
                    xycoords='axes fraction',
                    arrowprops=dict(facecolor='black', arrowstyle='-|>,widthA=0.4,widthB=1.4', linewidth=0.5)
                )

                ax2.set_ylabel(ry_label, labelpad=20, rotation=0)
                ax2.yaxis.set_label_coords(1.01, 1.03)

                plt.title(title, fontsize=20)

                plt.tight_layout()

                file_name = f"MES_{machine['name']}_Chart.png"
                # save_path = os.path.join("yearly_output")
                chart_img = os.path.join(save_path, file_name)

                plt.savefig(f"{chart_img}", dpi=100, bbox_inches="tight", pad_inches=0.45)
                plt.close('all')
            except Exception as e:
                print(f"{machine['name']}: {e}")
                pass

    def get_mach_list(self, plant):
        if plant == 'NBR':
            plant_ = 'NBR'
        else:
            plant_ = 'PVC1'

        sql = f"""
                        select name FROM [PMGMES].[dbo].[PMG_DML_DataModelList] 
                        where DataModelTypeId = 'DMT000003' and name like '%{plant_}%' 
                        order by name
                    """
        data = mes_database().select_sql_dict(sql)

        return data

    def main(self):
        for plant in self.plant_name:
            self.mach_list = self.get_mach_list(plant)
            self.generate_chart(plant)
            if self.mode == 'WEEKLY':
                 self.weekly_chart(plant)
            self.rate_chart(plant)
            self.generate_raw_excel(plant)

        self.send_email(self.file_list, self.image_buffers)

    def delete_counting_weekly_info_raw(self, plant, year, month_week):
        mes_olap = mes_olap_database()
        sql = f"""
            delete from [MES_OLAP].[dbo].[counting_weekly_info_raw]
            where Plant = '{plant}' and [Year] = {year} and MonthWeek = '{month_week}'
        """
        mes_olap.execute_sql(sql)

    def insert_counting_weekly_info_raw(self, plant, year, month_week, summary_df):
        mes_olap = mes_olap_database()
        # Only show data which activation not null
        df = summary_df.loc[summary_df["良率"].notna() & (summary_df["良率"] != "")]

        counting_sum = df["點數機數量"].sum()
        separate_sum = df["隔離品數量"].sum()
        target_sum = df["目標產能"].sum()
        onlinePacking_sum = df["包裝確認量"].sum()
        faulty_sum = df["二級品數量"].sum()
        scrap_sum = df["廢品數量"].sum()

        speed_avg = round(df["平均車速"].mean(), 0)
        # activation_avg = round(df["稼動率"].mean(), 3)
        activation_avg = 'null'
        capacity_avg = round(df["產能效率"].mean(), 3)
        yield_avg = round(df["良率"].mean(), 3)

        # oee_avg = round(activation_avg*capacity_avg*yield_avg, 3)
        oee_avg = 'null'

        separate_rate = round(separate_sum / onlinePacking_sum, 3)
        scrap_rate = round(scrap_sum / (counting_sum + faulty_sum + scrap_sum), 3)

        sql = f"""
        Insert into [MES_OLAP].[dbo].[counting_weekly_info_raw](Plant, [Year], MonthWeek, CountingQty, SeparateQty, AvgSpeed, 
        Activation, Capacity, Yield, OEE, SeparateRate, ScrapRate, Target, OnlinePacking)
        Values('{plant}', {year}, '{month_week}', {counting_sum}, {separate_sum}, {speed_avg}, 
        {activation_avg}, {capacity_avg},{yield_avg},{oee_avg},{separate_rate},{scrap_rate},{target_sum},{onlinePacking_sum})
        """
        print(sql)
        mes_olap.execute_sql(sql)

import argparse
from datetime import datetime, timedelta, date

mode = "WEEKLY"

report = mes_weekly_report(mode)
report.main()