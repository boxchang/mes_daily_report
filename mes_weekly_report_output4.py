import sys
import os
curPath = os.path.abspath(os.path.dirname(__file__))
rootPath = os.path.split(curPath)[0]
sys.path.append(rootPath)
import matplotlib
import time
from matplotlib.ticker import FuncFormatter
from openpyxl.utils import get_column_letter
from database import vnedc_database, mes_database, mes_olap_database, lkmes_database, lkmes_olap_database, lkedc_database
from factory import DataControl, ColumnControl, Factory, ConfigObject, SetReportLog
from lib.utils import Utils
import pandas as pd
import matplotlib.pyplot as plt
import logging
import math
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, NamedStyle, Font, Border, Side, PatternFill
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from email.mime.image import MIMEImage
from openpyxl.drawing.image import Image
import matplotlib.gridspec as gridspec
import numpy as np


class mes_weekly_report(object):
    this_start_date = ""
    this_end_date = ""
    last_end_date = ""
    last_start_date = ""
    db = None
    mode = ""
    save_path = ""
    date_mark = ""
    mach_list = ""
    year = None
    week_no = None

    # Define Style
    percent_style = NamedStyle(name='percent_style', number_format='0.00%')
    right_align_style = NamedStyle(name='right_align_style', alignment=Alignment(horizontal='right'))
    center_align_style = NamedStyle(name='center_align_style', alignment=Alignment(horizontal='center'))

    # Define Header
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal='center')
    header_border = Border(bottom=Side(style='thin'))

    def get_week_date_df_fix(self):

        sql = f"""
            SELECT *
            FROM [MES_OLAP].[dbo].[week_date] 
            where  [year] = 2025 and [month] =3 and week_no = 'W1'
             and enable = 1
        """

        raws = self.mes_olap_db.select_sql_dict(sql)

        return raws[0]

    def __init__(self):
        config_file = "mes_weekly_report.config"
        mail_config_file = 'mes_weekly_report_mail.config'

        self.config = ConfigObject(config_file, mail_config_file)

        SetReportLog()

        logging.info(f"Location {self.config.location} ......")

        if self.config.location in "GD":
            self.mes_db = mes_database()
            self.mes_olap_db = mes_olap_database()
            self.vnedc_db = vnedc_database()
        elif self.config.location in "LK":
            self.mes_db = lkmes_database()
            self.mes_olap_db = lkmes_olap_database()
            self.vnedc_db = lkedc_database()
        else:
            self.mes_db = None
            self.mes_olap_db = None
            self.vnedc_db = None

        week_date = Utils().get_week_date_dist(self.mes_olap_db)
        # week_date = self.get_week_date_df_fix()
        self.this_end_date = week_date['end_date']
        self.this_start_date = week_date['start_date']
        self.year = week_date['year']
        self.week_no = week_date['week_no']
        self.week_no = str(week_date['week_no'])
        fold_name = str(week_date['year']) + str(week_date['month']).zfill(2) + 'W' + str(week_date['week_no']).zfill(2)
        save_path = os.path.join("weekly_output", fold_name)

        self.save_path = save_path
        # Check folder to create
        if not os.path.exists(save_path):
            os.makedirs(save_path)

        tmp_start_date = datetime.strptime(self.this_start_date, "%Y-%m-%d").strftime("%m%d")
        tmp_end_date = datetime.strptime(self.this_end_date, "%Y-%m-%d").strftime("%m%d")
        self.date_mark = "{start_date}_{end_date}".format(start_date=tmp_start_date,
                                                          end_date=tmp_end_date)

    def main(self):
        mes_db = self.mes_db
        mes_olap_db = self.mes_olap_db
        vnedc_db = self.vnedc_db

        year = self.year
        week_no = self.week_no
        this_start_date = self.this_start_date
        this_end_date = self.this_end_date
        date_mark = self.date_mark
        save_path = self.save_path

        location = self.config.location
        plants = self.config.plants
        hour_output_limit = self.config.hour_output_limit
        fix_mode = self.config.fix_mode
        report_font = self.config.report_font

        file_list = []
        image_buffers = []

        print(f"Plant: {plants}, Hour Output Limit: {hour_output_limit}, Fix Mode: {fix_mode}")

        for plant in plants:

            logging.info(f"{plant} start running......")
            dr = WeeklyReport(mes_db, mes_olap_db, vnedc_db, location, plant,
                              year, week_no, this_start_date, this_end_date, date_mark,
                              hour_output_limit,
                              report_font, save_path, fix_mode, logging)

            logging.info(f"{plant} Get Machine List......")
            dr.mach_list = dr.get_mach_list(plant)

            logging.info(f"{plant} generate_main_df......")
            machine_groups, summary_chart_dict, scrap_chart_dict = dr.generate_main_df()

            logging.info(f"{plant} sorting_data......")
            summary_df, all_mach_sum_df = dr.sorting_data(machine_groups)

            logging.info(f"{plant} validate_data......")
            dr.validate_data()

            logging.info(f"{plant} weekly_chart......")
            dr.weekly_chart(self.save_path)

            # Generate Excel file
            logging.info(f"{plant} generate_excel......")
            file_name = f'MES_{plant}_Weekly_Report_{date_mark}.xlsx'
            excel_file = os.path.join(self.save_path, file_name)
            with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
                logging.info(f"{plant} generate_summary_excel......")
                dr.generate_summary_excel(writer, summary_df)  # Summary

                for machine_name, machine_df in machine_groups:
                    logging.info(f"{plant} generate_raw_excel......")
                    dr.generate_raw_excel(writer, machine_df, machine_name)  # 機台明細

                logging.info(f"{plant} generate_ipqc_ng_data_excel......")
                dr.generate_ipqc_ng_data_excel(writer, all_mach_sum_df)  # IPQC異常分析

                logging.info(f"{plant} delete_counting_weekly_info_raw......")
                dr.delete_counting_weekly_info_raw()
                logging.info(f"{plant} insert_counting_weekly_info_raw......")
                dr.insert_counting_weekly_info_raw(summary_df)
                logging.info(f"{plant} generate_12aspect_output_excel......")
                dr.generate_12aspect_output_excel(writer)  # 週累計

                logging.info(f"{plant} generate_cosmetic_data......")
                cosmetic_df = dr.generate_cosmetic_data()
                logging.info(f"{plant} generate_12aspect_cosmetic_excel......")
                dr.generate_12aspect_cosmetic_excel(writer, cosmetic_df)  # 外觀
                logging.info(f"{plant} generate_12aspect_cosmetic_summary_excel......")
                dr.generate_12aspect_cosmetic_summary_excel(writer)  # 外觀週累計

            if os.path.exists(excel_file):
                file_list.append({'file_name': file_name, 'excel_file': excel_file})

            # Generate Chart
            image_buffer = dr.generate_chart(summary_chart_dict)
            image_buffers.append(image_buffer)

            if 'PVC' in plant:
                image_buffer = dr.rate_chart(scrap_chart_dict)
            if 'NBR' in plant:
                image_buffer = dr.rate_chart2(scrap_chart_dict)
            image_buffers.append(image_buffer)

        if not fix_mode:
            logging.info(f"{self.config.location} send_email......")
            subject = f'[{self.config.location} Report] {self.year} 第{self.week_no}週達成率報表 {date_mark}'
            dr.send_email(self.config, subject, file_list, image_buffers, dr.msg_list, dr.error_list)



class WeeklyReport(Factory):
    file_list = []
    error_list = []
    msg_list = []
    mach_list = []


    def __init__(self, mes_db, mes_olap_db, vnedc_db, location, plant,
                 year, week_no, report_date1, report_date2, date_mark,
                 hour_output_limit, report_font, save_path, fix_mode, logging):
        self.location = location
        self.plant = plant
        self.year = year
        self.week_no = week_no
        self.report_date1 = report_date1
        self.report_date2 = report_date2
        self.report_date3 = (datetime.strptime(report_date2, "%Y-%m-%d") + timedelta(days=1)).strftime("%Y-%m-%d")
        self.hour_output_limit = int(hour_output_limit) if hour_output_limit else 1000
        self.report_font = report_font
        self.mes_db = mes_db
        self.mes_olap_db = mes_olap_db
        self.vnedc_db = vnedc_db
        self.logging = logging
        self.date_mark = date_mark
        self.save_path = save_path
        self.fix_mode = fix_mode

        if 'GD' in location:
            self.capacity_target = 0.99
        elif 'LK' in location:
            self.capacity_target = 0.97

        if 'NBR' in plant:
            self.plant_ = 'NBR'
            self.scrap_target = 0.009
        elif 'PVC' in plant:
            self.plant_ = 'PVC1'
            self.scrap_target = 0.0035

    def get_mach_list(self):
        start_time = time.time()

        sql = f"""
                        select name FROM [PMGMES].[dbo].[PMG_DML_DataModelList] 
                        where DataModelTypeId = 'DMT000003' and name like '%{self.plant_}%' 
                        order by name
                    """
        data = self.mes_db.select_sql_dict(sql)

        end_time = time.time()
        elapsed_time = end_time - start_time
        self.logging.info(f"Time taken: {elapsed_time:.2f} seconds.")

        return data

    # 離型不良率
    def get_dmf_rate(self):
        if "NBR" in self.plant:
            sql = f"""
            SELECT 
                FORMAT(CreationTime, 'yyyy-MM-dd') WorkDate,cd.MES_MACHINE Machine, cd.LINE Line, CAST(DATEPART(hour, CreationTime) as INT) Period, 
                CASE 
                    WHEN SUM(ModelQty2) = 0 THEN 0
                    ELSE ROUND(
                        (SUM(OverShortQty2) + SUM(OverLongQty2)) / SUM(ModelQty2), 
                        4
                    )
                END AS DMF_Rate
            FROM 
                [PMG_DEVICE].[dbo].[COUNTING_DATA] c
            JOIN 
                [PMG_DEVICE].[dbo].[COUNTING_DATA_MACHINE] cd 
                ON c.MachineName = cd.COUNTING_MACHINE
            WHERE 
                CreationTime BETWEEN 
                    CONVERT(DATETIME, '{self.report_date1} 06:00:00', 120) 
                    AND 
                    CONVERT(DATETIME, '{self.report_date3} 05:59:59', 120)
            GROUP BY 
                FORMAT(CreationTime, 'yyyy-MM-dd'),MES_MACHINE,LINE,CAST(DATEPART(hour, CreationTime) as INT)
            """

        elif "PVC" in self.plant:
            sql = f"""
            SELECT FORMAT(CreationTime, 'yyyy-MM-dd') WorkDate,cd.MES_MACHINE Machine, cd.LINE Line, CAST(DATEPART(hour, CreationTime) as INT) Period, CASE 
                    WHEN SUM(ModelQty2) = 0 THEN 0
                    ELSE ROUND(SUM(Qty2) / SUM(ModelQty2), 4)
                END AS DMF_Rate
              FROM [PMG_DEVICE].[dbo].[PVC_GRM_DATA] g
              JOIN [PMG_DEVICE].[dbo].[COUNTING_DATA_MACHINE] cd on g.MachineName = cd.COUNTING_MACHINE
              where CreationTime between CONVERT(DATETIME, '{self.report_date1} 06:00:00', 120) and CONVERT(DATETIME, '{self.report_date3} 05:59:59', 120)
              group by FORMAT(CreationTime, 'yyyy-MM-dd'),cd.MES_MACHINE, cd.LINE, CAST(DATEPART(hour, CreationTime) as INT) 

            """
        rows = self.mes_db.select_sql_dict(sql)

        df = pd.DataFrame(rows)
        df.loc[df['DMF_Rate'] >= 0.95, 'DMF_Rate'] = 0

        return df

    def generate_main_df(self):
        start_time = time.time()

        this_start_date = self.report_date1
        this_end_date = self.report_date2

        # region Counting Machine Data
        sql = f"""SELECT WorkDate, Machine, Line, Shift, WorkOrder, PartNo, ProductItem, StandardAQL, InspectedAQL,
                    Period, MaxSpeed, MinSpeed, AvgSpeed, StdSpeed, CountingQty, OnlinePacking, WIPPacking, Target, ScrapQuantity, FaultyQuantity, RunTime, StopTime, 60 as AllTime, wd.week_no as Week_No,
                    Tensile_Value,Tensile_Limit,Tensile_Status,Elongation_Value,Elongation_Limit,Elongation_Status,
                    Roll_Value,Roll_Limit,Roll_Status,Cuff_Value,Cuff_Limit,Cuff_Status,Palm_Value,Palm_Limit,Palm_Status,
                    Finger_Value,Finger_Limit,Finger_Status,FingerTip_Value,FingerTip_Limit,FingerTip_Status,
                    Length_Value,Length_Limit,Length_Status, Weight_Value, Weight_Limit, IsolationQty,
                    CASE WHEN Weight_Defect IS NULL THEN NULL WHEN Weight_Defect = 'LL2' THEN 'NG' ELSE 'PASS' END AS Weight_Light, 
                    CASE WHEN Weight_Defect IS NULL THEN NULL WHEN Weight_Defect = 'LL1' THEN 'NG' ELSE 'PASS' END AS Weight_Heavy,
                    Width_Value,Width_Limit,Width_Status,
                    Pinhole_Value,Pinhole_Limit,Pinhole_Status
                    FROM [MES_OLAP].[dbo].[counting_hourly_info_raw] c
                    left join [MES_OLAP].[dbo].[week_date] wd
                    on c.Week_No = wd.week_no
                    LEFT JOIN [MES_OLAP].[dbo].[mes_ipqc_data] ipqc on c.Runcard = ipqc.Runcard
                    where c.Year = {self.year} and c.Week_No = '{self.week_no}'  
                    and c.Machine like'%{self.plant}%'
                    and not (WorkOrder != '' and InspectedAQL ='')
                    --and StandardAQL is not Null and InspectedAQL is not null 
                    order by Machine, WorkDate, Cast(Period as Int), Line
                    """

        data = self.mes_olap_db.select_sql_dict(sql)
        df = pd.DataFrame(data)

        dmf_rate_df = self.get_dmf_rate()

        df = pd.merge(df, dmf_rate_df, on=['WorkDate', 'Machine', 'Line', 'Period'], how='left')

        # 設定IPQC欄位判斷條件
        df['IPQC'] = df.apply(lambda row: "" if pd.isna(row['WorkOrder']) or row['WorkOrder'] == ""
        else ('NG' if 'NG' in row[['Tensile_Status', 'Elongation_Status', 'Roll_Status', 'Cuff_Status',
                                   'Palm_Status', 'Finger_Status', 'FingerTip_Status', 'Length_Status',
                                   'Weight_Heavy', 'Weight_Light', 'Width_Status', 'Pinhole_Status']].values
              else 'PASS'), axis=1)

        df.loc[df['Weight_Value'].isna() | (df['Weight_Value'] == 0), 'Target'] = 0
        df.loc[df['Weight_Value'].isna() | (df['Weight_Value'] == 0), 'CountingQty'] = 0

        df['Period'] = df['Period'].apply(lambda x: f"{int(x):02}:00")
        df['Week_No'] = df['Week_No'].astype(str).apply(lambda x: 'W' + x if not str(x).startswith('W') else x)
        df["Weight_Value"] = df["Weight_Value"].apply(lambda x: f"{x:.2f}")
        df.loc[df['Weight_Value'].isna() | (df['Weight_Value'] == 0), 'Weight_Value'] = ''

        machine_groups_df = df.groupby('Machine')
        # endregion

        # region Summary Achievement Chart
        # Target只看有做IPQC的部分
        sql = f"""
                  SELECT pdd.Name name,
                              SUM(ISNULL( OnlinePacking, 0) + ISNULL(WIPPacking, 0)) AS this_time,
                              SUM(ISNULL( Target, 0)) AS target_this_time,
                              SUM(ISNULL( Target, 0)) - SUM(ISNULL( OnlinePacking, 0) + ISNULL(WIPPacking, 0)) AS this_unfinish
                  FROM [PMGMES].[dbo].[PMG_DML_DataModelList] pdd
                  LEFT JOIN (
                    SELECT chir.*
                       FROM [MES_OLAP].[dbo].[counting_hourly_info_raw] chir
                      INNER JOIN [MES_OLAP].[dbo].[mes_ipqc_data] ipqc ON ipqc.Runcard = chir.Runcard
                    WHERE 1 = 1
                         AND chir.belong_to BETWEEN '{this_start_date}' AND '{this_end_date}'
                         AND NOT (WorkOrder != '' AND InspectedAQL ='')
						 AND  Weight_Value > 0 AND CountingQty > 100
                  ) t1 ON pdd.Name = t1.Machine
                WHERE pdd.DataModelTypeId = 'DMT000003'
                      AND pdd.Name LIKE '%{self.plant}%'
                GROUP BY pdd.Name
                ORDER BY pdd.Name      
                        """
        summary_chart_dict = self.mes_olap_db.select_sql_dict(sql)
        # endregion

        # region Separation
        sql = f"""WITH raw_data AS (
                                SELECT Machine name, WorkDate, 
                                       CAST(ScrapQuantity AS FLOAT) AS Scrap, 
                                       CAST(FaultyQuantity AS FLOAT) AS SecondGrade,
                                       OnlinePacking+WIPPacking+ScrapQuantity+FaultyQuantity as sum_qty
                                FROM [MES_OLAP].[dbo].[counting_hourly_info_raw]
                                WHERE belong_to between '{this_start_date}' AND '{this_end_date}'
                            )
                            SELECT 
                                name,
                                SUM(Case when Scrap > 0 then Scrap else 0 end) AS scrap,
                                SUM(case when SecondGrade > 0 then SecondGrade else 0 end) AS secondgrade,
                                SUM(case when sum_qty > 0 then sum_qty else 0 end) as sum_qty
                            FROM raw_data
                            WHERE name like '%{self.plant_}%'
                            GROUP BY name
                            ORDER BY name;
                        """
        data2 = self.mes_olap_db.select_sql_dict(sql)
        scrap_chart_dict = {item['name']: item for item in data2}
        # endregion

        end_time = time.time()
        elapsed_time = end_time - start_time
        self.logging.info(f"Time taken: {elapsed_time:.2f} seconds.")

        return machine_groups_df, summary_chart_dict, scrap_chart_dict

    def generate_cosmetic_data(self):
        start_time = time.time()

        sql = f"""
                      WITH Customer AS (
                      SELECT CAST(CAST(KUNNR AS BIGINT) AS VARCHAR(50)) customer_code,SORTL customer_name,[VKBUR] SalePlaceCode
                      FROM [PMG_SAP].[dbo].[ZKNA1] where VKBUR in ('6100', '6200', '6300', '7000')
                      )
                                      
                      SELECT counting.Runcard runcard,counting.belong_to,counting.Machine,counting.Line,counting.Shift,counting.WorkOrder,counting.PartNo,counting.ProductItem,SalePlaceCode,counting.Period
                      FROM [MES_OLAP].[dbo].[counting_hourly_info_raw] counting
                      LEFT JOIN [MES_OLAP].[dbo].[mes_ipqc_data] ipqc on ipqc.Runcard = counting.Runcard
                      LEFT JOIN Customer cus on cus.customer_code = counting.CustomerCode
                      where counting.Year = {self.year} and counting.Week_No = '{self.week_no}'
                      and Machine like '%{self.plant}%'
                      and (OnlinePacking > 0 or WIPPacking > 0)
                      order by Machine, WorkDate, Cast(Period as Int)

                    """
        data = self.mes_olap_db.select_sql_dict(sql)
        df = pd.DataFrame(data)

        weight_sql = f"""
                      WITH Customer AS (
                      SELECT CAST(CAST(KUNNR AS BIGINT) AS VARCHAR(50)) customer_code,SORTL customer_name,[VKBUR] SalePlaceCode
                      FROM [PMG_SAP].[dbo].[ZKNA1] where VKBUR in ('6100', '6200', '6300', '7000')
                      )        
                
                    SELECT 
                        ipqc.Runcard AS runcard, 
                        CAST(SalePlaceCode AS VARCHAR) + Weight_Defect AS defect_code, 
                        1 AS qty
                    FROM [MES_OLAP].[dbo].[counting_hourly_info_raw] counting
                    LEFT JOIN [MES_OLAP].[dbo].[mes_ipqc_data] ipqc ON ipqc.Runcard = counting.Runcard
                    LEFT JOIN Customer cus ON cus.customer_code = counting.CustomerCode
                    WHERE counting.Year = {self.year}
                      AND counting.Week_No = '{self.week_no}'
                      AND Weight_Status = 'NG'
                      AND Machine LIKE '%{self.plant}%'
                      AND InspectedAQL IS NOT NULL
                    
                    UNION
                    
                    -- 補上 SalePlaceCode 固定值與 defect_code LL1/LL2 的組合（當沒資料時也強制出現）
                    SELECT 
                        NULL AS runcard, 
                        CAST(sale_place_code AS VARCHAR) + defect_code AS defect_code,
                        0 AS qty  -- 補齊用的資料，數量為 0
                    FROM (
                        VALUES 
                            (6100, 'LL1'),
                            (6100, 'LL2'),
                            (6200, 'LL1'),
                            (6200, 'LL2'),
                            (6300, 'LL1'),
                            (6300, 'LL2'),
                            (7000, 'LL1'),
                            (7000, 'LL2')
                    ) AS supplement(sale_place_code, defect_code)
                    WHERE NOT EXISTS (
                        SELECT 1
                        FROM [MES_OLAP].[dbo].[counting_hourly_info_raw] counting
                        LEFT JOIN [MES_OLAP].[dbo].[mes_ipqc_data] ipqc ON ipqc.Runcard = counting.Runcard
                        LEFT JOIN Customer cus ON cus.customer_code = counting.CustomerCode
                        WHERE counting.Year = {self.year}
                          AND counting.Week_No = '{self.week_no}'
                          AND Weight_Status = 'NG'
                          AND Machine LIKE '%{self.plant}%'
                          AND InspectedAQL IS NOT NULL
                          AND CAST(SalePlaceCode AS VARCHAR) + Weight_Defect = CAST(sale_place_code AS VARCHAR) + defect_code
                    )

                      --6100 美
                      --6200 歐
                      --6300 日
                      --LL1 過重
                      --LL2 過輕
                    """
        weight_data = self.mes_olap_db.select_sql_dict(weight_sql)
        weight_df = pd.DataFrame(weight_data)

        weight_df = weight_df.pivot(index="runcard", columns="defect_code", values="qty").reset_index()

        fixed_columns = ['runcard', '6100LL1', '6100LL2', '6200LL1', '6200LL2', '6300LL1', '6300LL2', '7000LL1',
                         '7000LL2']

        for col in fixed_columns:
            if col not in weight_df.columns:
                weight_df[col] = np.nan

        # 按照 target_columns 的順序排列欄位
        weight_df = weight_df[fixed_columns]

        cosmetic_sql = f"""
                     SELECT r.runcard, d.defect_code, sum(qty) cosmetic_qty,  max(cos.cosmetic_inspect_qty) cosmetic_inspect_qty
                      FROM [MES_OLAP].[dbo].[counting_hourly_info_raw] r
                      LEFT JOIN [MES_OLAP].[dbo].[mes_ipqc_cosmetic_data] cos on r.Runcard = cos.runcard
                      LEFT JOIN [MES_OLAP].[dbo].[mes_defect_define] d on d.defect_code = cos.defect_code
                      where r.Year = {self.year} and r.Week_No = '{self.week_no}' and r.runcard <> ''
                      group by r.runcard, defect_level, d.defect_type, d.defect_code, desc2
                    """
        cosmetic_data = self.mes_olap_db.select_sql_dict(cosmetic_sql)
        cosmetic_sample_df = pd.DataFrame(cosmetic_data)

        cosmetic_summary_df = cosmetic_sample_df.groupby('runcard', as_index=False).agg({
            'cosmetic_qty': 'max',
            'cosmetic_inspect_qty': 'max'
        }).copy()
        cosmetic_detail_df = cosmetic_sample_df.pivot(index="runcard", columns="defect_code",
                                                      values="cosmetic_qty").reset_index().copy()

        # 計算針孔Defect Code加總
        pinhole_sql = f"""
                    SELECT r.runcard, d.defect_code, sum(qty) sum_qty
                      FROM [MES_OLAP].[dbo].[counting_hourly_info_raw] r
                      JOIN [MES_OLAP].[dbo].[mes_ipqc_pinhole_data] cos on r.Runcard = cos.runcard
                      JOIN [MES_OLAP].[dbo].[mes_defect_define] d on d.defect_code = cos.defect_code
                      where r.Year = {self.year} and r.Week_No = '{self.week_no}'
                      group by r.runcard, defect_level, d.defect_code, desc2
                    """
        pinhole_data = self.mes_olap_db.select_sql_dict(pinhole_sql)
        pinhole_df = pd.DataFrame(pinhole_data)
        pinhole_pivot_df = pinhole_df.pivot(index="runcard", columns="defect_code", values="sum_qty").reset_index()
        pinhole_pivot_df['Pinhole'] = pinhole_pivot_df.iloc[:, 1:-1].notna().any(axis=1).astype(int)

        # 計算針孔樣本數
        pinhole_sample_sql = f"""
                    SELECT r.Id runcard,WorkCenterTypeName, AQL AS WO_AQL
                      FROM [PMGMES].[dbo].[PMG_MES_RunCard] r
                      JOIN [PMGMES].[dbo].[PMG_MES_RunCard_IPQCInspectIOptionMapping] m on r.Id = m.RunCardId
                      JOIN [PMGMES].[dbo].[PMG_MES_WorkOrder] w on r.WorkOrderId = w.Id
                      where GroupType = 'Pinhole' and r.InspectionDate between '{self.report_date1}' and '{self.report_date2}'
                    """
        pinhole_sample_data = self.mes_db.select_sql_dict(pinhole_sample_sql)
        pinhole_sample_df = pd.DataFrame(pinhole_sample_data)

        if "PVC" in self.plant:
            pinhole_sample_df['Pinhole_Sample'] = 25
        elif "NBR" in self.plant:
            pinhole_sample_df['Pinhole_Sample'] = np.where(pinhole_sample_df['WO_AQL'] == '1.0', 50, 25)

        df = pd.merge(df, weight_df, on=['runcard'], how='left')
        df = pd.merge(df, cosmetic_summary_df, on=['runcard'], how='left')
        df = pd.merge(df, pinhole_pivot_df, on=['runcard'], how='left')
        df = pd.merge(df, cosmetic_detail_df, on="runcard", how="left")
        df = pd.merge(df, pinhole_sample_df, on="runcard", how="left")
        df = df.fillna('')

        end_time = time.time()
        elapsed_time = end_time - start_time
        self.logging.info(f"Time taken: {elapsed_time:.2f} seconds.")

        return df

    def generate_raw_excel(self, writer, df, machine_name):
        start_time = time.time()

        # region 1. Column Define
        font = Font(name=self.report_font, size=10, bold=False)
        sheet = DataControl(self.fix_mode)
        sheet.add(ColumnControl('WorkDate', 'center', '@', '作業日期', font, hidden=False, width=18))
        sheet.add(ColumnControl('Machine', 'center', '@', '機台號', font, hidden=False, width=18))
        sheet.add(ColumnControl('Line', 'center', '@', '線別', font, hidden=False, width=9))
        sheet.add(ColumnControl('Shift', 'center', '@', '班別', font, hidden=False, width=9))
        sheet.add(ColumnControl('WorkOrder', 'center', '@', '工單', font, hidden=False, width=17))
        sheet.add(ColumnControl('PartNo', 'center', '@', '料號', font, hidden=False, width=17))
        sheet.add(ColumnControl('ProductItem', 'left', '@', '品項', font, hidden=False))
        sheet.add(ColumnControl('StandardAQL', 'center', '@', '工單AQL', font, hidden=False, width=9))
        sheet.add(ColumnControl('InspectedAQL', 'center', '@', '量測AQL', font, hidden=False, width=9))
        sheet.add(ColumnControl('Period', 'center', '@', 'Period', font, hidden=False, width=8))
        sheet.add(ColumnControl('MaxSpeed', 'right', '0', '車速(最高)', font, hidden=False, width=9.5))
        sheet.add(ColumnControl('MinSpeed', 'right', '0', '車速(最低)', font, hidden=False, width=9.5))
        sheet.add(ColumnControl('AvgSpeed', 'right', '0', '車速(平均)', font, hidden=False, width=9.5))
        sheet.add(ColumnControl('StdSpeed', 'right', '0', '標準車速', font, hidden=False, width=9.5))
        sheet.add(ColumnControl('CountingQty', 'right', '#,##0', '產量(加總)', font, hidden=False, width=11))
        sheet.add(ColumnControl('OnlinePacking', 'right', '#,##0', '包裝確認量', font, hidden=False, width=11))
        sheet.add(ColumnControl('WIPPacking', 'right', '#,##0', '半成品入庫量', font, hidden=False, width=11))
        sheet.add(ColumnControl('Target', 'center', '#,##0', '目標產能', font, hidden=False, width=11,
                                comment="60 * (標準車速上限/節距調整值)"))
        sheet.add(ColumnControl('ScrapQuantity', 'right', '#,##0', '廢品數量', font, hidden=False, width=11))
        sheet.add(ColumnControl('FaultyQuantity', 'right', '#,##0', '二級品數量', font, hidden=False, width=11))
        sheet.add(ColumnControl('RunTime', 'right', '0', '實際運轉時間', font, hidden=False, width=9.5))
        sheet.add(ColumnControl('StopTime', 'right', '0', '停機時間', font, hidden=False, width=9.5))
        sheet.add(ColumnControl('AllTime', 'right', '0', '可運轉時間', font, hidden=False, width=9.5))
        sheet.add(ColumnControl('Tensile_Value', 'right', '0', '抗拉強度值', font, hidden=False, width=10, level=1))
        sheet.add(ColumnControl('Tensile_Limit', 'center', '@', '抗拉強度值', font, hidden=False, width=10, level=1))
        sheet.add(ColumnControl('Tensile_Status', 'center', '@', '抗拉強度結果', font, hidden=False, width=10, level=1))
        sheet.add(ColumnControl('Elongation_Value', 'right', '0', '伸長率值', font, hidden=False, width=10, level=1))
        sheet.add(ColumnControl('Elongation_Limit', 'center', '@', '伸長率上下限', font, hidden=False, width=10, level=1))
        sheet.add(ColumnControl('Elongation_Status', 'center', '@', '伸長率結果', font, hidden=False, width=10, level=1))
        sheet.add(ColumnControl('Roll_Value', 'right', '0', '卷唇厚度值', font, hidden=False, width=10, level=1))
        sheet.add(ColumnControl('Roll_Limit', 'center', '@', '卷唇厚度上下限', font, hidden=False, width=10, level=1))
        sheet.add(ColumnControl('Roll_Status', 'center', '@', '卷唇厚度結果', font, hidden=False, width=10, level=1))
        sheet.add(ColumnControl('Cuff_Value', 'right', '0', '袖厚度值', font, hidden=False, width=10, level=1))
        sheet.add(ColumnControl('Cuff_Limit', 'center', '@', '袖厚度上下限', font, hidden=False, width=10, level=1))
        sheet.add(ColumnControl('Cuff_Status', 'center', '@', '袖厚度結果', font, hidden=False, width=10, level=1))
        sheet.add(ColumnControl('Palm_Value', 'right', '0', '掌厚度值', font, hidden=False, width=10, level=1))
        sheet.add(ColumnControl('Palm_Limit', 'center', '@', '掌厚度上下限', font, hidden=False, width=10, level=1))
        sheet.add(ColumnControl('Palm_Status', 'center', '@', '掌厚度結果', font, hidden=False, width=10, level=1))
        sheet.add(ColumnControl('Finger_Value', 'right', '0', '指厚度值', font, hidden=False, width=10, level=1))
        sheet.add(ColumnControl('Finger_Limit', 'center', '@', '指厚度上下限', font, hidden=False, width=10, level=1))
        sheet.add(ColumnControl('Finger_Status', 'center', '@', '指厚度結果', font, hidden=False, width=10, level=1))
        sheet.add(ColumnControl('FingerTip_Value', 'right', '0', '指尖厚度值', font, hidden=False, width=10, level=1))
        sheet.add(ColumnControl('FingerTip_Limit', 'center', '@', '指尖厚度上下限', font, hidden=False, width=10, level=1))
        sheet.add(ColumnControl('FingerTip_Status', 'center', '@', '指尖厚度結果', font, hidden=False, width=10, level=1))
        sheet.add(ColumnControl('Length_Value', 'right', '0', '長度值', font, hidden=False, width=10, level=1))
        sheet.add(ColumnControl('Length_Limit', 'center', '@', '長度上下限', font, hidden=False, width=10, level=1))
        sheet.add(ColumnControl('Length_Status', 'center', '@', '長度結果', font, hidden=False, width=10, level=1))
        sheet.add(ColumnControl('Weight_Value', 'right', '0', '重量值', font, hidden=False, width=10, level=1))
        sheet.add(ColumnControl('Weight_Limit', 'center', '@', '重量上下限', font, hidden=False, width=10, level=1))
        sheet.add(ColumnControl('Weight_Light', 'center', '@', '超輕檢驗', font, hidden=False, width=10, level=1))
        sheet.add(ColumnControl('Weight_Heavy', 'center', '@', '超重檢驗', font, hidden=False, width=10, level=1))
        sheet.add(ColumnControl('Width_Value', 'right', '0', '寬度值', font, hidden=False, width=10, level=1))
        sheet.add(ColumnControl('Width_Limit', 'center', '@', '寬度上下限', font, hidden=False, width=10, level=1))
        sheet.add(ColumnControl('Width_Status', 'center', '@', '寬度結果', font, hidden=False, width=10, level=1))
        sheet.add(ColumnControl('Pinhole_Value', 'right', '0', '針孔值', font, hidden=False, width=10, level=1))
        sheet.add(ColumnControl('Pinhole_Limit', 'center', '@', '針孔上下限', font, hidden=False, width=10, level=1))
        sheet.add(ColumnControl('Pinhole_Status', 'center', '@', '針孔結果', font, hidden=False, width=10, level=1))
        sheet.add(ColumnControl('IPQC', 'center', '@', 'IPQC', font, hidden=False, width=10))
        sheet.add(ColumnControl('IsolationQty', 'right', '#,##0', '隔離品數量', font, hidden=False, width=11))

        header_columns = sheet.header_columns
        column_letter = sheet.column_letter
        selected_columns = [col for col in sheet.column_names if col in df.columns]
        # endregion

        # region 2. DataFrame convert to Excel
        copy_df = df[selected_columns].copy()
        copy_df = copy_df.fillna("")

        copy_df.rename(columns=header_columns, inplace=True)

        namesheet = str(machine_name).split('_')[-1]

        save_path = self.save_path
        file_name = f"MES_{machine_name}_Chart.png"
        chart_img = os.path.join(save_path, file_name)

        header_row = 0
        data_start_row = 1

        # Write data to the Excel sheet
        copy_df.to_excel(writer, sheet_name=namesheet, index=False, startrow=header_row)

        workbook = writer.book
        worksheet = writer.sheets.get(namesheet)

        # Freeze the first row
        # worksheet.freeze_panes = worksheet['A'+str(data_start_row+1)]
        worksheet.freeze_panes = worksheet['A2']

        if not worksheet:
            worksheet = workbook.add_worksheet(namesheet)

        sheet.apply_formatting(worksheet)
        # endregion

        # region 3. Customize
        # # 設置欄讓其可以折疊/展開
        worksheet.column_dimensions.group(column_letter['Tensile_Value'], column_letter['Pinhole_Status'],
                                          hidden=True)

        try:
            img = Image(chart_img)
            img.height = 6 * 96
            img.width = 16 * 96
            img.anchor = 'A' + str(len(df) + 5)
            worksheet.add_image(img)
        except:
            print('No counting machine data yet!')
            pass
        # endregion

        end_time = time.time()
        elapsed_time = end_time - start_time
        self.logging.info(f"Time taken: {elapsed_time:.2f} seconds.")


    def ipqc_ng_data(self, df):
        # IPQC NG Data
        # Only check IPQC data
        df['Weight_Value'] = pd.to_numeric(df['Weight_Value'], errors='coerce')
        df = df[df['Weight_Value'] > 0].copy()

        # 計算 "NG" 的次數
        df["Length_NG_Count"] = (df["OnlinePacking"] + df["WIPPacking"]).where(df["Length_Status"] == "NG", 0)
        df["Width_NG_Count"] = (df["OnlinePacking"] + df["WIPPacking"]).where(df["Width_Status"] == "NG", 0)
        df["Weight_Light_NG_Count"] = (df["OnlinePacking"] + df["WIPPacking"]).where(df["Weight_Light"] == "NG", 0)
        df["Weight_Heavy_NG_Count"] = (df["OnlinePacking"] + df["WIPPacking"]).where(df["Weight_Heavy"] == "NG", 0)
        df["Pinhole_NG_Count"] = (df["OnlinePacking"] + df["WIPPacking"]).where(df["Pinhole_Status"] == "NG", 0)
        df["Tensile_NG_Count"] = (df["OnlinePacking"] + df["WIPPacking"]).where(df["Tensile_Status"] == "NG", 0)
        df["Elongation_NG_Count"] = (df["OnlinePacking"] + df["WIPPacking"]).where(df["Elongation_Status"] == "NG", 0)
        df["Roll_NG_Count"] = (df["OnlinePacking"] + df["WIPPacking"]).where(df["Roll_Status"] == "NG", 0)
        df["Cuff_NG_Count"] = (df["OnlinePacking"] + df["WIPPacking"]).where(df["Cuff_Status"] == "NG", 0)
        df["Palm_NG_Count"] = (df["OnlinePacking"] + df["WIPPacking"]).where(df["Palm_Status"] == "NG", 0)
        df["Finger_NG_Count"] = (df["OnlinePacking"] + df["WIPPacking"]).where(df["Finger_Status"] == "NG", 0)
        df["FingerTip_NG_Count"] = (df["OnlinePacking"] + df["WIPPacking"]).where(df["FingerTip_Status"] == "NG", 0)
        df["Thickness_NG_Count"] = df["Roll_NG_Count"] + df["Cuff_NG_Count"] + df["Palm_NG_Count"] + df[
            "Finger_NG_Count"] + df["FingerTip_NG_Count"]

        # 計算總 NG 數量
        df["Total_NG"] = df["Tensile_NG_Count"] + df["Elongation_NG_Count"] + df["Roll_NG_Count"] \
                         + df["Cuff_NG_Count"] + df["Palm_NG_Count"] + df["Finger_NG_Count"] + df[
                             "FingerTip_NG_Count"] \
                         + df["Length_NG_Count"] + df["Width_NG_Count"] \
                         + df["Weight_Light_NG_Count"] + df["Weight_Heavy_NG_Count"] + df["Pinhole_NG_Count"]

        sum_list = ["OnlinePacking", "WIPPacking", "ScrapQuantity", "FaultyQuantity", "Length_NG_Count", "Width_NG_Count", "Weight_Light_NG_Count",
                    "Weight_Heavy_NG_Count", "Pinhole_NG_Count", "Tensile_NG_Count", "Elongation_NG_Count", "Thickness_NG_Count"]
        ipqc_ng_df = df.groupby(["Machine", "Line", "ProductItem"], as_index=False)[sum_list].sum()

        return ipqc_ng_df

    def generate_ipqc_ng_data_excel(self, writer, df):
        start_time = time.time()

        # region 1. Column Define
        font = Font(name=self.report_font, size=10, bold=False)
        sheet = DataControl()
        sheet.add(ColumnControl('Machine', 'center', '@', '機台號', font, hidden=False, width=18))
        sheet.add(ColumnControl('Line', 'center', '@', '線別', font, hidden=False, width=17))
        sheet.add(ColumnControl('ProductItem', 'left', '@', '品項', font, hidden=False, width=30))
        sheet.add(ColumnControl('OnlinePacking', 'right', '#,##0', '包裝確認量', font, hidden=False, width=12))
        sheet.add(ColumnControl('WIPPacking', 'right', '#,##0', '半成品入庫量', font, hidden=False, width=12))
        sheet.add(ColumnControl('ScrapQuantity', 'right', '#,##0', '廢品數量', font, hidden=False, width=12))
        sheet.add(ColumnControl('FaultyQuantity', 'right', '#,##0', '二級品數量', font, hidden=False, width=12))
        sheet.add(ColumnControl('Length_NG_Count', 'right', '#,##0', '長度隔離', font, hidden=False, width=12))
        sheet.add(ColumnControl('Width_NG_Count', 'right', '#,##0', '寬度隔離', font, hidden=False, width=12))
        sheet.add(ColumnControl('Weight_Light_NG_Count', 'right', '#,##0', '過輕隔離', font, hidden=False, width=12))
        sheet.add(ColumnControl('Weight_Heavy_NG_Count', 'right', '#,##0', '過重隔離', font, hidden=False, width=12))
        sheet.add(ColumnControl('Pinhole_NG_Count', 'right', '#,##0', '針孔隔離', font, hidden=False, width=12))
        sheet.add(ColumnControl('Tensile_NG_Count', 'right', '#,##0', '拉力隔離', font, hidden=False, width=12))
        sheet.add(ColumnControl('Elongation_NG_Count', 'right', '#,##0', '伸長率隔離', font, hidden=False, width=12))
        sheet.add(ColumnControl('Thickness_NG_Count', 'right', '#,##0', '手套厚度隔離', font, hidden=False, width=12))

        selected_columns = [col for col in sheet.column_names if col in df.columns]
        header_columns = sheet.header_columns
        # endregion

        # region 2. DataFrame convert to Excel
        df = df[selected_columns].copy()

        df.rename(columns=header_columns, inplace=True)

        sheet_name = "IPQC異常分析"
        df.to_excel(writer, sheet_name=sheet_name, index=False)

        worksheet = writer.sheets[sheet_name]

        sheet.apply_formatting(worksheet)
        # endregion

        end_time = time.time()
        elapsed_time = end_time - start_time
        self.logging.info(f"Time taken: {elapsed_time:.2f} seconds.")

    # Sheet 外觀週累計
    def generate_12aspect_cosmetic_summary_excel(self, writer):
        start_time = time.time()

        # region 1. Column Define
        font = Font(name=self.report_font, size=10, bold=False)
        sheet = DataControl()
        sheet.add(ColumnControl('Plant', 'center', '@', '廠別', font, hidden=False, width=9))
        sheet.add(ColumnControl('Year', 'center', '@', '年', font, hidden=False, width=9))
        sheet.add(ColumnControl('Week_No', 'center', '@', '週別', font, hidden=False, width=9))
        sheet.add(ColumnControl('total_6100', 'right', '#,##0', '美線總時數', font, hidden=True, width=11))
        sheet.add(ColumnControl('LL1_6100', 'right', '#,##0', '美線超重時數', font, hidden=True, width=11))
        sheet.add(ColumnControl('LL1_6100_rate', 'right', '0.00%', '美線超重比例', font, hidden=False, width=11))
        sheet.add(ColumnControl('LL2_6100', 'right', '#,##0', '美線超輕時數', font, hidden=True, width=11))
        sheet.add(ColumnControl('LL2_6100_rate', 'right', '0.00%', '美線超輕比例', font, hidden=False, width=11))
        sheet.add(ColumnControl('total_6200', 'right', '#,##0', '歐線總時數', font, hidden=True, width=11))
        sheet.add(ColumnControl('LL1_6200', 'right', '#,##0', '歐線超重時數', font, hidden=True, width=11))
        sheet.add(ColumnControl('LL1_6200_rate', 'right', '0.00%', '歐線超重比例', font, hidden=False, width=11))
        sheet.add(ColumnControl('LL2_6200', 'right', '#,##0', '歐線超輕時數', font, hidden=True, width=11))
        sheet.add(ColumnControl('LL2_6200_rate', 'right', '0.00%', '歐線超輕比例', font, hidden=False, width=11))
        sheet.add(ColumnControl('total_6300', 'right', '#,##0', '日線總時數', font, hidden=True, width=11))
        sheet.add(ColumnControl('LL1_6300', 'right', '#,##0', '日線超重時數', font, hidden=True, width=11))
        sheet.add(ColumnControl('LL1_6300_rate', 'right', '0.00%', '日線超重比例', font, hidden=False, width=11))
        sheet.add(ColumnControl('LL2_6300', 'right', '#,##0', '日線超輕時數', font, hidden=True, width=11))
        sheet.add(ColumnControl('LL2_6300_rate', 'right', '0.00%', '日線超輕比例', font, hidden=False, width=11))
        sheet.add(ColumnControl('total_7000', 'right', '#,##0', 'OBM總時數', font, hidden=True, width=11))
        sheet.add(ColumnControl('LL1_7000', 'right', '#,##0', 'OBM超重時數', font, hidden=True, width=11))
        sheet.add(ColumnControl('LL1_7000_rate', 'right', '0.00%', 'OBM超重比例', font, hidden=False, width=11))
        sheet.add(ColumnControl('LL2_7000', 'right', '#,##0', 'OBM超輕時數', font, hidden=True, width=11))
        sheet.add(ColumnControl('LL2_7000_rate', 'right', '0.00%', 'OBM超輕比例', font, hidden=False, width=11))
        sheet.add(ColumnControl('critical_qty', 'right', '#,##0', 'Critical數量', font, hidden=True, width=11))
        sheet.add(ColumnControl('critical_rate', 'right', '0.00%', 'Critical比例', font, hidden=False, width=11))
        sheet.add(ColumnControl('critical_dpm', 'right', '#,##0', 'Critical DPM', font, hidden=False, width=11))
        sheet.add(ColumnControl('major_qty', 'right', '#,##0', 'Major數量', font, hidden=True, width=11))
        sheet.add(ColumnControl('major_rate', 'left', '0.00%', 'Major比例', font, hidden=False, width=11))
        sheet.add(ColumnControl('major_dpm', 'right', '#,##0', 'Major DPM', font, hidden=False, width=11))
        sheet.add(ColumnControl('minor_qty', 'right', '#,##0', 'Minor數量', font, hidden=True, width=11))
        sheet.add(ColumnControl('minor_rate', 'center', '0.00%', 'Minor比例', font, hidden=False, width=11))
        sheet.add(ColumnControl('minor_dpm', 'right', '#,##0', 'Minor DPM', font, hidden=False, width=11))
        sheet.add(ColumnControl('pinhole_qty', 'right', '0', '針孔數量', font, hidden=True, width=11))
        sheet.add(ColumnControl('pinhole_rate', 'center', '0.00%', '針孔比例', font, hidden=False, width=11))
        sheet.add(ColumnControl('pinhole_dpm', 'right', '#,##0', '針孔DPM', font, hidden=False, width=11))
        sheet.add(ColumnControl('cosmetic_check_qty', 'right', '#,##0', '外觀檢查總數量', font, hidden=False, width=13))

        header_columns = sheet.header_columns
        # endregion

        # region 2. DataFrame convert to Excel
        # 外觀週累計
        cosmetic_summary_sql = f"""
                    SELECT [Plant],r.Year,r.Week_No
                          ,[total_6100],[LL1_6100],[LL1_6100_rate],[LL2_6100],[LL2_6100_rate]
                          ,[total_6200],[LL1_6200],[LL1_6200_rate],[LL2_6200],[LL2_6200_rate]
                          ,[total_6300],[LL1_6300],[LL1_6300_rate],[LL2_6300],[LL2_6300_rate]
                          ,[total_7000],[LL1_7000],[LL1_7000_rate],[LL2_7000],[LL2_7000_rate]
                          ,[critical_qty],[critical_rate],[critical_dpm]
                          ,[major_qty],[major_rate],[major_dpm]
                          ,[minor_qty],[minor_rate],[minor_dpm]
                          ,[pinhole_qty],[pinhole_rate],[pinhole_dpm]
                          ,[cosmetic_check_qty]
                      FROM [MES_OLAP].[dbo].[appearance_weekly_info_raw] r
                      JOIN [MES_OLAP].[dbo].[week_date] w on r.Year = w.year and r.Week_No = w.week_no
                      where Plant = '{self.plant}' and w.enable = 1
                      order by CAST(w.week_no AS Int)
                    """
        cosmetic_summary_data = self.mes_olap_db.select_sql_dict(cosmetic_summary_sql)
        df = pd.DataFrame(cosmetic_summary_data)

        df = df.rename(columns=header_columns)

        sheet_name = "外觀週累計"
        df.to_excel(writer, sheet_name=sheet_name, index=False)

        worksheet = writer.sheets[sheet_name]

        sheet.apply_formatting(worksheet)

        worksheet.freeze_panes = worksheet['A2']
        # endregion

        end_time = time.time()
        elapsed_time = end_time - start_time
        self.logging.info(f"Time taken: {elapsed_time:.2f} seconds.")

    # Sheet 外觀
    def generate_12aspect_cosmetic_excel(self, writer, cosmetic_df):
        start_time = time.time()

        # region 1. Column Define
        font = Font(name=self.report_font, size=10, bold=False)
        sheet = DataControl()
        sheet.add(ColumnControl('runcard', 'center', '@', 'Runcard', font, hidden=False, width=14))
        sheet.add(ColumnControl('belong_to', 'center', '@', '作業日期', font, hidden=False, width=14))
        sheet.add(ColumnControl('Machine', 'center', '@', '機台', font, hidden=False, width=18))
        sheet.add(ColumnControl('Line', 'right', '@', '線別', font, hidden=False, width=8))
        sheet.add(ColumnControl('Shift', 'right', '@', '班別', font, hidden=False, width=8))
        sheet.add(ColumnControl('WorkOrder', 'center', '@', '工單', font, hidden=False, width=16))
        sheet.add(ColumnControl('PartNo', 'center', '@', '料號', font, hidden=False, width=14))
        sheet.add(ColumnControl('ProductItem', 'center', '@', '品項', font, hidden=False, width=28))
        sheet.add(ColumnControl('SalePlaceCode', 'center', '@', '銷售地點', font, hidden=False, width=8))
        sheet.add(ColumnControl('Period', 'center', '@', '', font, hidden=False, width=10))
        sheet.add(ColumnControl('6100LL1', 'center', '0', '美線過重', font, hidden=False, width=14, group='WEIGHT', apply_format=False))
        sheet.add(ColumnControl('6100LL2', 'center', '0', '美線過輕', font, hidden=False, width=14, group='WEIGHT', apply_format=False))
        sheet.add(ColumnControl('6200LL1', 'center', '0', '歐線過重', font, hidden=False, width=14, group='WEIGHT', apply_format=False))
        sheet.add(ColumnControl('6200LL2', 'center', '0', '歐線過輕', font, hidden=False, width=14, group='WEIGHT', apply_format=False))
        sheet.add(ColumnControl('6300LL1', 'center', '0', '日線過重', font, hidden=False, width=14, group='WEIGHT', apply_format=False))
        sheet.add(ColumnControl('6300LL2', 'center', '0', '日線過輕', font, hidden=False, width=14, group='WEIGHT', apply_format=False))
        sheet.add(ColumnControl('7000LL1', 'center', '0', 'OBM過重', font, hidden=False, width=14, group='WEIGHT', apply_format=False))
        sheet.add(ColumnControl('7000LL2', 'center', '0', 'OBM過輕', font, hidden=False, width=14, group='WEIGHT', apply_format=False))

        sql = f"""
                  SELECT distinct d.defect_level, d.defect_code, d.defect_code, d.desc1, d.desc2
                    FROM [MES_OLAP].[dbo].[counting_hourly_info_raw] r
                    LEFT JOIN [MES_OLAP].[dbo].[mes_ipqc_cosmetic_data] cos on r.Runcard = cos.runcard
                    LEFT JOIN [MES_OLAP].[dbo].[mes_defect_define] d on d.defect_code = cos.defect_code
                    where r.Year = {self.year} and r.Week_No = '{self.week_no}' and d.defect_type <> ''
                    order by d.defect_level, d.defect_code
                """
        cosmetic_rows = self.mes_olap_db.select_sql_dict(sql)

        critical_defect = []
        major_defect = []
        minor_defect = []
        defect_list = {'CRITICAL': critical_defect, 'MAJOR': major_defect, 'MINOR': minor_defect}

        for row in cosmetic_rows:
            defect_list[row['defect_level']].append(row['defect_code'])
            desc = row['desc1'] if row['desc1'] != '' else row['desc2']
            sheet.add(ColumnControl(row['defect_code'], 'center', '0', desc, font, hidden=False, width=14, group=row['defect_level'], apply_format=False))

        sheet.add(ColumnControl('cosmetic_qty', 'right', '#,##0', '缺陷手套數量', font, hidden=False, width=14, group='COSMETIC',
                          apply_format=False))

        sheet.add(ColumnControl('cosmetic_inspect_qty', 'right', '#,##0', '外觀檢查數量', font, hidden=False, width=14, group='COSMETIC', apply_format=False))

        sql = f"""
                SELECT distinct d.defect_code, d.desc1, d.desc2
                  FROM [MES_OLAP].[dbo].[counting_hourly_info_raw] r
                  JOIN [MES_OLAP].[dbo].[mes_ipqc_pinhole_data] cos on r.Runcard = cos.runcard
                  JOIN [MES_OLAP].[dbo].[mes_defect_define] d on d.defect_code = cos.defect_code
                  where r.Year = {self.year} and r.Week_No = '{self.week_no}'
                  order by d.defect_code
                        """
        pinhole_rows = self.mes_olap_db.select_sql_dict(sql)

        pinhole_defect = []
        defect_list['PINHOLE'] = pinhole_defect

        for row in pinhole_rows:
            defect_list['PINHOLE'].append(row['defect_code'])
            desc = row['desc1'] if row['desc1'] != '' else row['desc2']
            sheet.add(ColumnControl(row['defect_code'], 'center', '0', desc, font, hidden=False, width=14, group='COSMETIC', apply_format=False))

        sheet.add(ColumnControl('Pinhole', 'right', '0', '針孔數量', font, hidden=False, width=14, group='COSMETIC', apply_format=False))
        sheet.add(ColumnControl('Pinhole_Sample', 'right', '0', '針孔檢查數量', font, hidden=False, width=14, group='COSMETIC', apply_format=False))

        header_columns = sheet.header_columns
        column_letter = sheet.column_letter
        selected_columns = [col for col in sheet.column_names if col in cosmetic_df.columns]
        # endregion

        # region 2. DataFrame convert to Excel
        df = cosmetic_df[selected_columns].copy()

        df = df.rename(columns=header_columns)

        sheet_name = "外觀"
        df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=1)

        worksheet = writer.sheets[sheet_name]
        sheet.apply_formatting(worksheet)
        # endregion

        # region 3. Customize
        header_border = Border(
            top=Side(style='medium'),
            bottom=Side(style='medium'),
            left=Side(style='medium'),
            right=Side(style='medium')
        )
        fill_style = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

        # Weight Header
        worksheet.merge_cells(f"{column_letter['6100LL1']}1:{column_letter['7000LL2']}1")
        start_row, start_col = worksheet[f"{column_letter['6100LL1']}1"].row, worksheet[
            f"{column_letter['6100LL1']}1"].column
        end_row, end_col = worksheet[f"{column_letter['7000LL2']}1"].row, worksheet[
            f"{column_letter['7000LL2']}1"].column
        cell = worksheet[f"{column_letter['6100LL1']}1"]
        cell.value = "重量檢驗"
        cell.alignment = Alignment(horizontal="center", vertical="center")
        # cell.border = header_border
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.border = header_border
                cell.fill = fill_style

        # Cosmetic Header
        critical_start = defect_list['CRITICAL'][0]
        critical_end = defect_list['CRITICAL'][-1]
        worksheet.merge_cells(f"{column_letter[critical_start]}1:{column_letter[critical_end]}1")
        start_row, start_col = worksheet[f"{column_letter[critical_start]}1"].row, worksheet[
            f"{column_letter[critical_start]}1"].column
        end_row, end_col = worksheet[f"{column_letter[critical_end]}1"].row, worksheet[
            f"{column_letter[critical_end]}1"].column
        cell = worksheet[f"{column_letter[critical_start]}1"]
        cell.value = "外觀CRITICAL"
        cell.alignment = Alignment(horizontal="center", vertical="center")
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.border = header_border
                cell.fill = fill_style

        major_start = defect_list['MAJOR'][0]
        major_end = defect_list['MAJOR'][-1]
        worksheet.merge_cells(f"{column_letter[major_start]}1:{column_letter[major_end]}1")
        start_row, start_col = worksheet[f"{column_letter[major_start]}1"].row, worksheet[
            f"{column_letter[major_start]}1"].column
        end_row, end_col = worksheet[f"{column_letter[major_end]}1"].row, worksheet[
            f"{column_letter[major_end]}1"].column
        cell = worksheet[f"{column_letter[major_start]}1"]
        cell.value = "外觀MAJOR"
        cell.alignment = Alignment(horizontal="center", vertical="center")
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.border = header_border
                cell.fill = fill_style

        minor_start = defect_list['MINOR'][0]
        minor_end = defect_list['MINOR'][-1]
        worksheet.merge_cells(f"{column_letter[minor_start]}1:{column_letter[minor_end]}1")
        start_row, start_col = worksheet[f"{column_letter[minor_start]}1"].row, worksheet[
            f"{column_letter[minor_start]}1"].column
        end_row, end_col = worksheet[f"{column_letter[minor_end]}1"].row, worksheet[
            f"{column_letter[minor_end]}1"].column
        cell = worksheet[f"{column_letter[minor_start]}1"]
        cell.value = "外觀MINOR"
        cell.alignment = Alignment(horizontal="center", vertical="center")
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.border = header_border
                cell.fill = fill_style

        # Pinhole Header
        pinhole_start = defect_list['PINHOLE'][0]
        pinhole_end = defect_list['PINHOLE'][-1]
        worksheet.merge_cells(f"{column_letter[pinhole_start]}1:{column_letter[pinhole_end]}1")
        start_row, start_col = worksheet[f"{column_letter[pinhole_start]}1"].row, worksheet[
            f"{column_letter[pinhole_start]}1"].column
        end_row, end_col = worksheet[f"{column_letter[pinhole_end]}1"].row, worksheet[
            f"{column_letter[pinhole_end]}1"].column
        cell = worksheet[f"{column_letter[pinhole_start]}1"]
        cell.value = "針孔"
        cell.alignment = Alignment(horizontal="center", vertical="center")
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.border = header_border
                cell.fill = fill_style

        # 最後一行加總
        thin_top_border = Border(top=Side(style="thin"))
        last_row = worksheet.max_row + 1

        for col_idx in range(11, len(header_columns) + 1):
            col_letter = get_column_letter(col_idx)

            # 設定 SUM 公式
            sum_formula = f"=SUM({col_letter}2:{col_letter}{last_row-1})"
            worksheet[f"{col_letter}{last_row}"] = sum_formula  # ✅ 填入最後一行

            # 在最後一行上方（倒數第 2 行）加上框線
            if last_row > 1:  # 確保至少有兩行
                cell_above = worksheet[f"{col_letter}{last_row}"]
                cell_above.border = thin_top_border  # ✅ 設定上方框線
        # endregion

        # region 4. Keep Data
        sum_list = {}
        for col in sheet.columns:
            if col.group in ["WEIGHT", "COSMETIC", "CRITICAL", "MAJOR", "MINOR"]:
                cosmetic_df[col.name] = pd.to_numeric(cosmetic_df[col.name], errors='coerce').fillna(0)
                sum_list[col.name] = cosmetic_df[col.name].sum()

        sum_LL1_6100 = sum_list['6100LL1']
        sum_LL2_6100 = sum_list['6100LL2']
        count_6100_qty = (cosmetic_df['SalePlaceCode'] == 6100).sum()
        rate_LL1_6100 = round(sum_LL1_6100 / count_6100_qty, 4) if count_6100_qty > 0 else 0
        rate_LL2_6100 = round(sum_LL2_6100 / count_6100_qty, 4) if count_6100_qty > 0 else 0

        sum_LL1_6200 = sum_list['6200LL1']
        sum_LL2_6200 = sum_list['6200LL2']
        count_6200_qty = (cosmetic_df['SalePlaceCode'] == 6200).sum()
        rate_LL1_6200 = round(sum_LL1_6200 / count_6200_qty, 4) if count_6200_qty > 0 else 0
        rate_LL2_6200 = round(sum_LL2_6200 / count_6200_qty, 4) if count_6200_qty > 0 else 0

        sum_LL1_6300 = sum_list['6300LL1']
        sum_LL2_6300 = sum_list['6300LL2']
        count_6300_qty = (cosmetic_df['SalePlaceCode'] == 6300).sum()
        rate_LL1_6300 = round(sum_LL1_6300 / count_6300_qty, 4) if count_6300_qty > 0 else 0
        rate_LL2_6300 = round(sum_LL2_6300 / count_6300_qty, 4) if count_6300_qty > 0 else 0

        sum_LL1_7000 = sum_list['7000LL1']
        sum_LL2_7000 = sum_list['7000LL2']
        count_7000_qty = (cosmetic_df['SalePlaceCode'] == 7000).sum()
        rate_LL1_7000 = round(sum_LL1_7000 / count_7000_qty, 4) if count_7000_qty > 0 else 0
        rate_LL2_7000 = round(sum_LL2_7000 / count_7000_qty, 4) if count_7000_qty > 0 else 0

        sum_inspect_qty = sum_list['cosmetic_inspect_qty']

        sum_critical = sum(sum_list[col.name] for col in sheet.columns if col.group == "CRITICAL")
        critical_rate = round(sum_critical / sum_inspect_qty, 4)
        critical_dpm = round(critical_rate * 1000000, 0)

        sum_major = sum(sum_list[col.name] for col in sheet.columns if col.group == "MAJOR")
        major_rate = round(sum_major / sum_inspect_qty, 4)
        major_dpm = round(major_rate * 1000000, 0)

        sum_minor = sum(sum_list[col.name] for col in sheet.columns if col.group == "MINOR")
        minor_rate = round(sum_minor / sum_inspect_qty, 4)
        minor_dpm = round(minor_rate * 1000000, 0)

        sum_pinhole = sum_list['Pinhole']
        sum_pinhole_sample = sum_list['Pinhole_Sample']

        pinhole_rate = round(sum_pinhole / sum_pinhole_sample, 4)
        pinhole_dpm = round(pinhole_rate * 1000000, 0)

        delete_sql = f"""
        DELETE FROM [MES_OLAP].[dbo].[appearance_weekly_info_raw]
        WHERE Plant = '{self.plant}' AND [Year] = {self.year} AND Week_No = '{self.week_no}'
        """
        self.mes_olap_db.execute_sql(delete_sql)

        insert_sql = f"""
        INSERT INTO [MES_OLAP].[dbo].[appearance_weekly_info_raw] (Plant, Year, Week_No, 
        total_6100, LL1_6100, LL1_6100_rate, LL2_6100, LL2_6100_rate, 
        total_6200, LL1_6200, LL1_6200_rate, LL2_6200, LL2_6200_rate,
        total_6300, LL1_6300, LL1_6300_rate, LL2_6300, LL2_6300_rate,
        total_7000, LL1_7000, LL1_7000_rate, LL2_7000, LL2_7000_rate,
        critical_qty, critical_rate, critical_dpm,
        major_qty, major_rate, major_dpm, minor_qty, minor_rate, minor_dpm, pinhole_qty, pinhole_rate, pinhole_dpm,
        cosmetic_check_qty, pinhole_check_qty)
        VALUES('{self.plant}', {self.year}, '{self.week_no}', 
        {count_6100_qty}, {sum_LL1_6100}, {rate_LL1_6100}, {sum_LL2_6100}, {rate_LL2_6100}, 
        {count_6200_qty}, {sum_LL1_6200}, {rate_LL1_6200}, {sum_LL2_6200}, {rate_LL2_6200},
        {count_6300_qty}, {sum_LL1_6300}, {rate_LL1_6300}, {sum_LL2_6300}, {rate_LL2_6300}, 
        {count_7000_qty}, {sum_LL1_7000}, {rate_LL1_7000}, {sum_LL2_7000}, {rate_LL2_7000}, 
        {sum_critical}, {critical_rate}, {critical_dpm},
        {sum_major}, {major_rate}, {major_dpm}, {sum_minor}, {minor_rate}, {minor_dpm}, {sum_pinhole}, {pinhole_rate}, {pinhole_dpm},
        {sum_inspect_qty}, null)
        """
        self.mes_olap_db.execute_sql(insert_sql)
        # endregion

        end_time = time.time()
        elapsed_time = end_time - start_time
        self.logging.info(f"Time taken: {elapsed_time:.2f} seconds.")

    # Sheet 週累計
    def generate_12aspect_output_excel(self, writer):
        start_time = time.time()

        # region 1. Column Define
        font = Font(name=self.report_font, size=10, bold=False)
        sheet = DataControl()
        sheet.add(ColumnControl('Plant', 'center', '@', '廠別', font, hidden=False, width=9))
        sheet.add(ColumnControl('Year', 'center', '@', '年', font, hidden=False, width=9))
        sheet.add(ColumnControl('Week_No', 'center', '@', '週別', font, hidden=False, width=9))
        sheet.add(ColumnControl('CountingQty', 'right', '#,##0', '點數機數量', font, hidden=False, width=12))
        sheet.add(ColumnControl('IsolationQuantity', 'right', '#,##0', '隔離品數量', font, hidden=False, width=12))
        sheet.add(ColumnControl('AvgSpeed', 'center', '0', '平均車速', font, hidden=False, width=11))
        sheet.add(ColumnControl('Capacity', 'center', '0.00%', '產能效率', font, hidden=False, width=11))
        sheet.add(ColumnControl('Yield', 'center', '0.00%', '良率', font, hidden=False, width=11))
        sheet.add(ColumnControl('Activation', 'center', '0.00%', '稼動率', font, hidden=True, width=11))
        sheet.add(ColumnControl('OEE', 'center', '0.00%', 'OEE', font, hidden=True, width=11))
        sheet.add(ColumnControl('IsolationRate', 'center', '0.00%', '隔離品率', font, hidden=False, width=11))
        sheet.add(ColumnControl('ScrapRate', 'center', '0.00%', '廢品率', font, hidden=False, width=11))
        sheet.add(ColumnControl('Target', 'center', '#,##0', '目標產能', font, hidden=False, width=13))
        sheet.add(ColumnControl('OnlinePacking', 'center', '#,##0', '包裝確認量', font, hidden=False, width=13))
        sheet.add(ColumnControl('WIPPacking', 'center', '#,##0', '半成品入庫量', font, hidden=False, width=13))
        sheet.add(ColumnControl('Gap', 'center', '#,##0', '目標差異', font, hidden=False, width=13))

        sql = f"""
            SELECT [Plant], r.Year, w.week_no Week_No
              ,[CountingQty],[IsolationQty] IsolationQuantity,[AvgSpeed]
              ,[Capacity],[Yield],[Activation]
              ,[OEE],[IsolationRate],[ScrapRate],[Target],[OnlinePacking], [WIPPacking], Target-OnlinePacking-WIPPacking Gap
          FROM [MES_OLAP].[dbo].[counting_weekly_info_raw] r
          JOIN [MES_OLAP].[dbo].[week_date] w on r.Year = w.year and r.Week_No = w.week_no
          where Plant = '{self.plant}' and w.enable = 1
          Order by CAST(w.week_no AS Int)
        """
        data = self.mes_olap_db.select_sql_dict(sql)
        df = pd.DataFrame(data)

        header_columns = sheet.header_columns
        selected_columns = [col for col in sheet.column_names if col in df.columns]
        # endregion

        # region 2. DataFrame convert to Excel
        df = df[selected_columns].copy()

        df.rename(columns=header_columns, inplace=True)

        sheet_name = "週累計"
        df.to_excel(writer, sheet_name=sheet_name, index=False)

        worksheet = writer.sheets[sheet_name]

        sheet.apply_formatting(worksheet)

        # Freeze the first row
        worksheet.freeze_panes = worksheet['A2']
        # endregion

        end_time = time.time()
        elapsed_time = end_time - start_time
        self.logging.info(f"Time taken: {elapsed_time:.2f} seconds.")

    def sorting_data(self, machine_groups):
        start_time = time.time()

        summary_data = []
        mach_sum_list = []

        tmp_date = self.date_mark.replace('_', '~')
        tmp_week = f"{self.week_no} ({tmp_date})"

        for machine_name, machine_df in machine_groups:
            mach_sum_df = self.ipqc_ng_data(machine_df)  # OK
            mach_sum_list.append(mach_sum_df)

            for shift in machine_df['Shift'].unique():  # Loop through each unique Line
                for line in machine_df['Line'].unique():
                    filtered_df = machine_df[(machine_df['Line'] == line) & (machine_df['Shift'] == shift)]
                    countingQty = filtered_df['CountingQty'].sum()
                    faultyQty = filtered_df['FaultyQuantity'].sum()
                    scrapQty = filtered_df['ScrapQuantity'].sum()
                    isolationQty = filtered_df['IsolationQty'].sum()
                    onlinePacking = filtered_df['OnlinePacking'].sum()
                    wipPacking = filtered_df['WIPPacking'].sum()
                    runTime = filtered_df['RunTime'].sum()
                    stopTime = filtered_df['StopTime'].sum()
                    allTime = filtered_df['AllTime'].sum()
                    avgSpeed = round(filtered_df['StdSpeed'].mean(), 0)
                    target = filtered_df['Target'].sum()
                    rate = round(((int(onlinePacking)+int(wipPacking)) / int(target)), 3) if int(target) > 0 else 0

                    output = onlinePacking + wipPacking + faultyQty + scrapQty
                    isolationRate = round(isolationQty / output, 3) if int(output) > 0 else 0
                    dmf_rate = round(filtered_df['DMF_Rate'].mean(), 4)

                    summary_row = {
                        'Name': machine_name,
                        'Date': tmp_week,
                        'Shift': shift,
                        'Line': line,
                        'CountingQty': countingQty,
                        'FaultyQuantity': faultyQty,
                        'ScrapQuantity': scrapQty,
                        'IsolationQuantity': isolationQty,
                        'OnlinePacking': onlinePacking,
                        'WIPPacking': wipPacking,
                        'RunTime': runTime,
                        'AllTime': allTime,
                        'AvgSpeed': avgSpeed,
                        'Target': target,
                        'Capacity': '',
                        'Yield': '',
                        'Activation': '',
                        'OEE': '',
                        'Achievement Rate': rate,
                        'IsolationRate': isolationRate,
                        'DMF_Rate': dmf_rate
                    }
                    summary_data.append(summary_row)

            # Summary Row
            countingQty = sum(item['CountingQty'] for item in summary_data if item['Name'] == machine_name)
            faultyQty = sum(item['FaultyQuantity'] for item in summary_data if item['Name'] == machine_name)
            scrapQty = sum(item['ScrapQuantity'] for item in summary_data if item['Name'] == machine_name)
            isolationQty = sum(item['IsolationQuantity'] for item in summary_data if item['Name'] == machine_name)
            onlinePacking = sum(item['OnlinePacking'] for item in summary_data if item['Name'] == machine_name)
            wipPacking = sum(item['WIPPacking'] for item in summary_data if item['Name'] == machine_name)
            runTime = sum(item['RunTime'] for item in summary_data if item['Name'] == machine_name)
            allTime = sum(item['AllTime'] for item in summary_data if item['Name'] == machine_name)

            avgSpeed_values = [item['AvgSpeed'] for item in summary_data if item['Name'] == machine_name]
            avgSpeed = round(sum(avgSpeed_values) / len(avgSpeed_values), 0) if avgSpeed_values else 0

            target = sum(item['Target'] for item in summary_data if item['Name'] == machine_name)
            # 稼動率allTime必須扣除計劃性停機時間
            # activation = round(runTime/allTime, 3) if int(allTime) > 0 else 0
            activation = ''
            output = onlinePacking + wipPacking + faultyQty + scrapQty
            capacity = round(output / target, 3) if int(target) > 0 else 0
            _yield = round((onlinePacking + wipPacking - isolationQty) / output, 3) if int(output) > 0 else 0

            # 等有稼動率才能做OEE
            # oee = round(activation * capacity * _yield, 3)
            oee = ''
            rate = round((onlinePacking+wipPacking) / target, 3) if int(target) > 0 else 0
            isolationRate = round(isolationQty / output, 3) if int(output) > 0 else 0

            dmf_rate_values = [item['DMF_Rate'] for item in summary_data if item['Name'] == machine_name]
            dmf_rate = round(sum(dmf_rate_values) / len(dmf_rate_values), 4) if dmf_rate_values else 0

            summary_data.append({'Name': machine_name, 'Date': tmp_week, 'Shift': '', 'Line': '',
                                 'CountingQty': countingQty, 'FaultyQuantity': faultyQty, 'ScrapQuantity': scrapQty,
                                 'IsolationQuantity': isolationQty, 'OnlinePacking': onlinePacking, 'WIPPacking': wipPacking,
                                 'RunTime': runTime, 'AllTime': allTime, 'AvgSpeed': avgSpeed, 'Target': target,
                                 'Capacity': capacity, 'Yield': _yield, 'Activation': activation, 'OEE': oee,
                                 'Achievement Rate': rate, 'IsolationRate': isolationRate, 'DMF_Rate': dmf_rate})
        summary_df = pd.DataFrame(summary_data)

        all_mach_sum_df = pd.concat(mach_sum_list)

        end_time = time.time()
        elapsed_time = end_time - start_time
        self.logging.info(f"Time taken: {elapsed_time:.2f} seconds.")

        return summary_df, all_mach_sum_df

    def validate_data(self):
        pass

    # @Summary
    def generate_summary_excel(self, writer, summary_df):
        start_time = time.time()

        # region 1. Column Define
        font = Font(name=self.report_font, size=10, bold=False)
        sheet = DataControl()
        sheet.add(ColumnControl('Name', 'center', '@', '機台號', font, hidden=False, width=18))
        sheet.add(ColumnControl('Date', 'center', '@', '日期範圍', font, hidden=False, width=18))
        sheet.add(ColumnControl('Shift', 'center', '@', '班別', font, hidden=False, width=9))
        sheet.add(ColumnControl('Line', 'center', '@', '線別', font, hidden=False, width=9))
        sheet.add(ColumnControl('AvgSpeed', 'right', '0', '平均車速', font, hidden=False, width=10,
                                comment="車速標準下限~車速標準上限+2%", comment_width=200))

        sheet.add(ColumnControl('CountingQty', 'right', '#,##0', '生產總量', font, hidden=False, width=13, level=1,
                                comment="點數機數量", comment_width=400))

        sheet.add(ColumnControl('OnlinePacking', 'right', '#,##0', '包裝確認量', font, hidden=False, width=13, level=1))
        sheet.add(ColumnControl('WIPPacking', 'right', '#,##0', '半成品入庫量', font, hidden=False, width=13, level=1))
        sheet.add(ColumnControl('IsolationQuantity', 'right', '#,##0', '隔離品數量', font, hidden=False, width=13, level=1,
                                comment="MES輸入的隔離品數量", comment_width=400))
        sheet.add(ColumnControl('ScrapQuantity', 'right', '#,##0', '廢品數量', font, hidden=False, width=13, level=1))
        sheet.add(ColumnControl('FaultyQuantity', 'right', '#,##0', '二級品數量', font, hidden=False, width=13, level=1))
        sheet.add(ColumnControl('Target', 'right', '#,##0', '目標產能', font, hidden=False, width=10,
                                comment="生產時間(IPQC) * (標準車速上限/節距調整值)", comment_width=700, level=1))

        sheet.add(ColumnControl('Achievement Rate', 'right', '0.00%', '目標達成率', font, hidden=False, width=10,
                                comment="包裝確認量+半成品入庫量/目標產能", comment_width=700))

        sheet.add(ColumnControl('RunTime', 'right', '#,##0', '實際運轉時間', font, hidden=True, width=11))
        sheet.add(ColumnControl('AllTime', 'right', '#,##0', '可運轉時間', font, hidden=True, width=11))

        sheet.add(ColumnControl('Activation', 'right', '0.00%', '稼動率', font, hidden=True, width=10,
                                comment="有做IPQC的機台實際運轉時間/(可運轉時間-計劃性停機時間)", comment_width=700))

        sheet.add(ColumnControl('Capacity', 'right', '0.00%', '產能效率', font, hidden=False, width=10,
                                comment="(包裝確認量+半成品入庫量+二級品數量+廢品數量) / 目標產能", comment_width=700))
        sheet.add(ColumnControl('Yield', 'right', '0.00%', '良率', font, hidden=False, width=10,
                                comment="(包裝確認量+半成品入庫量-隔離品數量)/(包裝確認量+半成品入庫量+二級品數量+廢品數量)", comment_width=700))

        sheet.add(ColumnControl('OEE', 'right', '0.00%', 'OEE', font, hidden=True, width=10,
                                comment="稼動率 * 產能效率 * 良率", comment_width=600))

        sheet.add(ColumnControl('IsolationRate', 'right', '0.00%', '隔離率', font, hidden=False, width=10,
                                comment="隔離品數量/(包裝確認量+半成品入庫量+二級品數量+廢品數量)", comment_width=700))
        sheet.add(ColumnControl('DMF_Rate', 'center', '0.00%', '離型不良率', font, hidden=False, width=13))



        column_letter = sheet.column_letter
        header_columns = sheet.header_columns
        selected_columns = [col for col in sheet.column_names if col in summary_df.columns]
        # endregion

        # region 2. DataFrame convert to Excel
        df = summary_df[selected_columns].copy()
        # Change column names
        df.rename(columns=header_columns, inplace=True)
        summary_sheet_name = "Summary"
        df.to_excel(writer, sheet_name=summary_sheet_name, index=False)
        # Read the written Excel file
        workbook = writer.book
        worksheet = writer.sheets[summary_sheet_name]

        sheet.apply_formatting(worksheet)

        name_col = 1
        current_name = None
        group_start = None

        thin_border_top = Border(top=Side(style="thin"))
        thin_border_bottom = Border(bottom=Side(style="thin"))
        for row in range(2, len(df) + 2):  # Data starts from row 2 in Excel
            name = worksheet.cell(row=row, column=name_col).value

            if name != current_name:
                # Finalize the previous group
                if group_start is not None:
                    last_row = row - 1
                    for col in range(1, df.shape[1] + 1):  # Apply border to all columns
                        worksheet.cell(row=last_row, column=col).border = thin_border_top + thin_border_bottom
                    if row - group_start > 1:
                        worksheet.row_dimensions.group(group_start, row - 2, hidden=True)

                # Start a new group
                current_name = name
                group_start = row

        # Finalize the last group
        if group_start is not None:
            last_row = len(df) + 1
            for col in range(1, df.shape[1] + 1):
                worksheet.cell(row=last_row, column=col).border = thin_border_top + thin_border_bottom
            if len(df) + 1 - group_start > 1:
                worksheet.row_dimensions.group(group_start, len(df) + 1 - 1, hidden=True)

        worksheet.column_dimensions.group(column_letter['CountingQty'], column_letter['Target'], hidden=True)

        end_time = time.time()
        elapsed_time = end_time - start_time
        self.logging.info(f"Time taken: {elapsed_time:.2f} seconds.")

    def generate_chart(self, data):
        start_time = time.time()

        save_path = self.save_path
        yticks_labels = []

        plt.rcParams['font.sans-serif'] = ['Microsoft YaHei']

        # Output Bar Chart
        x_labels = [str(item['name']).split('_')[-1] for item in data]
        x_range = range(0, len(x_labels) * 2, 2)

        this_data = [int(item['this_time']) for item in data]
        this_unfinish = [int(item['this_unfinish']) if int(item['this_unfinish']) > 0 else 0 for item in data]

        max_data = max(this_data, default=0)
        step_data = 10

        this_rate = [round((item['this_time'] / item['target_this_time']) * 100, 2) if int(
            item['target_this_time']) > 0 else 0 for item in data]

        max_rate = max(this_rate, default=0)
        rounded_max_rate = (math.ceil(max_rate / 10) * 10)
        rounded_step_rate = 20

        bar_width = 0.6
        plt.figure(figsize=(16, 9))
        fig, ax1 = plt.subplots(figsize=(16, 9))

        this_week_bars = ax1.bar([i for i in x_range], this_data, width=bar_width,
                                 label=f"入庫量",
                                 align='center', color='#10ba81')
        unfinish_bars = ax1.bar([i for i in x_range], this_unfinish, width=bar_width, bottom=this_data,
                                label=f"週目標差異",
                                align='center', color='lightgreen')

        ax1.set_xticks(x_range)
        ax1.set_xticklabels(x_labels, rotation=0, ha="center", fontsize=12)

        yticks_positions = []
        if 'NBR' in self.plant:
            yticks_positions = [1000000, 2000000, 3000000, 4000000, 5000000, 6000000, 7000000, 10000000]
            yticks_labels = ['1百萬', '2百萬', '3百萬', '4百萬', '5百萬', '6百萬', '7百萬', '']
        elif 'PVC' in self.plant:
            yticks_positions = [1000000, 2000000, 3000000, 4000000, 5000000]
            yticks_labels = ['1百萬', '2百萬', '3百萬', '4百萬', '']

        for index, bar in enumerate(this_week_bars):
            height = bar.get_height()
            unfinish_height = unfinish_bars[index].get_height()

            if len(str(max_data)) > 7:
                show_text = f'{round(height/(10**(len(str(max_data)) - 2)),2)}'[:4] if height > 0 else ''
            elif 4 < len(str(max_data)) <= 7:
                show_text = f'{round(height/(10**(len(str(max_data)) - 1)),2)}'[:4] if height > 0 else ''

            ax1.text(
                bar.get_x() + bar.get_width() / 2,
                height + unfinish_height,
                show_text,
                ha='center', va='bottom', fontsize=12  # Align the text
            )

        ax1.set_yticks(yticks_positions)
        ax1.set_yticklabels(yticks_labels, fontsize=12)

        # Achievement Rate Line Chart (橘色的線)
        ax2 = ax1.twinx()

        # line_label = f"{this_start_date.strftime('%d/%m')}-{this_end_date.strftime('%d/%m')}"
        sr_achieve_rate = "本週達成率"
        filtered_data = [(x, rate) for x, rate in zip(x_range, this_rate) if rate != 0]  # 折線圖上的文字
        x_filtered, this_rate_filtered = zip(*filtered_data)
        this_rate_line = ax2.plot(x_filtered, this_rate_filtered,
                                  label=sr_achieve_rate,
                                  color='#ED7D31', marker='o', linewidth=1.5)

        target_rate = int(self.capacity_target*100)

        # Label Name
        sr_target = "達成率目標%"
        # Chart Label
        ry_label = "達成率(%)"
        ly_label = "Product (百萬)"

        name = self.date_mark
        title = f"\n{self.plant} {self.year} 第{self.week_no}週 ({name})目標達成率 (達成率目標 > {target_rate}%)\n"

        yticks_positions = list(range(0, rounded_max_rate + 1 * rounded_step_rate, rounded_step_rate))
        # if target_rate not in yticks_positions:
        #     yticks_positions.append(target_rate)
        yticks_positions.append(rounded_max_rate + 0.25 * rounded_step_rate)
        yticks_positions = sorted(yticks_positions)

        yticks_labels = [f"{i}" + '%' for i in yticks_positions]
        yticks_labels[-1] = ""
        ax2.set_yticks(yticks_positions)
        ax2.set_yticklabels(yticks_labels, fontsize=12)
        ax2.axhline(y=target_rate, color='red', linestyle='--', linewidth=1, label=sr_target)

        ax1.xaxis.set_label_coords(0.975, -0.014)
        ax1.set_ylabel('Product output', fontsize=12)
        ax2.set_ylabel('Archive rates', fontsize=12)

        for i, value in enumerate(this_rate):
            if value > 0 and value < target_rate:
                ax2.text(
                    x_range[i],
                    value + 1.5,
                    f"{value:.1f}%",
                    ha='center', va='bottom', fontsize=12, color='white',
                    bbox=dict(facecolor='#ED7D31', edgecolor='none', boxstyle='round,pad=0.3')
                )

                ax2.text(
                    x_range[i],
                    value + 1.5,
                    f"{value:.1f}%",
                    ha='center', va='bottom', fontsize=12, color='white',
                    bbox=dict(facecolor='none', edgecolor='red', boxstyle='circle,pad=1.5', linewidth=2)
                )
            else:
                ax2.text(
                    x_range[i],
                    value + 1.5,
                    f"{value:.1f}%" if value != 0 else '',
                    ha='center', va='bottom', fontsize=12, color='white',
                    bbox=dict(facecolor='#ED7D31', edgecolor='none', boxstyle='round,pad=0.3')
                )

        handles1, labels1 = ax1.get_legend_handles_labels()
        handles2, labels2 = ax2.get_legend_handles_labels()
        fig.legend(
            handles1 + handles2,
            labels1 + labels2,
            loc='center left',
            fontsize=12,
            title="Note",
            title_fontsize=12,
            bbox_to_anchor=(1.0, 0.5),
            ncol=1
        )

        # Customize axes and borders
        ax = plt.gca()  # Get current axes
        ax.spines['top'].set_color('white')
        ax.spines['top'].set_linewidth(1.5)

        ax1.tick_params(axis='y', colors='black')  # Y-axis ticks color
        y_tick_lines = ax1.get_yticklines()
        y_tick_lines[-2].set_visible(False)
        y_tick_lines[-1].set_visible(False)
        ax1.annotate(
            '',
            xy=(0, 1.0),
            xytext=(0, 0.97),
            xycoords='axes fraction',
            arrowprops=dict(facecolor='black', arrowstyle='-|>,widthA=0.4,widthB=1.4', linewidth=0.5)
        )
        ax1.set_ylabel(ly_label, labelpad=20, rotation=0)
        ax1.yaxis.set_label_coords(-0.01, 1.01)

        ax2.tick_params(axis='y', colors='black')
        y_tick_lines = ax2.get_yticklines()
        y_tick_lines[-2].set_visible(False)
        y_tick_lines[-1].set_visible(False)
        ax2.annotate(
            '',
            xy=(1, 1.0),
            xytext=(1, 0.97),
            xycoords='axes fraction',
            arrowprops=dict(facecolor='black', arrowstyle='-|>,widthA=0.4,widthB=1.4', linewidth=0.5)
        )

        ax2.set_ylabel(ry_label, labelpad=20, rotation=0)
        ax2.yaxis.set_label_coords(1.01, 1.03)

        plt.title(title, fontsize=20)

        plt.tight_layout()

        file_name = f'MES_{self.plant}_Weekly_{self.date_mark}_Chart.png'
        chart_img = os.path.join(save_path, file_name)

        plt.savefig(f"{chart_img}", dpi=100, bbox_inches="tight", pad_inches=0.45)
        plt.close()

        end_time = time.time()
        elapsed_time = end_time - start_time
        self.logging.info(f"Time taken: {elapsed_time:.2f} seconds.")

        return chart_img

    # 過濾掉無效數據 (scrap 和 secondgrade)
    def filter_valid_data(self, x_range, y_values):
        filtered_x = [x for i, x in enumerate(x_range) if
                      y_values[i] is not None and not np.isnan(y_values[i])]
        filtered_y = [y for y in y_values if y is not None and not np.isnan(y)]
        return filtered_x, filtered_y

    def rate_chart(self, scrap_dict):
        start_time = time.time()

        save_path = self.save_path

        data1 = self.mach_list

        data = []
        for item in data1:
            name = item['name']
            if name in scrap_dict:
                data.append(scrap_dict[name])
            else:
                data.append({'name': name, 'scrap': 0, 'secondgrade': 0, 'sum_qty': 0})

        x_labels = [str(item['name']).split('_')[-1] for item in data]
        x_range = range(len(x_labels))

        scrap = [round((item['scrap'] / (item['sum_qty'] + item['scrap'] + item['secondgrade'])) * 100, 2) if item[
                                                                                                                  'sum_qty'] > 0 else 0
                 for item in data]
        secondgrade = [
            round((item['secondgrade'] / (item['sum_qty'] + item['scrap'] + item['secondgrade'])) * 100, 2) if item[
                                                                                                                   'sum_qty'] > 0 else 0
            for item in
            data]

        # 過濾 scrap 數據
        filtered_x_scrap, filtered_scrap = self.filter_valid_data(x_range, scrap)

        # 過濾 secondgrade 數據
        filtered_x_secondgrade, filtered_secondgrade = self.filter_valid_data(x_range, secondgrade)

        # Create the figure and subplots
        plt.figure(figsize=(16, 9))
        fig = plt.figure(figsize=(16, 9))
        gs = gridspec.GridSpec(2, 1, height_ratios=[1, 1], hspace=0)  # No space between subplots

        # Subplot for scrap
        y_max1 = 3
        y_ticks = np.arange(0, y_max1, 0.4)  # 分成10個刻度
        y_ticks = y_ticks[:-1]

        ax1 = fig.add_subplot(gs[0])
        ax1.plot(filtered_x_scrap, filtered_scrap, label="廢品率 (%)", marker='o', linestyle='-', color='#ED7D31',
                 linewidth=2)
        scrap_target = round(self.scrap_target*100, 2)
        ax1.axhline(y=scrap_target, color='#ED7D31', linestyle='--', linewidth=1.5, label=f"廢品標準線({scrap_target})")
        ax1.set_xticks(x_range)
        ax1.set_xticklabels([])  # Hide x-axis labels for the top plot
        # ax1.set_ylabel('廢品率 (%)', fontsize=12, rotation=0)
        ax1.text(-0.1, 0.6, '廢品率 (%)', fontsize=12, rotation=0, ha='center', va='center', transform=ax1.transAxes)
        ax1.set_ylim(0, y_max1)
        ax1.set_yticks(y_ticks)
        ax1.yaxis.set_major_formatter(FuncFormatter(self.add_percent))

        offset = 0.03
        for i, scrap_val in enumerate(filtered_scrap):
            ax1.text(filtered_x_scrap[i], scrap_val + offset, f"{scrap_val:.2f}%", ha='center', va='bottom',
                     fontsize=12,
                     color='#ED7D31')

        # Subplot for secondgrade
        y_max2 = 1.6
        y_ticks = np.arange(0, y_max2, 0.2)  # 分成10個刻度
        y_ticks = y_ticks[:-1]

        ax2 = fig.add_subplot(gs[1])
        ax2.plot(filtered_x_secondgrade, filtered_secondgrade, label="二級品率 (%)", marker='o', linestyle='-',
                 color='#70AD47', linewidth=2)
        ax2.axhline(y=0.2, color='#70AD47', linestyle='--', linewidth=1.5, label="二級品標準線(0.2)")
        ax2.set_xticks(x_range)
        ax2.set_xticklabels(x_labels, rotation=0, ha="center", fontsize=12)
        # ax2.set_ylabel('二級品率 (%)', fontsize=12, rotation=0)
        ax2.text(-0.1, 0.6, '二級品率 (%)', fontsize=12, rotation=0, ha='center', va='center', transform=ax2.transAxes)
        ax2.set_ylim(0, y_max2)
        ax2.set_yticks(y_ticks)
        ax2.yaxis.set_major_formatter(FuncFormatter(self.add_percent))

        offset = 0.03
        for i, secondgrade_val in enumerate(filtered_secondgrade):
            ax2.text(filtered_x_secondgrade[i], secondgrade_val + offset, f"{secondgrade_val:.2f}%", ha='center',
                     va='bottom',
                     fontsize=12, color='#70AD47')

        # Add a title for the entire figure
        fig.suptitle(f"二級品及廢品率 ({self.plant})", fontsize=20)

        # Add legend for both plots
        handles1, labels1 = ax1.get_legend_handles_labels()
        handles2, labels2 = ax2.get_legend_handles_labels()
        fig.legend(
            handles1 + handles2,
            labels1 + labels2,
            loc='center left',
            fontsize=10,
            title="Note",
            title_fontsize=12,
            bbox_to_anchor=(1.0, 0.5),
            ncol=1
        )

        plt.tight_layout()

        file_name = f'MES_{self.location}_{self.plant}_{self.date_mark}_Rate_Chart_Line.png'
        chart_img = os.path.join(save_path, file_name)

        plt.savefig(chart_img, dpi=100, bbox_inches="tight", pad_inches=0.45)
        plt.close()

        end_time = time.time()
        elapsed_time = end_time - start_time
        self.logging.info(f"Time taken: {elapsed_time:.2f} seconds.")

        return chart_img

    def rate_chart2(self, scrap_dict):
        start_time = time.time()

        save_path = self.save_path

        data1 = self.mach_list

        data = []
        for item in data1:
            name = item['name']
            if name in scrap_dict:
                data.append(scrap_dict[name])
            else:
                data.append({'name': name, 'scrap': 0, 'secondgrade': 0, 'sum_qty': 0})

        x_labels = [str(item['name']).split('_')[-1] for item in data]
        x_range = range(len(x_labels))

        scrap = [round((item['scrap'] / (item['sum_qty'] + item['scrap'] + item['secondgrade'])) * 100, 2) if item[
                                                                                                                  'sum_qty'] > 0 else 0
                 for item in data]

        # 過濾 scrap 數據
        filtered_x_scrap, filtered_scrap = self.filter_valid_data(x_range, scrap)

        # Create the figure and subplots
        plt.figure(figsize=(16, 5))
        fig, ax1 = plt.subplots(figsize=(16, 5))

        # Subplot for scrap
        y_max1 = 4
        y_ticks = np.arange(0, y_max1, 0.4)  # 分成10個刻度
        y_ticks = y_ticks[:-1]

        ax1.plot(filtered_x_scrap, filtered_scrap, label="廢品率 (%)", marker='o', linestyle='-', color='#ED7D31',
                 linewidth=2)
        scrap_target = round(self.scrap_target*100, 2)
        ax1.axhline(y=scrap_target, color='#ED7D31', linestyle='--', linewidth=1.5, label=f"廢品標準線({scrap_target})")
        ax1.set_xticks(x_range)
        ax1.set_xticklabels(x_labels, rotation=0, ha="center", fontsize=12)

        # ax1.set_ylabel('廢品率 (%)', fontsize=12, rotation=0)
        ax1.text(-0.1, 0.6, '廢品率 (%)', fontsize=12, rotation=0, ha='center', va='center', transform=ax1.transAxes)
        ax1.set_ylim(0, y_max1)
        ax1.set_yticks(y_ticks)
        ax1.yaxis.set_major_formatter(FuncFormatter(self.add_percent))

        offset = 0.03
        for i, scrap_val in enumerate(filtered_scrap):
            x = filtered_x_scrap[i]
            y = scrap_val + offset
            text_str = f"{scrap_val:.2f}%"

            # 畫主文字
            ax1.text(
                x, y, text_str,
                ha='center', va='bottom',
                fontsize=12,
                color='#ED7D31'
            )


        # Add a title for the entire figure
        fig.suptitle(f"廢品率 ({self.plant})", fontsize=20)

        # Add legend for both plots
        handles1, labels1 = ax1.get_legend_handles_labels()

        fig.legend(
            handles1,
            labels1,
            loc='center left',
            fontsize=10,
            title="Note",
            title_fontsize=12,
            bbox_to_anchor=(1.0, 0.5),
            ncol=1
        )

        plt.tight_layout()

        file_name = f'MES_{self.location}_{self.plant}_{self.date_mark}_Rate_Chart_Line.png'
        chart_img = os.path.join(save_path, file_name)

        plt.savefig(chart_img, dpi=100, bbox_inches="tight", pad_inches=0.45)
        plt.close()

        end_time = time.time()
        elapsed_time = end_time - start_time
        self.logging.info(f"Time taken: {elapsed_time:.2f} seconds.")

        return chart_img

    def send_email(self, config, subject, file_list, image_buffers, msg_list, error_list):
        start_time = time.time()

        logging.info(f"Start to send Email")

        error_msg = '<br>'.join(error_list)
        if len(error_list) > 0:
            error_msg = error_msg + '<br>'
        normal_msg = '<br>'.join(msg_list)
        if len(msg_list) > 0:
            normal_msg = normal_msg + '<br>'

        max_reSend = 5
        reSent = 0
        while reSent < max_reSend:
            try:
                super().send_email(config, subject, file_list, image_buffers, error_msg, normal_msg=normal_msg)
                print('Email sent successfully')
                logging.info(f"Email sent successfully")
                break
            except Exception as e:
                print(f'Email sending failed: {e}')
                logging.info(f"Email sending failed: {e}")
                reSent += 1
                if reSent >= max_reSend:
                    print('Failed to send email after 5 retries')
                    logging.info(f"Failed to send email after 5 retries")
                    break
                time.sleep(180)  # seconds

        end_time = time.time()
        elapsed_time = end_time - start_time
        self.logging.info(f"Time taken: {elapsed_time:.2f} seconds.")

    def read_config(self, config_file):
        smtp_config = {}
        to_emails = []
        admin_emails = []

        current_section = None

        with open(config_file, 'r') as file:
            for line in file:
                line = line.strip()

                # Skip empty lines and comments
                if not line or line.startswith('#'):
                    continue

                # Detect section headers
                if line.startswith('[') and line.endswith(']'):
                    current_section = line[1:-1].lower()
                    continue

                # Read lines based on the current section
                if current_section == 'smtp':
                    if '=' in line:
                        key, value = line.split('=', 1)
                        smtp_config[key.strip()] = value.strip()
                elif current_section == 'recipients':
                    to_emails.append(line)
                elif current_section == 'admin_email':
                    admin_emails.append(line)

        return smtp_config, to_emails, admin_emails

    def add_percent(self, y, pos):
        return f"{y:.1f}%"  # 格式化為整數百分比

    def weekly_chart(self, save_path):
        start_time = time.time()

        data1 = self.mach_list
        matplotlib.rcParams['font.family'] = ['Microsoft YaHei']
        matplotlib.rcParams['axes.unicode_minus'] = False

        for machine in data1:
            try:
                this_data = []
                this_rate = []
                x_labels = []
                week_date = []
                unfinish_data = []

                today = datetime.now()

                if machine['name'] in ['VN_GD_PVC1_L01', 'VN_GD_PVC1_L02']:
                    start_date = date(2025, 1, 6)
                    start_week_number = 1
                elif machine['name'] in ['VN_GD_PVC1_L05', 'VN_GD_PVC1_L06']:
                    start_date = date(2024, 12, 2)
                    start_week_number = 49
                else:
                    # Get the starting week (Sep 30)
                    start_date = date(2024, 9, 30)
                    start_week_number = 40

                x_labels, week_date = Utils().generate_previous_weeks_with_dates(self.mes_olap_db, self.report_date2)

                for i, item in enumerate(week_date):
                    week_name = 'W' + x_labels[i]
                    if item[0] < start_date:
                        this_data.append(0)
                        this_rate.append(0)
                        unfinish_data.append(0)
                        continue

                    # Target只看有做IPQC的部分
                    sql = f"""
                                SELECT name,qty, target, (case when target-qty > 0 then target-qty else 0 end) unfinish_qty FROM 
                                (
                                        SELECT Machine name, SUM(OnlinePacking)+SUM(WIPPacking) AS qty,
                                        SUM(Target) AS target
                                        FROM [MES_OLAP].[dbo].[counting_hourly_info_raw] c
                                        JOIN [MES_OLAP].[dbo].[mes_ipqc_data] ipqc on c.Runcard = ipqc.Runcard
                                        WHERE Machine = '{machine['name']}'
                                        AND Weight_Value > 0
                                        AND belong_to BETWEEN '{item[0]}' AND '{item[1]}'
                                        GROUP BY Machine
                                ) A
                                """
                    # print(sql)
                    data_dict = self.mes_olap_db.select_sql_dict(sql)

                    if len(data_dict) == 0:
                        this_data.append(0)
                        this_rate.append(0)
                        unfinish_data.append(0)
                    else:
                        this_data.append(data_dict[0]['qty'])
                        unfinish_data.append(data_dict[0]['unfinish_qty'])

                        try:
                            rate = round(data_dict[0]['qty'] / data_dict[0]['target'], 3) * 100 if data_dict[0][
                                                                                             'target'] > 0 else 100
                        except Exception as e:
                            rate = 100
                            print(f"{e} at {machine['name']} at week start {item[0]} - end {item[1]}")
                            pass
                        this_rate.append(rate)
                    # x_labels.append(f'W {week}')

                x_range = range(0, len(x_labels) * 2, 2)

                tmp_data = [qty + unfinish for qty, unfinish in zip(this_data, unfinish_data)]

                max_data = max(tmp_data, default=0)
                step_data = 10
                yticks_positions, yticks_labels = Utils().chart_y_label(max_data, step_data)

                if len(yticks_labels) > 0:
                    yticks_labels[-1] = ""

                    max_rate = max(this_rate, default=0)
                    rounded_max_rate = (math.ceil(max_rate / 10) * 10)
                    rounded_step_rate = 20

                    bar_width = 0.6
                    plt.figure(figsize=(24, 9))
                    fig, ax1 = plt.subplots(figsize=(24, 9))
                    this_week_bars = ax1.bar([i for i in x_range], this_data, width=bar_width,
                                             label=f"包裝確認量週累計", align='center', color='#10ba81')
                    unfinish_bars = ax1.bar([i for i in x_range], unfinish_data, width=bar_width, bottom=this_data,
                                            label=f"週目標差異",
                                            align='center', color='lightgreen')
                    ax1.set_xticks(x_range)
                    ax1.set_xticklabels(x_labels, rotation=0, ha="center", fontsize=12)

                    for index, bar in enumerate(this_week_bars):
                        height = bar.get_height()
                        unfinish_height = unfinish_bars[index].get_height()
                        if len(str(max_data)) >= 7:
                            show_text = f'{round(height/1000000, 1)}'[:4] if height > 0 else ''
                        elif 4 < len(str(max_data)) < 7:
                            show_text = f'{round(height/10000, 1)}'[:4] if height > 0 else ''
                        ax1.text(
                            bar.get_x() + bar.get_width() / 2,
                            height + unfinish_height,
                            show_text,
                            ha='center', va='bottom', fontsize=12  # Align the text
                        )

                    ax1.set_yticks(yticks_positions)
                    ax1.set_yticklabels(yticks_labels, fontsize=12)

                    ax2 = ax1.twinx()
                    sr_achieve_rate = "達成率"

                    filtered_data = [(x, rate) for x, rate in zip(x_range, this_rate) if rate != 0]
                    x_filtered, this_rate_filtered = zip(*filtered_data)
                    this_rate_line = ax2.plot(x_filtered, this_rate_filtered,
                                              label=sr_achieve_rate,
                                              color='#ED7D31', marker='o', linewidth=1.5)
                    sr_target = "達成率目標"

                    ry_label = "達成率(%)"
                    ly_label = "Product (百萬)"

                    name = f"各週產出量"
                    title = f"\n{self.plant} ({machine['name']})\n"

                    target_rate = int(self.capacity_target*100)

                    yticks_positions = list(range(0, rounded_max_rate + 1 * rounded_step_rate, rounded_step_rate))
                    if target_rate not in yticks_positions:
                        yticks_positions.append(target_rate)
                    yticks_positions.append(rounded_max_rate + 0.25 * rounded_step_rate)
                    yticks_positions = sorted(yticks_positions)

                    yticks_labels = [f"{i}" + '%' for i in yticks_positions]
                    yticks_labels[-1] = ""
                    ax2.set_yticks(yticks_positions)
                    ax2.set_yticklabels(yticks_labels)
                    ax2.axhline(y=target_rate, color='red', linestyle='--', linewidth=1, label=sr_target)

                    ax1.xaxis.set_label_coords(0.975, -0.014)
                    ax1.set_ylabel('Product output', fontsize=12)
                    ax2.set_ylabel('Archive rates', fontsize=12)

                    for i, value in enumerate(this_rate):
                        ax2.text(
                            x_range[i],
                            value + 1.5,
                            f"{value:.1f}%" if value != 0 else '',
                            ha='center', va='bottom', fontsize=12, color='white',
                            bbox=dict(facecolor='#ED7D31', edgecolor='none', boxstyle='round,pad=0.3')
                        )

                    handles1, labels1 = ax1.get_legend_handles_labels()
                    handles2, labels2 = ax2.get_legend_handles_labels()
                    fig.legend(
                        handles1 + handles2,
                        labels1 + labels2,
                        loc='center left',
                        fontsize=12,
                        title="Note",
                        title_fontsize=12,
                        bbox_to_anchor=(1.0, 0.5),
                        ncol=1
                    )

                    # Customize axes and borders
                    ax = plt.gca()  # Get current axes
                    ax.spines['top'].set_color('white')
                    ax.spines['top'].set_linewidth(1.5)

                    ax1.tick_params(axis='y', colors='black')  # Y-axis ticks color
                    y_tick_lines = ax1.get_yticklines()
                    y_tick_lines[-2].set_visible(False)
                    y_tick_lines[-1].set_visible(False)
                    ax1.annotate(
                        '',
                        xy=(0, 1.0),
                        xytext=(0, 0.98),
                        xycoords='axes fraction',
                        arrowprops=dict(facecolor='black', arrowstyle='-|>,widthA=0.4,widthB=1.4', linewidth=0.5)
                    )
                    ax1.set_ylabel(ly_label, labelpad=20, rotation=0)
                    ax1.yaxis.set_label_coords(-0.01, 1.01)

                    ax2.tick_params(axis='y', colors='black')
                    y_tick_lines = ax2.get_yticklines()
                    y_tick_lines[-2].set_visible(False)
                    y_tick_lines[-1].set_visible(False)
                    ax2.annotate(
                        '',
                        xy=(1, 1.0),
                        xytext=(1, 0.98),
                        xycoords='axes fraction',
                        arrowprops=dict(facecolor='black', arrowstyle='-|>,widthA=0.4,widthB=1.4', linewidth=0.5)
                    )

                    ax2.set_ylabel(ry_label, labelpad=20, rotation=0)
                    ax2.yaxis.set_label_coords(1.01, 1.03)

                    plt.title(title, fontsize=20)
                    plt.rcParams['font.sans-serif'] = [self.report_font]
                    plt.tight_layout()

                    file_name = f"MES_{machine['name']}_Chart.png"
                    chart_img = os.path.join(save_path, file_name)

                    plt.savefig(f"{chart_img}", dpi=100, bbox_inches="tight", pad_inches=0.45)
                    plt.close('all')

            except Exception as e:
                print(f"{machine['name']}: {e}")
                pass
        end_time = time.time()
        elapsed_time = end_time - start_time
        self.logging.info(f"Time taken: {elapsed_time:.2f} seconds.")


    def get_mach_list(self, plant):
        if 'NBR' in plant:
            plant_ = plant
        elif 'PVC' in plant:
            plant_ = 'PVC1'

        sql = f"""
                        select name FROM [PMGMES].[dbo].[PMG_DML_DataModelList] 
                        where DataModelTypeId = 'DMT000003' and name like '%{plant_}%' 
                        order by name
                    """
        data = self.mes_db.select_sql_dict(sql)

        return data


    def delete_counting_weekly_info_raw(self):
        start_time = time.time()

        sql = f"""
            delete from [MES_OLAP].[dbo].[counting_weekly_info_raw]
            where Plant = '{self.plant}' and [Year] = {self.year} and Week_No = '{self.week_no}'
        """
        self.mes_olap_db.execute_sql(sql)

        end_time = time.time()
        elapsed_time = end_time - start_time
        self.logging.info(f"Time taken: {elapsed_time:.2f} seconds.")

    def insert_counting_weekly_info_raw(self, summary_df):
        start_time = time.time()
        # Only show data which activation not null
        df = summary_df.loc[summary_df["Yield"].notna() & (summary_df["Yield"] != "")]

        counting_sum = df["CountingQty"].sum()
        isolation_sum = df["IsolationQuantity"].sum()
        target_sum = df["Target"].sum()
        onlinePacking_sum = df["OnlinePacking"].sum()
        wipPacking_sum = df["WIPPacking"].sum()
        faulty_sum = df["FaultyQuantity"].sum()
        scrap_sum = df["ScrapQuantity"].sum()

        speed_avg = round(df["AvgSpeed"].mean(), 0)
        # activation_avg = round(df["稼動率"].mean(), 3)
        activation_avg = 'null'
        capacity_avg = round(df["Capacity"].mean(), 3)
        yield_avg = round(df["Yield"].mean(), 3)

        # oee_avg = round(activation_avg*capacity_avg*yield_avg, 3)
        oee_avg = 'null'

        isolation_rate = round(isolation_sum / onlinePacking_sum, 3) if onlinePacking_sum > 0 else 0
        scrap_rate = round(scrap_sum / (counting_sum + faulty_sum + scrap_sum), 3)

        sql = f"""
        Insert into [MES_OLAP].[dbo].[counting_weekly_info_raw](Plant, [Year], Week_No, CountingQty, IsolationQty, AvgSpeed, 
        Activation, Capacity, Yield, OEE, IsolationRate, ScrapRate, Target, OnlinePacking, WIPPacking)
        Values('{self.plant}', {self.year}, '{self.week_no}', {counting_sum}, {isolation_sum}, {speed_avg}, 
        {activation_avg}, {capacity_avg},{yield_avg},{oee_avg},{isolation_rate},{scrap_rate},{target_sum},{onlinePacking_sum},{wipPacking_sum})
        """
        self.mes_olap_db.execute_sql(sql)

        end_time = time.time()
        elapsed_time = end_time - start_time
        self.logging.info(f"Time taken: {elapsed_time:.2f} seconds.")


from datetime import datetime, timedelta, date

report = mes_weekly_report()
report.main()