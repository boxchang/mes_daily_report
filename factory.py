import sys
import os
curPath = os.path.abspath(os.path.dirname(__file__))
rootPath = os.path.split(curPath)[0]
sys.path.append(rootPath)
from openpyxl.formatting.rule import CellIsRule, FormulaRule
import configparser
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from email.mime.image import MIMEImage
from abc import ABC, abstractmethod
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, NamedStyle, Font, Border, Side, PatternFill
import logging
from openpyxl.comments import Comment


class Factory(ABC):
    def __init__(self):
        self.attachment_list = []  # Excel 附件清單

    @abstractmethod
    def generate_main_df(self, sql_query, connection_string):
        pass

    @abstractmethod
    def sorting_data(self, df):
        pass

    @abstractmethod
    def validate_data(self):
        pass

    @abstractmethod
    def generate_chart(self, df, filename):
        pass

    @abstractmethod
    def send_email(self, config, subject, file_list, image_buffers, error_msg, normal_msg, content=None):
        # SMTP Sever config setting
        smtp_server = config.smtp_config.get('smtp_server')
        smtp_port = int(config.smtp_config.get('smtp_port', 587))
        smtp_user = config.smtp_config.get('smtp_user')
        smtp_password = config.smtp_config.get('smtp_password')
        sender_alias = f"{config.location} Report"
        sender_email = smtp_user
        # Mail Info
        msg = MIMEMultipart()
        msg['From'] = f"{sender_alias} <{sender_email}>"
        msg['To'] = ', '.join(config.to_emails)
        msg['Subject'] = subject

        if len(error_msg) > 0:
            msg['Subject'] = msg['Subject'] + ' 資料異常, 取消派送'

        # Mail Content
        html = f"""\
                <html>
                  <body>
                  {normal_msg}
                  {error_msg}
                """
        for i in range(len(image_buffers)):
            html += f'<img src="cid:chart_image{i}"><br>'

        html += f"""\
                    {content}
                  </body>
                </html>
                """

        msg.attach(MIMEText(html, 'html'))

        # Attach Excel
        for file in file_list:
            excel_file = file['excel_file']
            file_name = file['file_name']
            with open(excel_file, 'rb') as attachment:
                part = MIMEBase('application', 'octet-stream')
                part.set_payload(attachment.read())
                encoders.encode_base64(part)
                part.add_header('Content-Disposition', f"attachment; filename= {file_name}")
                msg.attach(part)
        logging.info(f"Attached Excel files")

        # Attach Picture
        for idx, image_path in enumerate(image_buffers):
            if os.path.exists(image_path):
                with open(image_path, "rb") as img_file:
                    img = MIMEImage(img_file.read())
                img.add_header("Content-ID", f"<chart_image{idx}>")
                msg.attach(img)
            else:
                logging.info(f"Image not found: {image_path}")
                print(f"Image not found: {image_path}")
        logging.info(f"Attached Picture")

        # Send Email
        try:
            # server = smtplib.SMTP(smtp_server, smtp_port)
            # server.starttls()  # 启用 TLS 加密
            # server.login(smtp_user, smtp_password)  # 登录到 SMTP 服务器
            # server.sendmail(smtp_user, to_emails, msg.as_string())
            # server.quit()

            # 發送郵件（不進行密碼驗證）
            server = smtplib.SMTP(smtp_server, smtp_port)
            server.ehlo()  # 啟動與伺服器的對話
            if len(error_msg) > 0:
                server.sendmail(smtp_user, config.admin_emails, msg.as_string())
            else:
                server.sendmail(smtp_user, config.to_emails, msg.as_string())
            print("Sent Email Successfully")
        except Exception as e:
            print(f"Sent Email Fail: {e}")
            logging.info(f"Sent Email Fail: {e}")
        finally:
            attachment.close()


class ColumnControl:
    def __init__(self, name, align, format_style, header_name, font=None, hidden=False, width=None, data_type=None,
                 comment=None, comment_width=None, level=None, group=None, apply_format=None, limit=None):
        """
        :param name: 欄位名稱 (str)
        :param align: 對齊方式 ('left', 'center', 'right')
        :param format_style: 格式設定，例如 '0.0%'、'#,##0' 等
        :param header_name: 欄位的顯示名稱 (str)
        :param font: openpyxl.styles.Font 物件 (可選)
        :param hidden: 是否隱藏欄位 (bool)
        :param width: 欄位寬度 (float)，如果為 None，則根據內容調整
        :param data_type: 指定資料型別 (`str`, `int`, `float`...)，預設為 None
        :param comment: 欄位的 Excel 註解 (str)，可選
        :param comment_width: 註解的寬度 (int)，預設為 20
        :param level: Excel折疊Level
        :param group: 欄位歸屬群組，使用於合併儲存格
        :param apply_format: 是否套用format，使用於欄位太多的Excel
        :param limit: 該欄位值的上下限
        """
        self.name = name
        self.align = align.lower()
        self.format_style = format_style
        self.header_name = header_name
        self.font = font if font else Font(name="Calibri", size=11)
        self.hidden = hidden
        self.width = width
        self.data_type = data_type
        self.comment = comment
        self.comment_width = comment_width
        self.level = level
        self.group = group
        self.apply_format = apply_format
        self.limit = limit

    def get_alignment(self):
        """ 依據 align 設定對齊方式 """
        alignments = {"center": "center", "right": "right", "left": "left"}
        return Alignment(horizontal=alignments.get(self.align, "left"))

    def apply_data_format(self, cell):
        """ 套用字型、對齊、格式與註解 """
        cell.font = self.font
        cell.alignment = self.get_alignment()
        cell.number_format = self.format_style  # 這行設定格式，例如 '@' 為文字
        cell.value = self.convert_value(cell.value)

    def convert_value(self, value):
        """
        依據指定的 data_type 轉換數值
        """
        if self.data_type and value is not None:
            try:
                return self.data_type(value)
            except ValueError:
                return value  # 如果轉換失敗，回傳原始值
        return value

    def __repr__(self):
        return (f"ColumnControl(name='{self.name}', align='{self.align}', format_style='{self.format_style}', "
                f"header_name='{self.header_name}', font={self.font}, hidden={self.hidden}, "
                f"width={self.width}, data_type={self.data_type}, comment={self.comment}, level={self.level})")


class DataControl:
    def __init__(self, fix_mode=False):
        self.columns = []
        self.fix_mode = fix_mode

    def add(self, column):
        if isinstance(column, ColumnControl):
            self.columns.append(column)
            self.header_font = Font(bold=True)
            self.header_alignment = Alignment(horizontal='center')
            self.header_border = Border(bottom=Side(style='thin'))
        else:
            raise TypeError("Only ColumnControl objects can be added.")

    @property
    def column_names(self):
        """ 回傳所有欄位名稱 """
        return [col.name for col in self.columns]

    def apply_header_format(self, cell):
        """對指定的 Excel 單元格應用標題樣式"""
        cell.font = self.header_font
        cell.alignment = self.header_alignment
        cell.border = self.header_border

    @property
    def column_letter(self):
        """
        產生 {欄位名稱: Excel 欄位字母} 對應表
        """
        return {col.name: get_column_letter(i + 1) for i, col in enumerate(self.columns)}

    @property
    def column_index(self):
        """
        產生 {欄位名稱: Excel 欄位順序} 對應表
        """
        return {col.name: i for i, col in enumerate(self.columns)}

    def __repr__(self):
        return f"DataControl(columns={self.columns})"

    @property
    def header_columns(self):
        """
        產生 {欄位名稱: 顯示名稱} 對應表
        """
        return {col.name: col.header_name for col in self.columns}

    def apply_formatting(self, worksheet):
        red_fill = PatternFill(start_color="FF99CC", end_color="FF99CC", fill_type="solid")
        blue_fill = PatternFill(start_color="9FD7F9", end_color="9FD7F9", fill_type="solid")
        white_bold_font = Font(color='000000', bold=True)

        for i, col in enumerate(self.columns, start=1):  # Excel 欄位從 1 開始
            col_letter = get_column_letter(i)
            cell = worksheet.cell(row=1, column=i)
            cell.font = col.font  # 套用字型
            cell.alignment = Alignment(horizontal="center")
            self.apply_header_format(cell)

            if col.hidden:  # 隱藏欄位
                worksheet.column_dimensions[col_letter].hidden = True

            if col.level:
                worksheet.column_dimensions[col_letter].outlineLevel = col.level

            if col.width is not None:
                worksheet.column_dimensions[col_letter].width = col.width  # 套用指定欄寬
            else:
                max_length = max(len(str(cell.value)) for cell in worksheet[col_letter][1:])
                worksheet.column_dimensions[col_letter].width = max_length + 5

            if col.comment:  # 如果有提供註解
                comment = Comment(col.comment, "System")
                comment.width = col.comment_width  # 設定註解寬度
                cell.comment = comment

            if col.group == "AVG_SPEED":
                avg_speed_col = self.column_letter['avg_speed']
                lower_speed_col = self.column_letter['LineSpeedLower']
                upper_speed_col = self.column_letter['LineSpeedUpper']

                worksheet.conditional_formatting.add(f'{avg_speed_col}2:{avg_speed_col}{worksheet.max_row}',
                    FormulaRule(
                      formula=[f"IF(ISNUMBER({lower_speed_col}2), {lower_speed_col}2 > {avg_speed_col}2, FALSE)"],
                      fill=blue_fill,
                      font=white_bold_font
                    )
                )

                worksheet.conditional_formatting.add(f'{avg_speed_col}2:{avg_speed_col}{worksheet.max_row}',
                    FormulaRule(
                      formula=[f"=IF(ISNUMBER({upper_speed_col}2),{avg_speed_col}2 >= {upper_speed_col}2 * 1.02, FALSE)"],
                      fill=red_fill,
                      font=white_bold_font
                    )
                )

            # 使用 CellIsRule 設置條件格式，只需設定一次
            if col.limit:
                if col.limit[0]:
                    lower = col.limit[0]
                    worksheet.conditional_formatting.add(f'{col_letter}2:{col_letter}{worksheet.max_row}',
                    CellIsRule(operator='greaterThanOrEqual', formula=[lower], fill=red_fill, font=white_bold_font))
                if col.limit[1]:
                    upper = col.limit[1]
                    worksheet.conditional_formatting.add(f'{col_letter}2:{col_letter}{worksheet.max_row}',
                    CellIsRule(operator='lessThan', formula=[upper], fill=red_fill, font=white_bold_font))


        # 轉換數據型別
        if col.apply_format != False and not self.fix_mode:
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                for col, cell in zip(self.columns, row):
                    col.apply_data_format(cell)

    def __repr__(self):
        return f"DataControl(columns={self.columns})"


class ConfigObject:
    def read_config(self, config_file):
        config = configparser.ConfigParser()
        config.read(config_file, encoding='utf-8')

        self.location = config['Settings'].get('location')
        self.plants = config['Settings'].get('plants', '').split(',')
        self.hour_output_limit = int(config['Settings'].get('hour_output_limit', 0))
        self.fix_mode = config['Settings'].getboolean('fix_mode', False)
        self.report_font = config['Settings'].get('report_font', 'Arial')

    def read_mail_config(self, config_file):
        self.smtp_config = {}
        self.to_emails = []
        self.admin_emails = []

        current_section = None

        with open(config_file, 'r') as file:
            for line in file:
                line = line.strip()

                # Skip empty lines and comments
                if not line or line.startswith('#'):
                    continue

                # Detect section headers
                if line.startswith('[') and line.endswith(']'):
                    current_section = line[1:-1].lower()
                    continue

                # Read lines based on the current section
                if current_section == 'smtp':
                    if '=' in line:
                        key, value = line.split('=', 1)
                        self.smtp_config[key.strip()] = value.strip()
                elif current_section == 'recipients':
                    self.to_emails.append(line)
                elif current_section == 'admin_email':
                    self.admin_emails.append(line)

    def __init__(self, config_file, mail_config_file):
        self.read_config(config_file)
        self.read_mail_config(mail_config_file)


class SetReportLog(ABC):
    # 建立 logger
    logger = logging.getLogger()
    logger.setLevel(logging.INFO)

    # 建立輸出到終端機（命令視窗）的 handler
    console_handler = logging.StreamHandler(sys.stdout)
    console_handler.setLevel(logging.INFO)
    console_formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
    console_handler.setFormatter(console_formatter)

    # 建立輸出到檔案的 handler
    file_handler = logging.FileHandler('logfile.log', encoding='utf-8')
    file_handler.setLevel(logging.INFO)
    file_formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
    file_handler.setFormatter(file_formatter)

    # 清除預設 handler（避免重複 log）
    if logger.hasHandlers():
        logger.handlers.clear()

    # 加入兩個 handler
    logger.addHandler(console_handler)
    logger.addHandler(file_handler)