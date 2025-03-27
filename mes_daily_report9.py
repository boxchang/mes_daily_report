import sys
import os
import numpy as np
from matplotlib.ticker import FuncFormatter
from database import mes_database, mes_olap_database, vnedc_database
from openpyxl.formatting.rule import CellIsRule

curPath = os.path.abspath(os.path.dirname(__file__))
rootPath = os.path.split(curPath)[0]
sys.path.append(rootPath)
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from email.mime.image import MIMEImage
import pandas as pd
from openpyxl.styles import Alignment, NamedStyle, Font, Border, Side, PatternFill
import matplotlib.pyplot as plt
from io import BytesIO
import logging
import time
from openpyxl.comments import Comment


class mes_daily_report(object):
    report_date1 = ""
    report_date2 = ""

    # Define Style
    percent_style = NamedStyle(name='percent_style', number_format='0.00%')
    right_align_style = NamedStyle(name='right_align_style', alignment=Alignment(horizontal='right'))
    center_align_style = NamedStyle(name='center_align_style', alignment=Alignment(horizontal='center'))

    # Define Header
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal='center')
    header_border = Border(bottom=Side(style='thin'))
    header_columns = {
        'Name': '機台號',
        'WorkOrderId': '工單',
        'PartNo': '料號',
        'ProductItem': '品項',
        'Line': '線別',
        'Shift': '班別',
        'max_speed': '車速(最高)',
        'min_speed': '車速(最低)',
        'avg_speed': '車速(平均)',
        'sum_qty': '產量(加總)',
        'Ticket_Qty': '入庫量(加總)',
        'Good_Qty': '包裝確認量',
        'ProductionTime': '生產時間',
        'LineSpeedStd': '標準車速',
        'Target': '目標產能',
        'Separate': '針孔',
        'Scrap': '廢品',
        'SecondGrade': '二級品',
        'OverControl': '超內控',
        'WeightValue': 'IPQC克重',
        'WeightLower': '重量下限',
        'WeightUpper': '重量上限',
        'Activation': '稼動率',
        'OpticalNGRate': '光檢不良率',
        'WoStartDate': '工單開始時間',
        'WoEndDate': '工單結束時間'
    }

    # 配置日志记录器
    logging.basicConfig(
        level=logging.INFO,  # 设置日志级别为 DEBUG，这样所有级别的日志都会被记录
        format='%(asctime)s - %(levelname)s - %(message)s',  # 指定日志格式
        filename='app.log',  # 指定日志文件
        filemode='w'  # 写入模式，'w' 表示每次运行程序时会覆盖日志文件
    )

    def __init__(self, report_date1, report_date2):

        self.report_date1 = report_date1
        self.report_date2 = report_date2
        self.mes_db = mes_database()
        self.mes_olap_db = mes_olap_database()

    def read_config(self, config_file):
        smtp_config = {}
        to_emails = []
        admin_emails = []

        current_section = None

        with open(config_file, 'r') as file:
            for line in file:
                line = line.strip()

                # Skip empty lines and comments
                if not line or line.startswith('#'):
                    continue

                # Detect section headers
                if line.startswith('[') and line.endswith(']'):
                    current_section = line[1:-1].lower()
                    continue

                # Read lines based on the current section
                if current_section == 'smtp':
                    if '=' in line:
                        key, value = line.split('=', 1)
                        smtp_config[key.strip()] = value.strip()
                elif current_section == 'recipients':
                    to_emails.append(line)
                elif current_section == 'admin_email':
                    admin_emails.append(line)

        return smtp_config, to_emails, admin_emails

    def send_email(self, file_list, image_buffers, data_date, error_msg, normal_msg=None):
        logging.info(f"Start to send Email")
        smtp_config, to_emails, admin_emails = self.read_config('mes_daily_report_mail.config')

        # SMTP Sever config setting
        smtp_server = smtp_config.get('smtp_server')
        smtp_port = int(smtp_config.get('smtp_port', 587))
        smtp_user = smtp_config.get('smtp_user')
        smtp_password = smtp_config.get('smtp_password')
        sender_alias = "GD Report"
        sender_email = smtp_user
        # Mail Info
        msg = MIMEMultipart()
        msg['From'] = f"{sender_alias} <{sender_email}>"
        msg['To'] = ', '.join(to_emails)
        msg['Subject'] = f'[GD Report] 產量日報表 {data_date}'

        # Mail Content
        html = f"""\
                <html>
                  <body>
                  {normal_msg}
                """
        for i in range(len(image_buffers)):
            html += f'<img src="cid:chart_image{i}"><br>'

        html += """\
                  </body>
                </html>
                """

        msg.attach(MIMEText(html, 'html'))

        # Attach Excel
        for file in file_list:
            excel_file = file['excel_file']
            file_name = file['file_name']
            with open(excel_file, 'rb') as attachment:
                part = MIMEBase('application', 'octet-stream')
                part.set_payload(attachment.read())
                encoders.encode_base64(part)
                part.add_header('Content-Disposition', f"attachment; filename= {file_name}")
                msg.attach(part)
        logging.info(f"Attached Excel files")

        # Attach Picture
        for i, buffer in enumerate(image_buffers):
            image = MIMEImage(buffer.read())
            image.add_header('Content-ID', f'<chart_image{i}>')
            msg.attach(image)
        logging.info(f"Attached Picture")

        # Send Email
        try:
            # server = smtplib.SMTP(smtp_server, smtp_port)
            # server.starttls()  # 启用 TLS 加密
            # server.login(smtp_user, smtp_password)  # 登录到 SMTP 服务器
            # server.sendmail(smtp_user, to_emails, msg.as_string())
            # server.quit()

            # 發送郵件（不進行密碼驗證）
            server = smtplib.SMTP(smtp_server, smtp_port)
            server.ehlo()  # 啟動與伺服器的對話
            server.sendmail(smtp_user, to_emails, msg.as_string())
            print("Sent Email Successfully")
        except Exception as e:
            print(f"Sent Email Fail: {e}")
            logging.info(f"Sent Email Fail: {e}")
        finally:
            attachment.close()

    def send_admin_email(self, file_list, image_buffers, data_date, error_msg, normal_msg=None):
        logging.info(f"Start to send Email")
        smtp_config, to_emails, admin_emails = self.read_config('mes_daily_report_mail.config')

        # SMTP Sever config setting
        smtp_server = smtp_config.get('smtp_server')
        smtp_port = int(smtp_config.get('smtp_port', 587))
        smtp_user = smtp_config.get('smtp_user')
        smtp_password = smtp_config.get('smtp_password')
        sender_alias = "GD Report"
        sender_email = smtp_user
        # Mail Info
        msg = MIMEMultipart()
        msg['From'] = f"{sender_alias} <{sender_email}>"
        msg['To'] = ', '.join(admin_emails)
        msg['Subject'] = f'[GD Report] 產量日報表 {data_date} 資料異常, 取消派送'

        # Mail Content
        html = f"""\
                <html>
                  <body>
                  {error_msg}
                """
        for i in range(len(image_buffers)):
            html += f'<img src="cid:chart_image{i}"><br>'

        html += """\
                  </body>
                </html>
                """

        msg.attach(MIMEText(html, 'html'))

        # Attach Excel
        for file in file_list:
            excel_file = file['excel_file']
            file_name = file['file_name']
            with open(excel_file, 'rb') as attachment:
                part = MIMEBase('application', 'octet-stream')
                part.set_payload(attachment.read())
                encoders.encode_base64(part)
                part.add_header('Content-Disposition', f"attachment; filename= {file_name}")
                msg.attach(part)
        logging.info(f"Attached Excel files")

        # Attach Picture
        for i, buffer in enumerate(image_buffers):
            image = MIMEImage(buffer.read())
            image.add_header('Content-ID', f'<chart_image{i}>')
            msg.attach(image)
        logging.info(f"Attached Picture")

        # Send Email
        try:
            # server = smtplib.SMTP(smtp_server, smtp_port)
            # server.starttls()  # 启用 TLS 加密
            # server.login(smtp_user, smtp_password)  # 登录到 SMTP 服务器
            # server.sendmail(smtp_user, to_emails, msg.as_string())
            # server.quit()

            # 發送郵件（不進行密碼驗證）
            server = smtplib.SMTP(smtp_server, smtp_port)
            server.ehlo()  # 啟動與伺服器的對話
            server.sendmail(smtp_user, admin_emails, msg.as_string())
            print("Sent Email Successfully")
        except Exception as e:
            print(f"Sent Email Fail: {e}")
            logging.info(f"Sent Email Fail: {e}")
        finally:
            attachment.close()

    def is_special_date(self, date):
        result = False

        sql = f"""
        SELECT CONTROL_DATE FROM [MES_OLAP].[dbo].[special_date] WHERE job_name = 'mes_daily_report'
        AND CONTROL_DATE = CONVERT(DATE, '{date}', 112)
        """
        raws = self.mes_olap_db.select_sql_dict(sql)

        if len(raws) > 0:
            result = True

        return result

    def get_df_fix(self, report_date1, report_date2):

        sql = f"""
        SELECT WorkDate CountingDate, Machine Name, Line, Period, MinSpeed, MaxSpeed, AvgSpeed, CountingQty
          FROM [MES_OLAP].[dbo].[counting_daily_info_fix] where 
          WorkDate between '{report_date1}' and '{report_date2}'
        """

        raws = self.mes_olap_db.select_sql_dict(sql)
        df = pd.DataFrame(raws)

        return df



    def main(self, fix_mode):
        report_date1 = self.report_date1
        report_date2 = self.report_date2

        if self.is_special_date(report_date1):
            sys.exit()

        file_list = []
        error_list = []
        msg_list = []

        # Email Attachment
        image_buffers = []

        # Save Path media/daily_output/
        save_path = os.path.join("daily_output")

        # Check folder to create
        if not os.path.exists(save_path):
            os.makedirs(save_path)

        for plant in ['NBR', 'PVC']:
            try:
                special_output = False
                is_greater_than_24 = False

                logging.info(f"{plant} start running......")
                # Create Excel file
                file_name = f'MES_{plant}_DAILY_Report_{report_date1}.xlsx'
                excel_file = os.path.join(save_path, file_name)

                df_main = self.get_df_main(self.mes_db, report_date1, report_date2, plant)

                df_detail = self.get_df_detail(self.mes_db, report_date1, report_date2, plant)

                df_final = pd.merge(df_main, df_detail, on=['Name', 'Period', 'Line'], how='left')

                df_fix = self.get_df_fix(report_date1, report_date2)

                if len(df_fix) > 0:
                    df_final = pd.merge(df_final, df_fix, on=['CountingDate', 'Name', 'Period', 'Line'], how='left')

                    # 點數機資料修正
                    df_final.loc[
                        df_final["CountingQty"].notna(), ["max_speed", "min_speed", "avg_speed", "sum_qty"]] = \
                        df_final.loc[df_final["CountingQty"].notna(), ["MaxSpeed", "MinSpeed", "AvgSpeed", "CountingQty"]].values

                # 檢查欄位 LineSpeedStd 是否有空值
                df_filtered = df_main[df_main['Line'].notnull()]
                isNoStandard = df_filtered['LineSpeedStd'].isnull().any()

                df_selected = df_final[
                    ['Date', 'Name', 'Line', 'Shift', 'WorkOrderId', 'PartNo', 'ProductItem', 'AQL', 'ProductionTime',
                     'Period', 'max_speed', 'min_speed', 'avg_speed', 'LineSpeedStd', 'sum_qty', 'Ticket_Qty', 'Good_Qty', 'Separate',
                     'Target', 'Scrap', 'SecondGrade', 'OverControl', 'WeightValue', 'OpticalNGRate', 'WeightLower',
                     'WeightUpper', 'WoStartDate', 'WoEndDate', ]]

                df_with_subtotals, df_chart, df_activation = self.sorting_data(self.mes_db, df_selected)

                with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
                    self.generate_summary_excel(writer, df_with_subtotals)

                    machine_groups = df_selected.groupby('Name')
                    for machine_name, machine_df in machine_groups:
                        # 處理停機情況
                        if not machine_df['ProductItem'].iloc[0]:
                            continue

                        machine_clean_df = machine_df.sort_values(by=['Date', 'Shift', 'Period'])
                        self.generate_excel(writer, machine_clean_df, plant, machine_name)

                        # 因同時生產兩種尺寸的工單，使用舊點數機人工作業分類，故無法取得正確資料進行計算
                        printed_machines = set()
                        check_df = machine_df.groupby(['Name', 'Line', 'Date', 'Shift', 'Period'])[
                            'ProductItem'].nunique().reset_index()
                        conflict_rows = check_df[check_df['ProductItem'] > 1]
                        for _, row in conflict_rows.iterrows():
                            key = (row['Name'], row['Line'])
                            if key not in printed_machines:
                                msg_list.append(f"{row['Name']} {row['Line']} 邊因同時生產兩種尺寸的工單，使用舊點數機人工作業分類，故無法取得正確資料進行計算。")
                                printed_machines.add(key)
                    # 稼動率Raw Data
                    self.generate_activation_excel(writer, df_activation)

                file_list.append({'file_name': file_name, 'excel_file': excel_file})

                # Generate Chart
                image_buffer = self.generate_chart(save_path, plant, report_date1, df_chart)
                image_buffers.append(image_buffer)

                # 生產時間不可能超過24小時，防呆檢查
                numeric_production_time = df_with_subtotals['生產時間'].str.rstrip('H').astype(float)
                machines_exceeding_24 = df_with_subtotals.loc[numeric_production_time > 24, '機台號'].unique()
                if machines_exceeding_24.size > 0:
                    for machine in machines_exceeding_24:
                        for normal_msg in msg_list:
                            if machine in normal_msg:
                                break
                            else:
                                error_list.append(f"{machine}發生總時數超過24，可能IPQC有用錯RunCard的情況")

                # 判斷是否有用其他方式收貨，要去詢問產線異常原因
                for _, row in df_final.iterrows():
                    if not pd.isna(row['sum_qty']) and not pd.isna(row['Ticket_Qty']):
                        if int(row['sum_qty']) < 100 and int(row['Ticket_Qty']) > 1000:
                            abnormal_machine = row['Name']
                            # 判斷正常情況不歸屬點數機異常
                            for normal_msg in msg_list:
                                if abnormal_machine in normal_msg:
                                    break
                                else:
                                    error_list.append(f"{abnormal_machine} 點數機資料與SAP入庫資料差異過大，可能發生用舊點數機的情況")

                if isNoStandard:
                    error_list.append(f"有品項尚未維護標準值，無法計算目標產量")

                logging.info(f"{plant} save raw data")

            except Exception as e:
                logging.info(f"{e}")

        print("isCountingError Check")
        logging.info(f"isCountingError Check")

        isCountingErrorResult, error_device = self.isCountingError(report_date1, report_date2)
        if isCountingErrorResult:
            for device in error_device:
                # 判斷正常情況不歸屬點數機異常
                for normal_msg in msg_list:
                    if device in normal_msg:
                        break
                    else:
                        error_list.append(f"{device}點數機資料異常")
        error_msg = '<br>'.join(error_list)
        normal_msg = '<br>'.join(msg_list)
        normal_msg = normal_msg + '<br>'

        if len(error_list) > 0:
            if not fix_mode:
                self.send_admin_email(file_list, image_buffers, report_date1, error_msg)
                print('Admin Email sent successfully')
                logging.info(f"Admin Email sent successfully")
        else:
            # Send Email
            if not fix_mode:
                max_reSend = 5
                reSent = 0
                while reSent < max_reSend:
                    try:
                        self.send_email(file_list, image_buffers, report_date1, error_msg, normal_msg=normal_msg)
                        print('Email sent successfully')
                        logging.info(f"Email sent successfully")
                        break
                    except Exception as e:
                        print(f'Email sending failed: {e}')
                        logging.info(f"Email sending failed: {e}")
                        reSent += 1
                        if reSent >= max_reSend:
                            print('Failed to send email after 5 retries')
                            logging.info(f"Failed to send email after 5 retries")
                            break
                        time.sleep(180)  # seconds

    def isCountingError(self, report_date1, report_date2):
        result = False
        error_device = []

        sql = f"""
        SELECT *
        FROM [PMG_DEVICE].[dbo].[COUNTING_DATA] c, [PMG_DEVICE].[dbo].[COUNTING_DATA_MACHINE] m
        where CreationTime between CONVERT(DATETIME, '{report_date1} 06:00:00', 120) and CONVERT(DATETIME, '{report_date2} 05:59:59', 120)
        and c.MachineName = m.counting_machine 
        and qty2 > 1000
        """
        rows = self.mes_db.select_sql_dict(sql)

        if len(rows) > 0:
            for row in rows:
                error_device.append(row['device_group'])
            result = True
        else:
            result = False

        return result, error_device

    def delete_mes_olap(self, report_date1, report_date2, plant):

        table_name = '[MES_OLAP].[dbo].[mes_daily_report_raw]'
        sql_delete = f"""
            delete
            from {table_name}
            WHERE Name like '%{plant}%' and ((date = '{report_date2}' AND period BETWEEN 0 AND 5)
            OR (date = '{report_date1}' AND period BETWEEN 6 AND 23))
        """
        self.mes_olap_db.execute_sql(sql_delete)

    def insert_mes_olap(self, df):
        try:
            df = df.replace({np.nan: 'null'})

            table_name = '[MES_OLAP].[dbo].[mes_daily_report_raw]'
            for index, row in df.iterrows():
                Date = row['Date']
                Name = row['Name']
                Line = row['Line']
                Shift = row['Shift']
                WorkOrderId = row['WorkOrderId']
                PartNo = row['PartNo']
                ProductItem = row['ProductItem']
                AQL = row['AQL']
                ProductionTime = row['ProductionTime']
                Period = row['Period']
                max_speed = row['max_speed']
                min_speed = row['min_speed']
                avg_speed = row['avg_speed']
                LineSpeedStd = row['LineSpeedStd']
                sum_qty = row['sum_qty']
                Separate = row['Separate']
                Target = row['Target']
                Scrap = row['Scrap']
                SecondGrade = row['SecondGrade']
                OverControl = row['OverControl']
                WeightValue = row['WeightValue']
                WeightLower = row['WeightLower']
                WeightUpper = row['WeightUpper']
                ticket_qty = row['Ticket_Qty']
                good_qty = row['Good_Qty']
                Activation = 'null'
                OpticalNGRate = row['OpticalNGRate']

                if Date != 'null':
                    if int(Period) >= 0 and int(Period) <= 5:
                        belong_to = (datetime.strptime(Date, "%Y-%m-%d") - timedelta(days=1)).strftime("%Y-%m-%d")
                    else:
                        belong_to = Date

                    SN = str(row['Date']).replace('-', '') + str(int(Period)).zfill(2)
                    sql = f"""insert into {table_name}(Date,Name,Line,Shift,WorkOrderId,PartNo,ProductItem,AQL,ProductionTime,
                    Period,max_speed,min_speed,avg_speed,LineSpeedStd,sum_qty,Separate,Target,Scrap,SecondGrade,OverControl,
                    WeightValue,WeightLower,WeightUpper, Activation, update_time,SN,ticket_qty,belong_to, OpticalNGRate, good_qty) Values('{Date}','{Name}','{Line}',N'{Shift}','{WorkOrderId}','{PartNo}',
                    '{ProductItem}','{AQL}',{ProductionTime},{Period},{max_speed},{min_speed},{avg_speed},{LineSpeedStd},
                    {sum_qty},'{Separate}',{Target},'{Scrap}','{SecondGrade}','{OverControl}',{WeightValue},{WeightLower},{WeightUpper}, {Activation}, GETDATE(),{SN},{ticket_qty},'{belong_to}', {OpticalNGRate}, {good_qty})"""
                    # print(sql)
                    self.mes_olap_db.execute_sql(sql)
        except Exception as e:
            print(e)

    def shift(self, period):
        try:
            if 6 <= int(period) <= 17:
                return '早班'
            else:
                return '晚班'
        except Exception as ex:
            return ''

    # Work Order
    def get_df_main(self, db, report_date1, report_date2, plant):
        scada_table = ""
        upper_column = ""
        lower_column = ""

        if plant == "NBR":
            upper_column = "UpperLineSpeed_Min"
            lower_column = "LowerLineSpeed_Min"
            scada_table = "[PMGMES].[dbo].[PMG_MES_NBR_SCADA_Std]"
        elif plant == "PVC":
            upper_column = "UpperSpeed"
            lower_column = "LowerSpeed"
            scada_table = "[PMGMES].[dbo].[PMG_MES_PVC_SCADA_Std]"

        sql = f"""
            WITH WorkOrderInfo AS (
                SELECT 
                    w.MachineId,
                    dl.Name,
                    wi.LineId AS Line,
                    CAST(r.Period AS INT) AS Period,
                    wi.WorkOrderId,
                    wi.StartDate, 
                    wi.EndDate,
                    WorkOrderDate,
                    CustomerName,
                    w.ProductItem,
                    nss.{upper_column} LineSpeedStd,
                    m.COUNTING_MACHINE,
                    w.PartNo,
                    w.AQL,
                    w.PlanQty,
                    wi.Qty,
                    w.Status,
                    ipqc.InspectionStatus weight_result,
                    ipqc.InspectionValue weight_value,
					ipqc.Lower_InspectionValue weight_lower,
					ipqc.Upper_InspectionValue weight_upper,
					hole.InspectionStatus hole_result,
					r.Id runcard,
					r.InspectionDate
                FROM 
                    [PMGMES].[dbo].[PMG_MES_WorkOrderInfo] wi
                    JOIN [PMGMES].[dbo].[PMG_MES_WorkOrder] w ON wi.WorkOrderId = w.id
                    JOIN [PMGMES].[dbo].[PMG_DML_DataModelList] dl ON w.MachineId = dl.Id
                    JOIN [PMGMES].[dbo].[PMG_MES_RunCard] r ON r.WorkOrderId = w.Id AND r.LineName = wi.LineId
                    JOIN [PMGMES].[dbo].[PMG_MES_IPQCInspectingRecord] ipqc ON ipqc.RunCardId = r.Id
                    LEFT JOIN [PMGMES].[dbo].[PMG_MES_IPQCInspectingRecord] hole ON hole.RunCardId = r.Id and hole.OptionName = 'Pinhole'
                    LEFT JOIN {scada_table} nss ON nss.PartNo = w.PartNo
                    JOIN [PMG_DEVICE].[dbo].[COUNTING_DATA_MACHINE] m ON dl.Name = m.MES_MACHINE AND wi.LineId = m.LINE
                WHERE 
                    ((r.InspectionDate = '{report_date1}' AND Period between 6 and 23) or (r.InspectionDate = '{report_date2}' AND Period between 0 and 5))
                    AND ipqc.OptionName = 'Weight'
                    AND dl.Name LIKE '%{plant}%' AND COUNTING_MACHINE LIKE '%CountingMachine%'
            ),
            Pitch AS (
				SELECT Name, 
				CAST(ISNULL(attr1.AttrValue, 1) AS FLOAT) AS std_val, 
				CAST(attr2.AttrValue AS FLOAT) AS act_val, 
				ISNULL(CAST(CAST(attr2.AttrValue AS FLOAT) / CAST(ISNULL(attr1.AttrValue, 1) AS FLOAT) AS FLOAT), 1) AS pitch_rate
				  FROM [PMGMES].[dbo].[PMG_DML_DataModelList] dl
				  LEFT JOIN [PMGMES].[dbo].[PMG_DML_DataModelAttrList] attr1 on dl.Id = attr1.DataModelListId and attr1.AttrName = 'StandardPitch'
				  LEFT JOIN [PMGMES].[dbo].[PMG_DML_DataModelAttrList] attr2 on dl.Id = attr2.DataModelListId and attr2.AttrName = 'ActualPitch'
				  WHERE DataModelTypeId = 'DMT000003'
			),
            CountingData AS (
                SELECT 
                    m.mes_machine,
                    CAST(DATEPART(hour, c.CreationTime) AS INT) AS Period,
                    m.line,
                    COUNT(*) * 5 AS CountedValue
                FROM 
                    [PMG_DEVICE].[dbo].[COUNTING_DATA] c
                    JOIN [PMG_DEVICE].[dbo].[COUNTING_DATA_MACHINE] m ON c.MachineName = m.counting_machine
                WHERE 
                    m.mes_machine LIKE '%{plant}%' AND COUNTING_MACHINE LIKE '%CountingMachine%'
                    AND c.CreationTime BETWEEN CONVERT(DATETIME, '{report_date1} 06:00:00', 120) AND CONVERT(DATETIME, '{report_date2} 05:59:59', 120)
                    AND (c.speed < 60 OR c.speed IS NULL)
                GROUP BY 
                    m.mes_machine,
                    FORMAT(c.CreationTime, 'yyyy-MM-dd'),
                    DATEPART(hour, c.CreationTime),
                    m.line
            ),
			Machines as (
				select distinct MES_MACHINE Name from [PMG_DEVICE].[dbo].[COUNTING_DATA] c
                    join [PMG_DEVICE].[dbo].[COUNTING_DATA_MACHINE] m on c.MachineName = m.COUNTING_MACHINE
                    where creationTime between CONVERT(DATETIME, '{report_date1} 06:00:00', 120) AND CONVERT(DATETIME, '{report_date2} 05:59:59', 120)
                    and m.MES_MACHINE like '%{plant}%'
			),
			Optical as (
				SELECT MES_MACHINE, LINE,CAST(DATEPART(hour, Cdt) AS INT) AS Period, sum(CAST(OKQty AS BIGINT)) OKQty, sum(CAST(NGQty AS BIGINT)) NGQty, 
				CASE 
						 WHEN sum(CAST(NGQty AS BIGINT)) = 0 THEN 0
						 ELSE ROUND(CAST(sum(CAST(NGQty AS BIGINT)) AS FLOAT)/(sum(CAST(OKQty AS BIGINT))+sum(CAST(NGQty AS BIGINT))), 3) 
					   END OpticalNGRate
				  FROM [PMG_DEVICE].[dbo].[OpticalDevice] o
				  JOIN [PMG_DEVICE].[dbo].[COUNTING_DATA_MACHINE] m on o.DeviceId = m.COUNTING_MACHINE
				  where Cdt between CONVERT(DATETIME, '{report_date1} 06:00:00', 120) and CONVERT(DATETIME, '{report_date2} 06:00:00', 120)
				  group by MES_MACHINE, LINE, CAST(DATEPART(hour, Cdt) AS INT)
			),
			Faulty as (
			    SELECT r.InspectionDate,r.Period,f.CreationTime,m.Name,f.ActualQty,f.DefectCode1,c.Descr,r.WorkOrderId,r.LineName,r.Id runcardId
                  FROM [PMGMES].[dbo].[PMG_MES_Faulty] f
                  JOIN [PMGMES].[dbo].[PMG_MES_RunCard] r on f.RunCardId = r.Id
                  JOIN [PMGMES].[dbo].[PMG_DML_DataModelList] m on r.MachineId = m.Id
                  LEFT JOIN [PMGMES].[dbo].[PMG_DML_ConstValue] c on c.[Value] = f.DefectCode1
                  where ((r.InspectionDate = '{report_date1}' AND Period between 6 and 23) or (r.InspectionDate = '{report_date2}' AND Period between 0 and 5))
			),
			Scrap as (
				SELECT r.InspectionDate,r.Period,s.CreationTime,m.Name,s.ActualQty,r.WorkOrderId,r.LineName,r.Id runcardId
                  FROM [PMGMES].[dbo].[PMG_MES_Scrap] s
                  JOIN [PMGMES].[dbo].[PMG_MES_RunCard] r on s.RunCardId = r.Id
                  JOIN [PMGMES].[dbo].[PMG_DML_DataModelList] m on r.MachineId = m.Id
                  where ((r.InspectionDate = '{report_date1}' AND Period between 6 and 23) or (r.InspectionDate = '{report_date2}' AND Period between 0 and 5))
			),
			WorkInProcess as (
				SELECT RunCardId,sum(ActualQty) ActualQty from [PMGMES].[dbo].[PMG_MES_WorkInProcess] wip
				JOIN WorkOrderInfo w on w.runcard = wip.RunCardId
				group by RunCardId
			),
			GoodStock as (
				SELECT RunCardId,sum(ActualQty) ActualQty from [PMGMES].[dbo].[PMG_MES_WorkInProcess] wip
				JOIN WorkOrderInfo w on w.runcard = wip.RunCardId
				AND PackingType = 'OnlinePacking'
				group by RunCardId
			)

            SELECT 
                wo.MachineId,
                mach.Name,
                wo.Line,
                wo.Period,
                wo.StartDate, 
                wo.EndDate,
                wo.WorkOrderId,
                wo.WorkOrderDate,
                wo.CustomerName,
                wo.PartNo,
                wo.ProductItem,
                wo.AQL,
                wo.PlanQty,
                wo.Qty,
                wo.Status,
                CAST(wo.LineSpeedStd AS FLOAT) AS LineSpeedStd,
                60 AS ProductionTime,
                hole_result Separate,
                ISNULL(s.ActualQty, 0) Scrap,
                ISNULL(f.ActualQty, 0) SecondGrade,
                CAST(60 * wo.LineSpeedStd/pitch_rate AS INT) AS Target,
                weight_result OverControl,
                CAST(round(weight_value,2) AS DECIMAL(10, 2)) WeightValue,
                OpticalNGRate,
				CAST(round(weight_lower,2) AS DECIMAL(10, 2)) WeightLower,
				CAST(round(weight_upper,2) AS DECIMAL(10, 2)) WeightUpper,
                runcard,
                t.ActualQty Ticket_Qty,
                gs.ActualQty Good_Qty,
                wo.StartDate WoStartDate, 
                wo.EndDate WoEndDate,
				wo.InspectionDate AS Date
            FROM 
                Machines mach
                LEFT JOIN WorkOrderInfo wo ON mach.Name = wo.Name
                LEFT JOIN CountingData cd ON wo.Name = cd.mes_machine AND wo.Line = cd.line AND wo.Period = cd.Period
                LEFT JOIN Optical o ON wo.Name = o.MES_MACHINE AND wo.Line = o.LINE AND wo.Period = o.Period
                LEFT JOIN Faulty f ON wo.runcard = f.runcardId
                LEFT JOIN Scrap s ON wo.runcard = s.runcardId
                LEFT JOIN WorkInProcess t on wo.runcard = t.RunCardId
                LEFT JOIN GoodStock gs on wo.runcard = gs.RunCardId
                LEFT JOIN Pitch pc on pc.Name = wo.Name
                WHERE NOT (wo.WorkOrderId IS NOT NULL AND t.ActualQty IS NULL) --有小票才列入計算，主要是User會用錯RunCard，以有小票為主進行統計
            ORDER BY 
                mach.Name, 
                wo.Period, 
                wo.Line;
                """
        raws = db.select_sql_dict(sql)

        df_main = pd.DataFrame(raws)

        # Add Column Shift
        df_main['Shift'] = df_main['Period'].apply(self.shift)

        return df_main

    # Counting Machine Data
    def get_df_detail(self, db, report_date1, report_date2, plant):

        sql = f"""
                        SELECT FORMAT(CreationTime, 'yyyy-MM-dd') AS CountingDate,CAST(DATEPART(hour, CreationTime) as INT) Period ,m.mes_machine Name,m.line Line, max(Speed) max_speed,min(Speed) min_speed,round(avg(Speed),0) avg_speed,sum(Qty2) sum_qty
                          FROM [PMG_DEVICE].[dbo].[COUNTING_DATA] c, [PMG_DEVICE].[dbo].[COUNTING_DATA_MACHINE] m
                          where CreationTime between CONVERT(DATETIME, '{report_date1} 06:00:00', 120) and CONVERT(DATETIME, '{report_date2} 05:59:59', 120)
                          and c.MachineName = m.counting_machine and m.mes_machine like '%{plant}%'
                          group by m.mes_machine,FORMAT(CreationTime, 'yyyy-MM-dd'),DATEPART(hour, CreationTime),m.line
                          order by m.mes_machine,FORMAT(CreationTime, 'yyyy-MM-dd'),DATEPART(hour, CreationTime),m.line
                        """
        detail_raws = db.select_sql_dict(sql)
        df_detail = pd.DataFrame(detail_raws)

        return df_detail

    def generate_excel(self, writer, df, plant, machine_name):
        colmn_index = {'Date': 0, 'Name': 1, 'Line': 2, 'Shift': 3, 'WorkOrderId': 4, 'PartNo': 5, 'ProductItem': 6,
                       'AQL': 7, 'ProductionTime': 8, 'Period': 9, 'max_speed': 10, 'min_speed': 11,
                       'avg_speed': 12, 'LineSpeedStd': 13, 'sum_qty': 14, 'Ticket_Qty': 15, 'Good_Qty': 16, 'Separate': 17,
                       'Scrap': 18, 'SecondGrade': 19, 'OverControl': 20, 'WeightValue': 21,
                       'OpticalRate': 22, 'WeightLower': 23, 'WeightUpper': 24, 'WoStartDate': 25, 'WoEndDate': 26}
        colmn_letter = {'Date': 'A', 'Name': 'B', 'Line': 'C', 'Shift': 'D', 'WorkOrderId': 'E', 'PartNo': 'F',
                        'ProductItem': 'G',
                        'AQL': 'H', 'ProductionTime': 'I', 'Period': 'J', 'max_speed': 'K', 'min_speed': 'L',
                        'avg_speed': 'M', 'LineSpeedStd': 'N', 'sum_qty': 'O', 'Ticket_Qty': 'P', 'Good_Qty': 'Q',
                        'Separate': 'R', 'Target': 'S',
                        'Scrap': 'T', 'SecondGrade': 'U', 'OverControl': 'V', 'WeightValue': 'W', 'OpticalRate': 'X',
                        'WeightLower': 'Y', 'WeightUpper': 'Z', 'WoStartDate': 'AA', 'WoEndDate': 'AB'}

        # 轉出Excel前進行資料處理
        df['ProductionTime'] = (df['ProductionTime'] // 60).astype(str) + 'H'
        df['Period'] = df['Period'].apply(lambda x: f"{int(x):02}:00")

        # Change column names
        df.rename(columns=self.header_columns, inplace=True)

        namesheet = str(machine_name).split('_')[-1]
        # Write data to the Excel sheet with the machine name as the sheet name
        df.to_excel(writer, sheet_name=namesheet, index=False)

        # Read the written Excel file
        workbook = writer.book
        worksheet = writer.sheets[namesheet]

        # Freeze the first row
        worksheet.freeze_panes = worksheet['A2']

        # Apply Header Style
        for cell in worksheet[1]:  # First line is Header
            cell.font = self.header_font
            cell.alignment = self.header_alignment
            cell.border = self.header_border

        # Formatting
        for col in worksheet.columns:
            max_length = max(len(str(cell.value)) for cell in col)
            col_letter = col[0].column_letter

            worksheet.column_dimensions[col_letter].width = max_length + 5

            # Set alignment
            for cell in col:
                if col_letter in [colmn_letter['max_speed'], colmn_letter['min_speed'], colmn_letter['avg_speed'],
                                  colmn_letter['LineSpeedStd']]:  # Apply right alignment for specific columns
                    cell.alignment = self.right_align_style.alignment
                elif col_letter in [colmn_letter['sum_qty'], colmn_letter['Target'], colmn_letter['Good_Qty']]:
                    cell.number_format = '#,##0'
                    cell.alignment = self.right_align_style.alignment
                elif col_letter in [colmn_letter['WeightValue']]:
                    try:
                        cell.value = float(cell.value)
                    except ValueError:
                        pass
                elif col_letter in [colmn_letter['OpticalRate']]:
                    worksheet.column_dimensions[col_letter].width = 10
                    cell.alignment = self.center_align_style.alignment
                    cell.number_format = '0.0%'
                elif col_letter in [colmn_letter['WeightLower'], colmn_letter['WeightUpper']]:
                    worksheet.column_dimensions[col_letter].hidden = True
                elif col_letter in [colmn_letter['Ticket_Qty']]:
                    cell.number_format = '#,##0'
                    cell.alignment = self.right_align_style.alignment
                    worksheet.column_dimensions[col_letter].hidden = True
                elif col_letter in [colmn_letter['WoStartDate'], colmn_letter['WoEndDate']]:
                    cell.alignment = self.center_align_style.alignment
                    cell.number_format = 'yyyy/mm/dd hh:mm:ss'
                    worksheet.column_dimensions[col_letter].hidden = True
                else:
                    cell.alignment = self.center_align_style.alignment

        for row in range(2, worksheet.max_row + 1):  # 從第2行開始，因為第1行是標題
            weight_value_cell = worksheet[colmn_letter['WeightValue'] + str(row)]
            weight_lower_cell = worksheet[colmn_letter['WeightLower'] + str(row)].value
            weight_upper_cell = worksheet[colmn_letter['WeightUpper'] + str(row)].value

            if weight_lower_cell or weight_upper_cell:
                comment = Comment(text="IPQC範圍(" + weight_lower_cell + "-" + weight_upper_cell + ")",
                                  author="System")  # 創建註解
                weight_value_cell.comment = comment

        return workbook

    def generate_summary_excel(self, writer, df):

        colmn_index = {'Name': 0, 'ProductItem': 1, 'AQL': 2, 'Shift': 3, 'Line': 4, 'max_speed': 5, 'min_speed': 6,
                       'avg_speed': 7, 'LineSpeedStd': 8, 'ProductionTime': 9, 'sum_qty': 10, 'TicketQty': 11, 'Good_Qty': 12,
                       'Separate': 13, 'Scrap': 14, 'SecondGrade': 15, 'Target': 16, 'OverControl': 17,
                       'ActiveRate': 18, 'OpticalNGRate': 19, }
        colmn_letter = {'Name': 'A', 'ProductItem': 'B', 'AQL': 'C', 'Shift': 'D', 'Line': 'E',
                        'max_speed': 'F', 'min_speed': 'G', 'avg_speed': 'H', 'LineSpeedStd': 'I',
                        'ProductionTime': 'J', 'sum_qty': 'K', 'TicketQty': 'L', 'Good_Qty': 'M', 'Separate': 'N', 'Scrap': 'O',
                        'SecondGrade': 'P', 'Target': 'Q', 'OverControl': 'R', 'ActiveRate': 'S',
                        'OpticalNGRate': 'T', }

        # Create a bold font style
        bold_font = Font(bold=True)

        # Create a border style with a bold line above
        thick_border = Border(top=Side(style='thick'), bottom=Side(style='thick'))

        namesheet = "Summary"
        # Write data to the Excel sheet with the machine name as the sheet name
        df.to_excel(writer, sheet_name=namesheet, index=False)

        # Read the written Excel file
        workbook = writer.book
        worksheet = writer.sheets[namesheet]

        # Freeze the first row
        worksheet.freeze_panes = worksheet['A2']

        # Formatting
        for col in worksheet.columns:
            col_letter = col[0].column_letter
            max_length = max(len(str(cell.value)) for cell in col)

            for cell in col:
                if col_letter in [colmn_letter['Name']]:  # 检查是否为指定的列
                    worksheet.column_dimensions[col_letter].width = max_length + 5
                elif col_letter in [colmn_letter['ProductItem']]:
                    worksheet.column_dimensions[col_letter].width = max_length + 5
                    self.left_align_style = Alignment(horizontal='left')
                    cell.alignment = self.left_align_style
                elif col_letter in [colmn_letter['max_speed'], colmn_letter['min_speed'], colmn_letter['avg_speed'],
                                    colmn_letter['LineSpeedStd']]:  # 检查是否为指定的列
                    worksheet.column_dimensions[col_letter].width = 15
                    cell.alignment = self.right_align_style.alignment
                elif col_letter in [colmn_letter['ProductionTime']]:
                    worksheet.column_dimensions[col_letter].width = 15
                    cell.alignment = self.center_align_style.alignment
                elif col_letter in [colmn_letter['sum_qty'], colmn_letter['Target'], colmn_letter['Good_Qty']]:
                    worksheet.column_dimensions[col_letter].width = 15
                    cell.alignment = self.right_align_style.alignment
                    cell.number_format = '#,##0'
                elif col_letter in [colmn_letter['Separate'], colmn_letter['Scrap'], colmn_letter['SecondGrade'],
                                    colmn_letter['OverControl'], colmn_letter['OpticalNGRate']]:
                    worksheet.column_dimensions[col_letter].width = 10
                    cell.alignment = self.center_align_style.alignment
                    cell.number_format = '0.00%'  # 百分比格式，小數點 1 位
                elif col_letter in [colmn_letter['ActiveRate']]:
                    worksheet.column_dimensions[col_letter].width = 10
                    cell.alignment = self.center_align_style.alignment
                    cell.number_format = '0.0%'  # 百分比格式，小數點 1 位
                    worksheet.column_dimensions[col_letter].hidden = True
                elif col_letter in [colmn_letter['TicketQty']]:
                    worksheet.column_dimensions[col_letter].width = 15
                    cell.alignment = self.right_align_style.alignment
                    cell.number_format = '#,##0'
                    worksheet.column_dimensions[col_letter].hidden = True
                else:
                    cell.alignment = self.center_align_style.alignment

        # Search all lines, bold font and bold line above
        index_start = 2
        index_end = 1
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            if row[colmn_index['Line']].value != '':  # Line
                for cell in row[colmn_index['Line']:]:
                    cell.fill = PatternFill(start_color="FDE9D9", end_color="FDE9D9", fill_type="solid")

            if row[colmn_index['ProductItem']].value != '' and row[colmn_index['Shift']].value == '':
                index_end += 1

            if row[colmn_index['Shift']].value != '':  # Shift
                worksheet.row_dimensions.group(index_start, index_end, hidden=True, outline_level=2)

                index_start = index_end + 1
                index_end = index_end + 1
                worksheet.row_dimensions.group(index_start, index_end, hidden=True, outline_level=1)
                index_start = index_end + 1

                for cell in row[colmn_index['Shift']:]:
                    cell.font = bold_font
                    cell.border = Border(top=Side(style='thin'))

            elif row[colmn_index['Name']].value != '':  # Machine
                # Hide detailed data
                worksheet.row_dimensions.group(index_start, index_end, hidden=False, outline_level=0)
                index_start = index_end + 1

                for cell in row:
                    cell.font = bold_font
                    cell.border = thick_border

                # Add a note (comment) to the 'Optical' column
                if str(row[colmn_index['ProductItem']].value).startswith('V S'):
                    note_text = "Yellow Gloves."
                    author = 'System'
                    row[colmn_index['OpticalNGRate']].comment = Comment(note_text, author)

            # 設置欄的 outlineLevel 讓其可以折疊/展開
            worksheet.column_dimensions[colmn_letter['Shift']].outlineLevel = 1
            worksheet.column_dimensions[colmn_letter['Line']].outlineLevel = 1
            worksheet.column_dimensions[colmn_letter['max_speed']].outlineLevel = 1
            worksheet.column_dimensions[colmn_letter['min_speed']].outlineLevel = 1

            # 總共折疊的區域
            worksheet.column_dimensions.group(colmn_letter['Shift'], colmn_letter['min_speed'], hidden=True)

        # Header說明
        comment = Comment(text="點數機(A1B1)生產時間/工單預計生產時間", author="System")
        comment.width = 600
        worksheet[colmn_letter['ActiveRate']+"1"].comment = comment
        comment = Comment(text="標準上限", author="System")
        comment.width = 200
        worksheet[colmn_letter['LineSpeedStd'] + "1"].comment = comment


        return workbook

    def generate_activation_excel(self, writer, df):
        colmn_letter = {'CreationTime': 'A', 'MES_MACHINE': 'B', 'A1_Qty': 'C', 'A1_Speed': 'D', 'A2_Qty': 'E',
                        'A2_Spped': 'F', 'B1_Qty': 'G', 'B1_Speed': 'H', 'B2_Qty': 'I', 'B2_Speed': 'J', }
        namesheet = "稼動RawData"
        # Write data to the Excel sheet with the machine name as the sheet name
        df.to_excel(writer, sheet_name=namesheet, index=False)

        # Read the written Excel file
        workbook = writer.book
        worksheet = writer.sheets[namesheet]

        # Freeze the first row
        worksheet.freeze_panes = worksheet['A2']

        try:
            for col in worksheet.columns:
                col_letter = col[0].column_letter
                max_length = max(len(str(cell.value)) for cell in col)

                for cell in col:
                    if col_letter in [colmn_letter['CreationTime'], colmn_letter['MES_MACHINE']]:  # 检查是否为指定的列
                        worksheet.column_dimensions[col_letter].width = max_length + 5
                        cell.alignment = self.center_align_style.alignment

            # 設置條件格式為黃顏色填充
            yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

            # 使用 CellIsRule 設置條件格式，只需設定一次
            # 此範例將適用於 A1_Qty 和 B1_Qty 列
            for col_letter in [colmn_letter['A1_Qty'], colmn_letter['B1_Qty']]:
                # 假設您知道資料的範圍為 B2:B100，您可以根據實際情況修改範圍
                worksheet.conditional_formatting.add(f'{col_letter}2:{col_letter}65535',
                                                     CellIsRule(operator='lessThanOrEqual',
                                                                formula=['10'],
                                                                fill=yellow_fill))
        except Exception as e:
            print(e)

        return workbook

    # Sorting data
    def sorting_data(self, db, df):

        chart_rows = []

        def join_values(col):
            return '/'.join(map(str, sorted(set(col))))

        def min2hour(col):
            sum_val = col.sum()
            data_list = col.tolist()
            hours = sum_val / 60
            return hours

        def counting_ng_ratio(col):
            data_list = col.tolist()
            ng_count = data_list.count('NG')

            # 計算 NG 的比例
            ng_ratio = ng_count / len(data_list)
            return ng_ratio

        def calculate_activation(mach):
            try:
                wo_time = 86400

                sql = f"""
                    WITH A1 AS (
                    SELECT m.MES_MACHINE,Qty2,Speed,LINE,CreationTime
                                      FROM [PMG_DEVICE].[dbo].[COUNTING_DATA] d
                                      JOIN [PMG_DEVICE].[dbo].[COUNTING_DATA_MACHINE] m on d.MachineName = m.COUNTING_MACHINE
                                      where m.MES_MACHINE = '{mach}' and m.LINE = 'A1'
                                      and CreationTime between CONVERT(DATETIME, '{report_date1} 06:00:00', 120) and CONVERT(DATETIME, '{report_date2} 05:59:59', 120) 
                    ),
                    B1 AS (
                    SELECT m.MES_MACHINE,Qty2,Speed,LINE,CreationTime
                                      FROM [PMG_DEVICE].[dbo].[COUNTING_DATA] d
                                      JOIN [PMG_DEVICE].[dbo].[COUNTING_DATA_MACHINE] m on d.MachineName = m.COUNTING_MACHINE
                                      where m.MES_MACHINE = '{mach}' and m.LINE = 'B1'
                                      and CreationTime between CONVERT(DATETIME, '{report_date1} 06:00:00', 120) and CONVERT(DATETIME, '{report_date2} 05:59:59', 120) 
                    ),
                    A2 AS (
                    SELECT m.MES_MACHINE,Qty2,Speed,LINE,CreationTime
                                      FROM [PMG_DEVICE].[dbo].[COUNTING_DATA] d
                                      JOIN [PMG_DEVICE].[dbo].[COUNTING_DATA_MACHINE] m on d.MachineName = m.COUNTING_MACHINE
                                      where m.MES_MACHINE = '{mach}' and m.LINE = 'A2'
                                      and CreationTime between CONVERT(DATETIME, '{report_date1} 06:00:00', 120) and CONVERT(DATETIME, '{report_date2} 05:59:59', 120) 
                    ),
                    B2 AS (
                    SELECT m.MES_MACHINE,Qty2,Speed,LINE,CreationTime
                                      FROM [PMG_DEVICE].[dbo].[COUNTING_DATA] d
                                      JOIN [PMG_DEVICE].[dbo].[COUNTING_DATA_MACHINE] m on d.MachineName = m.COUNTING_MACHINE
                                      where m.MES_MACHINE = '{mach}' and m.LINE = 'B2'
                                      and CreationTime between CONVERT(DATETIME, '{report_date1} 06:00:00', 120) and CONVERT(DATETIME, '{report_date2} 05:59:59', 120) 
                    )

                    Select FORMAT(A1.CreationTime, 'yyyy-MM-dd HH:mm:ss') CreationTime, A1.MES_MACHINE, A1.Qty2 A1_Qty, A1.Speed A1_Speed, A2.Qty2 A2_Qty, A2.Speed A2_Spped, B1.Qty2 B1_Qty, B1.Speed B1_Speed, B2.Qty2 B2_Qty, B2.Speed B2_Speed from A1 
                    left join A2 on A1.CreationTime = A2.CreationTime
                    join B1 on A1.CreationTime = B1.CreationTime
                    left join B2 on A1.CreationTime = B2.CreationTime

                """
                detail_raws = db.select_sql_dict(sql)

                df = pd.DataFrame(detail_raws)
                filtered_df = df[(df['A1_Qty'] > 10) | (df['B1_Qty'] > 10)]
                run_time = len(filtered_df) * 300
                active_rate = run_time / wo_time
                return round(active_rate, 2), df
            except Exception as e:
                print(e)

        try:
            # Drop the 'Period' and 'Date' column from each group
            group_without_period = df.drop(columns=['Period', 'Date'])

            # Data group by 'Name and then calculating
            mach_grouped = group_without_period.groupby(['Name'])

            rows = []
            chart_rows = []
            activation_rows = []

            for mach_name, mach_group in mach_grouped:

                # 處理停機情況
                if not mach_group['ProductItem'].iloc[0]:
                    subtotal = {
                        'Name': join_values(mach_group['Name']),
                        'ProductItem': '',
                        'AQL': '',
                        'Shift': '',
                        'Line': '',
                        'max_speed': '',
                        'min_speed': '',
                        'avg_speed': '',
                        'LineSpeedStd': '',
                        'ProductionTime': '',
                        'sum_qty': 0,
                        'Ticket_Qty': 0,
                        'Good_Qty': 0,
                        'Separate': '',
                        'Scrap': '',
                        'SecondGrade': '',
                        'Target': 0,
                        'OverControl': '',
                    }
                    subtotal_df = pd.DataFrame([subtotal])
                    chart_rows.append(subtotal_df)
                    continue

                line_grouped = mach_group.groupby(['Shift', 'Line'])

                tmp_rows = []
                for line_name, line_group in line_grouped:
                    # line_sum_df = line_group.groupby(['Name', 'Shift', 'Line']).agg({
                    #     'min_speed': 'min',  # Min speed
                    #     'max_speed': 'max',  # Max speed
                    #     'avg_speed': 'mean',  # Average speed
                    #     'sum_qty': 'sum',
                    #     'ProductionTime': 'sum',
                    #     'UpperLineSpeed': 'mean',
                    #     'Target': 'sum',
                    # }).reset_index()
                    #
                    # # Add AQL Column to fit format
                    # line_sum_df.insert(line_sum_df.columns.get_loc('Name') + 1, 'ProductItem', None)
                    # line_sum_df.insert(line_sum_df.columns.get_loc('Name') + 2, 'AQL', None)


                    # Second Grade
                    line_sum_qty = line_group['sum_qty'].sum()
                    line_secondGrade_qty = line_group['SecondGrade'].sum()
                    line_second_rate = round(float(line_secondGrade_qty) / line_sum_qty, 3) if line_sum_qty > 0 else 0

                    line_scrap_qty = line_group['Scrap'].sum()
                    line_scrap_rate = round(float(line_scrap_qty) / line_sum_qty, 3) if line_sum_qty > 0 else 0

                    line_sum = {
                        'Name': '',
                        'ProductItem': join_values(line_group['ProductItem']),
                        'AQL': join_values(line_group['AQL']),
                        'Shift': join_values(line_group['Shift']),
                        'Line': join_values(line_group['Line']),
                        'max_speed': line_group['max_speed'].max(),
                        'min_speed': line_group['min_speed'].min(),
                        'avg_speed': line_group['avg_speed'].mean(),
                        'LineSpeedStd': join_values(line_group['LineSpeedStd']),
                        'ProductionTime': min2hour(line_group['ProductionTime']),
                        'sum_qty': line_group['sum_qty'].sum(),
                        'Ticket_Qty': line_group['Ticket_Qty'].sum(),
                        'Good_Qty': line_group['Good_Qty'].sum(),
                        'Separate': counting_ng_ratio(line_group['Separate']),
                        'Scrap': line_scrap_rate,
                        'SecondGrade': line_second_rate,
                        'Target': line_group['Target'].sum(),
                        'OverControl': counting_ng_ratio(line_group['OverControl']),
                    }
                    line_sum_df = pd.DataFrame([line_sum])

                    tmp_rows.append(line_sum_df)

                df_tmp = pd.concat(tmp_rows, ignore_index=True)

                # Sorting Data
                # Day Shift
                day_df = df_tmp[df_tmp['Shift'] == '早班']
                day_production_time = 0
                night_production_time = 0
                if not day_df.empty:
                    day_df = day_df.copy()

                    subtotal = {
                        'Name': '',
                        'ProductItem': '',
                        'AQL': '',
                        'Shift': join_values(day_df['Shift']),
                        'Line': '',
                        'max_speed': day_df['max_speed'].max(),
                        'min_speed': day_df['min_speed'].min(),
                        'avg_speed': day_df['avg_speed'].mean(),
                        'LineSpeedStd': join_values(day_df['LineSpeedStd']),
                        'ProductionTime': day_df['ProductionTime'].mean(),
                        'sum_qty': day_df['sum_qty'].sum(),
                        'Ticket_Qty': day_df['Ticket_Qty'].sum(),
                        'Good_Qty': day_df['Good_Qty'].sum(),
                        'Separate': day_df['Separate'].mean(),
                        'Scrap': day_df['Scrap'].mean(),
                        'SecondGrade': day_df['SecondGrade'].mean(),
                        'Target': day_df['Target'].sum(),
                        'OverControl': day_df['OverControl'].mean(),
                    }
                    subtotal_df = pd.DataFrame([subtotal])
                    subtotal_df['avg_speed'] = subtotal_df['avg_speed'].round(0)

                    day_df[['Name', 'Shift']] = ''
                    day_df['avg_speed'] = day_df['avg_speed'].round(0)

                    rows.append(day_df)  # Day row data
                    rows.append(subtotal_df)  # Day Shift total summary

                    if not np.isnan(day_df['ProductionTime'].mean()):
                        day_production_time = day_df['ProductionTime'].mean()

                # Night Shift
                night_df = df_tmp[df_tmp['Shift'] == '晚班']
                if not night_df.empty:
                    night_df = night_df.copy()

                    subtotal = {
                        'Name': '',
                        'ProductItem': '',
                        'AQL': '',
                        'Shift': join_values(night_df['Shift']),
                        'Line': '',
                        'max_speed': night_df['max_speed'].max(),
                        'min_speed': night_df['min_speed'].min(),
                        'avg_speed': night_df['avg_speed'].mean(),
                        'LineSpeedStd': join_values(night_df['LineSpeedStd']),
                        'ProductionTime': night_df['ProductionTime'].mean(),
                        'sum_qty': night_df['sum_qty'].sum(),
                        'Ticket_Qty': night_df['Ticket_Qty'].sum(),
                        'Good_Qty': night_df['Good_Qty'].sum(),
                        'Separate': night_df['Separate'].mean(),
                        'Scrap': night_df['Scrap'].mean(),
                        'SecondGrade': night_df['SecondGrade'].mean(),
                        'Target': night_df['Target'].sum(),
                        'OverControl': night_df['OverControl'].mean(),
                    }
                    subtotal_df = pd.DataFrame([subtotal])
                    subtotal_df['avg_speed'] = subtotal_df['avg_speed'].round(0)

                    night_df[['Name', 'Shift']] = ''
                    night_df['avg_speed'] = night_df['avg_speed'].round(0)

                    rows.append(night_df)  # Night row data
                    rows.append(subtotal_df)  # Night Shift total summary

                    if not np.isnan(night_df['ProductionTime'].mean()):
                        night_production_time = night_df['ProductionTime'].mean()

                # Machine total summary
                activation_rate, df_activation_row = calculate_activation(mach_name)

                # Second Grade
                sum_qty = mach_group['sum_qty'].sum()

                activation_rows.append(df_activation_row)

                tmp_scrap = mach_group['Scrap'].sum()/sum_qty if sum_qty > 0 else 0
                tmp_second = mach_group['SecondGrade'].sum()/sum_qty if sum_qty > 0 else 0

                subtotal = {
                    'Name': join_values(mach_group['Name']),
                    'ProductItem': join_values(mach_group['ProductItem']),
                    'AQL': join_values(mach_group['AQL']),
                    'Shift': '',
                    'Line': '',
                    'max_speed': mach_group['max_speed'].max(),
                    'min_speed': mach_group['min_speed'].min(),
                    'avg_speed': mach_group['avg_speed'].mean(),
                    'LineSpeedStd': join_values(mach_group['LineSpeedStd']),
                    'ProductionTime': day_production_time + night_production_time,
                    'sum_qty': sum_qty,
                    'Ticket_Qty': mach_group['Ticket_Qty'].sum(),
                    'Good_Qty': mach_group['Good_Qty'].sum(),
                    'Separate': counting_ng_ratio(mach_group['Separate']),
                    'Scrap': tmp_scrap,
                    'SecondGrade': tmp_second,
                    'Target': mach_group['Target'].sum(),
                    'OverControl': counting_ng_ratio(mach_group['OverControl']),
                    'Activation': activation_rate,
                    'OpticalNGRate': mach_group['OpticalNGRate'].mean(),
                }
                subtotal_df = pd.DataFrame([subtotal])
                subtotal_df['avg_speed'] = subtotal_df['avg_speed'].round(0)
                rows.append(subtotal_df)  # Machine total summary
                chart_rows.append(subtotal_df)

            # Combine the grouped data into a DataFrame
            df_with_subtotals = pd.concat(rows, ignore_index=True)

            # ProductionTime加上小時的文字
            df_with_subtotals['ProductionTime'] = df_with_subtotals['ProductionTime'].astype(str) + 'H'
            # Change column names
            df_with_subtotals.rename(columns=self.header_columns, inplace=True)

            # Group the total quantity of each machine into a DataFrame
            df_chart = pd.concat(chart_rows, ignore_index=True)

            df_activation = pd.concat(activation_rows, ignore_index=True)

        except Exception as e:
            print(f"An error occurred: {e}")

        return df_with_subtotals, df_chart, df_activation

    def generate_chart(self, save_path, plant, report_date, df_chart):
        # Create Chart
        fig, ax1 = plt.subplots(figsize=(10, 6))

        plt.rcParams['font.sans-serif'] = ['Microsoft YaHei']

        # Only substring Name right 3 characters
        df_chart['Name_short'] = df_chart['Name'].apply(lambda x: x[-3:])

        df_chart['Unfinished'] = (df_chart['Target'] - df_chart['sum_qty']).apply(lambda x: max(x, 0))  # 未達成數量, 負數為0
        df_chart['Achievement Rate'] = df_chart['sum_qty'] / df_chart['Target'] * 100  # 達成率（百分比）
        df_chart.loc[df_chart['Target'] == 0, 'Achievement Rate'] = None  # 當 Target 為 0，將達成率設為 None

        # Draw Bar Chart
        bar_width = 0.6
        bars = ax1.bar(df_chart['Name_short'], df_chart['sum_qty'], width=bar_width, color='lightcoral', label='日目標達成率')
        ax1.bar(df_chart['Name_short'], df_chart['Unfinished'], width=bar_width, bottom=df_chart['sum_qty'],
                color='lightgreen')

        # Create a second Y axis
        # ax2 = ax1.twinx()

        # Draw Line Chart (speed)
        # ax2.plot(df_chart['Name_short'], df_chart['Achievement Rate'], color='red', marker='o')
        # ax2.set_ylabel('Achievement Rate (%)', color='red')

        # Set the X-axis label and the Y-axis label
        ax1.set_xlabel('機台')
        ax1.set_ylabel('日產量')
        # 設置 Y 軸的上限為 120 萬
        if plant == "PVC":
            ax1.set_ylim(0, 800000)
        else:
            ax1.set_ylim(0, 1200000)

        # 自定義 Y 軸以 10 萬為單位
        def y_formatter(x, pos):
            return f'{int(x/10000)}萬'  # 將數值轉換為「萬」的單位顯示

        ax1.yaxis.set_major_formatter(FuncFormatter(y_formatter))

        achieve_rate = 95 if plant == "NBR" else 98

        # 在每個長條圖上方顯示達成率百分比
        for bar, unfinished, rate in zip(bars, df_chart['Unfinished'], df_chart['Achievement Rate']):
            if pd.notnull(rate):  # 僅顯示達成率不為 None 的數值
                height = bar.get_height() + unfinished  # 計算長條的總高度
                if rate < achieve_rate:
                    ax1.text(bar.get_x() + bar.get_width() / 2, height + 20000, f'{rate:.1f}%', ha='center', va='bottom',
                             fontsize=10, color='red',
                             bbox=dict(boxstyle="circle", edgecolor='red', facecolor='none', linewidth=1.5))

                else:
                    ax1.text(bar.get_x() + bar.get_width() / 2, height + 20000, f'{rate:.1f}%', ha='center', va='bottom',
                             fontsize=10)

        plt.title(f'{plant} {report_date} 日產量與日目標達成率 (達成率目標 > {achieve_rate}%)')

        # Display the legend of the bar chart and line chart together
        fig.legend(loc="upper right", bbox_to_anchor=(1, 1), bbox_transform=ax1.transAxes)

        # Save the image to a local file
        image_file = f'{plant}_bar_chart_{report_date}.png'
        image_file = os.path.join(save_path, image_file)

        plt.savefig(image_file)

        # Save the image to a BytesIO object
        image_stream = BytesIO()
        plt.savefig(image_stream, format='png')
        image_stream.seek(0)  # Move the pointer to the beginning of the file
        plt.close()  # Close the image to free up memory

        return image_stream


import argparse
from datetime import datetime, timedelta, date

parser = argparse.ArgumentParser(description="解析外部参数")

parser.add_argument("--fix_mode", action="store_true", help="是否启用修复模式")
parser.add_argument("--start_date", type=str, default=None, help="开始日期 (格式: YYYYMMDD)")
parser.add_argument("--end_date", type=str, default=None, help="结束日期 (格式: YYYYMMDD)")

args = parser.parse_args()

if (args.start_date or args.end_date) and not args.fix_mode:
    parser.error("如果提供了 --start_date 或 --end_date，必须启用 --fix_mode")

print(f"fix_mode: {args.fix_mode}")
print(f"start_date: {args.start_date}")
print(f"end_date: {args.end_date}")

if args.start_date:
    start_date = datetime.strptime(args.start_date, "%Y%m%d")
    print(f"开始日期: {start_date}")
if args.end_date:
    end_date = datetime.strptime(args.end_date, "%Y%m%d")
    print(f"结束日期: {end_date}")

fix_mode = args.fix_mode

if fix_mode:
    start_date = datetime.strptime(args.start_date, "%Y%m%d") if args.start_date else datetime.today()-timedelta(days=1)
    end_date = datetime.strptime(args.end_date, "%Y%m%d") if args.end_date else datetime.today()-timedelta(days=1)
    # start_date = date(2024, 11, 4)
    # end_date = date(2024, 11, 5)

    current_date = start_date
    while current_date <= end_date:
        report_date1 = current_date
        report_date2 = report_date1 + timedelta(days=1)

        report_date1 = report_date1.strftime('%Y%m%d')
        report_date2 = report_date2.strftime('%Y%m%d')
        report = mes_daily_report(report_date1, report_date2)
        report.main(fix_mode)
        current_date += timedelta(days=1)
else:
    report_date1 = datetime.today() - timedelta(days=1)
    report_date1 = report_date1.strftime('%Y%m%d')

    report_date2 = datetime.today()
    report_date2 = report_date2.strftime('%Y%m%d')

    # report_date1 = "20250324"
    # report_date2 = "20250325"

    report = mes_daily_report(report_date1, report_date2)
    report.main(fix_mode)
