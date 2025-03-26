import sys
import os
from PIL import Image as PILImage
from matplotlib.ticker import MultipleLocator, FuncFormatter
import calendar
from database import vnedc_database, mes_database

curPath = os.path.abspath(os.path.dirname(__file__))
rootPath = os.path.split(curPath)[0]
sys.path.append(rootPath)
import pandas as pd
import matplotlib.pyplot as plt
import logging
from datetime import datetime, timedelta
import math
from matplotlib import font_manager
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, NamedStyle, Font, Border, Side, PatternFill
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from email.mime.image import MIMEImage
from openpyxl.drawing.image import Image
from dateutil.relativedelta import relativedelta
import matplotlib.gridspec as gridspec
import numpy as np
from matplotlib import rcParams
from scipy.interpolate import make_interp_spline


class mes_weekly_report(object):
    this_start_date = ""
    this_end_date = ""
    last_end_date = ""
    last_start_date = ""
    db = None
    plant_name = ['NBR', 'PVC']
    mode = ""
    save_path = ""
    date_mark = ""
    file_list = []
    image_buffers = []
    mach_list = ""
    week_range = 15

    # Define Style
    percent_style = NamedStyle(name='percent_style', number_format='0.00%')
    right_align_style = NamedStyle(name='right_align_style', alignment=Alignment(horizontal='right'))
    center_align_style = NamedStyle(name='center_align_style', alignment=Alignment(horizontal='center'))

    # Define Header
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal='center')
    header_border = Border(bottom=Side(style='thin'))
    header_columns = {
        'Name': '機台號',
        'WorkOrderId': '工單',
        'PartNo': '料號',
        'ProductItem': '品項',
        'Line': '線別',
        'Shift': '班別',
        'max_speed': '車速(最高)',
        'min_speed': '車速(最低)',
        'avg_speed': '車速(平均)',
        'sum_qty': '產量(加總)',
        'Ticket_Qty': '入庫量(加總)',
        'ProductionTime': '生產時間',
        'LineSpeedStd': '標準車速',
        'Target': '目標產能',
        'Separate': '隔離',
        'Scrap': '廢品',
        'SecondGrade': '二級品',
        'OverControl': '超內控',
        'WeightValue': 'IPQC克重',
        'WeightLower': '重量下限',
        'WeightUpper': '重量上限',
        'Activation': '稼動率',
        'OpticalNGRate': '光檢不良率',
        'Achievement Rate': '目標達成率',
        'Date': '日期範圍'
    }

    # 配置日志记录器
    logging.basicConfig(
        level=logging.INFO,  # 设置日志级别为 DEBUG，这样所有级别的日志都会被记录
        format='%(asctime)s - %(levelname)s - %(message)s',  # 指定日志格式
        filename='weekly.log',  # 指定日志文件
        filemode='w'  # 写入模式，'w' 表示每次运行程序时会覆盖日志文件
    )

    def __init__(self, mode):
        self.db = vnedc_database()
        today = datetime.now().date()
        self.mode = mode

        last_week_date = today - timedelta(weeks=1)
        self.last_week_info = last_week_date.isocalendar()[1]

        if mode == "MONTHLY":
            this_start_date = today.replace(day=1)

            last_end_date = this_start_date - timedelta(days=1)
            self.last_end_date = last_end_date
            last_start_date = last_end_date.replace(day=1)
            self.last_start_date = last_start_date

            this_end_date = this_start_date - timedelta(days=1)
            self.this_end_date = this_end_date
            this_start_date = last_end_date.replace(day=1)
            self.this_start_date = this_start_date

            fold_name = this_start_date.strftime('%Y%m').zfill(2)
            save_path = os.path.join("monthly_output", fold_name)
        elif mode == "WEEKLY":
            sql = f"""
                WITH CurrentAndPrevious AS (
                    SELECT 
                        *,
                        LAG(start_date) OVER (ORDER BY start_date) AS prev_start_date,
                        LAG(end_date) OVER (ORDER BY start_date) AS prev_end_date
                    FROM 
                        [VNEDC].[dbo].[week_date]
                ),
                Filtered AS (
                    SELECT * 
                    FROM CurrentAndPrevious
                    WHERE GETDATE() BETWEEN start_date AND end_date
                )
                SELECT 
                    *
                FROM 
                    CurrentAndPrevious
                WHERE 
                    start_date = (SELECT prev_start_date FROM Filtered)
                    OR GETDATE() BETWEEN start_date AND end_date;
"""
            date = vnedc_database().select_sql_dict(sql)
            this_end_date = datetime.strptime(date[0]['end_date'], '%Y-%m-%d').date()
            self.this_end_date = this_end_date
            this_start_date = datetime.strptime(date[0]['start_date'], '%Y-%m-%d').date()
            self.this_start_date = this_start_date

            last_end_date = datetime.strptime(date[0]['prev_end_date'], '%Y-%m-%d').date()
            self.last_end_date = last_end_date
            last_start_date = datetime.strptime(date[0]['prev_start_date'], '%Y-%m-%d').date()
            self.last_start_date = last_start_date

            # days_to_sunday = today.weekday()
            # this_end_date = today - timedelta(days=days_to_sunday + 1)
            # self.this_end_date = this_end_date
            # this_start_date = this_end_date - timedelta(days=6)
            # self.this_start_date = this_start_date
            #
            # last_end_date = this_start_date - timedelta(days=1)
            # self.last_end_date = last_end_date
            # last_start_date = last_end_date - timedelta(days=6)
            # self.last_start_date = last_start_date

            fold_name = 'W' + str(self.last_week_info).zfill(2)
            save_path = os.path.join("weekly_output", fold_name)

        self.save_path = save_path
        # Check folder to create
        if not os.path.exists(save_path):
            os.makedirs(save_path)

        self.date_mark = "{start_date}_{end_date}".format(start_date=this_start_date.strftime("%m%d"),
                                                          end_date=this_end_date.strftime("%m%d"))

    def generate_excel(self, writer, df, machine_name):
        colmn_letter = {'Date': 'A', 'Name': 'B', 'Line': 'C', 'Shift': 'D', 'WorkOrderId': 'E', 'PartNo': 'F',
                        'ProductItem': 'G',
                        'AQL': 'H', 'ProductionTime': 'I', 'Period': 'J', 'max_speed': 'K', 'min_speed': 'L',
                        'avg_speed': 'M', 'LineSpeedStd': 'N', 'sum_qty': 'O', 'Ticket_Qty': 'P', 'Separate': 'Q',
                        'Target': 'R',
                        'Scrap': 'S', 'SecondGrade': 'T', 'OverControl': 'U', 'WeightValue': 'V', 'WeightLower': 'W',
                        'WeightUpper': 'X', 'Activation': 'Y', 'OpticalNGRate': 'Z'}

        # Preprocess data for Excel
        df['ProductionTime'] = (df['ProductionTime'] // 60).astype(str) + 'H'
        df['Period'] = df['Period'].apply(lambda x: f"{int(x):02}:00")

        # Rename columns
        df.rename(columns=self.header_columns, inplace=True)

        namesheet = str(machine_name).split('_')[-1]
        save_path = self.save_path
        file_name = f"MES_{machine_name}_Chart.png"
        chart_img = os.path.join(save_path, file_name)
        if os.path.exists(chart_img):
            header_row = 31
            data_start_row = 32
        else:
            header_row = 0
            data_start_row = 1

        # Write data to the Excel sheet
        df.to_excel(writer, sheet_name=namesheet, index=False, startrow=header_row)

        workbook = writer.book
        worksheet = writer.sheets.get(namesheet)

        if not worksheet:
            worksheet = workbook.add_worksheet(namesheet)

        try:
            img = Image(chart_img)
            img.height = 6 * 96
            img.width = 16 * 96
            img.anchor = 'A1'
            worksheet.add_image(img)
        except:
            print('No counting machine data yet!')
            pass
        # Freeze the first row of data
        # worksheet.freeze_panes = worksheet[f'A{data_start_row+1}']

        # Apply Header Style
        for cell in worksheet[data_start_row]:
            cell.font = self.header_font
            cell.alignment = self.header_alignment
            cell.border = self.header_border

        # Adjust column formatting
        for col in worksheet.columns:
            max_length = max(len(str(cell.value)) for cell in col)
            col_letter = col[0].column_letter

            worksheet.column_dimensions[col_letter].width = max_length + 5

            for cell in col:
                if col_letter in [colmn_letter['max_speed'], colmn_letter['min_speed'], colmn_letter['avg_speed'],
                                  colmn_letter['LineSpeedStd']]:
                    cell.alignment = self.right_align_style.alignment
                elif col_letter in [colmn_letter['sum_qty'], colmn_letter['Target']]:
                    cell.number_format = '#,##0'
                    cell.alignment = self.right_align_style.alignment
                elif col_letter in [colmn_letter['WeightValue']]:
                    try:
                        cell.value = float(cell.value)
                    except:
                        pass
                elif col_letter in [colmn_letter['OpticalNGRate']]:
                    worksheet.column_dimensions[col_letter].width = 10
                    cell.alignment = self.center_align_style.alignment
                    cell.number_format = '0.0%'
                elif col_letter in [colmn_letter['WeightLower'], colmn_letter['WeightUpper']]:
                    worksheet.column_dimensions[col_letter].hidden = True
                elif col_letter in [colmn_letter['Ticket_Qty']]:
                    cell.number_format = '#,##0'
                    cell.alignment = self.right_align_style.alignment
                    worksheet.column_dimensions[col_letter].hidden = True
                else:
                    cell.alignment = self.center_align_style.alignment

        # Add comments to WeightValue cells based on WeightLower and WeightUpper
        for row in range(data_start_row, worksheet.max_row + 1):
            weight_value_cell = worksheet[colmn_letter['WeightValue'] + str(row)]
            weight_lower_cell = worksheet[colmn_letter['WeightLower'] + str(row)].value
            weight_upper_cell = worksheet[colmn_letter['WeightUpper'] + str(row)].value

            if weight_lower_cell or weight_upper_cell:
                comment = Comment(text=f"IPQC範圍({weight_lower_cell}-{weight_upper_cell})", author="System")
                weight_value_cell.comment = comment

        return workbook

    def generate_raw_excel(self, plant):
        this_start_date = self.this_start_date
        this_end_date = self.this_end_date
        save_path = self.save_path
        date_mark = self.date_mark
        mode = self.mode

        file_name = f'MES_{plant}_{mode}_Report_{date_mark}.xlsx'
        excel_file = os.path.join(save_path, file_name)
        with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
            sql = f"""SELECT belong_to
                      ,[Name]
                      ,[Line]
                      ,[Shift]
                      ,[WorkOrderId]
                      ,[PartNo]
                      ,[ProductItem]
                      ,[AQL]
                      ,[ProductionTime]
                      ,[Period]
                      ,[max_speed]
                      ,[min_speed]
                      ,[avg_speed]
                      ,[LineSpeedStd]
                      ,[sum_qty]
                      ,[ticket_qty]
                      ,CASE 
                           WHEN [Separate] = 'null' THEN '' 
                           ELSE [Separate] 
                       END AS [Separate]
                      ,[Target]
                      ,[Scrap]
                      ,[SecondGrade]
                      ,[OverControl]
                      ,[WeightValue]
                      ,[WeightLower]
                      ,[WeightUpper]
                      ,[Activation]
                      ,[OpticalNGRate]
                        FROM [MES_OLAP].[dbo].[mes_daily_report_raw]
                        where name like '%{plant}%' and belong_to between '{this_start_date}' and '{this_end_date}'
                        and name not in ('VN_GD_PVC1_L03','VN_GD_PVC1_L04')
                        order by name, belong_to, line, shift, period
                        """
            data = self.db.select_sql_dict(sql)
            df = pd.DataFrame(data)

            machine_groups = df.groupby('Name')
            self.generate_summary(writer, machine_groups)
            self.generate_chart_sheet(writer, plant)

            for machine_name, machine_df in machine_groups:
                self.generate_excel(writer, machine_df, machine_name)

        self.file_list.append(excel_file)

    def generate_chart_sheet(self, writer, plant):
        workbook = writer.book
        x_labels, week_date = self.generate_previous_weeks_with_dates()
        lasted_week = x_labels[-1]
        # chart2
        chart2_worksheet = workbook.create_sheet('產量差異')
        writer.sheets['chart 2'] = chart2_worksheet
        chart2_worksheet.sheet_state = "hidden"

        save_path = self.save_path
        file_name = f"GD-{plant}面向生产管理板.png"
        chart_img = os.path.join(save_path, file_name)

        try:
            img = Image(chart_img)
            img.height = 6 * 96
            img.width = 12 * 96
            img.anchor = 'A1'
            chart2_worksheet.add_image(img)
        except Exception as e:
            print(e)
            pass

        sql = f"""SELECT [WorkOrder]
              ,[PartNo]
              ,[ProductItem]
              ,[WorkDate]
              ,[Machine]
              ,[Line]
              ,[Runcard]
              ,[Period]
              ,[LowSpeed]
              ,[UpSpeed]
              ,[StdSpeed]
              ,[RunTime]
              ,[StopTime]
              ,[CountingQty]
              ,[SAPQuantity]
              ,[Target]
              ,[FaultyQuantity]
              ,[ScrapQuantity]
              ,[InspectionValue]
              ,[InspectionStatus]
              ,[DefectCode]
          FROM [MES_OLAP].[dbo].[counting_daily_info_raw] where MonthWeek = '{lasted_week}' and branch like '%GD{plant}%'
        """
        data = vnedc_database().select_sql_dict(sql)
        df = pd.DataFrame(data)


        header = ['工單', '料號', '品項', '資料日期', '機台', '線別', 'Runcard', '檢驗時間', '最低車速',
                  '最高車速', '標準車速', '生產時間', '停機時間', '點數機數量', '包裝數量', '預估產量',
                  '二級品數量', '廢品數量', '重量', '判定結果', '異常代碼']

        header_row = 32
        for i, row in enumerate(header, start=1):
            chart2_worksheet.cell(row=header_row, column=i, value=row)

        start_row = 33
        for i, row in enumerate(df.values, start=start_row):
            for j, value in enumerate(row, start=1):
                chart2_worksheet.cell(row=i, column=j, value=value)

        # # chart6a
        # chart3_worksheet = workbook.create_sheet('外觀DPM')
        # writer.sheets['chart 6a'] = chart3_worksheet
        # save_path = self.save_path
        # file_name = f"GD-{plant}外觀DPM.png"
        # chart_img = os.path.join(save_path, file_name)
        #
        # try:
        #     img = Image(chart_img)
        #     img.height = 6 * 96
        #     img.width = 12 * 96
        #     img.anchor = 'A1'
        #     chart3_worksheet.add_image(img)
        # except Exception as e:
        #     print(e)
        #     pass

        # chart8a
        chart2_worksheet = workbook.create_sheet('二級品率')
        writer.sheets['chart 8a'] = chart2_worksheet
        chart2_worksheet.sheet_state = "hidden"

        save_path = self.save_path
        file_name = f"GD-{plant}二級品率.png"
        chart_img = os.path.join(save_path, file_name)

        try:
            img = Image(chart_img)
            img.height = 6 * 96
            img.width = 12 * 96
            img.anchor = 'A1'
            chart2_worksheet.add_image(img)
        except Exception as e:
            print(e)
            pass

        # chart8b
        chart2_worksheet = workbook.create_sheet('廢品率')
        writer.sheets['chart 8b'] = chart2_worksheet
        chart2_worksheet.sheet_state = "hidden"

        save_path = self.save_path
        file_name = f"GD-{plant}废品率%.png"
        chart_img = os.path.join(save_path, file_name)

        try:
            img = Image(chart_img)
            img.height = 6 * 96
            img.width = 12 * 96
            img.anchor = 'A1'
            chart2_worksheet.add_image(img)
        except Exception as e:
            print(e)
            pass

    def generate_summary(self, writer, machine_groups):
        colmn_letter = {'Name': 'A', 'Week': 'B', 'shift': 'C', 'Line': 'D', 'Output': 'E', 'Target': 'F',
                        'Achievement Rate': 'G'}
        summary_data = []
        tmp_date = self.date_mark.replace('_', '~')
        mode = self.mode
        this_start_date = self.this_start_date

        if mode == 'WEEKLY':
            tmp_week = f"第{self.last_week_info}週({tmp_date})"
        else:
            tmp_week = f"{this_start_date.strftime('%Y %m')}({tmp_date})"

        thin_border_top = Border(top=Side(style="thin"))
        thin_border_bottom = Border(bottom=Side(style="thin"))
        for machine_name, machine_df in machine_groups:
            for shift in machine_df['Shift'].unique():  # Loop through each unique Line
                for line in machine_df['Line'].unique():
                    filtered_df = machine_df[(machine_df['Line'] == line) & (machine_df['Shift'] == shift)]
                    total_output = filtered_df['sum_qty'].sum()
                    total_target = filtered_df['Target'].sum()
                    rate = round((int(total_output) / int(total_target)), 3) if int(total_target) > 0 else 0

                    summary_row = {
                        'Name': machine_name,
                        'Date': tmp_week,
                        'Shift': shift,
                        'Line': line,
                        'sum_qty': total_output,
                        'Target': total_target,
                        'Achievement Rate': rate
                    }
                    summary_data.append(summary_row)
            avg_rate = [item['Achievement Rate'] for item in summary_data if item['Name'] == machine_name]
            sum_qty = sum(item['sum_qty'] for item in summary_data if item['Name'] == machine_name)
            sum_target = sum(item['Target'] for item in summary_data if item['Name'] == machine_name)

            summary_data.append({'Name': machine_name, 'Date': tmp_week, 'Shift': '', 'Line': '', 'sum_qty': sum_qty,
                                 'Target': sum_target, 'Achievement Rate': round(sum_qty / sum_target, 3)})
        summary_df = pd.DataFrame(summary_data)
        # Change column names
        summary_df.rename(columns=self.header_columns, inplace=True)
        summary_sheet_name = "Summary"
        summary_df.to_excel(writer, sheet_name=summary_sheet_name, index=False)
        # Read the written Excel file
        workbook = writer.book
        worksheet = writer.sheets[summary_sheet_name]

        # Apply Header Style
        for cell in worksheet[1]:  # First line is Header
            cell.font = self.header_font
            cell.alignment = self.header_alignment
            cell.border = self.header_border
            # Formatting

        for col in worksheet.columns:
            max_length = max(len(str(cell.value)) for cell in col)
            col_letter = col[0].column_letter

            worksheet.column_dimensions[col_letter].width = max_length + 5

            # Set alignment
            for cell in col:
                if col_letter in [colmn_letter['Output'],
                                  colmn_letter['Target']]:  # Apply right alignment for specific columns
                    cell.number_format = '#,##0'
                    cell.alignment = self.right_align_style.alignment
                elif col_letter in [colmn_letter['Achievement Rate']]:
                    worksheet.column_dimensions[col_letter].width = 15
                    cell.alignment = self.center_align_style.alignment
                    cell.number_format = '0.0%'
                else:
                    cell.alignment = self.center_align_style.alignment

        name_col = 1
        current_name = None
        group_start = None

        for row in range(2, len(summary_df) + 2):  # Data starts from row 2 in Excel
            name = worksheet.cell(row=row, column=name_col).value

            if name != current_name:
                # Finalize the previous group
                if group_start is not None:
                    last_row = row - 1
                    for col in range(1, summary_df.shape[1] + 1):  # Apply border to all columns
                        worksheet.cell(row=last_row, column=col).border = thin_border_top + thin_border_bottom
                    if row - group_start > 1:
                        worksheet.row_dimensions.group(group_start, row - 2, hidden=True)

                # Start a new group
                current_name = name
                group_start = row

        # Finalize the last group
        if group_start is not None:
            last_row = len(summary_df) + 1
            for col in range(1, summary_df.shape[1] + 1):
                worksheet.cell(row=last_row, column=col).border = thin_border_top + thin_border_bottom
            if len(summary_df) + 1 - group_start > 1:
                worksheet.row_dimensions.group(group_start, len(summary_df) + 1 - 1, hidden=True)

    def generate_chart(self, plant):
        this_start_date = self.this_start_date
        this_end_date = self.this_end_date
        last_start_date = self.last_start_date
        last_end_date = self.last_end_date
        mode = self.mode
        save_path = self.save_path
        date_mark = self.date_mark
        last_week_info = self.last_week_info
        yticks_labels = []

        plt.rcParams['font.sans-serif'] = ['Microsoft YaHei']

        sql = f"""SELECT name,
                    sum(case when cast(belong_to as date) between '{this_start_date}' and '{this_end_date}' then sum_qty else 0 end) as this_time,
                    sum(case when cast(belong_to as date) between '{this_start_date}' and '{this_end_date}' then target else 0 end) as target_this_time,
                    sum(case when cast(belong_to as date) between '{last_start_date}' and '{last_end_date}' then sum_qty else 0 end) as last_time,
                    sum(case when cast(belong_to as date) between '{last_start_date}' and '{last_end_date}' then target else 0 end) as target_last_time,
                    (sum(case when cast(belong_to as date) between '{this_start_date}' and '{this_end_date}' then target else 0 end) -
                    sum(case when cast(belong_to as date) between '{this_start_date}' and '{this_end_date}' then sum_qty else 0 end)) as this_unfinish,
                    (sum(case when cast(belong_to as date) between '{last_start_date}' and '{last_end_date}' then target else 0 end) -
                    sum(case when cast(belong_to as date) between '{last_start_date}' and '{last_end_date}' then sum_qty else 0 end)) as last_unfinish
                    FROM [MES_OLAP].[dbo].[mes_daily_report_raw]
                    where name like '%{plant}%'
                    and name not in ('VN_GD_PVC1_L03','VN_GD_PVC1_L04')
                    group by name
                    order by name"""
        data = self.db.select_sql_dict(sql)

        # Output Bar Chart
        x_labels = [str(item['name']).split('_')[-1] for item in data]
        x_range = range(0, len(x_labels) * 2, 2)

        this_data = [int(item['this_time']) for item in data]
        this_unfinish = [int(item['this_unfinish']) if int(item['this_unfinish']) > 0 else 0 for item in data]

        max_data = max(this_data, default=0)
        step_data = 10

        rounded_max_data = int(
            (((max_data / (10 ** (len(str(max_data)) - 2))) // step_data) * step_data + step_data) * (
                    10 ** (len(str(max_data)) - 2)))
        rounded_step_data = step_data * (10 ** (len(str(max_data)) - 2))

        this_rate = [round((item['this_time'] / item['target_this_time']) * 100, 2) if int(
            item['target_this_time']) > 0 else 0 for item in data]
        last_rate = [round((item['last_time'] / item['target_last_time']) * 100, 2) if int(
            item['target_last_time']) > 0 else 0 for item in data]
        max_rate = max(max(this_rate, default=0), max(last_rate, default=0))
        rounded_max_rate = (math.ceil(max_rate / 10) * 10)
        rounded_step_rate = 20

        bar_width = 0.6
        plt.figure(figsize=(16, 9))
        fig, ax1 = plt.subplots(figsize=(16, 9))
        if mode == "WEEKLY":
            this_month_bars = ax1.bar([i for i in x_range], this_data, width=bar_width,
                                      label=f"週產量",
                                      align='center', color='#10ba81')
            unfinish_bars = ax1.bar([i for i in x_range], this_unfinish, width=bar_width, bottom=this_data,
                                    label=f"週目標差異",
                                    align='center', color='lightgreen')
        if mode == "MONTHLY":
            this_month_bars = ax1.bar([i for i in x_range], this_data, width=bar_width,
                                      label=f"{this_start_date.strftime('%Y %m')}月產量", align='center', color='#10ba81')

            unfinish_bars = ax1.bar([i for i in x_range], this_unfinish, width=bar_width, bottom=this_data,
                                    label=f"{this_start_date.strftime('%Y %m')}月目標差異", align='center',
                                    color='lightgreen')

        ax1.set_xticks(x_range)
        ax1.set_xticklabels(x_labels, rotation=0, ha="center", fontsize=12)

        if len(str(max_data)) > 7:
            yticks_positions = list(range(0, rounded_max_data + 1 * rounded_step_data, rounded_step_data))
            yticks_positions.append(int(rounded_max_data + 2 * rounded_step_data))
            yticks_labels = [f"{int(i//(10**(len(str(max_data)))))}" + '百萬' if len(str(i)) > 6 else f"{i}" for i
                             in yticks_positions]

            for index, bar in enumerate(this_month_bars):
                height = bar.get_height()
                unfinish_height = unfinish_bars[index].get_height()
                ax1.text(
                    bar.get_x() + bar.get_width() / 2,
                    height + unfinish_height,
                    f'{round(height/(10**(len(str(max_data)) - 2)),2)}'[:4] if height > 0 else '',
                    ha='center', va='bottom', fontsize=12  # Align the text
                )
        elif 4 < len(str(max_data)) <= 7:
            yticks_positions = list(range(0, rounded_max_data + 1 * rounded_step_data, rounded_step_data))
            yticks_positions.append(int(rounded_max_data + 4 * rounded_step_data))
            yticks_labels = [f"{int(i//(10**(len(str(max_data)) - 1)))}" + '百萬' if len(str(i)) > 4 and int(
                i // (10 ** (len(str(max_data))))) % 60 == 0 else "" for i in yticks_positions]

            for index, bar in enumerate(this_month_bars):
                height = bar.get_height()
                unfinish_height = unfinish_bars[index].get_height()
                ax1.text(
                    bar.get_x() + bar.get_width() / 2,
                    height + unfinish_height,
                    f'{round(height/(10**(len(str(max_data)) - 1)),2)}'[:4] if height > 0 else '',
                    ha='center', va='bottom', fontsize=12  # Align the text
                )
        yticks_labels[-1] = ""
        ax1.set_yticks(yticks_positions)
        ax1.set_yticklabels(yticks_labels, fontsize=12)

        # Achievement Rate Line Chart (橘色的線)
        ax2 = ax1.twinx()
        if mode == "WEEKLY":
            # line_label = f"{this_start_date.strftime('%d/%m')}-{this_end_date.strftime('%d/%m')}"
            sr_achieve_rate = "本週達成率"
            name = f"{this_start_date.strftime('%d/%m')}-{this_end_date.strftime('%d/%m')}"
            # name1 = f"{last_start_date.strftime('%d/%m')}-{last_end_date.strftime('%d/%m')}"
            # filtered_data = [(x, rate) for x, rate in zip(x_range, last_rate) if rate != 0]
            # x_filtered, last_rate_filtered = zip(*filtered_data)
            # last_rate_line = ax2.plot(x_filtered, last_rate_filtered,
            #                           label=f"{last_start_date.strftime('%d/%m')}-{last_end_date.strftime('%d/%m')}",
            #                           color='#F8CBAD', marker='o', linewidth=1.5)
            filtered_data = [(x, rate) for x, rate in zip(x_range, this_rate) if rate != 0]  # 折線圖上的文字
            x_filtered, this_rate_filtered = zip(*filtered_data)
            this_rate_line = ax2.plot(x_filtered, this_rate_filtered,
                                      label=sr_achieve_rate,
                                      color='#ED7D31', marker='o', linewidth=1.5)

        if mode == "MONTHLY":
            # line_label = f"{this_start_date.strftime('%B %Y')}"
            sr_achieve_rate = "本月達成率"
            name = f"{this_start_date.strftime('%Y %m')}"
            # name1 = f"{last_start_date.strftime('%B %Y')}"
            # filtered_data = [(x, rate) for x, rate in zip(x_range, last_rate) if rate != 0]
            # x_filtered, last_rate_filtered = zip(*filtered_data)
            # last_rate_line = ax2.plot(x_filtered, last_rate_filtered, label=f"{last_start_date.strftime('%B %Y')}",
            #                           color='#F8CBAD', marker='o', linewidth=1.5)
            filtered_data = [(x, rate) for x, rate in zip(x_range, this_rate) if rate != 0]
            x_filtered, this_rate_filtered = zip(*filtered_data)
            this_rate_line = ax2.plot(x_filtered, this_rate_filtered,
                                      label=sr_achieve_rate,
                                      color='#ED7D31', marker='o', linewidth=1.5)

        # Label Name
        sr_target = "達成率目標%"
        # Chart Label
        ry_label = "達成率(%)"
        ly_label = "Product (百萬)"
        if self.mode == "WEEKLY":
            name = f"{this_start_date.strftime('%m/%d')}-{this_end_date.strftime('%m/%d')}"
            title = f"\n{plant} 第{last_week_info}週({name})目標達成率\n"
        else:
            name = f"{this_start_date.strftime('%Y %m')}"
            title = f"\n{plant} {name}月目標達成率\n"

        # Achievement Rate Standard Line
        if plant == "NBR":
            target_rate = 95
        else:
            target_rate = 98
        yticks_positions = list(range(0, rounded_max_rate + 1 * rounded_step_rate, rounded_step_rate))
        if target_rate not in yticks_positions:
            yticks_positions.append(target_rate)
        yticks_positions.append(rounded_max_rate + 0.25 * rounded_step_rate)
        yticks_positions = sorted(yticks_positions)

        yticks_labels = [f"{i}" + '%' for i in yticks_positions]
        yticks_labels[-1] = ""
        ax2.set_yticks(yticks_positions)
        ax2.set_yticklabels(yticks_labels, fontsize=12)
        ax2.axhline(y=target_rate, color='red', linestyle='--', linewidth=1, label=sr_target)
        # ax2.axhline(y=95, color='red', linestyle='--', linewidth=1)

        # ax1.set_xlabel(f'{plant} (line)', labelpad=10, fontsize=12)
        ax1.xaxis.set_label_coords(0.975, -0.014)
        ax1.set_ylabel('Product output', fontsize=12)
        ax2.set_ylabel('Archive rates', fontsize=12)

        # Achievement Rate Label 線上的數字
        # for i, value in enumerate(last_rate):
        #     ax2.text(
        #         x_range[i],
        #         value + 1.5 if value > this_rate[i] else value - 3.5,
        #         f"{value:.1f}" if value != 0 else '',
        #         ha='center', va='bottom', fontsize=8, color='white',
        #         bbox=dict(facecolor='#F8CBAD', edgecolor='none', boxstyle='round,pad=0.3')
        #     )

        for i, value in enumerate(this_rate):
            ax2.text(
                x_range[i],
                value + 1.5 if value > last_rate[i] else value - 3.5,
                f"{value:.1f}%" if value != 0 else '',
                ha='center', va='bottom', fontsize=12, color='white',
                bbox=dict(facecolor='#ED7D31', edgecolor='none', boxstyle='round,pad=0.3')
            )

        handles1, labels1 = ax1.get_legend_handles_labels()
        handles2, labels2 = ax2.get_legend_handles_labels()
        fig.legend(
            handles1 + handles2,
            labels1 + labels2,
            loc='center left',
            fontsize=12,
            title="Note",
            title_fontsize=12,
            bbox_to_anchor=(1.0, 0.5),
            ncol=1
        )

        # Customize axes and borders
        ax = plt.gca()  # Get current axes
        ax.spines['top'].set_color('white')
        ax.spines['top'].set_linewidth(1.5)

        ax1.tick_params(axis='y', colors='black')  # Y-axis ticks color
        y_tick_lines = ax1.get_yticklines()
        y_tick_lines[-2].set_visible(False)
        y_tick_lines[-1].set_visible(False)
        ax1.annotate(
            '',
            xy=(0, 1.0),
            xytext=(0, 0.98),
            xycoords='axes fraction',
            arrowprops=dict(facecolor='black', arrowstyle='-|>,widthA=0.4,widthB=1.4', linewidth=0.5)
        )
        ax1.set_ylabel(ly_label, labelpad=20, rotation=0)
        ax1.yaxis.set_label_coords(-0.01, 1.01)

        ax2.tick_params(axis='y', colors='black')
        y_tick_lines = ax2.get_yticklines()
        y_tick_lines[-2].set_visible(False)
        y_tick_lines[-1].set_visible(False)
        ax2.annotate(
            '',
            xy=(1, 1.0),
            xytext=(1, 0.98),
            xycoords='axes fraction',
            arrowprops=dict(facecolor='black', arrowstyle='-|>,widthA=0.4,widthB=1.4', linewidth=0.5)
        )

        ax2.set_ylabel(ry_label, labelpad=20, rotation=0)
        ax2.yaxis.set_label_coords(1.01, 1.03)

        # plt.text(
        #     x_range[-1] / 2,
        #     -rounded_max_rate * 0.125,
        #     title,
        #     fontsize=16, color='black', ha='center', va='center'
        # )
        plt.title(title, fontsize=20)

        plt.tight_layout()

        file_name = f'MES_{plant}_{mode}_{date_mark}_Chart.png'
        chart_img = os.path.join(save_path, file_name)

        plt.savefig(f"{chart_img}", dpi=100, bbox_inches="tight", pad_inches=0.45)
        plt.close()
        self.image_buffers.append(chart_img)

    # 過濾掉無效數據 (scrap 和 secondgrade)
    def filter_valid_data(self, x_range, y_values):
        filtered_x = [x for i, x in enumerate(x_range) if
                      y_values[i] is not None and not np.isnan(y_values[i])]
        filtered_y = [y for y in y_values if y is not None and not np.isnan(y)]
        return filtered_x, filtered_y

    def rate_chart(self, plant):
        this_start_date = self.this_start_date
        this_end_date = self.this_end_date
        date_mark = self.date_mark
        save_path = self.save_path

        if plant == 'NBR':
            plant_ = 'NBR'
        else:
            plant_ = 'PVC1'

        data1 = self.mach_list

        sql = f"""WITH raw_data AS (
                        SELECT name, date, 
                               CAST(Scrap AS FLOAT) AS Scrap, 
                               CAST(SecondGrade AS FLOAT) AS SecondGrade,
                               CAST(sum_qty as Float) as sum_qty
                        FROM [MES_OLAP].[dbo].[mes_daily_report_raw]
                        WHERE belong_to between '{this_start_date}' AND '{this_end_date}'
                    )
                    SELECT 
                        name,
                        SUM(Case when Scrap > 0 then Scrap else 0 end) AS scrap,
                        SUM(case when SecondGrade > 0 then SecondGrade else 0 end) AS secondgrade,
                        SUM(case when sum_qty > 0 then sum_qty else 0 end) as sum_qty
                    FROM raw_data
                    WHERE name like '%{plant_}%'
                    GROUP BY name
                    ORDER BY name;
                """
        data2 = vnedc_database().select_sql_dict(sql)
        data2_dict = {item['name']: item for item in data2}
        data = []
        for item in data1:
            name = item['name']
            if name in data2_dict:
                data.append(data2_dict[name])
            else:
                data.append({'name': name, 'scrap': 0, 'secondgrade': 0, 'sum_qty': 0})

        x_labels = [str(item['name']).split('_')[-1] for item in data]
        x_range = range(len(x_labels))

        scrap = [round((item['scrap'] / item['sum_qty']) * 100, 2) if item['sum_qty'] > 0 else 0 for item in data]
        secondgrade = [round((item['secondgrade'] / item['sum_qty']) * 100, 2) if item['sum_qty'] > 0 else 0 for item in
                       data]

        # 過濾 scrap 數據
        filtered_x_scrap, filtered_scrap = self.filter_valid_data(x_range, scrap)

        # 過濾 secondgrade 數據
        filtered_x_secondgrade, filtered_secondgrade = self.filter_valid_data(x_range, secondgrade)

        # Create the figure and subplots
        plt.figure(figsize=(16, 9))
        fig = plt.figure(figsize=(16, 9))
        gs = gridspec.GridSpec(2, 1, height_ratios=[1, 1], hspace=0)  # No space between subplots

        # Subplot for scrap
        y_max1 = 3
        y_ticks = np.arange(0, y_max1, 0.4)  # 分成10個刻度
        y_ticks = y_ticks[:-1]

        ax1 = fig.add_subplot(gs[0])
        ax1.plot(filtered_x_scrap, filtered_scrap, label="廢品率 (%)", marker='o', linestyle='-', color='#ED7D31',
                 linewidth=2)
        ax1.axhline(y=0.8, color='#ED7D31', linestyle='--', linewidth=1.5, label="廢品標準線(0.8)")
        ax1.set_xticks(x_range)
        ax1.set_xticklabels([])  # Hide x-axis labels for the top plot
        # ax1.set_ylabel('廢品率 (%)', fontsize=12, rotation=0)
        ax1.text(-0.1, 0.6, '廢品率 (%)', fontsize=12, rotation=0, ha='center', va='center', transform=ax1.transAxes)
        ax1.set_ylim(0, y_max1)
        ax1.set_yticks(y_ticks)
        ax1.yaxis.set_major_formatter(FuncFormatter(self.add_percent))

        offset = 0.03
        for i, scrap_val in enumerate(filtered_scrap):
            ax1.text(filtered_x_scrap[i], scrap_val + offset, f"{scrap_val:.2f}%", ha='center', va='bottom',
                     fontsize=12,
                     color='#ED7D31')

        # Subplot for secondgrade
        y_max2 = 1.6
        y_ticks = np.arange(0, y_max2, 0.2)  # 分成10個刻度
        y_ticks = y_ticks[:-1]

        ax2 = fig.add_subplot(gs[1])
        ax2.plot(filtered_x_secondgrade, filtered_secondgrade, label="二級品率 (%)", marker='o', linestyle='-',
                 color='#70AD47', linewidth=2)
        ax2.axhline(y=0.2, color='#70AD47', linestyle='--', linewidth=1.5, label="二級品標準線(0.2)")
        ax2.set_xticks(x_range)
        ax2.set_xticklabels(x_labels, rotation=0, ha="center", fontsize=12)
        # ax2.set_ylabel('二級品率 (%)', fontsize=12, rotation=0)
        ax2.text(-0.1, 0.6, '二級品率 (%)', fontsize=12, rotation=0, ha='center', va='center', transform=ax2.transAxes)
        ax2.set_ylim(0, y_max2)
        ax2.set_yticks(y_ticks)
        ax2.yaxis.set_major_formatter(FuncFormatter(self.add_percent))

        offset = 0.03
        for i, secondgrade_val in enumerate(filtered_secondgrade):
            ax2.text(filtered_x_secondgrade[i], secondgrade_val + offset, f"{secondgrade_val:.2f}%", ha='center',
                     va='bottom',
                     fontsize=12, color='#70AD47')

        # Add a title for the entire figure
        fig.suptitle(f"二級品及廢品率 ({plant})", fontsize=20)

        # Add legend for both plots
        handles1, labels1 = ax1.get_legend_handles_labels()
        handles2, labels2 = ax2.get_legend_handles_labels()
        fig.legend(
            handles1 + handles2,
            labels1 + labels2,
            loc='center left',
            fontsize=10,
            title="Note",
            title_fontsize=12,
            bbox_to_anchor=(1.0, 0.5),
            ncol=1
        )

        plt.tight_layout()

        file_name = f'MES_{plant}_{date_mark}_Rate_Chart_Line.png'
        chart_img = os.path.join(save_path, file_name)

        plt.savefig(chart_img, dpi=100, bbox_inches="tight", pad_inches=0.45)
        plt.close()

        self.image_buffers.append(chart_img)

    def chart_table(self, plant, x_labels, secondgrade, scrap):
        date_mark = self.date_mark
        save_path = self.save_path

        table_data = [
            [plant] + x_labels,
            ["廢品(%)"] + scrap,
            ["二級品(%)"] + secondgrade,
        ]

        # 創建圖表
        fig, ax = plt.subplots(figsize=(12, 3))  # 調整尺寸以適配內容
        ax.axis("tight")
        ax.axis("off")

        # 添加表格
        table = ax.table(cellText=table_data, loc="center", cellLoc="center")

        font_path = font_manager.findfont("Microsoft YaHei")
        font_prop = font_manager.FontProperties(fname=font_path, size=14)
        header_color = "#0A4E9B"
        row_colors = ["#E8F3FF", "#FFFFFF"]

        for (row, col), cell in table.get_celld().items():
            cell.set_width(0.1)
            # Header styling
            if row == 0:
                cell.set_facecolor(header_color)
                cell.set_text_props(color="white", weight="bold", fontproperties=font_prop)
            # Row styling
            else:
                cell.set_facecolor(row_colors[row % 2])
                cell.set_text_props(fontproperties=font_prop)
            # Adjust height for all cells
            cell.set_height(0.2)

        plt.tight_layout(rect=[0, 0, 0, 0])
        file_name = f'MES_{plant}_{date_mark}_Rate_Chart_Table.png'
        chart_img = os.path.join(save_path, file_name)
        plt.savefig(chart_img, bbox_inches="tight", dpi=100, pad_inches=0)
        plt.close()

        self.image_buffers.append(chart_img)

    def send_email(self, file_list, image_buffers):
        mode = self.mode
        date_mark = self.date_mark
        this_start_date = self.this_start_date
        logging.info(f"Start to send Email")
        smtp_config, to_emails, admin_emails = self.read_config('mes_weekly_report_mail.config')

        # SMTP Sever config setting
        smtp_server = smtp_config.get('smtp_server')
        smtp_port = int(smtp_config.get('smtp_port', 587))
        smtp_user = smtp_config.get('smtp_user')
        smtp_password = smtp_config.get('smtp_password')
        sender_alias = "GD Report"
        sender_email = smtp_user
        # Mail Info
        msg = MIMEMultipart()
        msg['From'] = f"{sender_alias} <{sender_email}>"
        msg['To'] = ', '.join(to_emails)

        if mode == "WEEKLY":
            msg['Subject'] = f'[GD Report] 第{self.last_week_info}週達成率報表 {date_mark}'
        elif mode == "MONTHLY":
            name = f"{this_start_date.strftime('%Y %m')}"
            msg['Subject'] = f'[GD Report] {name}月達成率報表 {date_mark}'

        # Mail Content
        html = """\
                <html>
                  <body>
                """
        for i in range(len(image_buffers)):
            html += f'<img src="cid:chart_image{i}"><br>'

        html += """\
                  </body>
                </html>
                """

        msg.attach(MIMEText(html, 'html'))

        # Attach Excel
        for file_path in file_list:
            if os.path.exists(file_path):
                with open(file_path, "rb") as attachment:
                    part = MIMEBase("application", "octet-stream")
                    part.set_payload(attachment.read())
                encoders.encode_base64(part)
                part.add_header(
                    "Content-Disposition",
                    f"attachment; filename={os.path.basename(file_path)}"
                )
                msg.attach(part)
            else:
                print(f"File not found: {file_path}")
        logging.info(f"Attached Excel files")

        # Attach Picture
        for idx, image_path in enumerate(image_buffers):
            if os.path.exists(image_path):
                with open(image_path, "rb") as img_file:
                    img = MIMEImage(img_file.read())
                img.add_header("Content-ID", f"<chart_image{idx}>")
                msg.attach(img)
            else:
                logging.info(f"Image not found: {image_path}")
                print(f"Image not found: {image_path}")
        logging.info(f"Attached Picture")

        # Send Email
        try:
            # server = smtplib.SMTP(smtp_server, smtp_port)
            # server.starttls()  # 启用 TLS 加密
            # server.login(smtp_user, smtp_password)  # 登录到 SMTP 服务器
            # server.sendmail(smtp_user, to_emails, msg.as_string())
            # server.quit()

            # 發送郵件（不進行密碼驗證）
            server = smtplib.SMTP(smtp_server, smtp_port)
            server.ehlo()  # 啟動與伺服器的對話
            server.sendmail(smtp_user, to_emails, msg.as_string())
            print("Sent Email Successfully")
        except Exception as e:
            print(f"Sent Email Fail: {e}")
            logging.info(f"Sent Email Fail: {e}")
        finally:
            attachment.close()

    def read_config(self, config_file):
        smtp_config = {}
        to_emails = []
        admin_emails = []

        current_section = None

        with open(config_file, 'r') as file:
            for line in file:
                line = line.strip()

                # Skip empty lines and comments
                if not line or line.startswith('#'):
                    continue

                # Detect section headers
                if line.startswith('[') and line.endswith(']'):
                    current_section = line[1:-1].lower()
                    continue

                # Read lines based on the current section
                if current_section == 'smtp':
                    if '=' in line:
                        key, value = line.split('=', 1)
                        smtp_config[key.strip()] = value.strip()
                elif current_section == 'recipients':
                    to_emails.append(line)
                elif current_section == 'admin_email':
                    admin_emails.append(line)

        return smtp_config, to_emails, admin_emails

    def generate_previous_weeks_with_dates(self):
        sql = f"""
        WITH CurrentMonthRows AS (
            SELECT month, month_week, start_date, end_date
            FROM [VNEDC].[dbo].[week_date]
            WHERE month = MONTH(GETDATE()) AND YEAR(start_date) = YEAR(GETDATE()) and GETDATE() > start_date
        ),
        PreviousMonthRows AS (
            SELECT TOP (13 - (SELECT COUNT(*) FROM CurrentMonthRows)) month, month_week, start_date, end_date
            FROM [VNEDC].[dbo].[week_date]
            WHERE (month < MONTH(GETDATE()) AND YEAR(start_date) = YEAR(GETDATE())) OR (YEAR(start_date) < YEAR(GETDATE()))
            ORDER BY YEAR(start_date) DESC, month DESC, month_week DESC, start_date DESC
        )
        SELECT * FROM (SELECT * FROM PreviousMonthRows UNION ALL SELECT * FROM CurrentMonthRows) AS CombinedResults
        ORDER BY YEAR(start_date), month, month_week, start_date;
        """
        rows = vnedc_database().select_sql_dict(sql)
        weeks_list = []
        week_dates = []
        for row in rows:
            weeks_list.append(f"{row['month']}{row['month_week']}")
            week_dates.append([datetime.strptime(row['start_date'], '%Y-%m-%d').date(),
                               datetime.strptime(row['end_date'], '%Y-%m-%d').date()])
        return weeks_list[:-1], week_dates[:-1]

    def get_week_data(self):
        vnedc_db = vnedc_database()
        today = datetime.now()
        seven_days_ago = (today - timedelta(days=7)).strftime('%Y-%m-%d')

        sql = f"""
              SELECT * FROM (
                 SELECT TOP(15) *  FROM [VNEDC].[dbo].[week_date]
                     WHERE CONVERT(DATETIME, '{seven_days_ago}', 120) > end_date
                              ORDER BY year desc, month desc, month_week
                              ) A
                              ORDER BY year, month, month_week
        """

        print(sql)
        raws = vnedc_db.select_sql_dict(sql)

        return raws

    def add_percent(self, y, pos):
        return f"{y:.1f}%"  # 格式化為整數百分比

    def weekly_chart(self, plant):
        save_path = self.save_path

        data1 = self.mach_list

        for machine in data1:
            try:
                this_data = []
                this_rate = []
                x_labels = []
                week_date = []
                unfinish_data = []

                today = datetime.now()

                if machine['name'] in ['VN_GD_PVC1_L01', 'VN_GD_PVC1_L02']:
                    start_date = date(2025, 1, 6)
                    start_week_number = 1
                elif machine['name'] in ['VN_GD_PVC1_L05', 'VN_GD_PVC1_L06']:
                    start_date = date(2024, 12, 2)
                    start_week_number = 49
                else:
                    # Get the starting week (Sep 30)
                    start_date = date(2024, 9, 30)
                    start_week_number = 40

                x_labels, week_date = self.generate_previous_weeks_with_dates()

                print(machine['name'])
                for i, item in enumerate(week_date):
                    week_name = x_labels[i]
                    if item[0] < start_date:
                        this_data.append(0)
                        this_rate.append(0)
                        unfinish_data.append(0)
                        continue

                    print(f"Week {week_name} - start {item[0]} - end - {item[1]}")
                    sql = f"""
                            SELECT name,qty, target, (case when target-qty > 0 then target-qty else 0 end) unfinish_qty FROM (
                            SELECT name, sum(case when sum_qty > 0 then sum_qty else 0 end) as qty, sum(case when target > 0 then target else 0 end) as target
                            FROM [MES_OLAP].[dbo].[mes_daily_report_raw] where Name = '{machine['name']}'
                            and (belong_to between '{item[0]}' and '{item[1]}')
                            group by name) A"""
                    rows = vnedc_database().select_sql_dict(sql)
                    if len(rows) == 0:
                        this_data.append(0)
                        this_rate.append(0)
                        unfinish_data.append(0)
                    else:
                        this_data.append(rows[0]['qty'])
                        unfinish_data.append(rows[0]['unfinish_qty'])
                        try:
                            rate = int((rows[0]['qty'] / rows[0]['target']) * 100) if rows[0]['target'] > 0 else 100
                        except Exception as e:
                            rate = 100
                            print(f"{e} at {machine['name']} at week start {item[0]} - end {item[1]}")
                            pass
                        this_rate.append(rate)
                    # x_labels.append(f'W {week}')

                x_range = range(0, len(x_labels) * 2, 2)

                max_data = max(this_data, default=0)
                step_data = 10
                rounded_max_data = int(
                    (((max_data / (10 ** (len(str(max_data)) - 2))) // step_data) * step_data + step_data) * (
                            10 ** (len(str(max_data)) - 2)))
                rounded_step_data = step_data * (10 ** (len(str(max_data)) - 2))
                max_rate = max(this_rate, default=0)
                rounded_max_rate = (math.ceil(max_rate / 10) * 10)
                rounded_step_rate = 20

                bar_width = 0.6
                plt.figure(figsize=(24, 9))
                fig, ax1 = plt.subplots(figsize=(24, 9))
                this_month_bars = ax1.bar([i for i in x_range], this_data, width=bar_width,
                                          label=f"週產量", align='center', color='#10ba81')
                unfinish_bars = ax1.bar([i for i in x_range], unfinish_data, width=bar_width, bottom=this_data,
                                        label=f"週目標差異",
                                        align='center', color='lightgreen')
                ax1.set_xticks(x_range)
                ax1.set_xticklabels(x_labels, rotation=0, ha="center", fontsize=12)
                if len(str(max_data)) > 7:
                    yticks_positions = list(range(0, rounded_max_data + 1 * rounded_step_data, rounded_step_data))
                    yticks_positions.append(int(rounded_max_data + 2 * rounded_step_data))
                    yticks_labels = [
                        f"{int(i//(10**(len(str(max_data)) - 2)))}" + '百萬' if len(str(i)) > 6 else f"{i}" for i
                        in yticks_positions]
                    for index, bar in enumerate(this_month_bars):
                        height = bar.get_height()
                        unfinish_height = unfinish_bars[index].get_height()
                        ax1.text(
                            bar.get_x() + bar.get_width() / 2,
                            height + unfinish_height,
                            f'{round(height/(10**(len(str(max_data)) - 2)),2)}'[:4] if height > 0 else '',
                            ha='center', va='bottom', fontsize=12  # Align the text
                        )
                elif 4 < len(str(max_data)) <= 7:
                    yticks_positions = list(range(0, rounded_max_data, rounded_step_data))
                    yticks_positions.append(int(rounded_max_data + 3 * rounded_step_data))
                    # yticks_labels = [f"{int(i//(10**(len(str(max_data)) - 3)))}" + '萬 PCS' if len(str(i)) > 4 and int(
                    #     i // (10 ** (len(str(max_data)) - 3))) % 60 == 0 else "" for i in yticks_positions]
                    yticks_labels = [f"{int(i/1000000)} 百萬" if i > 0 else 0 for i in yticks_positions]
                    for index, bar in enumerate(this_month_bars):
                        height = bar.get_height()
                        unfinish_height = unfinish_bars[index].get_height()
                        ax1.text(
                            bar.get_x() + bar.get_width() / 2,
                            height + unfinish_height,
                            f'{round(height/(10**(len(str(max_data)) - 1)),2)}'[:4] if height > 0 else '',
                            ha='center', va='bottom', fontsize=12  # Align the text
                        )

                yticks_labels[-1] = ""
                ax1.set_yticks(yticks_positions)
                ax1.set_yticklabels(yticks_labels, fontsize=12)

                ax2 = ax1.twinx()
                sr_achieve_rate = "達成率"

                filtered_data = [(x, rate) for x, rate in zip(x_range, this_rate) if rate != 0]
                x_filtered, this_rate_filtered = zip(*filtered_data)
                this_rate_line = ax2.plot(x_filtered, this_rate_filtered,
                                          label=sr_achieve_rate,
                                          color='#ED7D31', marker='o', linewidth=1.5)
                sr_target = "達成率目標"

                ry_label = "達成率(%)"
                ly_label = "Product (百萬)"

                name = f"各週產出量"
                title = f"\n{plant} ({machine['name']})\n"

                if plant == "NBR":
                    target_rate = 95
                else:
                    target_rate = 98

                yticks_positions = list(range(0, rounded_max_rate + 1 * rounded_step_rate, rounded_step_rate))
                if target_rate not in yticks_positions:
                    yticks_positions.append(target_rate)
                yticks_positions.append(rounded_max_rate + 0.25 * rounded_step_rate)
                yticks_positions = sorted(yticks_positions)

                yticks_labels = [f"{i}" + '%' for i in yticks_positions]
                yticks_labels[-1] = ""
                ax2.set_yticks(yticks_positions)
                ax2.set_yticklabels(yticks_labels)
                ax2.axhline(y=target_rate, color='red', linestyle='--', linewidth=1, label=sr_target)

                ax1.xaxis.set_label_coords(0.975, -0.014)
                ax1.set_ylabel('Product output', fontsize=12)
                ax2.set_ylabel('Archive rates', fontsize=12)

                for i, value in enumerate(this_rate):
                    ax2.text(
                        x_range[i],
                        value + 1.5,
                        f"{value:.1f}%" if value != 0 else '',
                        ha='center', va='bottom', fontsize=12, color='white',
                        bbox=dict(facecolor='#ED7D31', edgecolor='none', boxstyle='round,pad=0.3')
                    )

                handles1, labels1 = ax1.get_legend_handles_labels()
                handles2, labels2 = ax2.get_legend_handles_labels()
                fig.legend(
                    handles1 + handles2,
                    labels1 + labels2,
                    loc='center left',
                    fontsize=12,
                    title="Note",
                    title_fontsize=12,
                    bbox_to_anchor=(1.0, 0.5),
                    ncol=1
                )

                # Customize axes and borders
                ax = plt.gca()  # Get current axes
                ax.spines['top'].set_color('white')
                ax.spines['top'].set_linewidth(1.5)

                ax1.tick_params(axis='y', colors='black')  # Y-axis ticks color
                y_tick_lines = ax1.get_yticklines()
                y_tick_lines[-2].set_visible(False)
                y_tick_lines[-1].set_visible(False)
                ax1.annotate(
                    '',
                    xy=(0, 1.0),
                    xytext=(0, 0.98),
                    xycoords='axes fraction',
                    arrowprops=dict(facecolor='black', arrowstyle='-|>,widthA=0.4,widthB=1.4', linewidth=0.5)
                )
                ax1.set_ylabel(ly_label, labelpad=20, rotation=0)
                ax1.yaxis.set_label_coords(-0.01, 1.01)

                ax2.tick_params(axis='y', colors='black')
                y_tick_lines = ax2.get_yticklines()
                y_tick_lines[-2].set_visible(False)
                y_tick_lines[-1].set_visible(False)
                ax2.annotate(
                    '',
                    xy=(1, 1.0),
                    xytext=(1, 0.98),
                    xycoords='axes fraction',
                    arrowprops=dict(facecolor='black', arrowstyle='-|>,widthA=0.4,widthB=1.4', linewidth=0.5)
                )

                ax2.set_ylabel(ry_label, labelpad=20, rotation=0)
                ax2.yaxis.set_label_coords(1.01, 1.03)

                plt.title(title, fontsize=20)

                plt.tight_layout()

                file_name = f"MES_{machine['name']}_Chart.png"
                # save_path = os.path.join("yearly_output")
                chart_img = os.path.join(save_path, file_name)

                plt.savefig(f"{chart_img}", dpi=100, bbox_inches="tight", pad_inches=0.45)
                plt.close('all')
            except Exception as e:
                print(f"{machine['name']}: {e}")
                pass

    def monthly_chart(self, plant):
        save_path = self.save_path

        data1 = self.mach_list

        for machine in data1:
            print(f"Machine {machine}")
            try:
                today = datetime.today()
                month_date = []
                this_data = []
                this_rate = []
                x_labels = []
                unfinish_data = []

                for i in range(12):
                    end_date = (today - relativedelta(months=i)).replace(day=1) - timedelta(days=1)
                    start_date = end_date.replace(day=1)
                    month_date.append(
                        [start_date.strftime('%Y-%m-%d'), end_date.strftime('%Y-%m-%d'), start_date.strftime("%B %Y")])
                    x_labels.append(start_date.strftime("%Y-%m"))
                month_date.reverse()
                x_labels.reverse()
                for i, item in enumerate(month_date):
                    print(f"Month {item[2]} - start {item[0]} - end - {item[1]}")
                    sql = f"""
                            SELECT name,qty, target, (case when target-qty > 0 then target-qty then 0 end) unfinish_qty FROM (
                            SELECT name, sum(case when sum_qty > 0 then sum_qty else 0 end) as qty, sum(case when target > 0 then target else 0 end) as target
                            FROM [MES_OLAP].[dbo].[mes_daily_report_raw] where Name = '{machine['name']}'
                            and (belong_to between '{item[0]}' and '{item[1]}')
                            group by name) A"""
                    rows = vnedc_database().select_sql_dict(sql)
                    if len(rows) == 0:
                        this_data.append(0)
                        this_rate.append(0)
                        unfinish_data.append(0)
                    else:
                        this_data.append(rows[0]['qty'])
                        unfinish_data.append(rows[0]['unfinish_qty'])
                        try:
                            rate = int((rows[0]['qty'] / rows[0]['target']) * 100) if rows[0]['target'] > 0 else 100
                        except Exception as e:
                            rate = 100
                            print(f"{e} at {machine['name']} at {item[2]} start {item[0]} - end {item[1]}")
                            pass
                        this_rate.append(rate)

                x_range = range(0, len(x_labels) * 2, 2)

                max_data = max(this_data, default=0)
                step_data = 5
                rounded_max_data = int(
                    (((max_data / (10 ** (len(str(max_data)) - 2))) // step_data) * step_data + step_data) * (
                            10 ** (len(str(max_data)) - 2)))
                rounded_step_data = step_data * (10 ** (len(str(max_data)) - 2))
                max_rate = max(this_rate, default=0)
                rounded_max_rate = (math.ceil(max_rate / 10) * 10)
                rounded_step_rate = 20

                bar_width = 0.6
                plt.figure(figsize=(16, 9))
                fig, ax1 = plt.subplots(figsize=(16, 9))
                this_month_bars = ax1.bar([i for i in x_range], this_data, width=bar_width,
                                          label=f"月產量", align='center', color='#10ba81')
                unfinish_bars = ax1.bar([i for i in x_range], unfinish_data, width=bar_width, bottom=this_data,
                                        label=f"月目標差異",
                                        align='center', color='lightgreen')
                ax1.set_xticks(x_range)
                ax1.set_xticklabels(x_labels, rotation=0, ha="center", fontsize=12)
                if len(str(max_data)) > 7:
                    yticks_positions = list(range(0, rounded_max_data + 1 * rounded_step_data, rounded_step_data))
                    yticks_positions.append(int(rounded_max_data + 2 * rounded_step_data))
                    yticks_labels = [
                        f"{int(i//(10**(len(str(max_data)) - 2)))}" + '百萬' if len(str(i)) > 6 else f"{i}" for i
                        in yticks_positions]
                    for index, bar in enumerate(this_month_bars):
                        height = bar.get_height()
                        unfinish_height = unfinish_bars[index].get_height()
                        ax1.text(
                            bar.get_x() + bar.get_width() / 2,
                            height + unfinish_height,
                            f'{round(height/(10**(len(str(max_data)) - 2)),2)}'[
                            :4] if height > 0 else '',
                            ha='center', va='bottom', fontsize=12  # Align the text
                        )
                elif 4 < len(str(max_data)) <= 7:
                    yticks_positions = list(range(0, rounded_max_data, rounded_step_data))
                    yticks_positions.append(int(rounded_max_data + 3 * rounded_step_data))
                    # yticks_labels = [f"{int(i//(10**(len(str(max_data)) - 3)))}" + '萬 PCS' if len(str(i)) > 4 and int(
                    #     i // (10 ** (len(str(max_data)) - 3))) % 60 == 0 else "" for i in yticks_positions]
                    yticks_labels = [f"{int(i/1000000)} 百萬" if i > 0 else 0 for i in yticks_positions]
                    for index, bar in enumerate(this_month_bars):
                        height = bar.get_height()
                        unfinish_height = unfinish_bars[index].get_height()
                        ax1.text(
                            bar.get_x() + bar.get_width() / 2,
                            height + unfinish_height,
                            f'{int(height/(10**(len(str(max_data)) - 3)))}'[:4] if height > 0 else '',
                            ha='center', va='bottom', fontsize=12  # Align the text
                        )

                yticks_labels[-1] = ""
                ax1.set_yticks(yticks_positions)
                ax1.set_yticklabels(yticks_labels, fontsize=12)

                ax2 = ax1.twinx()
                sr_achieve_rate = "達成率"

                filtered_data = [(x, rate) for x, rate in zip(x_range, this_rate) if rate != 0]
                x_filtered, this_rate_filtered = zip(*filtered_data)
                this_rate_line = ax2.plot(x_filtered, this_rate_filtered,
                                          label=sr_achieve_rate,
                                          color='#ED7D31', marker='o', linewidth=1.5)
                sr_target = "達成率目標"

                ry_label = "達成率(%)"
                ly_label = "Product (百萬)"

                name = f"各週產出量"
                title = f"\n{plant} ({machine['name']})\n"

                if plant == "NBR":
                    target_rate = 95
                else:
                    target_rate = 98

                yticks_positions = list(range(0, rounded_max_rate + 1 * rounded_step_rate, rounded_step_rate))
                if target_rate not in yticks_positions:
                    yticks_positions.append(target_rate)
                yticks_positions.append(rounded_max_rate + 0.25 * rounded_step_rate)
                yticks_positions = sorted(yticks_positions)

                yticks_labels = [f"{i}" + '%' for i in yticks_positions]
                yticks_labels[-1] = ""
                ax2.set_yticks(yticks_positions)
                ax2.set_yticklabels(yticks_labels)
                ax2.axhline(y=target_rate, color='red', linestyle='--', linewidth=1, label=sr_target)

                ax1.xaxis.set_label_coords(0.975, -0.014)
                ax1.set_ylabel('Product output', fontsize=12)
                ax2.set_ylabel('Archive rates', fontsize=12)

                for i, value in enumerate(this_rate):
                    ax2.text(
                        x_range[i],
                        value + 1.5,
                        f"{value:.1f}%" if value != 0 else '',
                        ha='center', va='bottom', fontsize=10, color='white',
                        bbox=dict(facecolor='#ED7D31', edgecolor='none', boxstyle='round,pad=0.3')
                    )

                handles1, labels1 = ax1.get_legend_handles_labels()
                handles2, labels2 = ax2.get_legend_handles_labels()
                fig.legend(
                    handles1 + handles2,
                    labels1 + labels2,
                    loc='center left',
                    fontsize=10,
                    title="Note",
                    title_fontsize=12,
                    bbox_to_anchor=(1.0, 0.5),
                    ncol=1
                )

                # Customize axes and borders
                ax = plt.gca()  # Get current axes
                ax.spines['top'].set_color('white')
                ax.spines['top'].set_linewidth(1.5)

                ax1.tick_params(axis='y', colors='black')  # Y-axis ticks color
                y_tick_lines = ax1.get_yticklines()
                y_tick_lines[-2].set_visible(False)
                y_tick_lines[-1].set_visible(False)
                ax1.annotate(
                    '',
                    xy=(0, 1.0),
                    xytext=(0, 0.98),
                    xycoords='axes fraction',
                    arrowprops=dict(facecolor='black', arrowstyle='-|>,widthA=0.4,widthB=1.4', linewidth=0.5)
                )
                ax1.set_ylabel(ly_label, labelpad=20, rotation=0)
                ax1.yaxis.set_label_coords(-0.01, 1.01)

                ax2.tick_params(axis='y', colors='black')
                y_tick_lines = ax2.get_yticklines()
                y_tick_lines[-2].set_visible(False)
                y_tick_lines[-1].set_visible(False)
                ax2.annotate(
                    '',
                    xy=(1, 1.0),
                    xytext=(1, 0.98),
                    xycoords='axes fraction',
                    arrowprops=dict(facecolor='black', arrowstyle='-|>,widthA=0.4,widthB=1.4', linewidth=0.5)
                )

                ax2.set_ylabel(ry_label, labelpad=20, rotation=0)
                ax2.yaxis.set_label_coords(1.01, 1.03)

                plt.title(title, fontsize=20)

                plt.tight_layout()

                file_name = f"MES_{machine['name']}_Chart.png"
                # save_path = os.path.join("yearly_output")
                chart_img = os.path.join(save_path, file_name)

                plt.savefig(f"{chart_img}", dpi=100, bbox_inches="tight", pad_inches=0.45)
                plt.close('all')
            except:
                pass

    def excel_chart_2(self, plant):
        try:
            rcParams['font.sans-serif'] = ['Microsoft YaHei']
            rcParams['axes.unicode_minus'] = False
            save_path = self.save_path

            x_labels, week_date = self.generate_previous_weeks_with_dates()
            sap_qty = []
            pdt_qty = []
            dif_qty = []

            for date in week_date:
                sql = f"""
                    SELECT sum(Target) as qty, sum(SAPQuantity) as sap_qty
                    FROM [MES_OLAP].[dbo].[counting_daily_info_raw]
					WHERE Machine like '%{plant}%' AND belong_to between '{date[0]}' and '{date[1]}'
                    AND CountingQty IS NOT NULL
                """
                qty = vnedc_database().select_sql_dict(sql)
                sap_qty.append(int(qty[0]['sap_qty']) if qty[0]['sap_qty'] is not None else 0)
                pdt_qty.append(int(qty[0]['qty']) if qty[0]['qty'] is not None else 0)
                dif_qty.append((int(qty[0]['qty']) - int(qty[0]['sap_qty'])) if (
                            qty[0]['sap_qty'] is not None and qty[0]['qty'] is not None) else 0)

            x_range = range(0, len(x_labels) * 3, 3)
            max_data = max(max(sap_qty, default=0), max(pdt_qty, default=0))
            rounded_max_data = int(max_data * 1.2)
            if plant == 'NBR':
                rounded_step_data = 20000000
            elif plant == 'PVC':
                rounded_step_data = 10000000

            rounded_min_data = (min(dif_qty) // rounded_step_data) * rounded_step_data

            bar_width = 0.6
            plt.figure(figsize=(18, 9))
            fig, ax1 = plt.subplots(figsize=(18, 9))
            pdt_qty_bars = ax1.bar([i - bar_width - 0.15 for i in x_range], pdt_qty, width=bar_width,
                                   label=f"预估包装产量(A)", align='center', color='#156082')
            sap_qty_bars = ax1.bar([i for i in x_range], sap_qty, width=bar_width,
                                   label=f" 包装确认量(B)", align='center', color='#e97132')
            dif_qty_bars = ax1.bar([i + bar_width + 0.15 for i in x_range], dif_qty, width=bar_width,
                                   label=f"差异(A-B)", align='center', color='#196b24')

            ax1.set_xticks(x_range)
            ax1.set_xticklabels(x_labels, rotation=0, ha="center", fontsize=12)
            if len(str(max_data)) > 7:
                yticks_positions = list(range(rounded_min_data, 0, 10000000)) + list(
                    range(0, rounded_max_data, rounded_step_data))
                yticks_labels = [
                    f"{int(i//(1000000))}" + '百萬' if len(str(i)) > 6 else f"{i}" for i
                    in yticks_positions]
                for index, bar in enumerate(pdt_qty_bars):
                    height = bar.get_height()
                    ax1.text(
                        bar.get_x() + bar.get_width() / 2,
                        height,
                        f'{round(height/1000000,2)}' if height > 0 else '',
                        ha='center', va='bottom', fontsize=10  # Align the text
                    )
                for index, bar in enumerate(sap_qty_bars):
                    height = bar.get_height()
                    ax1.text(
                        bar.get_x() + bar.get_width() / 2,
                        height,
                        f'{round(height/1000000,2)}' if height > 0 else '',
                        ha='center', va='bottom', fontsize=10  # Align the text
                    )
                for index, bar in enumerate(dif_qty_bars):
                    height = bar.get_height()
                    ax1.text(
                        bar.get_x() + bar.get_width() / 2,
                        (
                            height - 4800000 if plant == 'NBR' and height < 0 else height - 2800000) if height < 0 else height,
                        f'{round(height/1000000,2)}' if abs(round(height / 10000000, 2)) != 0 else '',
                        ha='center', va='bottom', fontsize=10  # Align the text
                    )

            yticks_labels[-1] = ""
            yticks_labels[0] = ""

            yticks_labels = yticks_labels[1:-1]
            ax1.set_yticks(yticks_positions[1:-1])
            ax1.set_yticklabels(yticks_labels, fontsize=12)
            ax1.axhline(y=0, color='black', linestyle='-', linewidth=1)
            ax1.set_ylim(rounded_min_data, rounded_max_data)
            ax1.spines['top'].set_visible(False)
            ax1.spines['right'].set_visible(False)
            ax1.spines['left'].set_visible(True)
            ax1.spines['bottom'].set_visible(False)
            pvc_sx = "(当责：制造-Nguyen Van A - 负责：制造-Nguyen Van C + Thach A)"
            nbr_sx = "(当责：制造-Nguyen Van B - 负责：制造-Nguyen Van D + Thach B)"
            title = f"""2.GD-{plant} 面向生产管理板\n{pvc_sx if plant=='PVC' else nbr_sx}"""
            handles1, labels1 = ax1.get_legend_handles_labels()
            fig.legend(
                handles1,
                labels1,
                loc='center left',
                fontsize=10,
                title_fontsize=14,
                bbox_to_anchor=(1.0, 0.5),
                ncol=1
            )

            plt.title(title, fontsize=20, fontweight='bold', color='#0070c0')

            plt.tight_layout()

            file_name = f"""GD-{plant}面向生产管理板"""
            chart_img = os.path.join(save_path, file_name)

            plt.savefig(f"{chart_img}", dpi=100, bbox_inches="tight", pad_inches=0.45)
            plt.close('all')
        except Exception as e:
            print(e)
            pass

    def excel_chart_6a(self, plant):
        try:
            rcParams['font.sans-serif'] = ['Microsoft YaHei']
            rcParams['axes.unicode_minus'] = False
            save_path = self.save_path

            vnedc_db = vnedc_database()
            sql = f"""
            WITH defect_total AS (
                SELECT Year,MonthWeek, SUM(cosmetic_inspect_qty) AS total_qty 
                FROM (
                    SELECT DISTINCT Year,MonthWeek, c.Runcard, cosmetic_inspect_qty 
                    FROM [MES_OLAP].[dbo].[counting_daily_info_raw] c
                    JOIN [MES_OLAP].[dbo].[mes_ipqc_cosmetic_data] ipqc 
                    ON c.Runcard = ipqc.runcard
                    WHERE c.branch like '%{plant}%'
                ) A
                GROUP BY Year, MonthWeek
            ), 
            detail AS (
                SELECT Year, MonthWeek, defect_code, SUM(qty) AS qty 
                FROM [MES_OLAP].[dbo].[counting_daily_info_raw] c
                JOIN [MES_OLAP].[dbo].[mes_ipqc_cosmetic_data] ipqc 
                ON c.Runcard = ipqc.runcard
                WHERE c.branch like '%{plant}%'
                GROUP BY Year,MonthWeek, defect_code
            )

            SELECT Year,MonthWeek,defect_level,sum(DPM) DPM from (
            SELECT 
                d.Year,
                d.MonthWeek,
                d.defect_code,
                defect.defect_level,
                defect.desc2,
                d.qty,
                t.total_qty,
                ROUND(CAST(d.qty AS FLOAT) / CAST(t.total_qty AS FLOAT), 6) * 1000000 AS DPM
            FROM defect_total t
            JOIN detail d ON d.Year = t.Year and d.MonthWeek = t.MonthWeek
            JOIN [MES_OLAP].[dbo].mes_defect_data defect ON defect.defect_code = d.defect_code) A
            GROUP BY Year,MonthWeek,defect_level
            Order by Year, MonthWeek, defect_level
            """
            raws = vnedc_db.select_sql_dict(sql)
            raw_df = pd.DataFrame(raws)

            week_dict = self.get_week_data()
            critical_qty = []
            major_qty = []
            minor_qty = []
            x_labels = []

            for week in week_dict:
                filter_df = raw_df[(raw_df['Year'] == week['year']) &
                                   (raw_df['MonthWeek'] == str(week['month']) + week['month_week'])]

                critical_qty.append(filter_df[filter_df['defect_level'] == 'CRITICAL']['DPM'].iloc[0])
                major_qty.append(filter_df[filter_df['defect_level'] == 'MAJOR']['DPM'].iloc[0])
                minor_qty.append(filter_df[filter_df['defect_level'] == 'MINOR']['DPM'].iloc[0])
                x_labels.append(str(week['month']) + week['month_week'])
            x_range = range(0, len(x_labels) * 3, 3)
            max_data = max(max(critical_qty, default=0), max(major_qty, default=0), max(minor_qty, default=0))
            rounded_max_data = int(max_data * 1.2)
            if plant == 'NBR':
                rounded_step_data = 40000
            elif plant == 'PVC':
                rounded_step_data = 10000

            bar_width = 0.6
            plt.figure(figsize=(18, 9))
            fig, ax1 = plt.subplots(figsize=(18, 9))
            critical_qty_bars = ax1.bar([i - bar_width - 0.15 for i in x_range], critical_qty, width=bar_width,
                                        label=f"CRITICAL", align='center', color='#156082')
            major_qty_bars = ax1.bar([i for i in x_range], major_qty, width=bar_width,
                                     label=f"MAJOR", align='center', color='#e97132')
            minor_qty_bars = ax1.bar([i + bar_width + 0.15 for i in x_range], minor_qty, width=bar_width,
                                     label=f"MINOR", align='center', color='#196b24')

            ax1.set_xticks(x_range)
            ax1.set_xticklabels(x_labels, rotation=0, ha="center", fontsize=12)
            if len(str(int(max_data))) > 5:
                yticks_positions = list(range(0, rounded_max_data, rounded_step_data))
                yticks_labels = [
                    f"{int(i//(100000))}" + '萬' if len(str(i)) > 6 else f"{i}" for i
                    in yticks_positions]
                for index, bar in enumerate(critical_qty_bars):
                    height = bar.get_height()
                    ax1.text(
                        bar.get_x() + bar.get_width() / 2,
                        height,
                        f'{round(height/1000,2)}' if height > 0 else '',
                        ha='center', va='bottom', fontsize=10  # Align the text
                    )
                for index, bar in enumerate(major_qty_bars):
                    height = bar.get_height()
                    ax1.text(
                        bar.get_x() + bar.get_width() / 2,
                        height,
                        f'{round(height/1000,2)}' if height > 0 else '',
                        ha='center', va='bottom', fontsize=10  # Align the text
                    )
                for index, bar in enumerate(minor_qty_bars):
                    height = bar.get_height()
                    ax1.text(
                        bar.get_x() + bar.get_width() / 2,
                        height,
                        f'{round(height/1000,2)}' if height > 0 else '',
                        ha='center', va='bottom', fontsize=10  # Align the text
                    )

            yticks_labels[-1] = ""
            yticks_labels[0] = ""

            yticks_labels = yticks_labels[1:-1]
            ax1.set_yticks(yticks_positions[1:-1])
            ax1.set_yticklabels(yticks_labels, fontsize=12)
            ax1.axhline(y=0, color='black', linestyle='-', linewidth=1)
            ax1.set_ylim(0, rounded_max_data)
            ax1.spines['top'].set_visible(False)
            ax1.spines['right'].set_visible(False)
            ax1.spines['left'].set_visible(True)
            ax1.spines['bottom'].set_visible(False)
            pvc_sx = "(当责：制造-Nguyen Van A - 负责：制造-Nguyen Van C + Thach A)"
            nbr_sx = "(当责：制造-Nguyen Van B - 负责：制造-Nguyen Van D + Thach B)"
            title = f"""GD-{plant}-外觀DPM\n{pvc_sx if plant=='PVC' else nbr_sx}"""
            handles1, labels1 = ax1.get_legend_handles_labels()
            fig.legend(
                handles1,
                labels1,
                loc='center left',
                fontsize=10,
                title_fontsize=14,
                bbox_to_anchor=(1.0, 0.5),
                ncol=1
            )

            plt.title(title, fontsize=20, fontweight='bold', color='#0070c0')

            plt.tight_layout()

            file_name = f"""GD-{plant}外觀DPM"""
            chart_img = os.path.join(save_path, file_name)

            plt.savefig(f"{chart_img}", dpi=100, bbox_inches="tight", pad_inches=0.45)
            plt.close('all')
        except Exception as e:
            print(e)
            pass

    def excel_chart_8b(self, plant):
        try:
            rcParams['font.sans-serif'] = ['Microsoft YaHei']
            rcParams['axes.unicode_minus'] = False
            save_path = self.save_path

            x_labels, week_date = self.generate_previous_weeks_with_dates()
            rate = []
            for label in x_labels:
                sql = f"""
                SELECT SUM(CAST(CASE WHEN ScrapQuantity IS NOT NULL THEN ScrapQuantity ELSE 0 END AS INT)) AS scrap, 
                SUM(CAST(CASE WHEN CountingQty IS NOT NULL THEN CountingQty ELSE 0 END AS BIGINT)) 
                + SUM(CAST(CASE WHEN ScrapQuantity IS NOT NULL THEN ScrapQuantity ELSE 0 END AS INT)) 
                + SUM(CAST(CASE WHEN FaultyQuantity IS NOT NULL THEN FaultyQuantity ELSE 0 END AS INT)) AS total
                FROM [MES_OLAP].[dbo].[counting_daily_info_raw]
                where Machine like '%{plant}%' and MonthWeek = '{label}'
                """
                row = vnedc_database().select_sql_dict(sql)
                rate.append(
                    round((row[0]['scrap'] / row[0]['total']) * 100, 2) if row[0]['scrap'] is not None and row[0][
                        'total'] is not None else -1)

            filtered_data = [(i, r) for i, r in enumerate(rate) if r != -1]
            x_filtered, valid_rate = zip(*filtered_data) if filtered_data else ([], [])

            max_data = max(valid_rate, default=0)
            rounded_max_data = int(max_data * 1.5) if int(max_data * 1.5) >= 1 else 1
            rounded_step_data = 0.25 if int(max_data * 1.5) >= 1 else 0.1

            plt.figure(figsize=(18, 9))
            fig, ax1 = plt.subplots(figsize=(18, 9))
            ax1.set_xticks(range(len(x_labels)))
            ax1.set_xticklabels(x_labels, rotation=0, ha="center", fontsize=12)
            if plant == 'NBR':
                ax1.axhline(y=1, color='#ffdf7f', linestyle='-', linewidth=2, label='废品率上限')
            else:
                ax1.axhline(y=0.5, color='#ffdf7f', linestyle='-', linewidth=2, label='废品率上限')
            x_smooth = np.linspace(min(x_filtered), max(x_filtered), 300)
            spline = make_interp_spline(x_filtered, valid_rate, k=2)  # Quadratic spline for smoother corners
            y_smooth = spline(x_smooth)
            ax1.plot(x_smooth, y_smooth, color='#0f9ed5', label='废品%', linewidth=2)
            ax1.scatter(x_filtered, valid_rate, color='#0f9ed5', label='', zorder=5)

            # ax1.plot(x_filtered, valid_rate, marker='o', linestyle='-', color='#0f9ed5', label='废品%', linewidth=2)
            for x, y in zip(x_filtered, valid_rate):
                ax1.annotate(f'{y:.2f}%', xy=(x, y), xytext=(0, 5), textcoords='offset points', ha='center',
                             fontsize=10, bbox=dict(facecolor='#dbdbdc', edgecolor='none', boxstyle='round,pad=0.3'))
            ax1.set_ylim(0, rounded_max_data + rounded_step_data)
            yticks = [round(i, 2) for i in np.arange(0, rounded_max_data + rounded_step_data, rounded_step_data)]
            ax1.set_yticks(yticks)
            ax1.set_yticklabels([f'{i}%' for i in yticks])
            ax1.set_ylabel("废品%", fontsize=14)
            ax1.set_xlim(-0.5, len(x_labels) - 0.5)
            ax1.spines['top'].set_visible(False)
            ax1.spines['right'].set_visible(False)
            ax1.spines['left'].set_visible(True)
            ax1.spines['bottom'].set_visible(True)

            pvc_sx = "(当责：制造-Nguyen Van A - 负责：制造-Nguyen Van C + Thach A)"
            nbr_sx = "(当责：制造-Nguyen Van B - 负责：制造-Nguyen Van D + Thach B)"
            title = f"""GD-{plant} 废品率%\n{pvc_sx if plant == 'PVC' else nbr_sx}"""
            ax1.set_title(title, fontsize=20, fontweight='bold', color='#0070c0')

            handles1, labels1 = ax1.get_legend_handles_labels()
            fig.legend(
                handles1,
                labels1,
                loc='center left',
                fontsize=10,
                title_fontsize=14,
                bbox_to_anchor=(1.0, 0.5),
                ncol=1
            )

            plt.tight_layout()
            os.makedirs(save_path, exist_ok=True)
            file_name = f"GD-{plant}废品率%"
            chart_img = os.path.join(save_path, f"{file_name}.png")
            plt.savefig(chart_img, dpi=100, bbox_inches="tight", pad_inches=0.45)
            plt.close('all')
        except Exception as e:
            print(e)
            pass

    def excel_chart_8a(self, plant):
        try:
            rcParams['font.sans-serif'] = ['Microsoft YaHei']
            rcParams['axes.unicode_minus'] = False
            save_path = self.save_path

            x_labels, week_date = self.generate_previous_weeks_with_dates()
            rate = []
            for label in x_labels:
                sql = f"""
                   SELECT SUM(CAST(CASE WHEN FaultyQuantity IS NOT NULL THEN FaultyQuantity ELSE 0 END AS INT)) AS faulty, 
                   SUM(CAST(CASE WHEN CountingQty IS NOT NULL THEN CountingQty ELSE 0 END AS BIGINT)) 
                   + SUM(CAST(CASE WHEN ScrapQuantity IS NOT NULL THEN ScrapQuantity ELSE 0 END AS INT)) 
                   + SUM(CAST(CASE WHEN FaultyQuantity IS NOT NULL THEN FaultyQuantity ELSE 0 END AS INT)) AS total
                   FROM [MES_OLAP].[dbo].[counting_daily_info_raw]
                   where Machine like '%{plant}%' and MonthWeek = '{label}'
                   """
                row = vnedc_database().select_sql_dict(sql)
                rate.append(
                    round((row[0]['faulty'] / row[0]['total']) * 100, 2) if row[0]['faulty'] is not None and row[0][
                        'total'] is not None else -1)

            filtered_data = [(i, r) for i, r in enumerate(rate) if r != -1]
            x_filtered, valid_rate = zip(*filtered_data) if filtered_data else ([], [])

            max_data = max(valid_rate, default=0)
            rounded_max_data = int(max_data * 1.5) if int(max_data * 1.5) >= 1 else 1
            rounded_step_data = 0.25 if int(max_data * 1.5) >= 1 else 0.1

            plt.figure(figsize=(18, 9))
            fig, ax1 = plt.subplots(figsize=(18, 9))
            ax1.set_xticks(range(len(x_labels)))
            ax1.set_xticklabels(x_labels, rotation=0, ha="center", fontsize=12)

            x_smooth = np.linspace(min(x_filtered), max(x_filtered), 300)
            spline = make_interp_spline(x_filtered, valid_rate, k=2)  # Quadratic spline for smoother corners
            y_smooth = spline(x_smooth)
            ax1.plot(x_smooth, y_smooth, color='#0f9ed5', label='二級品率%', linewidth=2)
            ax1.scatter(x_filtered, valid_rate, color='#0f9ed5', label='', zorder=5)

            # ax1.plot(x_filtered, valid_rate, marker='o', linestyle='-', color='#0f9ed5', label='废品%', linewidth=2)
            for x, y in zip(x_filtered, valid_rate):
                ax1.annotate(f'{y:.2f}%', xy=(x, y), xytext=(0, 10), textcoords='offset points', ha='center',
                             fontsize=10, bbox=dict(facecolor='#dbdbdc', edgecolor='none', boxstyle='round,pad=0.3'))
            ax1.set_ylim(0, rounded_max_data + rounded_step_data)
            yticks = [round(i, 2) for i in np.arange(0, rounded_max_data + rounded_step_data, rounded_step_data)]
            ax1.set_yticks(yticks)
            ax1.set_yticklabels([f'{i}%' for i in yticks])
            ax1.set_ylabel("废品%", fontsize=14)
            ax1.set_xlim(-0.5, len(x_labels) - 0.5)
            ax1.spines['top'].set_visible(False)
            ax1.spines['right'].set_visible(False)
            ax1.spines['left'].set_visible(True)
            ax1.spines['bottom'].set_visible(True)

            pvc_sx = "(当责：制造-Nguyen Van A - 负责：制造-Nguyen Van C + Thach A)"
            nbr_sx = "(当责：制造-Nguyen Van B - 负责：制造-Nguyen Van D + Thach B)"
            title = f"""8b.GD-{plant} 二級品率%\n{pvc_sx if plant == 'PVC' else nbr_sx}"""
            ax1.set_title(title, fontsize=20, fontweight='bold', color='#0070c0')

            handles1, labels1 = ax1.get_legend_handles_labels()
            fig.legend(
                handles1,
                labels1,
                loc='center left',
                fontsize=10,
                title_fontsize=14,
                bbox_to_anchor=(1.0, 0.5),
                ncol=1
            )

            plt.tight_layout()
            os.makedirs(save_path, exist_ok=True)
            file_name = f"GD-{plant}二級品率"
            chart_img = os.path.join(save_path, f"{file_name}.png")
            plt.savefig(chart_img, dpi=100, bbox_inches="tight", pad_inches=0.45)
            plt.close('all')
        except Exception as e:
            print(e)
            pass

    def get_mach_list(self, plant):
        if plant == 'NBR':
            plant_ = 'NBR'
        else:
            plant_ = 'PVC1'

        sql = f"""
                        select name FROM [PMGMES].[dbo].[PMG_DML_DataModelList] 
                        where DataModelTypeId = 'DMT000003' and name like '%{plant_}%' 
                        and name not in ('VN_GD_PVC1_L03','VN_GD_PVC1_L04')
                        order by name
                    """
        data = mes_database().select_sql_dict(sql)

        return data

    def main(self):
        for plant in self.plant_name:
            self.mach_list = self.get_mach_list(plant)
            self.generate_chart(plant)
            if self.mode == 'WEEKLY':
                self.excel_chart_2(plant)
                # self.excel_chart_6a(plant)
                self.excel_chart_8a(plant)
                self.excel_chart_8b(plant)
                self.weekly_chart(plant)
            if self.mode == 'MONTHLY':
                self.monthly_chart(plant)
            self.rate_chart(plant)
            self.generate_raw_excel(plant)

        # self.send_email(self.file_list, self.image_buffers)


import argparse
from datetime import datetime, timedelta, date

parser = argparse.ArgumentParser(description="解析外部参数")
parser.add_argument("--mode", choices=['WEEKLY', 'MONTHLY'], help="MONTHLY OR WEEKLY")
args = parser.parse_args()
mode = args.mode

if not mode:
    mode = "WEEKLY"

report = mes_weekly_report(mode)
report.main()