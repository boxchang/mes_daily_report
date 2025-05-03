import sys
import os

from openpyxl.formatting.rule import CellIsRule

curPath = os.path.abspath(os.path.dirname(__file__))
rootPath = os.path.split(curPath)[0]
sys.path.append(rootPath)
import time
from matplotlib.ticker import FuncFormatter
import configparser
import ast
from factory import Factory, DataControl, ColumnControl, ConfigObject, SetReportLog
import numpy as np
from openpyxl.utils import get_column_letter
from database import mes_database, mes_olap_database, lkmes_database, lkmes_olap_database
import pandas as pd
from openpyxl.styles import Alignment, NamedStyle, Font, Border, Side, PatternFill
import logging
from openpyxl.comments import Comment
import matplotlib.pyplot as plt
from io import BytesIO
from datetime import datetime, timedelta, date


class mes_daily_report(object):
    def is_special_date(self, date):
        result = False

        sql = f"""
        SELECT CONTROL_DATE FROM [MES_OLAP].[dbo].[special_date] WHERE job_name = 'mes_daily_report'
        AND CONTROL_DATE = CONVERT(DATE, '{date}', 112)
        """
        raws = self.mes_olap_db.select_sql_dict(sql)

        if len(raws) > 0:
            result = True

        return result

    def __init__(self, report_date1, report_date2):
        config_file = "mes_daily_report.config"
        mail_config_file = 'mes_daily_report_mail.config'

        self.config = ConfigObject(config_file, mail_config_file)
        self.report_date1 = report_date1
        self.report_date2 = report_date2
        if self.config.location in "GD":
            self.mes_db = mes_database()
            self.mes_olap_db = mes_olap_database()
        elif self.config.location in "LK":
            self.mes_db = lkmes_database()
            self.mes_olap_db = lkmes_olap_database()
        else:
            self.mes_db = None
            self.mes_olap_db = None

        # Save Path media/daily_output/
        fold_name = self.report_date1
        self.save_path = os.path.join("daily_output", fold_name)

        # Check folder to create
        if not os.path.exists(self.save_path):
            os.makedirs(self.save_path)

        SetReportLog()

    def main(self):
        mes_db = self.mes_db
        mes_olap_db = self.mes_olap_db
        location = self.config.location
        plants = self.config.plants
        hour_output_limit = self.config.hour_output_limit
        fix_mode = self.config.fix_mode
        report_font = self.config.report_font

        print(f"Plant: {plants}, Hour Output Limit: {hour_output_limit}, Fix Mode: {fix_mode}")

        report_date1 = self.report_date1
        report_date2 = self.report_date2

        if self.is_special_date(report_date1):
            sys.exit()

        file_list = []
        image_buffers = []

        dr = DailyReport(mes_db, mes_olap_db, location, report_date1, report_date2, hour_output_limit,
                         report_font, logging)

        logging.info(f"precheck......")
        dr.Precheck()

        for plant in plants:
            dr.plant = plant

            logging.info(f"{plant} start running......")
            dr.Target_Setting(location, plant)

            logging.info(f"{plant} generate_main_df......")
            main_df, cosmetic_df = dr.generate_main_df()

            logging.info(f"{plant} fix_main_df......")
            fixed_main_df = dr.fix_main_df(main_df)

            logging.info(f"{plant} sorting_data......")
            subtotals_df, chart_df, activation_df = dr.sorting_data(fixed_main_df, cosmetic_df)

            logging.info(f"{plant} validate_data......")
            dr.validate_data(fixed_main_df, subtotals_df)

            # Generate Excel file
            logging.info(f"{plant} generate_excel......")
            file_name = f'MES_{location}_{plant}_DAILY_Report_{report_date1}.xlsx'
            excel_file = os.path.join(self.save_path, file_name)

            dr.generate_excel(fixed_main_df, subtotals_df, activation_df, cosmetic_df, excel_file)
            if os.path.exists(excel_file):
                file_list.append({'file_name': file_name, 'excel_file': excel_file})

            # Generate Chart
            logging.info(f"{plant} generate_chart......")
            image_file = f'MES_{location}_{plant}_bar_chart_{report_date1}.png'
            image_file = os.path.join(self.save_path, image_file)

            dr.generate_chart(chart_df, image_file)
            image_buffers.append(image_file)

        if not fix_mode:
            logging.info(f"{location} send_email......")
            subject = f'[{location} Report] 產量日報表 {self.report_date1}'
            dr.send_email(self.config, subject, file_list, image_buffers, dr.msg_list, dr.error_list)


class DailyReport(Factory):
    file_list = []
    error_list = []
    msg_list = []

    def Target_Setting(self, location, plant):
        if "GD" in location:
            self.capacity_target = 0.99
            self.yield_target = 0.97
            self.activation_target = 0.99
            self.oee_target = 0.95
        elif "LK" in location:
            self.capacity_target = 0.97
            self.yield_target = 0.95
            self.activation_target = 0.98
            self.oee_target = 0.90

        self.isolation_target = 0.05
        self.pinhole_target = 0.05
        self.weight_target = 0.1
        if 'NBR' in plant:
            self.scrap_target = 0.008
            self.faulty_target = 0.002
            self.former_miss_target = 0.015
        elif 'PVC' in plant:
            self.scrap_target = 0.0035
            self.faulty_target = 0
            self.former_miss_target = 0.001


    def __init__(self, mes_db, mes_olap_db, location, report_date1, report_date2, hour_output_limit, report_font,
                 logger):
        self.location = location
        self.report_date1 = report_date1
        self.report_date2 = report_date2
        self.hour_output_limit = int(hour_output_limit) if hour_output_limit else 1000
        self.report_font = report_font
        self.mes_db = mes_db
        self.mes_olap_db = mes_olap_db
        self.logger = logger

    def Precheck(self):
        start_time = time.time()

        sql = f"""
        SELECT distinct cos.defect_code
          FROM [MES_OLAP].[dbo].[counting_hourly_info_raw] r
          LEFT JOIN [MES_OLAP].[dbo].[mes_ipqc_cosmetic_data] cos on r.Runcard = cos.runcard
          LEFT JOIN [MES_OLAP].[dbo].[mes_defect_define] d on d.defect_code = cos.defect_code
          where r.belong_to = '{self.report_date1}' and cos.defect_code <> ''
		  and d.defect_code is null
        """
        rows = self.mes_olap_db.select_sql_dict(sql)

        for row in rows:
            defect_code = row['defect_code']
            msg = f"MES_OLAP mes_defect_define {defect_code} 沒有設定缺點"
            if msg not in self.error_list:
                self.error_list.append(msg)

        end_time = time.time()
        elapsed_time = end_time - start_time
        self.logger.info(f"Time taken: {elapsed_time:.2f} seconds.")

    def get_df_main(self):
        scada_table = ""
        upper_column = ""
        lower_column = ""

        if "NBR" in self.plant:
            upper_column = "UpperLineSpeed_Min"
            lower_column = "LowerLineSpeed_Min"
            scada_table = "[PMG_MES_NBR_SCADA_Std]"
        elif "PVC" in self.plant:
            upper_column = "UpperSpeed"
            lower_column = "LowerSpeed"
            scada_table = "[PMGMES].[dbo].[PMG_MES_PVC_SCADA_Std]"

        report_date1 = self.report_date1
        report_date2 = self.report_date2
        plant = self.plant

        sql = f"""
            WITH WorkOrderInfo AS (
                SELECT 
                    w.MachineId,
                    dl.Name,
                    wi.LineId AS Line,
                    CAST(r.Period AS INT) AS Period,
                    wi.WorkOrderId,
                    wi.StartDate, 
                    wi.EndDate,
                    WorkOrderDate,
                    CustomerName,
                    w.ProductItem,
                    nss.{lower_column} LineSpeedLower,
                    nss.{upper_column} LineSpeedUpper,
                    m.COUNTING_MACHINE,
                    w.PartNo,
                    w.AQL,
                    w.PlanQty,
                    wi.Qty,
                    w.Status,
                    ipqc.InspectionStatus weight_result,
                    ipqc.InspectionValue weight_value,
                    ipqc.Lower_InspectionValue weight_lower,
                    ipqc.Upper_InspectionValue weight_upper,
                    hole.InspectionStatus hole_result,
                    r.Id runcard,
                    r.InspectionDate
                FROM 
                    [PMGMES].[dbo].[PMG_MES_WorkOrderInfo] wi
                    JOIN [PMGMES].[dbo].[PMG_MES_WorkOrder] w ON wi.WorkOrderId = w.id
                    JOIN [PMGMES].[dbo].[PMG_DML_DataModelList] dl ON w.MachineId = dl.Id
                    JOIN [PMGMES].[dbo].[PMG_MES_RunCard] r ON r.WorkOrderId = w.Id AND r.LineName = wi.LineId
                    JOIN [PMGMES].[dbo].[PMG_MES_IPQCInspectingRecord] ipqc ON ipqc.RunCardId = r.Id
                    LEFT JOIN [PMGMES].[dbo].[PMG_MES_IPQCInspectingRecord] hole ON hole.RunCardId = r.Id and hole.OptionName = 'Pinhole'
                    LEFT JOIN {scada_table} nss ON nss.PartNo = w.PartNo
                    JOIN [PMG_DEVICE].[dbo].[COUNTING_DATA_MACHINE] m ON dl.Name = m.MES_MACHINE AND wi.LineId = m.LINE
                WHERE 
                    ((r.InspectionDate = '{report_date1}' AND Period between 6 and 23) or (r.InspectionDate = '{report_date2}' AND Period between 0 and 5))
                    AND ipqc.OptionName = 'Weight'
                    AND dl.Name LIKE '%{plant}%' AND COUNTING_MACHINE LIKE '%CountingMachine%'
            ),
            Pitch AS (
                SELECT Name, 
                CAST(ISNULL(attr1.AttrValue, 1) AS FLOAT) AS std_val, 
                CAST(attr2.AttrValue AS FLOAT) AS act_val, 
                ISNULL(CAST(CAST(attr2.AttrValue AS FLOAT) / CAST(ISNULL(attr1.AttrValue, 1) AS FLOAT) AS FLOAT), 1) AS pitch_rate
                  FROM [PMGMES].[dbo].[PMG_DML_DataModelList] dl
                  LEFT JOIN [PMGMES].[dbo].[PMG_DML_DataModelAttrList] attr1 on dl.Id = attr1.DataModelListId and attr1.AttrName = 'StandardPitch'
                  LEFT JOIN [PMGMES].[dbo].[PMG_DML_DataModelAttrList] attr2 on dl.Id = attr2.DataModelListId and attr2.AttrName = 'ActualPitch'
                  WHERE DataModelTypeId = 'DMT000003'
            ),
            Machines as (
                select distinct MES_MACHINE Name from [PMG_DEVICE].[dbo].[COUNTING_DATA] c
                    join [PMG_DEVICE].[dbo].[COUNTING_DATA_MACHINE] m on c.MachineName = m.COUNTING_MACHINE
                    where creationTime between CONVERT(DATETIME, '{report_date1} 06:00:00', 120) AND CONVERT(DATETIME, '{report_date2} 05:59:59', 120)
                    and m.MES_MACHINE like '%{plant}%'
            ),
            Optical as (
                SELECT MES_MACHINE, LINE,CAST(DATEPART(hour, Cdt) AS INT) AS Period, sum(CAST(OKQty AS BIGINT)) OKQty, sum(CAST(NGQty AS BIGINT)) NGQty, 
                CASE 
                         WHEN sum(CAST(NGQty AS BIGINT)) = 0 THEN 0
                         ELSE ROUND(CAST(sum(CAST(NGQty AS BIGINT)) AS FLOAT)/(sum(CAST(OKQty AS BIGINT))+sum(CAST(NGQty AS BIGINT))), 3) 
                       END OpticalNGRate
                  FROM [PMG_DEVICE].[dbo].[OpticalDevice] o
                  JOIN [PMG_DEVICE].[dbo].[COUNTING_DATA_MACHINE] m on o.DeviceId = m.COUNTING_MACHINE
                  where Cdt between CONVERT(DATETIME, '{report_date1} 06:00:00', 120) and CONVERT(DATETIME, '{report_date2} 06:00:00', 120)
                  group by MES_MACHINE, LINE, CAST(DATEPART(hour, Cdt) AS INT)
            ),
            Faulty as (
                SELECT r.InspectionDate,r.Period,f.CreationTime,m.Name,f.ActualQty,f.DefectCode1,c.Descr,r.WorkOrderId,r.LineName,r.Id runcardId
                  FROM [PMGMES].[dbo].[PMG_MES_Faulty] f
                  JOIN [PMGMES].[dbo].[PMG_MES_RunCard] r on f.RunCardId = r.Id
                  JOIN [PMGMES].[dbo].[PMG_DML_DataModelList] m on r.MachineId = m.Id
                  LEFT JOIN [PMGMES].[dbo].[PMG_DML_ConstValue] c on c.[Value] = f.DefectCode1
                  where ((r.InspectionDate = '{report_date1}' AND Period between 6 and 23) or (r.InspectionDate = '{report_date2}' AND Period between 0 and 5))
            ),
            Scrap as (
                SELECT r.InspectionDate,r.Period,s.CreationTime,m.Name,s.ActualQty,r.WorkOrderId,r.LineName,r.Id runcardId
                  FROM [PMGMES].[dbo].[PMG_MES_Scrap] s
                  JOIN [PMGMES].[dbo].[PMG_MES_RunCard] r on s.RunCardId = r.Id
                  JOIN [PMGMES].[dbo].[PMG_DML_DataModelList] m on r.MachineId = m.Id
                  where ((r.InspectionDate = '{report_date1}' AND Period between 6 and 23) or (r.InspectionDate = '{report_date2}' AND Period between 0 and 5))
            ),
            WIPPacking as (
                SELECT RunCardId,InspectedAQL,sum(ActualQty) ActualQty from [PMGMES].[dbo].[PMG_MES_WorkInProcess] wip
                JOIN WorkOrderInfo w on w.runcard = wip.RunCardId
                AND PackingType = 'WIPPacking'
                group by RunCardId,InspectedAQL
            ),
            OnlinePacking as (
                SELECT RunCardId,InspectedAQL,sum(ActualQty) ActualQty from [PMGMES].[dbo].[PMG_MES_WorkInProcess] wip
                JOIN WorkOrderInfo w on w.runcard = wip.RunCardId
                AND PackingType = 'OnlinePacking'
                group by RunCardId,InspectedAQL
            ),
            Ticket as (
                SELECT RunCardId,InspectedAQL,sum(ActualQty) ActualQty from [PMGMES].[dbo].[PMG_MES_WorkInProcess] wip
                JOIN WorkOrderInfo w on w.runcard = wip.RunCardId
                group by RunCardId,InspectedAQL
            ),
			PMG_MES_Isolation2 as (
				SELECT RunCardId,sum(ActualQty) ActualQty from [PMGMES].[dbo].[PMG_MES_Isolation] i
                JOIN WorkOrderInfo w on w.runcard = i.RunCardId
                group by RunCardId
			)

            SELECT 
                wo.MachineId,
                mach.Name,
                wo.Line,
                wo.Period,
                wo.StartDate, 
                wo.EndDate,
                wo.WorkOrderId,
                wo.WorkOrderDate,
                wo.CustomerName,
                wo.PartNo,
                wo.ProductItem,
                wo.AQL,
                t.InspectedAQL,
                wo.PlanQty,
                wo.Qty,
                wo.Status,
                CAST(wo.LineSpeedLower/pitch_rate AS INT) AS LineSpeedLower,
                CAST(wo.LineSpeedUpper/pitch_rate AS INT) AS LineSpeedUpper,
                60 AS ProductionTime,
                hole_result Separate,
                ISNULL(s.ActualQty, 0) Scrap,
                ISNULL(f.ActualQty, 0) SecondGrade,
                CAST(60 * wo.LineSpeedUpper/pitch_rate AS INT) AS Target,
                weight_result OverControl,
                CAST(round(weight_value,2) AS DECIMAL(10, 2)) WeightValue,
                OpticalNGRate,
                CAST(round(weight_lower,2) AS DECIMAL(10, 2)) WeightLower,
                CAST(round(weight_upper,2) AS DECIMAL(10, 2)) WeightUpper,
                runcard,
                wp.ActualQty WIPPacking,
                op.ActualQty OnlinePacking,
                t.ActualQty Ticket_Qty,
                isn.ActualQty Isolation_Qty,
                wo.StartDate WoStartDate, 
                wo.EndDate WoEndDate,
                wo.InspectionDate AS Date
            FROM 
                Machines mach
                LEFT JOIN WorkOrderInfo wo ON mach.Name = wo.Name
                LEFT JOIN Optical o ON wo.Name = o.MES_MACHINE AND wo.Line = o.LINE AND wo.Period = o.Period
                LEFT JOIN Faulty f ON wo.runcard = f.runcardId
                LEFT JOIN Scrap s ON wo.runcard = s.runcardId
                LEFT JOIN WIPPacking wp on wo.runcard = wp.RunCardId
                LEFT JOIN OnlinePacking op on wo.runcard = op.RunCardId
                LEFT JOIN Pitch pc on pc.Name = wo.Name
                LEFT JOIN PMG_MES_Isolation2 isn on isn.RunCardId = wo.runcard
                LEFT JOIN Ticket t on wo.runcard = t.RunCardId
                WHERE NOT (wo.WorkOrderId IS NOT NULL AND t.ActualQty IS NULL) --有小票才列入計算，主要是User會用錯RunCard，以有小票為主進行統計
            ORDER BY 
                mach.Name, 
                wo.Period, 
                wo.Line;
                """
        raws = self.mes_db.select_sql_dict(sql)

        df_main = pd.DataFrame(raws)

        return df_main

    # Counting Machine Data
    def get_df_detail(self):

        sql = f"""
                        SELECT FORMAT(CreationTime, 'yyyy-MM-dd') AS CountingDate,CAST(DATEPART(hour, CreationTime) as INT) Period ,m.mes_machine Name,m.line Line, max(Speed) max_speed,min(Speed) min_speed,round(avg(Speed),0) avg_speed,sum(Qty2) sum_qty
                          FROM [PMG_DEVICE].[dbo].[COUNTING_DATA] c, [PMG_DEVICE].[dbo].[COUNTING_DATA_MACHINE] m
                          where CreationTime between CONVERT(DATETIME, '{self.report_date1} 06:00:00', 120) and CONVERT(DATETIME, '{self.report_date2} 05:59:59', 120)
                          and c.MachineName = m.counting_machine and m.mes_machine like '%{self.plant}%'
                          group by m.mes_machine,FORMAT(CreationTime, 'yyyy-MM-dd'),DATEPART(hour, CreationTime),m.line
                          order by m.mes_machine,FORMAT(CreationTime, 'yyyy-MM-dd'),DATEPART(hour, CreationTime),m.line
                        """
        detail_raws = self.mes_db.select_sql_dict(sql)
        df_detail = pd.DataFrame(detail_raws)

        return df_detail

    def get_df_ipqc(self):
        sql = f"""
        WITH Customer AS (
        SELECT distinct CAST(CAST(substring([KUNNR],6,6) AS Int) AS varchar) customer_code,SORTL customer_name,[VKBUR] SalePlaceCode
        FROM [PMG_SAP].[dbo].[ZKNA1]
        )
        
        select c.Runcard runcard,cu.SalePlaceCode
          ,[Tensile_Value]
          ,[Tensile_Limit]
          ,[Tensile_Status]
          ,[Tensile_Defect]
          ,[Elongation_Value]
          ,[Elongation_Limit]
          ,[Elongation_Status]
          ,[Elongation_Defect]
          ,[Roll_Value]
          ,[Roll_Limit]
          ,[Roll_Status]
          ,[Roll_Defect]
          ,[Cuff_Value]
          ,[Cuff_Limit]
          ,[Cuff_Status]
          ,[Cuff_Defect]
          ,[Palm_Value]
          ,[Palm_Limit]
          ,[Palm_Status]
          ,[Palm_Defect]
          ,[Finger_Value]
          ,[Finger_Limit]
          ,[Finger_Status]
          ,[Finger_Defect]
          ,[FingerTip_Value]
          ,[FingerTip_Limit]
          ,[FingerTip_Status]
          ,[FingerTip_Defect]
          ,[Length_Value]
          ,[Length_Limit]
          ,[Length_Status]
          ,[Length_Defect]
          ,[Weight_Value]
          ,[Weight_Limit]
          ,[Weight_Status]
          ,[Weight_Defect]
          ,CASE WHEN Weight_Defect IS NULL THEN NULL WHEN Weight_Defect = 'LL2' THEN 'NG' ELSE 'PASS' END AS Weight_Light
          ,CASE WHEN Weight_Defect IS NULL THEN NULL WHEN Weight_Defect = 'LL1' THEN 'NG' ELSE 'PASS' END AS Weight_Heavy
          ,[Width_Value]
          ,[Width_Limit]
          ,[Width_Status]
          ,[Width_Defect]
          ,[Pinhole_Value]
          ,[Pinhole_Limit]
          ,[Pinhole_Status]
          ,[Pinhole_Defect]
          ,[Cosmetic_Value]
          ,[Cosmetic_Status]
        from MES_OLAP.dbo.counting_hourly_info_raw c
        JOIN MES_OLAP.dbo.mes_ipqc_data ipqc on c.Runcard = ipqc.Runcard
        LEFT JOIN Customer cu on cu.customer_code = c.CustomerCode
        where c.belong_to = '{self.report_date1}'
        """
        rows = self.mes_olap_db.select_sql_dict(sql)
        df = pd.DataFrame(rows)

        return df

    def get_df_cosmetic(self):

        sql = f"""
          WITH Customer AS (
          SELECT distinct CAST(CAST(KUNNR AS BIGINT) AS VARCHAR(50)) customer_code,SORTL customer_name,[VKBUR] SalePlaceCode
          FROM [PMG_SAP].[dbo].[ZKNA1]
          )        
            
          SELECT counting.Runcard runcard,counting.belong_to,counting.Machine,counting.Line,counting.Shift,counting.WorkOrder,counting.PartNo,counting.ProductItem,SalePlaceCode,counting.Period
          FROM [MES_OLAP].[dbo].[counting_hourly_info_raw] counting
          LEFT JOIN [MES_OLAP].[dbo].[mes_ipqc_data] ipqc on ipqc.Runcard = counting.Runcard
          LEFT JOIN Customer cus on cus.customer_code = counting.CustomerCode
          where counting.belong_to = '{self.report_date1}'
          and Machine like '%{self.plant}%'
          and (OnlinePacking > 0 or WIPPacking > 0)
          order by Machine, WorkDate, Cast(Period as Int)

        """
        data = self.mes_olap_db.select_sql_dict(sql)
        df = pd.DataFrame(data)

        weight_sql = f"""
        WITH Customer AS (
        SELECT distinct CAST(CAST(KUNNR AS BIGINT) AS VARCHAR(50)) customer_code,SORTL customer_name,[VKBUR] SalePlaceCode
        FROM [PMG_SAP].[dbo].[ZKNA1]
        )         
        
        SELECT ipqc.Runcard runcard, Cast(SalePlaceCode as varchar)+Weight_Defect defect_code, 1 qty 
          FROM [MES_OLAP].[dbo].[counting_hourly_info_raw] counting
          LEFT JOIN [MES_OLAP].[dbo].[mes_ipqc_data] ipqc on ipqc.Runcard = counting.Runcard
          LEFT JOIN Customer cus on cus.customer_code = counting.CustomerCode
          where counting.belong_to = '{report_date1}'
          and Weight_Status = 'NG'
          and Machine like '%{self.plant}%'
          and InspectedAQL is not Null
          and cus.customer_code is not null

          --6100 美
          --6200 歐
          --6300 日
          --LL1 過重
          --LL2 過輕
        """
        weight_data = self.mes_olap_db.select_sql_dict(weight_sql)
        weight_df = pd.DataFrame(weight_data)
        weight_df = weight_df.pivot(index="runcard", columns="defect_code", values="qty").reset_index()

        fixed_columns = ['runcard', '6100LL1', '6100LL2', '6200LL1', '6200LL2', '6300LL1', '6300LL2', '7000LL1', '7000LL2']

        for col in fixed_columns:
            if col not in weight_df.columns:
                weight_df[col] = np.nan

        # 按照 target_columns 的順序排列欄位
        weight_df = weight_df[fixed_columns]

        cosmetic_sql = f"""
         SELECT r.runcard, d.defect_code, sum(qty) cosmetic_qty,  max(cos.cosmetic_inspect_qty) cosmetic_inspect_qty
          FROM [MES_OLAP].[dbo].[counting_hourly_info_raw] r
          LEFT JOIN [MES_OLAP].[dbo].[mes_ipqc_cosmetic_data] cos on r.Runcard = cos.runcard
          LEFT JOIN [MES_OLAP].[dbo].[mes_defect_define] d on d.defect_code = cos.defect_code
          where r.belong_to = '{self.report_date1}' and r.runcard<>''
          group by r.runcard, defect_level, d.defect_code, desc2
        """
        cosmetic_data = self.mes_olap_db.select_sql_dict(cosmetic_sql)
        cosmetic_sample_df = pd.DataFrame(cosmetic_data)

        cosmetic_summary_df = cosmetic_sample_df.groupby('runcard', as_index=False).agg({
                'cosmetic_qty': 'max',
                'cosmetic_inspect_qty': 'max'
            }).copy()
        cosmetic_detail_df = cosmetic_sample_df.pivot(index="runcard", columns="defect_code",
                                                     values="cosmetic_qty").reset_index().copy()

        # 計算針孔Defect Code加總
        pinhole_sql = f"""
        SELECT r.runcard, d.defect_code, sum(qty) pinhole_sum_qty
          FROM [MES_OLAP].[dbo].[counting_hourly_info_raw] r
          JOIN [MES_OLAP].[dbo].[mes_ipqc_pinhole_data] cos on r.Runcard = cos.runcard
          JOIN [MES_OLAP].[dbo].[mes_defect_define] d on d.defect_code = cos.defect_code
          where r.belong_to = '{report_date1}'
          group by r.runcard, defect_level, d.defect_code, desc2
        """
        pinhole_data = self.mes_olap_db.select_sql_dict(pinhole_sql)
        pinhole_df = pd.DataFrame(pinhole_data)
        pinhole_pivot_df = pinhole_df.pivot(index="runcard", columns="defect_code",
                                            values="pinhole_sum_qty").reset_index()
        pinhole_pivot_df['Pinhole'] = pinhole_pivot_df.iloc[:, 1:].notna().any(axis=1).astype(int)

        # 計算針孔樣本數
        pinhole_sample_sql = f"""
        SELECT r.Id runcard,WorkCenterTypeName, AQL AS WO_AQL
          FROM [PMGMES].[dbo].[PMG_MES_RunCard] r
          JOIN [PMGMES].[dbo].[PMG_MES_RunCard_IPQCInspectIOptionMapping] m on r.Id = m.RunCardId
          JOIN [PMGMES].[dbo].[PMG_MES_WorkOrder] w on r.WorkOrderId = w.Id
          where GroupType = 'Pinhole' and r.InspectionDate between '{self.report_date1}' and '{self.report_date2}'
        """
        pinhole_sample_data = self.mes_db.select_sql_dict(pinhole_sample_sql)
        pinhole_sample_df = pd.DataFrame(pinhole_sample_data)

        if "PVC" in self.plant:
            pinhole_sample_df['Pinhole_Sample'] = 25
        elif "NBR" in self.plant:
            pinhole_sample_df['Pinhole_Sample'] = np.where(pinhole_sample_df['WO_AQL'] == '1.0', 50, 25)

        df = pd.merge(df, weight_df, on=['runcard'], how='left')
        df = pd.merge(df, cosmetic_summary_df, on=['runcard'], how='left')
        df = pd.merge(df, pinhole_pivot_df, on=['runcard'], how='left')
        df = pd.merge(df, cosmetic_detail_df, on="runcard", how="left")
        df = pd.merge(df, pinhole_sample_df, on="runcard", how="left")
        df = df.fillna('')

        return df

    def generate_main_df(self):
        start_time = time.time()

        df_main = self.get_df_main()

        df_detail = self.get_df_detail()

        df_ipqc = self.get_df_ipqc()

        df_dmf = self.get_dmf_rate()

        df_lost_mold = self.get_lost_mold_rate()

        final_df = pd.merge(df_main, df_detail, on=['Name', 'Period', 'Line'], how='left')

        final_df = pd.merge(final_df, df_ipqc, on=['runcard'], how='left')

        final_df = pd.merge(final_df, df_dmf, on=['Name', 'Period', 'Line'], how='left')

        final_df = pd.merge(final_df, df_lost_mold, on=['Name', 'Period', 'Line'], how='left')

        final_df['OverControlQty'] = final_df.apply(
            lambda row: row['Ticket_Qty'] if row['Weight_Status'] == 'NG' else 0, axis=1)

        cosmetic_df = self.get_df_cosmetic()

        end_time = time.time()
        elapsed_time = end_time - start_time
        self.logger.info(f"Time taken: {elapsed_time:.2f} seconds.")

        return final_df, cosmetic_df

    def get_df_fix(self):

        sql = f"""
        SELECT WorkDate CountingDate, Machine Name, Line, Period, MinSpeed, MaxSpeed, AvgSpeed, CountingQty
          FROM [MES_OLAP].[dbo].[counting_hourly_info_fix] where 
          WorkDate between '{self.report_date1}' and '{self.report_date2}'
        """

        raws = self.mes_olap_db.select_sql_dict(sql)
        df = pd.DataFrame(raws)

        return df

    def fix_main_df(self, main_df):
        start_time = time.time()

        fix_df = self.get_df_fix()

        if len(fix_df) > 0:
            main_df = pd.merge(main_df, fix_df, on=['CountingDate', 'Name', 'Period', 'Line'], how='left')

            # 點數機資料修正
            main_df.loc[
                main_df["CountingQty"].notna(), ["max_speed", "min_speed", "avg_speed", "sum_qty"]] = \
                main_df.loc[main_df["CountingQty"].notna(), ["MaxSpeed", "MinSpeed", "AvgSpeed", "CountingQty"]].values

        main_df['avg_speed'] = main_df['avg_speed'].fillna(0).round().astype('int')
        main_df['LineSpeedLower'] = main_df['LineSpeedLower'].fillna(0).round().astype('int')
        main_df['LineSpeedUpper'] = main_df['LineSpeedUpper'].fillna(0).round().astype('int')

        main_df['Date'] = pd.to_datetime(main_df['Date'])
        main_df['Date'] = main_df['Date'].dt.strftime('%Y/%m/%d')

        end_time = time.time()
        elapsed_time = end_time - start_time
        self.logger.info(f"Time taken: {elapsed_time:.2f} seconds.")

        return main_df

    def sorting_data(self, df, cosmetic_df):
        start_time = time.time()

        def join_values(col):
            return '/'.join(map(str, sorted(set(col))))

        def min2hour(col):
            sum_val = col.sum()
            data_list = col.tolist()
            hours = sum_val / 60
            return hours

        def counting_ng_ratio(col):
            data_list = col.tolist()
            ng_count = data_list.count('NG')

            # 計算 NG 的比例
            ng_ratio = ng_count / len(data_list)
            return ng_ratio

        def shift(period):
            try:
                if 6 <= int(period) <= 17:
                    return '早班'
                else:
                    return '晚班'
            except Exception as ex:
                return ''

        # Add Column Shift
        df['Shift'] = df['Period'].apply(shift)
        cosmetic_df['Pinhole'] = pd.to_numeric(cosmetic_df['Pinhole'], errors='coerce').fillna(0)
        cosmetic_df['Pinhole_Sample'] = pd.to_numeric(cosmetic_df['Pinhole_Sample'], errors='coerce').fillna(0)

        # 設定IPQC欄位判斷條件
        df['IPQC'] = df.apply(lambda row: "" if pd.isna(row['WorkOrderId']) or row['WorkOrderId'] == ""
        else ('NG' if 'NG' in row[['Tensile_Status', 'Elongation_Status', 'Roll_Status', 'Cuff_Status',
                                   'Palm_Status', 'Finger_Status', 'FingerTip_Status', 'Length_Status',
                                   'Weight_Heavy', 'Weight_Light', 'Width_Status', 'Pinhole_Status']].values
              else 'PASS'), axis=1)

        # Drop the 'Period' and 'Date' column from each group
        group_without_period = df.drop(columns=['Period', 'Date'])

        # Data group by 'Name and then calculating
        mach_grouped = group_without_period.groupby(['Name'])

        rows = []
        chart_rows = []
        activation_rows = []

        for mach_name, mach_group in mach_grouped:
            mach_cosmetic_dt = cosmetic_df[(cosmetic_df['Machine'] == mach_name)]

            # 處理停機情況
            if not mach_group['ProductItem'].iloc[0]:
                subtotal = {
                    'Name': join_values(mach_group['Name']),
                    'ProductItem': '',
                    'AQL': '',
                    'Shift': '',
                    'Line': '',
                    'max_speed': '',
                    'min_speed': '',
                    'avg_speed': '',
                    'LineSpeedLower': '',
                    'LineSpeedUpper': '',
                    'ProductionTime': '',
                    'sum_qty': 0,
                    'OnlinePacking': 0,
                    'WIPPacking': 0,
                    'Isolation_Qty': 0,
                    'DMF_Rate': 0,
                    'Lost_Mold_Rate': 0,
                    'PinholeRate': '',
                    'Scrap': '',
                    'SecondGrade': '',
                    'Target': 0,
                    'OverControl': '',
                }
                subtotal_df = pd.DataFrame([subtotal])
                chart_rows.append(subtotal_df)
                continue

            line_grouped = mach_group.groupby(['Shift', 'Line'])

            tmp_rows = []
            for (shift, line_name), line_group in line_grouped:

                line_sum_qty = line_group['OnlinePacking'].sum() + line_group['WIPPacking'].sum()
                line_tmp_qty = line_sum_qty + line_group['Scrap'].sum() + line_group['SecondGrade'].sum()

                line_secondGrade_qty = line_group['SecondGrade'].sum()
                line_second_rate = round(float(line_secondGrade_qty) / line_tmp_qty, 3) if line_tmp_qty > 0 else 0

                line_scrap_qty = line_group['Scrap'].sum()
                line_scrap_rate = round(float(line_scrap_qty) / line_tmp_qty, 3) if line_tmp_qty > 0 else 0

                line_over_control_qty = line_group['OverControlQty'].sum()
                line_over_control_rate = round(float(line_over_control_qty) / line_tmp_qty, 3) if line_tmp_qty > 0 else 0

                line_cosmetic_dt = mach_cosmetic_dt[(mach_cosmetic_dt['Shift'] == shift) &
                                                    (mach_cosmetic_dt['Line'] == line_name)
                                                    ]

                line_pinhole_dpm = self.calculate_pinhole_dpm(line_group, line_cosmetic_dt)

                try:
                    line_pinhole_rate = round(line_pinhole_dpm / 1_000_000, 4)
                except Exception as e:
                    print(e)

                line_sum = {
                    'Name': '',
                    'ProductItem': join_values(line_group['ProductItem']),
                    'AQL': join_values(line_group['AQL']),
                    'Shift': join_values(line_group['Shift']),
                    'Line': join_values(line_group['Line']),
                    'max_speed': line_group['max_speed'].max(),
                    'min_speed': line_group['min_speed'].min(),
                    'avg_speed': line_group['avg_speed'].mean(),
                    'LineSpeedLower': join_values(line_group['LineSpeedLower']),
                    'LineSpeedUpper': join_values(line_group['LineSpeedUpper']),
                    'ProductionTime': min2hour(line_group['ProductionTime']),
                    'sum_qty': line_group['sum_qty'].sum(),
                    'OnlinePacking': line_group['OnlinePacking'].sum(),
                    'WIPPacking': line_group['WIPPacking'].sum(),
                    'Isolation_Qty': line_group['Isolation_Qty'].sum(),
                    'DMF_Rate': line_group['DMF_Rate'].mean(),
                    'Lost_Mold_Rate': line_group['Lost_Mold_Rate'].mean(),
                    'PinholeRate': line_pinhole_rate,
                    'Scrap': line_scrap_rate,
                    'SecondGrade': line_second_rate,
                    'Target': line_group['Target'].sum(),
                    'OverControl': line_over_control_rate,
                    'Pinhole_DPM': line_pinhole_dpm,
                }
                line_sum_df = pd.DataFrame([line_sum])

                tmp_rows.append(line_sum_df)

            df_tmp = pd.concat(tmp_rows, ignore_index=True)

            # Sorting Data
            # Day Shift
            day_df = df_tmp[df_tmp['Shift'] == '早班'].copy()
            day_production_time = 0
            night_production_time = 0
            if not day_df.empty:
                subtotal = {
                    'Name': '',
                    'ProductItem': '',
                    'AQL': '',
                    'Shift': join_values(day_df['Shift']),
                    'Line': '',
                    'max_speed': day_df['max_speed'].max(),
                    'min_speed': day_df['min_speed'].min(),
                    'avg_speed': day_df['avg_speed'].mean(),
                    'LineSpeedLower': join_values(day_df['LineSpeedLower']),
                    'LineSpeedUpper': join_values(day_df['LineSpeedUpper']),
                    'ProductionTime': day_df['ProductionTime'].mean(),
                    'sum_qty': day_df['sum_qty'].sum(),
                    'OnlinePacking': day_df['OnlinePacking'].sum(),
                    'WIPPacking': day_df['WIPPacking'].sum(),
                    'Isolation_Qty': day_df['Isolation_Qty'].sum(),
                    'DMF_Rate': day_df['DMF_Rate'].mean(),
                    'Lost_Mold_Rate': day_df['Lost_Mold_Rate'].mean(),
                    'PinholeRate': round(day_df['Pinhole_DPM'].sum()/1_000_000, 4),
                    'Scrap': day_df['Scrap'].mean(),
                    'SecondGrade': day_df['SecondGrade'].mean(),
                    'Target': day_df['Target'].sum(),
                    'OverControl': day_df['OverControl'].mean(),
                    'Pinhole_DPM': day_df['Pinhole_DPM'].sum(),
                }
                subtotal_df = pd.DataFrame([subtotal])
                subtotal_df['avg_speed'] = subtotal_df['avg_speed'].round(0)

                day_df[['Name', 'Shift']] = ''
                day_df['avg_speed'] = day_df['avg_speed'].round(0)

                rows.append(day_df)  # Day row data
                rows.append(subtotal_df)  # Day Shift total summary

                if not np.isnan(day_df['ProductionTime'].mean()):
                    day_production_time = day_df['ProductionTime'].mean()

            # Night Shift
            night_df = df_tmp[df_tmp['Shift'] == '晚班'].copy()
            if not night_df.empty:
                subtotal = {
                    'Name': '',
                    'ProductItem': '',
                    'AQL': '',
                    'Shift': join_values(night_df['Shift']),
                    'Line': '',
                    'max_speed': night_df['max_speed'].max(),
                    'min_speed': night_df['min_speed'].min(),
                    'avg_speed': night_df['avg_speed'].mean(),
                    'LineSpeedLower': join_values(night_df['LineSpeedLower']),
                    'LineSpeedUpper': join_values(night_df['LineSpeedUpper']),
                    'ProductionTime': night_df['ProductionTime'].mean(),
                    'sum_qty': night_df['sum_qty'].sum(),
                    'OnlinePacking': night_df['OnlinePacking'].sum(),
                    'WIPPacking': night_df['WIPPacking'].sum(),
                    'Isolation_Qty': night_df['Isolation_Qty'].sum(),
                    'DMF_Rate': night_df['DMF_Rate'].mean(),
                    'Lost_Mold_Rate': night_df['Lost_Mold_Rate'].mean(),
                    'PinholeRate': round(night_df['Pinhole_DPM'].sum()/1_000_000, 4),
                    'Scrap': night_df['Scrap'].mean(),
                    'SecondGrade': night_df['SecondGrade'].mean(),
                    'Target': night_df['Target'].sum(),
                    'OverControl': night_df['OverControl'].mean(),
                    'Pinhole_DPM': night_df['Pinhole_DPM'].sum(),
                }
                subtotal_df = pd.DataFrame([subtotal])
                subtotal_df['avg_speed'] = subtotal_df['avg_speed'].round(0)

                night_df[['Name', 'Shift']] = ''
                night_df['avg_speed'] = night_df['avg_speed'].round(0)

                rows.append(night_df)  # Night row data
                rows.append(subtotal_df)  # Night Shift total summary

                if not np.isnan(night_df['ProductionTime'].mean()):
                    night_production_time = night_df['ProductionTime'].mean()

            # Machine total summary
            activation_rate, df_activation_row = self.calculate_activation(mach_name)

            # Second Grade
            sum_qty = mach_group['OnlinePacking'].sum() + mach_group['WIPPacking'].sum()
            scrap_qty = mach_group['Scrap'].sum()
            second_qty = mach_group['SecondGrade'].sum()
            target = mach_group['Target'].sum()
            online_qty = mach_group['OnlinePacking'].sum()
            isolation_qty = mach_group['Isolation_Qty'].sum()

            activation_rows.append(df_activation_row)

            tmp_qty = sum_qty + scrap_qty + second_qty
            tmp_scrap = scrap_qty / tmp_qty if sum_qty > 0 else 0
            tmp_second = second_qty / tmp_qty if sum_qty > 0 else 0
            capacity_rate = tmp_qty / target if target > 0 else 0
            yield_rate = (sum_qty - isolation_qty) / tmp_qty if tmp_qty > 0 else 0
            isolation_rate = isolation_qty / tmp_qty if tmp_qty > 0 else 0
            oee_rate = activation_rate * capacity_rate * yield_rate

            cosmetic_dpm = self.calculate_cosmetic_dpm(mach_group, cosmetic_df)
            pinhole_dpm = self.calculate_pinhole_dpm(mach_group, cosmetic_df)
            pinhole_rate = round(pinhole_dpm / 1_000_000, 4)

            over_control_qty = mach_group['OverControlQty'].sum()
            over_control_rate = round(over_control_qty/tmp_qty, 2) if tmp_qty > 0 else 0

            subtotal = {
                'Name': join_values(mach_group['Name']),
                'ProductItem': join_values(mach_group['ProductItem']),
                'AQL': join_values(mach_group['AQL']),
                'Shift': '',
                'Line': '',
                'max_speed': mach_group['max_speed'].max(),
                'min_speed': mach_group['min_speed'].min(),
                'avg_speed': mach_group['avg_speed'].mean(),
                'LineSpeedLower': join_values(mach_group['LineSpeedLower']),
                'LineSpeedUpper': join_values(mach_group['LineSpeedUpper']),
                'ProductionTime': day_production_time + night_production_time,
                'sum_qty': sum_qty,
                'OnlinePacking': mach_group['OnlinePacking'].sum(),
                'WIPPacking': mach_group['WIPPacking'].sum(),
                'Isolation_Qty': mach_group['Isolation_Qty'].sum(),
                'DMF_Rate': mach_group['DMF_Rate'].mean(),  # 離型不良率
                'Lost_Mold_Rate': mach_group['Lost_Mold_Rate'].mean(),  # 缺模率
                'PinholeRate': pinhole_rate,
                'Scrap': tmp_scrap,
                'SecondGrade': tmp_second,
                'Target': mach_group['Target'].sum(),
                'OverControl': over_control_rate,
                'Activation': activation_rate,
                'Capacity': capacity_rate,
                'Yield': yield_rate,
                'OEE': oee_rate,
                'Isolation': isolation_rate,
                'OpticalNGRate': mach_group['OpticalNGRate'].mean(),
                'Cosmetic_DPM': cosmetic_dpm,
                'Pinhole_DPM': pinhole_dpm,
            }
            subtotal_df = pd.DataFrame([subtotal])
            subtotal_df['avg_speed'] = subtotal_df['avg_speed'].round(0)
            rows.append(subtotal_df)  # Machine total summary
            chart_rows.append(subtotal_df)

        # Combine the grouped data into a DataFrame
        with_subtotals_df = pd.concat(rows, ignore_index=True)

        # ProductionTime加上小時的文字
        with_subtotals_df['ProductionTime'] = with_subtotals_df['ProductionTime'].astype(str) + 'H'

        # Group the total quantity of each machine into a DataFrame
        chart_df = pd.concat(chart_rows, ignore_index=True)

        activation_df = pd.concat(activation_rows, ignore_index=True)

        end_time = time.time()
        elapsed_time = end_time - start_time
        self.logger.info(f"Time taken: {elapsed_time:.2f} seconds.")
        return with_subtotals_df, chart_df, activation_df

    def validate_data(self, fixed_main_df, subtotals_df):
        start_time = time.time()
        # 檢查欄位 LineSpeedUpper 是否有空值
        df_filtered = fixed_main_df[fixed_main_df['Line'].notnull()]
        isNoStandard = df_filtered['LineSpeedUpper'].isnull().any()
        if isNoStandard:
            self.error_list.append(f"有品項尚未維護標準值，無法計算目標產量")

        # 生產時間不可能超過24小時，防呆檢查
        numeric_production_time = subtotals_df['ProductionTime'].str.rstrip('H').astype(float)
        machines_exceeding_24 = subtotals_df.loc[numeric_production_time > 24, 'Name'].unique()
        if machines_exceeding_24.size > 0:
            for machine in machines_exceeding_24:
                self.error_list.append(f"{machine}發生總時數超過24，可能IPQC有用錯RunCard的情況")

        # # 廢品資料尚未完成輸入
        # machine_qty_sum = fixed_main_df.groupby('Name')['Qty'].sum().reset_index()
        # machine_qty_sum = machine_qty_sum[machine_qty_sum['Qty'] > 0]
        # scrap_df = fixed_main_df[(fixed_main_df['Period'] == 5) & (fixed_main_df['Scrap'] > 0)]['Name'].unique()
        # machine_qty_sum['Scrap_Zero_in_P5'] = ~machine_qty_sum['Name'].isin(scrap_df)
        #
        # for machine, flag in zip(machine_qty_sum['Name'], machine_qty_sum['Scrap_Zero_in_P5']):
        #     if flag:
        #         self.msg_list.append(f"{machine} 尚未完成廢品資料輸入")


        # # 判斷是否有用其他方式收貨，要去詢問產線異常原因
        # for _, row in fixed_main_df.iterrows():
        #     if not pd.isna(row['sum_qty']) and not pd.isna(row['Ticket_Qty']):
        #         if int(row['sum_qty']) < 100 and int(row['Ticket_Qty']) > 1000:
        #             abnormal_machine = row['Name']
        #             self.error_list.append(f"{abnormal_machine} 點數機資料與SAP入庫資料差異過大，可能發生用舊點數機的情況")

        # 因同時生產兩種尺寸的工單，使用舊點數機人工作業分類，故無法取得正確資料進行計算
        printed_machines = set()
        check_df = fixed_main_df.groupby(['Name', 'Line', 'Date', 'Shift', 'Period'])[
            'ProductItem'].nunique().reset_index()
        conflict_rows = check_df[check_df['ProductItem'] > 1]
        for _, row in conflict_rows.iterrows():
            key = (row['Name'], row['Line'])
            if key not in printed_machines:
                self.msg_list.append(f"{row['Name']} {row['Line']} 邊因同時生產兩種尺寸的工單，使用舊點數機人工作業分類，故無法取得正確資料進行計算。")
                printed_machines.add(key)

        end_time = time.time()
        elapsed_time = end_time - start_time
        self.logger.info(f"Time taken: {elapsed_time:.2f} seconds.")

    # @Summary
    def generate_summary_excel(self, writer, df):
        # region 1. Column Define
        font = Font(name=self.report_font, size=10, bold=False)
        sheet = DataControl()
        sheet.add(ColumnControl('Name', 'center', '@', '機台號', font, hidden=False, width=19))
        sheet.add(ColumnControl('ProductItem', 'left', '@', '品項', font, hidden=False, width=30))
        sheet.add(ColumnControl('AQL', 'center', '@', 'AQL', font, hidden=False, width=9))
        sheet.add(ColumnControl('Shift', 'center', '@', '班別', font, hidden=False, width=9))
        sheet.add(ColumnControl('Line', 'center', '@', '線別', font, hidden=False, width=17))
        sheet.add(ColumnControl('avg_speed', 'right', '0', '平均車速', font, hidden=False, width=10, data_type=int, group="AVG_SPEED",
                                comment="車速標準下限~車速標準上限+2%", comment_width=300))
        sheet.add(ColumnControl('LineSpeedLower', 'right', '0', '標準車速下限', font, hidden=True, width=10, data_type=int,
                                comment="車速標準下限", comment_width=200))
        sheet.add(ColumnControl('LineSpeedUpper', 'right', '0', '標準車速上限', font, hidden=True, width=10, data_type=int,
                                comment="車速標準上限", comment_width=200))
        sheet.add(ColumnControl('ProductionTime', 'center', '@', '生產時間', font, hidden=False, width=10.5))
        sheet.add(ColumnControl('sum_qty', 'right', '#,##0', '生產總量', font, hidden=False, width=14,
                                comment="點數機數量", comment_width=200))
        sheet.add(ColumnControl('OnlinePacking', 'right', '#,##0', '包裝確認量', font, hidden=False, width=14))
        sheet.add(ColumnControl('WIPPacking', 'right', '#,##0', '半成品入庫量', font, hidden=False, width=14))
        sheet.add(ColumnControl('Target', 'right', '#,##0', '目標產能', font, hidden=False, width=14,
                                comment="生產時間(IPQC) * (標準車速上限/節距調整值)", comment_width=600))
        sheet.add(ColumnControl('Activation', 'right', '0.00%', f'稼動率≥ {self.activation_target*100:g}%', font, hidden=True, width=12, limit=[None, self.activation_target],
                                comment="點數機(A1B1)生產時間 / 工單預計生產時間"))
        sheet.add(ColumnControl('Capacity', 'right', '0.00%', f'產能效率≥ {self.capacity_target*100:g}%', font, hidden=False, width=12, limit=[None, self.capacity_target],
                                comment="(包裝確認量+半成品數量+二級品數量+廢品數量)/目標產能"))
        sheet.add(ColumnControl('Yield', 'right', '0.00%', f'良率≥ {self.yield_target*100:g}%', font, hidden=False, width=10, limit=[None, self.yield_target],
                                comment="(包裝確認量+半成品數量-隔離品數量) / (包裝確認量+半成品數量+二級品數量+廢品數量)"))
        sheet.add(ColumnControl('OEE', 'right', '0.00%', f'OEE≥ {self.oee_target*100:g}%', font, hidden=True, width=10, limit=[None, self.oee_target],
                                comment="稼動率 x 產能效率 x 良率"))
        sheet.add(ColumnControl('Isolation_Qty', 'right', '#,##0', '隔離品數量', font, hidden=False, width=13,
                                comment="MES輸入的隔離品數量", comment_width=300))
        sheet.add(ColumnControl('Isolation', 'right', '0.00%', f'隔離品率≤ {round(self.isolation_target*100,2):g}%', font, hidden=False, width=13, limit=[self.isolation_target, None],
                                comment="隔離品數量/(包裝確認量+半成品數量+二級品數量+廢品數量)"))

        sheet.add(ColumnControl('PinholeRate', 'right', '0.00%', f'美醫針孔不良率≤ {round(self.pinhole_target*100,2):g}%', font, hidden=False, width=17, limit=[self.pinhole_target, None]))
        sheet.add(ColumnControl('Scrap', 'right', '0.00%', f'廢品≤ {round(self.scrap_target*100,2):g}%', font, hidden=False, width=13, limit=[self.scrap_target, None],
                                comment="廢品數量/(包裝確認量+半成品數量+二級品數量+廢品數量)", comment_width=200))
        sheet.add(ColumnControl('SecondGrade', 'right', '0.00%', f'二級品≤ {round(self.faulty_target*100,2):g}%', font, hidden=False, width=13, limit=[self.faulty_target, None],
                                comment="二級品數量/(包裝確認量+半成品數量+二級品數量+廢品數量)", comment_width=200))

        sheet.add(ColumnControl('OverControl', 'right', '0.00%', f'超內控≤ {round(self.weight_target*100,2):g}%', font, hidden=False, width=13, limit=[self.weight_target, None]))
        sheet.add(ColumnControl('Lost_Mold_Rate', 'center', '0.00%', f'缺模率≤ {round(self.former_miss_target*100,2):g}%', font,
                          hidden=True, width=13, limit=[self.former_miss_target, None]))
        sheet.add(ColumnControl('OpticalNGRate', 'center', '0.00%', '光檢不良率', font, hidden=False, width=15))
        sheet.add(ColumnControl('DMF_Rate', 'center', '0.00%', '離型不良率', font, hidden=True, width=13))
        sheet.add(ColumnControl('Cosmetic_DPM', 'right', '#,##0', '外觀DPM', font, hidden=False, width=10))
        sheet.add(ColumnControl('Pinhole_DPM', 'right', '#,##0', '針孔DPM', font, hidden=False, width=10))

        column_index = sheet.column_index
        column_letter = sheet.column_letter
        header_columns = sheet.header_columns
        selected_columns = [col for col in sheet.column_names if col in df.columns]
        # endregion

        # region 2. DataFrame convert to Excel
        df = df[selected_columns].copy()

        # Change column names
        df.rename(columns=header_columns, inplace=True)

        namesheet = "Summary"
        # Write data to the Excel sheet with the machine name as the sheet name
        df.to_excel(writer, sheet_name=namesheet, index=False)

        # Read the written Excel file
        workbook = writer.book
        worksheet = writer.sheets[namesheet]

        sheet.apply_formatting(worksheet)

        # Freeze the first row
        worksheet.freeze_panes = worksheet['F2']
        # endregion

        # region 3. Customize
        bold_font = Font(name=self.report_font, size=10, bold=False)
        thick_border = Border(top=Side(style='thick'), bottom=Side(style='thick'))
        # Search all lines, bold font and bold line above
        index_start = 2
        index_end = 1
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            if row[column_index['Line']].value != '':  # Line
                for cell in row[column_index['Line']:]:
                    cell.fill = PatternFill(start_color="FDE9D9", end_color="FDE9D9", fill_type="solid")

            if row[column_index['ProductItem']].value != '' and row[column_index['Shift']].value == '':
                index_end += 1

            if row[column_index['Shift']].value != '':  # Shift
                worksheet.row_dimensions.group(index_start, index_end, hidden=True, outline_level=2)

                index_start = index_end + 1
                index_end = index_end + 1
                worksheet.row_dimensions.group(index_start, index_end, hidden=True, outline_level=1)
                index_start = index_end + 1

                for cell in row[column_index['Shift']:]:
                    cell.font = bold_font
                    cell.border = Border(top=Side(style='thin'))

            elif row[column_index['Name']].value != '':  # Machine
                # Hide detailed data
                worksheet.row_dimensions.group(index_start, index_end, hidden=False, outline_level=0)
                index_start = index_end + 1

                for cell in row:
                    cell.font = bold_font
                    cell.border = thick_border

                # Add a note (comment) to the 'Optical' column
                if str(row[column_index['ProductItem']].value).startswith('V S'):
                    note_text = "Yellow Gloves."
                    author = 'System'
                    row[column_index['OpticalNGRate']].comment = Comment(note_text, author)

            # 設置欄的 outlineLevel 讓其可以折疊/展開
            worksheet.column_dimensions[column_letter['Shift']].outlineLevel = 1
            worksheet.column_dimensions[column_letter['Line']].outlineLevel = 1

            # 總共折疊的區域
            worksheet.column_dimensions.group(column_letter['Shift'], column_letter['Line'], hidden=True)
        # endregion

        return workbook

    def generate_machine_excel(self, writer, df, machine_name):
        # region 1. Column Define
        font = Font(name=self.report_font, size=10, bold=False)
        sheet = DataControl()
        sheet.add(ColumnControl('Date', 'center', '@', '作業日期', font, hidden=False, width=14))
        sheet.add(ColumnControl('Name', 'center', '@', '機台號', font, hidden=False, width=20))
        sheet.add(ColumnControl('Line', 'center', '@', '線別', font, hidden=False, width=9))
        sheet.add(ColumnControl('Shift', 'center', '@', '班別', font, hidden=False, width=9))
        sheet.add(ColumnControl('WorkOrderId', 'center', '@', '工單', font, hidden=False, width=17,
                                comment="IPQC工單", comment_width=200))
        sheet.add(ColumnControl('PartNo', 'center', '@', '料號', font, hidden=False, width=17))
        sheet.add(ColumnControl('ProductItem', 'center', '@', '品項', font, hidden=False, width=25))
        sheet.add(ColumnControl('AQL', 'center', '@', '工單AQL', font, hidden=False, width=9))
        sheet.add(ColumnControl('InspectedAQL', 'center', '@', '量測AQL', font, hidden=False, width=9))
        sheet.add(ColumnControl('SalePlaceCode', 'center', '@', '銷售地點代碼', font, hidden=False, width=9,
                                comment="有派工的工單", comment_width=200))
        sheet.add(ColumnControl('ProductionTime', 'center', '@', '生產時間', font, hidden=False, width=8))
        sheet.add(ColumnControl('Period', 'center', '@', 'Period', font, hidden=False, width=8))
        sheet.add(ColumnControl('max_speed', 'right', '0', '車速(最高)', font, hidden=False, width=9.5))
        sheet.add(ColumnControl('min_speed', 'right', '0', '車速(最低)', font, hidden=False, width=9.5))
        sheet.add(ColumnControl('avg_speed', 'right', '0', '車速(平均)', font, hidden=False, width=9.5))
        sheet.add(ColumnControl('LineSpeedLower', 'right', '0', '標準下限', font, hidden=False, width=9.5))
        sheet.add(ColumnControl('LineSpeedUpper', 'right', '0', '標準上限', font, hidden=False, width=9.5))
        sheet.add(ColumnControl('sum_qty', 'right', '#,##0', '產量(加總)', font, hidden=False, width=11))
        sheet.add(ColumnControl('OnlinePacking', 'right', '#,##0', '包裝確認量', font, hidden=False, width=11))
        sheet.add(ColumnControl('WIPPacking', 'right', '#,##0', '半成品入庫量', font, hidden=False, width=11))
        sheet.add(ColumnControl('Separate', 'center', '@', '針孔', font, hidden=False, width=9))
        sheet.add(ColumnControl('Target', 'center', '#,##0', '目標產能', font, hidden=False, width=11,
                                comment="有IPQC的機台運作分鐘數 * (標準車速上限/節距調整值)"))
        sheet.add(ColumnControl('Scrap', 'right', '#,##0', '廢品', font, hidden=False, width=11))
        sheet.add(ColumnControl('SecondGrade', 'right', '#,##0', '二級品', font, hidden=False, width=11))
        sheet.add(ColumnControl('Isolation_Qty', 'right', '#,##0', '隔離品數量', font, hidden=False, width=11))
        sheet.add(ColumnControl('DMF_Rate', 'right', '0.00%', '離型不良率', font, hidden=True, width=12))
        sheet.add(ColumnControl('Lost_Mold_Rate', 'right', '0.00%', '缺模率', font, hidden=True, width=12))
        sheet.add(ColumnControl('OverControl', 'center', '@', '超內控', font, hidden=False, width=9))
        sheet.add(ColumnControl('WeightValue', 'center', '0.00', 'IPQC克重', font, hidden=False, width=11, data_type=float))
        sheet.add(ColumnControl('OpticalNGRate', 'center', '0.00%', '光檢不良率', font, hidden=False, width=10))
        sheet.add(ColumnControl('Tensile_Value', 'center', '0.00', '抗拉強度值', font, hidden=True, width=11, level=1))
        sheet.add(ColumnControl('Tensile_Limit', 'center', '@', '抗拉強度上下限', font, hidden=True, width=11, level=1))
        sheet.add(ColumnControl('Tensile_Status', 'center', '@', '抗拉強度結果', font, hidden=True, width=11, level=1))
        sheet.add(ColumnControl('Elongation_Value', 'center', '0.00', '伸長率值', font, hidden=True, width=11, level=1))
        sheet.add(ColumnControl('Elongation_Limit', 'center', '@', '伸長率上下限', font, hidden=True, width=11, level=1))
        sheet.add(ColumnControl('Elongation_Status', 'center', '@', '伸長率結果', font, hidden=True, width=11, level=1))
        sheet.add(ColumnControl('Roll_Value', 'center', '0.00', '卷唇厚度值', font, hidden=True, width=11, level=1))
        sheet.add(ColumnControl('Roll_Limit', 'center', '@', '卷唇厚度上下限', font, hidden=True, width=11, level=1))
        sheet.add(ColumnControl('Roll_Status', 'center', '@', '卷唇厚度結果', font, hidden=True, width=11, level=1))
        sheet.add(ColumnControl('Cuff_Value', 'center', '0.00', '袖厚度值', font, hidden=True, width=11, level=1))
        sheet.add(ColumnControl('Cuff_Limit', 'center', '@', '袖厚度上下限', font, hidden=True, width=11, level=1))
        sheet.add(ColumnControl('Cuff_Status', 'center', '@', '袖厚度結果', font, hidden=True, width=11, level=1))
        sheet.add(ColumnControl('Palm_Value', 'center', '0.00', '掌厚度值', font, hidden=True, width=11, level=1))
        sheet.add(ColumnControl('Palm_Limit', 'center', '@', '掌厚度上下限', font, hidden=True, width=11, level=1))
        sheet.add(ColumnControl('Palm_Status', 'center', '@', '掌厚度結果', font, hidden=True, width=11, level=1))
        sheet.add(ColumnControl('Finger_Value', 'center', '0.00', '指厚度值', font, hidden=True, width=11, level=1))
        sheet.add(ColumnControl('Finger_Limit', 'center', '@', '指厚度上下限', font, hidden=True, width=11, level=1))
        sheet.add(ColumnControl('Finger_Status', 'center', '@', '指厚度結果', font, hidden=True, width=11, level=1))
        sheet.add(ColumnControl('FingerTip_Value', 'center', '0.00', '指尖厚度值', font, hidden=True, width=11, level=1))
        sheet.add(ColumnControl('FingerTip_Limit', 'center', '@', '指尖厚度上下限', font, hidden=True, width=11, level=1))
        sheet.add(ColumnControl('FingerTip_Status', 'center', '@', '指尖厚度結果', font, hidden=True, width=11, level=1))
        sheet.add(ColumnControl('Length_Value', 'center', '0.00', '長度值', font, hidden=True, width=11, level=1))
        sheet.add(ColumnControl('Length_Limit', 'center', '@', '長度上下限', font, hidden=True, width=11, level=1))
        sheet.add(ColumnControl('Length_Status', 'center', '@', '長度結果', font, hidden=True, width=11, level=1))
        sheet.add(ColumnControl('Weight_Value', 'center', '0.00', '重量值', font, hidden=True, width=11, level=1))
        sheet.add(ColumnControl('Weight_Limit', 'center', '@', '重量上下限', font, hidden=True, width=11, level=1))
        sheet.add(ColumnControl('Weight_Light', 'center', '@', '超輕檢驗', font, hidden=True, width=11, level=1))
        sheet.add(ColumnControl('Weight_Heavy', 'center', '@', '超重檢驗', font, hidden=True, width=11, level=1))
        sheet.add(ColumnControl('Width_Value', 'center', '0.00', '寬度值', font, hidden=True, width=11, level=1))
        sheet.add(ColumnControl('Width_Limit', 'center', '@', '寬度上下限', font, hidden=True, width=11, level=1))
        sheet.add(ColumnControl('Width_Status', 'center', '@', '寬度結果', font, hidden=True, width=11, level=1))
        sheet.add(ColumnControl('Pinhole_Value', 'center', '0.00', '針孔值', font, hidden=True, width=11, level=1))
        sheet.add(ColumnControl('Pinhole_Limit', 'center', '@', '針孔上下限', font, hidden=True, width=11, level=1))
        sheet.add(ColumnControl('Pinhole_Status', 'center', '@', '針孔結果', font, hidden=True, width=11, level=1))
        sheet.add(ColumnControl('IPQC', 'center', '@', 'IPQC', font, hidden=False, width=11))
        sheet.add(ColumnControl('WoStartDate', 'center', 'yyyy/mm/dd hh:mm:ss', '工單開始時間', font, hidden=True, width=11))
        sheet.add(ColumnControl('WoEndDate', 'center', 'yyyy/mm/dd hh:mm:ss', '工單結束時間', font, hidden=True, width=11))

        header_columns = sheet.header_columns
        column_letter = sheet.column_letter
        selected_columns = [col for col in sheet.column_names if col in df.columns]
        # endregion

        # region 2. DataFrame convert to Excel
        df = df[selected_columns].copy()

        # 轉出Excel前進行資料處理
        df['ProductionTime'] = (df['ProductionTime'] // 60).astype(str) + 'H'
        df['Period'] = df['Period'].apply(lambda x: f"{int(x):02}:00")

        # Change column names
        df.rename(columns=header_columns, inplace=True)

        namesheet = str(machine_name).split('_')[-1]
        # Write data to the Excel sheet with the machine name as the sheet name
        df.to_excel(writer, sheet_name=namesheet, index=False)

        # Read the written Excel file
        workbook = writer.book
        worksheet = writer.sheets[namesheet]

        # Freeze the first row
        worksheet.freeze_panes = worksheet['H2']

        sheet.apply_formatting(worksheet)
        # endregion

        # region 3. Customize
        # # 設置欄讓其可以折疊/展開
        worksheet.column_dimensions.group(column_letter['Tensile_Value'], column_letter['Pinhole_Status'], hidden=True)

        for row in range(2, worksheet.max_row + 1):  # 從第2行開始，因為第1行是標題
            weight_value_cell = worksheet[column_letter['WeightValue'] + str(row)]
            weight_limit_value = worksheet[column_letter['Weight_Limit'] + str(row)].value
            comment = Comment(text="IPQC範圍(" + weight_limit_value + ")",
                              author="System")  # 創建註解
            weight_value_cell.comment = comment

        # endregion

        return workbook

    def generate_activation_excel(self, writer, df):
        # region 1. Column Define
        font = Font(name=self.report_font, size=10, bold=False)
        sheet = DataControl()
        sheet.add(ColumnControl('CreationTime', 'center', '@', '點數機時間', font, hidden=False, width=20))
        sheet.add(ColumnControl('MES_MACHINE', 'left', '@', '機台號', font, hidden=False, width=20))
        sheet.add(ColumnControl('A1_Qty', 'center', '#,##0', 'A1數量', font, hidden=False, width=15))
        sheet.add(ColumnControl('A1_Speed', 'center', '0', 'A1車速', font, hidden=False, width=15))
        sheet.add(ColumnControl('A2_Qty', 'center', '#,##0', 'A2數量', font, hidden=False, width=15))
        sheet.add(ColumnControl('A2_Spped', 'center', '0', 'A2車速', font, hidden=False, width=15))
        sheet.add(ColumnControl('B1_Qty', 'right', '#,##0', 'B1數量', font, hidden=False, width=15))
        sheet.add(ColumnControl('B1_Speed', 'center', '0', 'B1車速', font, hidden=False, width=15))
        sheet.add(ColumnControl('B2_Qty', 'right', '#,##0', 'B2數量', font, hidden=False, width=15))
        sheet.add(ColumnControl('B2_Speed', 'center', '0', 'B2車速', font, hidden=False, width=15))

        column_letter = sheet.column_letter
        selected_columns = [col for col in sheet.column_names if col in df.columns]
        # endregion

        # region 2. DataFrame convert to Excel
        df = df[selected_columns].copy()

        namesheet = "點數機"
        # Write data to the Excel sheet with the machine name as the sheet name
        df.to_excel(writer, sheet_name=namesheet, index=False)

        # Read the written Excel file
        workbook = writer.book
        worksheet = writer.sheets[namesheet]

        sheet.apply_formatting(worksheet)

        # Freeze the first row
        worksheet.freeze_panes = worksheet['A2']
        # endregion

        # region 3. Customize
        try:
            # 設置條件格式為黃顏色填充
            yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

            # 使用 CellIsRule 設置條件格式，只需設定一次
            # 此範例將適用於 A1_Qty 和 B1_Qty 列
            for col_letter in [column_letter['A1_Qty'], column_letter['B1_Qty']]:
                # 假設您知道資料的範圍為 B2:B100，您可以根據實際情況修改範圍
                worksheet.conditional_formatting.add(f'{col_letter}2:{col_letter}65535',
                                                     CellIsRule(operator='lessThanOrEqual',
                                                                formula=['10'],
                                                                fill=yellow_fill))
        except Exception as e:
            print(e)
        # endregion

        return workbook

    def generate_cosmetic_excel(self, writer, cosmetic_df):
        # region 1. Column Define
        font = Font(name=self.report_font, size=10, bold=False)
        sheet = DataControl()
        sheet.add(ColumnControl('runcard', 'center', '@', 'Runcard', font, hidden=False, width=14))
        sheet.add(ColumnControl('belong_to', 'center', '@', '作業日期', font, hidden=False, width=14))
        sheet.add(ColumnControl('Machine', 'center', '@', '機台', font, hidden=False, width=18))
        sheet.add(ColumnControl('Line', 'right', '@', '線別', font, hidden=False, width=8))
        sheet.add(ColumnControl('Shift', 'right', '@', '班別', font, hidden=False, width=8))
        sheet.add(ColumnControl('WorkOrder', 'center', '@', '工單', font, hidden=False, width=16))
        sheet.add(ColumnControl('PartNo', 'center', '@', '料號', font, hidden=False, width=14))
        sheet.add(ColumnControl('ProductItem', 'center', '@', '品項', font, hidden=False, width=28))
        sheet.add(ColumnControl('SalePlaceCode', 'center', '@', '銷售地點', font, hidden=False, width=8))
        sheet.add(ColumnControl('Period', 'center', '@', '', font, hidden=False, width=10))
        sheet.add(ColumnControl('6100LL1', 'center', '0', '美線過重', font, hidden=False, width=14, group='WEIGHT',
                                apply_format=False))
        sheet.add(ColumnControl('6100LL2', 'center', '0', '美線過輕', font, hidden=False, width=14, group='WEIGHT',
                                apply_format=False))
        sheet.add(ColumnControl('6200LL1', 'center', '0', '歐線過重', font, hidden=False, width=14, group='WEIGHT',
                                apply_format=False))
        sheet.add(ColumnControl('6200LL2', 'center', '0', '歐線過輕', font, hidden=False, width=14, group='WEIGHT',
                                apply_format=False))
        sheet.add(ColumnControl('6300LL1', 'center', '0', '日線過重', font, hidden=False, width=14, group='WEIGHT',
                                apply_format=False))
        sheet.add(ColumnControl('6300LL2', 'center', '0', '日線過輕', font, hidden=False, width=14, group='WEIGHT',
                                apply_format=False))
        sheet.add(ColumnControl('7000LL1', 'center', '0', 'OBM過重', font, hidden=False, width=14, group='WEIGHT',
                                apply_format=False))
        sheet.add(ColumnControl('7000LL2', 'center', '0', 'OBM過輕', font, hidden=False, width=14, group='WEIGHT',
                                apply_format=False))

        sql = f"""
                          SELECT distinct d.defect_level, d.defect_code, d.defect_code, d.desc1, d.desc2
                            FROM [MES_OLAP].[dbo].[counting_hourly_info_raw] r
                            LEFT JOIN [MES_OLAP].[dbo].[mes_ipqc_cosmetic_data] cos on r.Runcard = cos.runcard
                            LEFT JOIN [MES_OLAP].[dbo].[mes_defect_define] d on d.defect_code = cos.defect_code
                            where r.belong_to = '{self.report_date1}' and d.defect_type <> ''
                            order by d.defect_level, d.defect_code
                        """
        cosmetic_rows = self.mes_olap_db.select_sql_dict(sql)

        critical_defect = []
        major_defect = []
        minor_defect = []
        defect_list = {'CRITICAL': critical_defect, 'MAJOR': major_defect, 'MINOR': minor_defect}

        for row in cosmetic_rows:
            defect_list[row['defect_level']].append(row['defect_code'])
            desc = row['desc1'] if row['desc1'] != '' else row['desc2']
            sheet.add(ColumnControl(row['defect_code'], 'center', '0', desc, font, hidden=False, width=14,
                                    group=row['defect_level'], apply_format=False))

        sheet.add(ColumnControl('cosmetic_qty', 'right', '#,##0', '缺陷手套數量', font, hidden=False, width=14, group='COSMETIC',
                                apply_format=False))
        sheet.add(ColumnControl('cosmetic_inspect_qty', 'right', '#,##0', '外觀檢查數量', font, hidden=False, width=14, group='COSMETIC',
                                apply_format=False))

        sql = f"""
                        SELECT distinct d.defect_code, d.desc1, d.desc2
                          FROM [MES_OLAP].[dbo].[counting_hourly_info_raw] r
                          JOIN [MES_OLAP].[dbo].[mes_ipqc_pinhole_data] cos on r.Runcard = cos.runcard
                          JOIN [MES_OLAP].[dbo].[mes_defect_define] d on d.defect_code = cos.defect_code
                          where r.belong_to = '{self.report_date1}'
                          order by d.defect_code
                                """
        pinhole_rows = self.mes_olap_db.select_sql_dict(sql)

        pinhole_defect = []
        defect_list['PINHOLE'] = pinhole_defect

        for row in pinhole_rows:
            defect_list['PINHOLE'].append(row['defect_code'])
            desc = row['desc1'] if row['desc1'] != '' else row['desc2']
            sheet.add(
                ColumnControl(row['defect_code'], 'center', '0', desc, font, hidden=False, width=14, group='COSMETIC',
                              apply_format=False))

        sheet.add(ColumnControl('Pinhole', 'right', '0', '針孔數量', font, hidden=False, width=14, group='COSMETIC',
                                apply_format=False))
        sheet.add(
            ColumnControl('Pinhole_Sample', 'right', '0', '針孔檢查數量', font, hidden=False, width=14, group='COSMETIC',
                          apply_format=False))

        header_columns = sheet.header_columns
        column_letter = sheet.column_letter
        selected_columns = [col for col in sheet.column_names if col in cosmetic_df.columns]
        # endregion

        # region 2. DataFrame convert to Excel
        df = cosmetic_df[selected_columns].copy()

        df = df.rename(columns=header_columns)

        sheet_name = "外觀"
        df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=1)

        worksheet = writer.sheets[sheet_name]
        sheet.apply_formatting(worksheet)
        # endregion

        # region 3. Customize
        header_border = Border(
            top=Side(style='medium'),
            bottom=Side(style='medium'),
            left=Side(style='medium'),
            right=Side(style='medium')
        )
        fill_style = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

        # Weight Header
        worksheet.merge_cells(f"{column_letter['6100LL1']}1:{column_letter['7000LL2']}1")
        start_row, start_col = worksheet[f"{column_letter['6100LL1']}1"].row, worksheet[
            f"{column_letter['6100LL1']}1"].column
        end_row, end_col = worksheet[f"{column_letter['7000LL2']}1"].row, worksheet[
            f"{column_letter['7000LL2']}1"].column
        cell = worksheet[f"{column_letter['6100LL1']}1"]
        cell.value = "重量檢驗"
        cell.alignment = Alignment(horizontal="center", vertical="center")
        # cell.border = header_border
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.border = header_border
                cell.fill = fill_style

        # Cosmetic Header
        critical_start = defect_list['CRITICAL'][0]
        critical_end = defect_list['CRITICAL'][-1]
        worksheet.merge_cells(f"{column_letter[critical_start]}1:{column_letter[critical_end]}1")
        start_row, start_col = worksheet[f"{column_letter[critical_start]}1"].row, worksheet[
            f"{column_letter[critical_start]}1"].column
        end_row, end_col = worksheet[f"{column_letter[critical_end]}1"].row, worksheet[
            f"{column_letter[critical_end]}1"].column
        cell = worksheet[f"{column_letter[critical_start]}1"]
        cell.value = "外觀CRITICAL"
        cell.alignment = Alignment(horizontal="center", vertical="center")
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.border = header_border
                cell.fill = fill_style

        major_start = defect_list['MAJOR'][0]
        major_end = defect_list['MAJOR'][-1]
        worksheet.merge_cells(f"{column_letter[major_start]}1:{column_letter[major_end]}1")
        start_row, start_col = worksheet[f"{column_letter[major_start]}1"].row, worksheet[
            f"{column_letter[major_start]}1"].column
        end_row, end_col = worksheet[f"{column_letter[major_end]}1"].row, worksheet[
            f"{column_letter[major_end]}1"].column
        cell = worksheet[f"{column_letter[major_start]}1"]
        cell.value = "外觀MAJOR"
        cell.alignment = Alignment(horizontal="center", vertical="center")
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.border = header_border
                cell.fill = fill_style

        minor_start = defect_list['MINOR'][0]
        minor_end = defect_list['MINOR'][-1]
        worksheet.merge_cells(f"{column_letter[minor_start]}1:{column_letter[minor_end]}1")
        start_row, start_col = worksheet[f"{column_letter[minor_start]}1"].row, worksheet[
            f"{column_letter[minor_start]}1"].column
        end_row, end_col = worksheet[f"{column_letter[minor_end]}1"].row, worksheet[
            f"{column_letter[minor_end]}1"].column
        cell = worksheet[f"{column_letter[minor_start]}1"]
        cell.value = "外觀MINOR"
        cell.alignment = Alignment(horizontal="center", vertical="center")
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.border = header_border
                cell.fill = fill_style

        # Pinhole Header
        pinhole_start = defect_list['PINHOLE'][0]
        pinhole_end = defect_list['PINHOLE'][-1]
        worksheet.merge_cells(f"{column_letter[pinhole_start]}1:{column_letter[pinhole_end]}1")
        start_row, start_col = worksheet[f"{column_letter[pinhole_start]}1"].row, worksheet[
            f"{column_letter[pinhole_start]}1"].column
        end_row, end_col = worksheet[f"{column_letter[pinhole_end]}1"].row, worksheet[
            f"{column_letter[pinhole_end]}1"].column
        cell = worksheet[f"{column_letter[pinhole_start]}1"]
        cell.value = "針孔"
        cell.alignment = Alignment(horizontal="center", vertical="center")
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.border = header_border
                cell.fill = fill_style

        # 最後一行加總
        thin_top_border = Border(top=Side(style="thin"))
        last_row = worksheet.max_row + 1

        for col_idx in range(11, len(header_columns) + 1):
            col_letter = get_column_letter(col_idx)

            # 設定 SUM 公式
            sum_formula = f"=SUM({col_letter}2:{col_letter}{last_row-1})"
            worksheet[f"{col_letter}{last_row}"] = sum_formula  # ✅ 填入最後一行

            # 在最後一行上方（倒數第 2 行）加上框線
            if last_row > 1:  # 確保至少有兩行
                cell_above = worksheet[f"{col_letter}{last_row}"]
                cell_above.border = thin_top_border  # ✅ 設定上方框線
        # endregion

    def generate_excel(self, fix_main_df, subtotals_df, activation_df, cosmetic_df, excel_file):
        start_time = time.time()
        with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
            self.generate_summary_excel(writer, subtotals_df)

            machine_groups = fix_main_df.groupby('Name')
            for machine_name, machine_df in machine_groups:
                # 處理停機情況
                if not machine_df['ProductItem'].iloc[0]:
                    continue

                machine_clean_df = machine_df.sort_values(by=['Date', 'Shift', 'Period'])
                self.generate_machine_excel(writer, machine_clean_df, machine_name)

            # 稼動率Raw Data
            self.generate_activation_excel(writer, activation_df)

            # 外觀
            self.generate_cosmetic_excel(writer, cosmetic_df)

        end_time = time.time()
        elapsed_time = end_time - start_time
        self.logger.info(f"Time taken: {elapsed_time:.2f} seconds.")

    def generate_chart(self, chart_df, image_file):
        start_time = time.time()

        # Create Chart
        fig, ax1 = plt.subplots(figsize=(10, 6))

        plt.rcParams['font.sans-serif'] = ['Microsoft YaHei']
        plt.rcParams['axes.unicode_minus'] = False  # 避免負號顯示錯誤

        # Only substring Name right 3 characters
        chart_df['Name_short'] = chart_df['Name'].apply(lambda x: x[-3:])

        chart_df['SecondGrade'] = chart_df['SecondGrade'].replace('', 0).fillna(0).astype(float)
        qty_sum = chart_df['OnlinePacking']+chart_df['WIPPacking']+chart_df['SecondGrade']

        chart_df['Unfinished'] = (
                chart_df['Target']-qty_sum
        ).clip(lower=0)
        chart_df['Achievement Rate'] = qty_sum / chart_df['Target'] * 100  # 達成率（百分比）
        chart_df.loc[chart_df['Target'] == 0, 'Achievement Rate'] = None  # 當 Target 為 0，將達成率設為 None

        # Draw Bar Chart
        bar_width = 0.6
        bars = ax1.bar(chart_df['Name_short'], qty_sum, width=bar_width, color='lightcoral', label='日目標達成率')
        ax1.bar(chart_df['Name_short'], chart_df['Unfinished'], width=bar_width, bottom=qty_sum,
                color='lightgreen')

        # Set the X-axis label and the Y-axis label
        ax1.set_xlabel('機台')
        ax1.set_ylabel('日產量')
        # 設置 Y 軸的上限為 120 萬
        if "PVC" in self.plant:
            ax1.set_ylim(0, 800000)
        else:
            ax1.set_ylim(0, 1200000)

        # 自定義 Y 軸以 10 萬為單位
        def y_formatter(x, pos):
            return f'{int(x/10000)}萬'  # 將數值轉換為「萬」的單位顯示

        ax1.yaxis.set_major_formatter(FuncFormatter(y_formatter))

        achieve_rate = self.capacity_target

        # 在每個長條圖上方顯示達成率百分比
        for bar, unfinished, rate in zip(bars, chart_df['Unfinished'], chart_df['Achievement Rate']):
            if pd.notnull(rate):  # 僅顯示達成率不為 None 的數值
                height = bar.get_height() + unfinished  # 計算長條的總高度
                if rate < achieve_rate*100:
                    ax1.text(bar.get_x() + bar.get_width() / 2, height + 20000, f'{rate:.1f}%', ha='center',
                             va='bottom',
                             fontsize=10, color='red',
                             bbox=dict(boxstyle="circle", edgecolor='red', facecolor='none', linewidth=1.5))

                else:
                    ax1.text(bar.get_x() + bar.get_width() / 2, height + 20000, f'{rate:.1f}%', ha='center',
                             va='bottom',
                             fontsize=10)

        plt.title(f'{self.location} {self.plant} {self.report_date1} 日產量與日目標達成率 (達成率目標 > {achieve_rate*100:g}%)')

        # Display the legend of the bar chart and line chart together
        fig.legend(loc="upper right", bbox_to_anchor=(1, 1), bbox_transform=ax1.transAxes)

        plt.savefig(image_file)

        # Save the image to a BytesIO object
        image_stream = BytesIO()
        plt.savefig(image_stream, format='png')
        image_stream.seek(0)  # Move the pointer to the beginning of the file
        plt.close()  # Close the image to free up memory

        end_time = time.time()
        elapsed_time = end_time - start_time
        self.logger.info(f"Time taken: {elapsed_time:.2f} seconds.")

    def calculate_activation(self, mach):
        try:
            wo_time = 86400

            sql = f"""
                WITH A1 AS (
                SELECT m.MES_MACHINE,Qty2,Speed,LINE,CreationTime
                                  FROM [PMG_DEVICE].[dbo].[COUNTING_DATA] d
                                  JOIN [PMG_DEVICE].[dbo].[COUNTING_DATA_MACHINE] m on d.MachineName = m.COUNTING_MACHINE
                                  where m.MES_MACHINE = '{mach}' and m.LINE = 'A1'
                                  and CreationTime between CONVERT(DATETIME, '{report_date1} 06:00:00', 120) and CONVERT(DATETIME, '{report_date2} 05:59:59', 120) 
                ),
                B1 AS (
                SELECT m.MES_MACHINE,Qty2,Speed,LINE,CreationTime
                                  FROM [PMG_DEVICE].[dbo].[COUNTING_DATA] d
                                  JOIN [PMG_DEVICE].[dbo].[COUNTING_DATA_MACHINE] m on d.MachineName = m.COUNTING_MACHINE
                                  where m.MES_MACHINE = '{mach}' and m.LINE = 'B1'
                                  and CreationTime between CONVERT(DATETIME, '{report_date1} 06:00:00', 120) and CONVERT(DATETIME, '{report_date2} 05:59:59', 120) 
                ),
                A2 AS (
                SELECT m.MES_MACHINE,Qty2,Speed,LINE,CreationTime
                                  FROM [PMG_DEVICE].[dbo].[COUNTING_DATA] d
                                  JOIN [PMG_DEVICE].[dbo].[COUNTING_DATA_MACHINE] m on d.MachineName = m.COUNTING_MACHINE
                                  where m.MES_MACHINE = '{mach}' and m.LINE = 'A2'
                                  and CreationTime between CONVERT(DATETIME, '{report_date1} 06:00:00', 120) and CONVERT(DATETIME, '{report_date2} 05:59:59', 120) 
                ),
                B2 AS (
                SELECT m.MES_MACHINE,Qty2,Speed,LINE,CreationTime
                                  FROM [PMG_DEVICE].[dbo].[COUNTING_DATA] d
                                  JOIN [PMG_DEVICE].[dbo].[COUNTING_DATA_MACHINE] m on d.MachineName = m.COUNTING_MACHINE
                                  where m.MES_MACHINE = '{mach}' and m.LINE = 'B2'
                                  and CreationTime between CONVERT(DATETIME, '{report_date1} 06:00:00', 120) and CONVERT(DATETIME, '{report_date2} 05:59:59', 120) 
                )

                Select FORMAT(A1.CreationTime, 'yyyy-MM-dd HH:mm:ss') CreationTime, A1.MES_MACHINE, A1.Qty2 A1_Qty, A1.Speed A1_Speed, A2.Qty2 A2_Qty, A2.Speed A2_Spped, B1.Qty2 B1_Qty, B1.Speed B1_Speed, B2.Qty2 B2_Qty, B2.Speed B2_Speed from A1 
                left join A2 on A1.CreationTime = A2.CreationTime
                join B1 on A1.CreationTime = B1.CreationTime
                left join B2 on A1.CreationTime = B2.CreationTime

            """
            detail_raws = self.mes_db.select_sql_dict(sql)

            df = pd.DataFrame(detail_raws)
            filtered_df = df[(df['A1_Qty'] > 10) | (df['B1_Qty'] > 10)]
            run_time = len(filtered_df) * 300
            active_rate = run_time / wo_time
            return round(active_rate, 2), df
        except Exception as e:
            print(e)

    def calculate_cosmetic_dpm(self, mach_df, cosmetic_df):
        dpm = None
        try:
            df = pd.merge(mach_df, cosmetic_df, on=['runcard'], how='left')
            df['cosmetic_inspect_qty'] = df['cosmetic_inspect_qty'].replace(r'^\s*$', np.nan, regex=True).fillna(0)
            df['cosmetic_qty'] = df['cosmetic_qty'].replace(r'^\s*$', np.nan, regex=True).fillna(0)
            inspect_total = df['cosmetic_inspect_qty'].sum()
            sum_qty_total = df['cosmetic_qty'].sum()

            if inspect_total > 0:
                dpm = round(sum_qty_total * 1_000_000 / inspect_total, 0)
            else:
                dpm = None  # 或設為 0 或 NaN，視你的需求而定
        except Exception as e:
            print(e)

        return dpm

    def calculate_pinhole_dpm(self, mach_df, cosmetic_df):
        df = pd.merge(mach_df, cosmetic_df, on=['runcard'], how='left')
        df['Pinhole_Sample'] = pd.to_numeric(df['Pinhole_Sample'], errors='coerce')
        df['Pinhole'] = pd.to_numeric(df['Pinhole'], errors='coerce')
        inspect_total = df['Pinhole_Sample'].sum()
        sum_qty_total = df['Pinhole'].sum()

        if inspect_total > 0:
            dpm = round(sum_qty_total * 1_000_000 / inspect_total, 0)
        else:
            dpm = 0  # 或設為 0 或 NaN，視你的需求而定

        return dpm

    # 缺模率
    def get_lost_mold_rate(self):
        sql = f"""
        SELECT cd.MES_MACHINE Name, cd.LINE Line, CAST(DATEPART(hour, CreationTime) as INT) Period, round(sum(ModelLostQty) / SUM(c.ModelQty+c.ModelLostQty), 4) Lost_Mold_Rate
          FROM [PMG_DEVICE].[dbo].[COUNTING_DATA] c
          JOIN [PMG_DEVICE].[dbo].[COUNTING_DATA_MACHINE] cd on c.MachineName = cd.COUNTING_MACHINE
          where ModelLostQty > 0 and ModelLostQty < 1000
          and CreationTime between CONVERT(DATETIME, '{self.report_date1} 06:00:00', 120) and CONVERT(DATETIME, '{self.report_date2} 05:59:59', 120)
          group by cd.MES_MACHINE, cd.LINE, CAST(DATEPART(hour, CreationTime) as INT)
        """

        rows = self.mes_db.select_sql_dict(sql)

        df = pd.DataFrame(rows)

        return df

    # 離型不良率
    def get_dmf_rate(self):
        if "NBR" in self.plant:
            sql = f"""
            SELECT 
                cd.MES_MACHINE Name, cd.LINE Line, CAST(DATEPART(hour, CreationTime) as INT) Period, 
                CASE 
                    WHEN SUM(ModelQty2) = 0 THEN 0
                    ELSE ROUND(
                        (SUM(OverShortQty2) + SUM(OverLongQty2)) / SUM(ModelQty2), 
                        4
                    )
                END AS DMF_Rate
            FROM 
                [PMG_DEVICE].[dbo].[COUNTING_DATA] c
            JOIN 
                [PMG_DEVICE].[dbo].[COUNTING_DATA_MACHINE] cd 
                ON c.MachineName = cd.COUNTING_MACHINE
            WHERE 
                CreationTime BETWEEN 
                    CONVERT(DATETIME, '{self.report_date1} 06:00:00', 120) 
                    AND 
                    CONVERT(DATETIME, '{self.report_date2} 05:59:59', 120)
            GROUP BY 
                MES_MACHINE,LINE,CAST(DATEPART(hour, CreationTime) as INT)
            """

        elif "PVC" in self.plant:
            sql = f"""
            SELECT cd.MES_MACHINE Name, cd.LINE Line, CAST(DATEPART(hour, CreationTime) as INT) Period, CASE 
                    WHEN SUM(ModelQty2) = 0 THEN 0
                    ELSE ROUND(SUM(Qty2) / SUM(ModelQty2), 4)
                END AS DMF_Rate
              FROM [PMG_DEVICE].[dbo].[PVC_GRM_DATA] g
              JOIN [PMG_DEVICE].[dbo].[COUNTING_DATA_MACHINE] cd on g.MachineName = cd.COUNTING_MACHINE
              where CreationTime between CONVERT(DATETIME, '{self.report_date1} 06:00:00', 120) and CONVERT(DATETIME, '{self.report_date2} 05:59:59', 120)
              group by cd.MES_MACHINE, cd.LINE, CAST(DATEPART(hour, CreationTime) as INT) 
            
            """
        rows = self.mes_db.select_sql_dict(sql)

        df = pd.DataFrame(rows)

        return df

    def send_email(self, config, subject, file_list, image_buffers, msg_list, error_list):
        logging.info(f"Start to send Email")

        error_msg = '<br>'.join(error_list)
        if len(error_list) > 0:
            error_msg = error_msg + '<br>'
        normal_msg = '<br>'.join(msg_list)
        if len(msg_list) > 0:
            normal_msg = normal_msg + '<br>'

        content = self.get_mail_content()

        max_reSend = 5
        reSent = 0
        while reSent < max_reSend:
            try:
                super().send_email(config, subject, file_list, image_buffers, error_msg, normal_msg=normal_msg, content=content)
                print('Email sent successfully')
                logging.info(f"Email sent successfully")
                break
            except Exception as e:
                print(f'Email sending failed: {e}')
                logging.info(f"Email sending failed: {e}")
                reSent += 1
                if reSent >= max_reSend:
                    print('Failed to send email after 5 retries')
                    logging.info(f"Failed to send email after 5 retries")
                    break
                time.sleep(180)  # seconds

    def get_mail_content(self):
        result = f"""
        <hr />
        <table border="1" cellspacing="0" cellpadding="6" style="border-collapse: collapse; font-family: Arial; font-size: 14px; width:100%;background-color:#FFF">
            <thead style="background-color:#f0f0f0;">
                <tr>
                  <th style="width:200px;">大目標(G)</th>
                  <th style="width:500px;">小目標KPI(P)</th>
                  <th style="width:150px;">執行階段日期</th>
                </tr>
              </thead>
              <tbody>
                <tr>
                  <td>設備綜合效率OEE<br>KPI:≧95%</td>
                  <td>
                    產能效率(KPI:≧97%) * 產品良率(KPI:≧95%) * 設備妥善率(KPI:≧98%)<br>
                    <span style="color:red;">OEE = 90%</span>
                  </td>
                  <td>2025/04-12</td>
                </tr>
                <tr>
                  <td rowspan="3">NBR廢品率<br>KPI:≦0.4%</td>
                  <td>廢品率(KPI:≦0.8%)</td>
                  <td>2025/01-06</td>
                </tr>
                <tr>
                  <td>廢品率(KPI:≦0.6%)</td>
                  <td>2025/07-10</td>
                </tr>
                <tr>
                  <td>廢品率(KPI:≦0.4%)</td>
                  <td>2025/11-12</td>
                </tr>
                <tr>
                  <td rowspan="3">PVC廢品率<br>KPI:≦0.2%</td>
                  <td>廢品率(KPI:≦0.35%)</td>
                  <td>2025/01-06</td>
                </tr>
                <tr>
                  <td>廢品率(KPI:≦0.3%)</td>
                  <td>2025/07-10</td>
                </tr>
                <tr>
                  <td>廢品率(KPI:≦0.2%)</td>
                  <td>2025/11-12</td>
                </tr>
              </tbody>
        </table>
        """
        return result

report_date1 = datetime.today() - timedelta(days=1)
report_date1 = report_date1.strftime('%Y%m%d')

report_date2 = datetime.today()
report_date2 = report_date2.strftime('%Y%m%d')

# report_date1 = "20250418"
# report_date2 = "20250419"

report = mes_daily_report(report_date1, report_date2)
report.main()